from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.utils import range_boundaries
from openpyxl.utils import column_index_from_string
#from openpyxl.utils import coordinate_from_string
from openpyxl.utils import coordinate_to_tuple
from openpyxl.worksheet.table import Table, TableStyleInfo

net_file = r"./nettestcase.txt"
outfile = "h132_testcase.csv"
excelpath = r'C:\Users\sytung\OneDrive - Synopsys, Inc\Desktop\py\H132_UCIe_TC_Bump_coordination.xlsx'
wb = load_workbook(excelpath, data_only=True)
ws = wb['UCIe_Mapping_connection']
# with open(net_file,'r') as netfile:
#             for line in netfile :
#                 net = str(line).strip()
#                 print(line)
#                 for row in range(3,1254):
#                         if 
with open('outfile.csv','w') as out:
    file = open(net_file, 'r')
    i = 1
    while True:
        line = file.readline()
        net =str(line).strip()
        # out.writelines("v Tung 0 dc='vdd'"+"\n")
        if not line:
            break
        for row in range(3,1254):
            if (net == ws.cell(column=1, row=row).value):
                mapping = ws.cell(column=2, row=row).value
                print(mapping)
                out.writelines(f"{net},{mapping},input,output,ubump,ubump,net{i}\n")
                i+=1
    print("Completed!!")
        # print(line)
  
    # out.writelines("\n")