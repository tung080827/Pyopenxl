from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.utils import range_boundaries
from openpyxl.utils import column_index_from_string
#from openpyxl.utils import coordinate_from_string
from openpyxl.utils import coordinate_to_tuple
from openpyxl.worksheet.table import Table, TableStyleInfo
# from tkinter import *
from tkinter import messagebox
from tkinter import ttk
import getcolumn
from array import *
import tkinter as tk
from tkinter import *
from ttkthemes import ThemedTk, THEMES
from PIL import Image
from PIL import ImageTk, Image
from tkinter.font import Font as tkfont
from tkinter import filedialog
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection
from openpyxl.worksheet.worksheet import Worksheet



def setcell(col,row):
    return get_column_letter(col)+str(row)

def gettable(ws: Worksheet,tb_name: str):
    isfound = 0
    found_cnt = 0
    for merge in ws.merged_cells.ranges:
        idx = str(merge).find(':')
        cell_begin = str(merge)[:idx]
        cell_end = str(merge)[idx+1:]
        print(f"cell begin: {cell_begin} \ncell end: {cell_end}")
        if (ws[cell_begin].value == tb_name):
            isfound = 1
            found_cnt += 1
            row_begin = coordinate_to_tuple(cell_begin)[0]
            col_begin = coordinate_to_tuple(cell_begin)[1]
            col_end = coordinate_to_tuple(cell_end)[1]
            print(f"col begin: {col_begin}\ncol end: {col_end}\nrow begin: {row_begin}")
            bor: Border = ws[setcell(col_begin, row_begin+1)].border.left.style
            row_end = row_begin + 1
            while(bor is not None):
                row_end += 1
                bor = ws[setcell(col_begin, row_end)].border.left.style
            row_end -= 1    
        
        else:
             pass
    if(isfound == 0):
        print(f"The table {tb_name} is not found") 
        return None
    elif(isfound ==1 and found_cnt > 1):
        print(f"More than 1 table \"{tb_name}\" present")
        return None
    else:
        print(f" col: {col_begin} {col_end} \nrow: {row_begin} {row_end}")
        return col_begin, col_end, row_begin, row_end
    
def generate_bump_table(excel_path, bump_visual_sheet, package_type, die_table, die_coor, dummybump, die_params, int_die_tb, int_gen, int_die_cnt, tc_sr):


    
    root.update_idletasks()
    mynotif("Loading the ploc file...")
    root.update_idletasks()
    try:
        # wb_d = load_workbook(excel_path, data_only=True)
        print("Opening excel file...")
        mynotif("Opening excel file...")
        wb_f = load_workbook(excel_path)
        print(wb_f)   
    except:
        print("Wrong Ploc path or Ploc file is openning. Please recheck/close the PLOC file before generate :(")
        show_error("Wrong Ploc path or Ploc file is openning. Please recheck/close the PLOC file before generate :(")
        progress_bar(0)
        mynotif("Error")
        root.update_idletasks()
        return
    
    # ws = wb_f.create_sheet('Tung')
    try:
        sheet_list = wb_f.sheetnames
       #wsvisual_d = wb_d[bump_visual_sheet]  # use for further function
        wsvisual_f = wb_f[bump_visual_sheet]
      # wsdiebump_d = wb_d[die_table['sheet']] # use for further function
       
        if die_table['sheet'] in sheet_list:
            wsdiebump_f = wb_f[die_table['sheet']]
        else:
            msg_ws = messagebox.askquestion('Create Sheet', 'The ' + die_table['sheet'] + ' doesn\'t exist. Do you want to create it?',icon='question')
            # mynotif("")
            mynotif("The " + die_table['sheet'] + " doesn't exist.")
            if(msg_ws == 'yes'):
                # mynotif("")
                mynotif('Creating the sheet...')
                wsdiebump_f = wb_f.create_sheet(die_table['sheet'])
            else:
                # mynotif("")
                progress_bar(0)
                return
        if(int_gen == 1):
            if int_die_tb['sheet'] in sheet_list:
                wsintbump_f = wb_f[int_die_tb['sheet']]
            else:
                # mynotif("")
                mynotif("The " + int_die_tb['sheet'] + " doesn't exist.")
                msg_ws = messagebox.askquestion('Create Sheet', 'The ' + int_die_tb['sheet'] + ' doesn\'t exist. Do you want to create it?', icon='question')
            
                if(msg_ws == 'yes'):
                    wsintbump_f = wb_f.create_sheet(int_die_tb['sheet'])
                    # mynotif("")
                    mynotif('Creating the sheet...')
                else:
                    # mynotif("")
                    progress_bar(0)
                    return
            
      
        
       
       #wsintbump_d = wb_d[int_die_tb['sheet']] # use for further function
       
    except:
        print("Sheet " + bump_visual_sheet + " doesn't exist")
        show_error("Sheet " + bump_visual_sheet + " doesn't exist")
        progress_bar(0)
        mynotif("Error")
        root.update_idletasks()
        return
    
    

    #----- Create dummy bump at 4 corner 140x140u for advance package (CoWos)-----------#
    ymin = coordinate_to_tuple(die_coor['window1'])[0]
    xmin = coordinate_to_tuple(die_coor['window1'])[1]
    ymax = coordinate_to_tuple(die_coor['window2'])[0]
    xmax = coordinate_to_tuple(die_coor['window2'])[1]

    print(xmin,xmax)
    print(ymin,ymax)
    progress_bar(60)
    
    try:
        #----- Create table from bump map-----------#
        die_tb_x = coordinate_to_tuple(die_table['location'])[1]
        die_tb_y = coordinate_to_tuple(die_table['location'])[0]
        int_tb_x = coordinate_to_tuple(int_die_tb['int_tb_location'])[1]
        int_tb_y = coordinate_to_tuple(int_die_tb['int_tb_location'])[0]
        

        r_die = die_tb_y + 2
        r_int = int_tb_y + 2

        title_bg_fill = PatternFill(patternType='solid', fgColor='9e42f5')
        subtil_bg_fill = PatternFill(patternType='solid',fgColor='0e7bf0')
        wsdiebump_f[die_table['location']].value = die_table['name']
        
        wsdiebump_f.merge_cells(die_table['location'] + ":" + get_column_letter(die_tb_x + 2) + str(die_tb_y))
        for c1 in range(0,3):
            wsdiebump_f[get_column_letter(die_tb_x + c1) + str(die_tb_y)].fill = title_bg_fill
            wsdiebump_f[get_column_letter(die_tb_x + c1) + str(die_tb_y)].alignment = Alignment(horizontal='center')
            wsdiebump_f[get_column_letter(die_tb_x + c1) + str(die_tb_y)].border = Border(left=Side(style='thin'),right=Side(style='thin'),top=Side(style='thin'),bottom=Side(style='thin'))

        wsdiebump_f[get_column_letter(die_tb_x) + str(die_tb_y + 1)].value = "X"

        wsdiebump_f[get_column_letter(die_tb_x + 1) + str(die_tb_y + 1)].value = "Y"

        wsdiebump_f[get_column_letter(die_tb_x + 2)  + str(str(die_tb_y + 1))].value = "Bump name"
        for c2 in range(0,3):
            wsdiebump_f[get_column_letter(die_tb_x + c2) + str(die_tb_y + 1)].fill = subtil_bg_fill
            wsdiebump_f[get_column_letter(die_tb_x + c2) + str(die_tb_y + 1)].alignment = Alignment(horizontal='center')
            wsdiebump_f[get_column_letter(die_tb_x + c2) + str(die_tb_y + 1)].border = Border(left=Side(style='thin'),right=Side(style='thin'),top=Side(style='thin'),bottom=Side(style='thin'))
        if(int(tc_sr['isTC'] == 1)):
            wsdiebump_f[get_column_letter(die_tb_x + 4) + str(die_tb_y)].value = tc_sr['sr_tb']
            wsdiebump_f.merge_cells(get_column_letter(die_tb_x + 4) + str(die_tb_y) + ":" + get_column_letter(die_tb_x + 6) + str(die_tb_y))
            for c1 in range(0,3):
                wsdiebump_f[get_column_letter(die_tb_x + 4 + c1) + str(die_tb_y)].fill = title_bg_fill
                wsdiebump_f[get_column_letter(die_tb_x + 4 + c1) + str(die_tb_y)].alignment = Alignment(horizontal='center')
                wsdiebump_f[get_column_letter(die_tb_x + 4 + c1) + str(die_tb_y)].border = Border(left=Side(style='thin'),right=Side(style='thin'),top=Side(style='thin'),bottom=Side(style='thin'))
            wsdiebump_f[get_column_letter(die_tb_x + 4) + str(die_tb_y + 1)].value = "X"
            wsdiebump_f[get_column_letter(die_tb_x + 5) + str(die_tb_y + 1)].value = "Y"
            wsdiebump_f[get_column_letter(die_tb_x + 6)  + str(str(die_tb_y + 1))].value = "Bump name"
            for c1 in range(0,3):
                wsdiebump_f[get_column_letter(die_tb_x + 4 + c1) + str(die_tb_y + 1)].fill = subtil_bg_fill
                wsdiebump_f[get_column_letter(die_tb_x + 4 + c1) + str(die_tb_y + 1)].alignment = Alignment(horizontal='center')
                wsdiebump_f[get_column_letter(die_tb_x + 4 + c1) + str(die_tb_y + 1)].border = Border(left=Side(style='thin'),right=Side(style='thin'),top=Side(style='thin'),bottom=Side(style='thin'))

        if(int_gen == 1):
            die1_list = int_die_tb['Die1_name'].split()
            die2_list = int_die_tb['Die2_name'].split()
            die1_xoffset_list = die_params['die1_xoffset'].split()
            die1_yoffset_list = die_params['die1_yoffset'].split()
            die2_xoffset_list = die_params['die2_xoffset'].split()
            die2_yoffset_list = die_params['die2_yoffset'].split()
            if(len(die1_list) != int(int_die_cnt)/2 or len(die2_list) != int(int_die_cnt)/2 or len(die1_xoffset_list) != int(int_die_cnt)/2 or len(die2_xoffset_list) != int(int_die_cnt)/2 or len(die1_yoffset_list) != int(int_die_cnt)/2 or len(die2_yoffset_list) != int(int_die_cnt)/2):
                show_error('The input die parameters incorrect. Please re-check it')
                int_input_correct = 0
                mynotif('The input die parameters incorrect. Please re-check it')
                # mynotif("")
                progress_bar(0)
                return
            else:
                wsintbump_f[get_column_letter(int_tb_x ) + str(int_tb_y)].value = "Die Flipped by Y axis"
                wsintbump_f.merge_cells(get_column_letter(int_tb_x) + str(int_tb_y) + ":" + get_column_letter(int_tb_x + 2) + str(int_tb_y))
                for c1 in range(0,3):
                    wsintbump_f[get_column_letter(int_tb_x + c1) + str(int_tb_y)].fill = title_bg_fill
                    wsintbump_f[get_column_letter(int_tb_x + c1) + str(int_tb_y)].alignment = Alignment(horizontal='center',wrapText=True)
                    wsintbump_f[get_column_letter(int_tb_x + c1) + str(int_tb_y)].border = Border(left=Side(style='thin'),right=Side(style='thin'),top=Side(style='thin'),bottom=Side(style='thin'))
                wsintbump_f[get_column_letter(int_tb_x) + str(int_tb_y + 1)].value = "X"
                wsintbump_f[get_column_letter(int_tb_x + 1) + str(int_tb_y + 1)].value = "Y"
                wsintbump_f[get_column_letter(int_tb_x + 2)  + str(str(int_tb_y + 1))].value = "Bump name"
                wsintbump_f.freeze_panes = 'A' + str(int_tb_y + 1)
                for c2 in range(0,3):
                    wsintbump_f[get_column_letter(int_tb_x + c2) + str(int_tb_y + 1)].fill = subtil_bg_fill
                    wsintbump_f[get_column_letter(int_tb_x + c2) + str(int_tb_y + 1)].alignment = Alignment(horizontal='center')
                    wsintbump_f[get_column_letter(int_tb_x + c2) + str(int_tb_y + 1)].border = Border(left=Side(style='thin'),right=Side(style='thin'),top=Side(style='thin'),bottom=Side(style='thin'))
                tbidx = 4
                for tb in range(0,int(int_die_cnt/2)):

                    wsintbump_f[get_column_letter(int_tb_x + tbidx) + str(int_tb_y)].value =  str(die1_list[tb]) + " = Die Flipped, Rotate -90 + Offset" + "(" + str(die1_xoffset_list[tb]) + "," + str(die1_yoffset_list[tb]) + ")"
                    wsintbump_f.merge_cells(get_column_letter(int_tb_x + tbidx) + str(int_tb_y) + ":" + get_column_letter(int_tb_x + tbidx + 2) + str(int_tb_y))
                    
                    for c1 in range(0,3):
                        wsintbump_f[get_column_letter(int_tb_x + tbidx + c1) + str(int_tb_y)].fill = title_bg_fill
                        wsintbump_f[get_column_letter(int_tb_x + tbidx + c1) + str(int_tb_y)].alignment = Alignment(horizontal='center', wrapText=True)
                        wsintbump_f[get_column_letter(int_tb_x + tbidx + c1) + str(int_tb_y)].border = Border(left=Side(style='thin'),right=Side(style='thin'),top=Side(style='thin'),bottom=Side(style='thin'))
                    wsintbump_f[get_column_letter(int_tb_x + tbidx) + str(int_tb_y + 1)].value = "X"
                    wsintbump_f[get_column_letter(int_tb_x + tbidx + 1) + str(int_tb_y + 1)].value = "Y"
                    wsintbump_f[get_column_letter(int_tb_x + tbidx + 2)  + str(str(int_tb_y + 1))].value = "Bump name"
                    for c2 in range(0,3):
                        wsintbump_f[get_column_letter(int_tb_x + tbidx + c2) + str(int_tb_y + 1)].fill = subtil_bg_fill
                        wsintbump_f[get_column_letter(int_tb_x + tbidx + c2) + str(int_tb_y + 1)].alignment = Alignment(horizontal='center')
                        wsintbump_f[get_column_letter(int_tb_x + tbidx + c2) + str(int_tb_y + 1)].border = Border(left=Side(style='thin'),right=Side(style='thin'),top=Side(style='thin'),bottom=Side(style='thin'))

                    wsintbump_f[get_column_letter(int_tb_x + tbidx + 4) + str(int_tb_y)].value = str(die2_list[tb]) + " = Die Flipped, Rotate +90 + Offset" + "(" + str(die2_xoffset_list[tb]) + "," + str(die2_yoffset_list[tb]) + ")"
                    wsintbump_f.merge_cells(get_column_letter(int_tb_x + tbidx + 4) + str(int_tb_y) + ":" + get_column_letter(int_tb_x + tbidx + 6) + str(int_tb_y))
                    for c1 in range(0,3):
                        wsintbump_f[get_column_letter(int_tb_x + tbidx + c1 + 4) + str(int_tb_y)].fill = title_bg_fill
                        wsintbump_f[get_column_letter(int_tb_x + tbidx + c1 + 4) + str(int_tb_y)].alignment = Alignment(horizontal='center', wrapText=True)
                        wsintbump_f[get_column_letter(int_tb_x + tbidx + c1 + 4) + str(int_tb_y)].border = Border(left=Side(style='thin'),right=Side(style='thin'),top=Side(style='thin'),bottom=Side(style='thin'))
                    wsintbump_f[get_column_letter(int_tb_x + tbidx + 4) + str(int_tb_y + 1)].value = "X"
                    wsintbump_f[get_column_letter(int_tb_x + tbidx + 5) + str(int_tb_y + 1)].value = "Y"
                    wsintbump_f[get_column_letter(int_tb_x + tbidx + 6)  + str(str(int_tb_y + 1))].value = "Bump name"
                    for c2 in range(0,3):
                        wsintbump_f[get_column_letter(int_tb_x + tbidx + c2 + 4) + str(int_tb_y + 1)].fill = subtil_bg_fill
                        wsintbump_f[get_column_letter(int_tb_x + tbidx + c2 + 4) + str(int_tb_y + 1)].alignment = Alignment(horizontal='center')
                        wsintbump_f[get_column_letter(int_tb_x + tbidx + c2 + 4) + str(int_tb_y + 1)].border = Border(left=Side(style='thin'),right=Side(style='thin'),top=Side(style='thin'),bottom=Side(style='thin'))
                    tbidx += 8
                int_input_correct = 1
        else:
            pass
        # xwidth = float (ws_f[get_column_letter(xmax) + die_coor["xcoor"]].value)
        # minxval = float (ws_f[get_column_letter(xmin) + die_coor["xcoor"]].value)
        # ywidth = float (ws_f[die_coor["ycoor"] + str(ymin)].value)
        # minyval = float (ws_f[die_coor["ycoor"] + str(ymax)].value)
        # xwidth = ws_f[get_column_letter(xmax) + die_coor["xcoor"]].value
        # minxval = ws_f[get_column_letter(xmin) + die_coor["xcoor"]].value
        # ywidth = ws_f[die_coor["ycoor"] + str(ymin)].value
        # minyval = ws_f[die_coor["ycoor"] + str(ymax)].value
        if (package_type == 1):
            
            print("Generate for Advance Package")
            mynotif("Generate for Advance Package")
            dm_bump_coor= []
            dm_cnt=0
            # mynotif("")
            root.update_idletasks()
            mynotif("Generating Dummy bump...")
            root.update_idletasks()
            for dm_bump in dummybump:
                bump = list(dummybump[dm_bump].values())
                    
                ymin_dm = coordinate_to_tuple(bump[0])[0]
                xmin_dm = coordinate_to_tuple(bump[0])[1]
                ymax_dm = coordinate_to_tuple(bump[1])[0]
                xmax_dm = coordinate_to_tuple(bump[1])[1]
                xcoor_dm = str(bump[2])
                ycoor_dm = str(bump[3])

                print(xmin_dm,xmax_dm)
                print(ymin_dm,ymax_dm)
                print(dummybump)

                for dummycol1 in range(xmin_dm, xmax_dm + 1):
                    for dummyrow1 in range(ymin_dm, ymax_dm + 1):
                        col_dm = get_column_letter(dummycol1)
                        if (wsvisual_f[col_dm + str(dummyrow1)].value != None):
                            
                            print("Processing for Dummy bump at: " + col_dm + str(dummyrow1))
                            mynotif("Processing for Dummy bump at: " + col_dm + str(dummyrow1))
                            # Gen dummy bump table
                            wsdiebump_f[get_column_letter(die_tb_x)+str(r_die)].value = f"='{bump_visual_sheet}'!{col_dm + xcoor_dm}"
                            wsdiebump_f[get_column_letter(die_tb_x + 1)+str(r_die)].value = f"='{bump_visual_sheet}'!{ycoor_dm + str(dummyrow1)}"
                            wsdiebump_f[get_column_letter(die_tb_x + 2)+str(r_die)].value =  f"='{bump_visual_sheet}'!{col_dm+ str(dummyrow1)}"
                            if(int(tc_sr['isTC'] == 1)):
                                wsdiebump_f[get_column_letter(die_tb_x + 4)+str(r_die)].value = f"=('{bump_visual_sheet}'!{col_dm + xcoor_dm})-({tc_sr['sr_w']})"
                                wsdiebump_f[get_column_letter(die_tb_x + 5)+str(r_die)].value = f"=('{bump_visual_sheet}'!{ycoor_dm + str(dummyrow1)})-({tc_sr['sr_w']}) "
                                wsdiebump_f[get_column_letter(die_tb_x + 6)+str(r_die)].value =  f"='{bump_visual_sheet}'!{col_dm+ str(dummyrow1)}"
                            r_die += 1
                            coor = col_dm + str(dummyrow1)
                            dm_bump_coor.append(coor)
                            dm_cnt += 1

                            if(int_gen == 1 and int_input_correct == 1):

                                wsintbump_f[get_column_letter(int_tb_x)+str(r_int)].value = f"=({str(die_params['chip_width']).replace('=','')})-('{bump_visual_sheet}'!{str(col_dm + xcoor_dm)})" # Flip Y axis
                                wsintbump_f[get_column_letter(int_tb_x + 1)+str(r_int)].value = f"='{bump_visual_sheet}'!{ycoor_dm + str(dummyrow1)}" # Flip Y axis
                                wsintbump_f[get_column_letter(int_tb_x + 2)+str(r_int)].value =  f"='{bump_visual_sheet}'!{col_dm+ str(dummyrow1)}" #Flip Y axis
                                tbidx2 = 0
                                # r_current = 

                                for tb in range(0,int(int_die_cnt/2)):
                                # #----------------------------Flip bump map in y axis - Rotate -90 - Rotate +90---------------------------
                                    
                                    wsintbump_f[get_column_letter(int_tb_x + tbidx2 + 5)+str(r_int)].value = f"=({str(die_params['chip_width']).replace('=','')})-('{bump_visual_sheet}'!{str(col_dm + xcoor_dm)})+({str(die1_yoffset_list[tb])})" # Rotate -90
                                    wsintbump_f[get_column_letter(int_tb_x + tbidx2 + 9)+str(r_int)].value = f"=('{bump_visual_sheet}'!{str(col_dm + xcoor_dm)})+({str(die1_yoffset_list[tb]).replace('=','')})" # Rotate +90

                                    wsintbump_f[get_column_letter(int_tb_x + tbidx2 + 4)+str(r_int)].value = f"=({str(die_params['chip_height']).replace('=','')})-('{bump_visual_sheet}'!{ycoor_dm + str(dummyrow1)})+({str(die1_xoffset_list[tb])})" # Rotate -90
                                    wsintbump_f[get_column_letter(int_tb_x + tbidx2 + 8)+str(r_int)].value = f"=('{bump_visual_sheet}'!{ycoor_dm + str(dummyrow1)})+({str(die2_xoffset_list[tb]).replace('=','')})" # Rotate +90

                                
                                    if(wsvisual_f[col_dm+ str(dummyrow1)].value == "VSS"):
                                        wsintbump_f[get_column_letter(int_tb_x + tbidx2 + 6)+str(r_int)].value = f"='{bump_visual_sheet}'!{col_dm+ str(dummyrow1)}" # Rotate -90
                                        wsintbump_f[get_column_letter(int_tb_x + tbidx2 + 10)+str(r_int)].value = f"='{bump_visual_sheet}'!{col_dm+ str(dummyrow1)}" # Rotate +90
                                    else:
                                        wsintbump_f[get_column_letter(int_tb_x + tbidx2 + 6)+str(r_int)].value = f"=\"{die1_list[tb]}_\"&'{bump_visual_sheet}'!{col_dm+ str(dummyrow1)}" # Rotate -90
                                        wsintbump_f[get_column_letter(int_tb_x + tbidx2 + 10)+str(r_int)].value = f"=\"{die2_list[tb]}_\"&'{bump_visual_sheet}'!{col_dm+ str(dummyrow1)}" # Rotate +90
                                    tbidx2 += 8
                                
                                r_int += 1

            #---------Create Die bump exclued dummy bump at 4 corner-----------#

            match = 0
            # mynotif("")
            root.update_idletasks()
            mynotif("Generating Die bump...")
            root.update_idletasks()
            for col in range(xmin, xmax + 1):
                for row in range(ymin, ymax + 1):       
                    col_l = get_column_letter(col)
                    #print(col_l)
                    i = 0 
                    while(i < len(dm_bump_coor)):
                        xy = col_l + str(row)
                        if(xy ==  dm_bump_coor[i]):
                            match = 1
                        else:
                            match = 0
                        if(match == 1):
                            break
                        i += 1
                    if (match == 0 and wsvisual_f[col_l + str(row)].value != None):
                        print("Processing for Die bump at: " + col_l + str(row))
                        mynotif("Processing for Die bump at: " + col_l + str(row))
                        #  get the X value from Visual bump sheet
                        if (wsvisual_f[col_l + str(row)].value != None):
                        
                            #  get the X value from Visual bump sheet
                        
                            wsdiebump_f[get_column_letter(die_tb_x)+str(r_die)].value = f"='{bump_visual_sheet}'!{col_l + str(die_coor['xcoor'])}"
                        
                            # #  get the Y value from Visual bump sheet
                        
                            wsdiebump_f[get_column_letter(die_tb_x + 1)+str(r_die)].value = f"='{bump_visual_sheet}'!{die_coor['ycoor'] + str(row)}"
                            
                            #  get the Bump name from Visual bump sheet
                            wsdiebump_f[get_column_letter(die_tb_x + 2)+str(r_die)].value =  f"='{bump_visual_sheet}'!{col_l+ str(row)}"
                            if(int(tc_sr['isTC'] == 1)):
                               wsdiebump_f[get_column_letter(die_tb_x + 4)+str(r_die)].value = f"=('{bump_visual_sheet}'!{col_l + str(die_coor['xcoor'])})-({tc_sr['sr_w']})" 
                               wsdiebump_f[get_column_letter(die_tb_x + 5)+str(r_die)].value = f"=('{bump_visual_sheet}'!{die_coor['ycoor'] + str(row)})-({tc_sr['sr_w']})"
                               wsdiebump_f[get_column_letter(die_tb_x + 6)+str(r_die)].value =  f"='{bump_visual_sheet}'!{col_l+ str(row)}"
                            r_die += 1
                            
                            if(int_gen == 1 and int_input_correct == 1):

                                wsintbump_f[get_column_letter(int_tb_x )+str(r_int)].value = f"=({str(die_params['chip_width']).replace('=','')})-('{bump_visual_sheet}'!{col_l + str(die_coor['xcoor'])})" # Flip Y axis
                                wsintbump_f[get_column_letter(int_tb_x + 1)+str(r_int)].value = f"='{bump_visual_sheet}'!{die_coor['ycoor'] + str(row)}" # Flip Y axis
                                wsintbump_f[get_column_letter(int_tb_x + 2)+str(r_int)].value =  f"='{bump_visual_sheet}'!{col_l+ str(row)}" #Flip Y axis

                                tbidx2 = 0
                                # r_current = 

                                for tb in range(0,int(int_die_cnt/2)):
                                # #----------------------------Flip bump map in y axis - Rotate -90 - Rotate +90---------------------------
                                    
                                    wsintbump_f[get_column_letter(int_tb_x + tbidx2 + 5)+str(r_int)].value = f"=({str(die_params['chip_width']).replace('=','')})-('{bump_visual_sheet}'!{col_l + str(die_coor['xcoor'])})+({str(die1_yoffset_list[tb])})" # Rotate -90
                                    wsintbump_f[get_column_letter(int_tb_x + tbidx2 + 9)+str(r_int)].value = f"=('{bump_visual_sheet}'!{col_l + str(die_coor['xcoor'])})+({str(die2_yoffset_list[tb]).replace('=','')})" # Rotate +90
                                    
                                    wsintbump_f[get_column_letter(int_tb_x + tbidx2 + 4)+str(r_int)].value = f"=({str(die_params['chip_height']).replace('=','')})-('{bump_visual_sheet}'!{die_coor['ycoor']+str(row)})+({str(die1_xoffset_list[tb])})" # Rotate -90
                                    wsintbump_f[get_column_letter(int_tb_x + tbidx2 + 8)+str(r_int)].value = f"=('{bump_visual_sheet}'!{die_coor['ycoor'] + str(row)})+({str(die2_xoffset_list[tb]).replace('=','')})" # Rotate +90
                                
                                    if(wsvisual_f[col_l+ str(row)].value == "VSS"):
                                        wsintbump_f[get_column_letter(int_tb_x + tbidx2 + 6)+str(r_int)].value = f"='{bump_visual_sheet}'!{col_l+ str(row)}" # Rotate -90
                                        wsintbump_f[get_column_letter(int_tb_x + tbidx2 + 10)+str(r_int)].value = f"='{bump_visual_sheet}'!{col_l+ str(row)}" # Rotate +90
                                    else:
                                        wsintbump_f[get_column_letter(int_tb_x + tbidx2 + 6)+str(r_int)].value = f"=\"{die1_list[tb]}_\"&'{bump_visual_sheet}'!{col_l+ str(row)}" # Rotate -90
                                        wsintbump_f[get_column_letter(int_tb_x + tbidx2 + 10)+str(r_int)].value = f"=\"{die2_list[tb]}_\"&'{bump_visual_sheet}'!{col_l+ str(row)}" # Rotate +90
                                    tbidx2 += 8
                                
                                r_int += 1
                                
                                
        else:
            process_notify("Generating Die bump...")
            for col in range(xmin, xmax + 1):
                    for row in range(ymin , ymax + 1):       
                        col_l = get_column_letter(col)
                        #print(col_l)
                        if (wsvisual_f[col_l + str(row)].value != None):
                            print("Processing for Die bump at: " + col_l + str(row))
                            mynotif("Processing for Die bump at: " + col_l + str(row))
                            #  get the X value from Visual bump sheet
                        
                            wsdiebump_f[get_column_letter(die_tb_x)+str(r_die)].value = f"='{bump_visual_sheet}'!{col_l + str(die_coor['xcoor'])}"
                        
                            # #  get the Y value from Visual bump sheet
                        
                            wsdiebump_f[get_column_letter(die_tb_x + 1)+str(r_die)].value = f"='{bump_visual_sheet}'!{die_coor['ycoor'] + str(row)}"
                            
                            #  get the Bump name from Visual bump sheet
                            wsdiebump_f[get_column_letter(die_tb_x + 2)+str(r_die)].value =  f"='{bump_visual_sheet}'!{col_l+ str(row)}"
                            if(int(tc_sr['isTC'] == 1)):
                               wsdiebump_f[get_column_letter(die_tb_x + 4)+str(r_die)].value = f"=('{bump_visual_sheet}'!{col_l + str(die_coor['xcoor'])})-({tc_sr['sr_w']})" 
                               wsdiebump_f[get_column_letter(die_tb_x + 5)+str(r_die)].value = f"=('{bump_visual_sheet}'!{die_coor['ycoor'] + str(row)})-({tc_sr['sr_w']})"
                               wsdiebump_f[get_column_letter(die_tb_x + 6)+str(r_die)].value =  f"='{bump_visual_sheet}'!{col_l+ str(row)}"
                            r_die += 1
                            if(int_gen == 1 and int_input_correct == 1):
                                
                                wsintbump_f[get_column_letter(int_tb_x )+str(r_int)].value = f"=({str(die_params['chip_width']).replace('=','')})-('{bump_visual_sheet}'!{col_l + str(die_coor['xcoor'])})" # Flip Y axis
                                wsintbump_f[get_column_letter(int_tb_x + 1)+str(r_int)].value = f"='{bump_visual_sheet}'!{die_coor['ycoor'] + str(row)}" # Flip Y axis
                                wsintbump_f[get_column_letter(int_tb_x + 2)+str(r_int)].value =  f"='{bump_visual_sheet}'!{col_l+ str(row)}" #Flip Y axis

                                tbidx2 = 0
                                # r_current = 

                                for tb in range(0,int(int_die_cnt/2)):
                                # #----------------------------Flip bump map in y axis - Rotate -90 - Rotate +90---------------------------
                                    
                                    wsintbump_f[get_column_letter(int_tb_x + tbidx2 + 5)+str(r_int)].value = f"=({str(die_params['chip_width']).replace('=','')})-('{bump_visual_sheet}'!{col_l + str(die_coor['xcoor'])})+({str(die1_yoffset_list[tb])})" # Rotate -90
                                    wsintbump_f[get_column_letter(int_tb_x + tbidx2 + 9)+str(r_int)].value = f"=('{bump_visual_sheet}'!{col_l + str(die_coor['xcoor'])})+({str(die2_yoffset_list[tb]).replace('=','')})" # Rotate +90
                                
                                    
                                    wsintbump_f[get_column_letter(int_tb_x + tbidx2 + 4)+str(r_int)].value = f"=({str(die_params['chip_height']).replace('=','')})-('{bump_visual_sheet}'!{die_coor['ycoor']+str(row)})+({str(die1_xoffset_list[tb])})" # Rotate -90
                                    wsintbump_f[get_column_letter(int_tb_x + tbidx2 + 8)+str(r_int)].value = f"=('{bump_visual_sheet}'!{die_coor['ycoor'] + str(row)})+({str(die2_xoffset_list[tb]).replace('=','')})" # Rotate +90
                                
                                
                                    if(wsvisual_f[col_l+ str(row)].value == "VSS"):
                                        wsintbump_f[get_column_letter(int_tb_x + tbidx2 + 6)+str(r_int)].value = f"='{bump_visual_sheet}'!{col_l+ str(row)}" # Rotate -90
                                        wsintbump_f[get_column_letter(int_tb_x + tbidx2 + 10)+str(r_int)].value = f"='{bump_visual_sheet}'!{col_l+ str(row)}" # Rotate +90
                                    else:
                                        wsintbump_f[get_column_letter(int_tb_x + tbidx2 + 6)+str(r_int)].value = f"=\"{die1_list[tb]}_\"&'{bump_visual_sheet}'!{col_l+ str(row)}" # Rotate -90
                                        wsintbump_f[get_column_letter(int_tb_x + tbidx2 + 10)+str(r_int)].value = f"=\"{die2_list[tb]}_\"&'{bump_visual_sheet}'!{col_l+ str(row)}" # Rotate +90
                                    tbidx2 += 8
                                

                                r_int += 1
            # tab = Table(displayName="Table1", ref="O65:Q500")
            # ws_f.add_table(tab)
        
        progress_bar(80)   
        print("Saving excel...") 
        mynotif("Saving excel file...")
        wb_f.save(excel_path)
        progress_bar(100)
        mynotif("Successed!!!")
        print("Completed!!!")
        popup("PLOC generated successful!!!")
        entry_disable(text)
        
    except (ValueError):
        print ("Wrong input, Please check and regenerate")
        show_error("Wrong input, Please check and regenerate")
        progress_bar(0)
        mynotif("Error")
        root.update_idletasks()
        return
    except:
        print('Error!!!')
        
        show_error("There are an error in caculations, Please recheck and make sure the input is correct!")
        progress_bar(0)
        mynotif("Error")
        root.update_idletasks()
        return