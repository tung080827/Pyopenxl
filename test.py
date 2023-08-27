from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.utils import range_boundaries
from openpyxl.utils import column_index_from_string
#from openpyxl.utils import coordinate_from_string
from openpyxl.utils import coordinate_to_tuple
from openpyxl.worksheet.table import Table, TableStyleInfo
# from tkinter import *
from tkinter import messagebox
from tkinter import ttk
import getcolumn
from array import *
import tkinter as tk
from tkinter import *
from ttkthemes import ThemedTk, THEMES
from PIL import Image
from PIL import ImageTk, Image
from tkinter.font import Font as tkfont
from tkinter import filedialog
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection
from openpyxl.worksheet.worksheet import Worksheet

def getstring(string: str,c1: str, c2: str):
    cell = string
    idx1 = cell.find(c1)
    idx2 = cell.find(c2)
    if(idx1 == -1 or idx2 == -1):
        return "NA","NA"
    else:
        str_wo_c = cell[idx1+1:idx2]
        str_w_c = cell[idx1:idx2+1]
        return str_wo_c,str_w_c
import gui_function as gui
root = ThemedTk()

wb = load_workbook(r"C:\Users\sytung\OneDrive - Synopsys, Inc\Desktop\py\Test3.xlsx")
ws = wb['Bump Visual']
tb_name = 'Bump'

merge_lsit = ws.merged_cells.ranges
def setcell(col,row):
    return get_column_letter(col)+str(row)
def gettable(ws: Worksheet,tb_name: str):
    isfound = 0
    found_cnt = 0
    for merge in ws.merged_cells.ranges:
        idx = str(merge).find(':')
        cell_begin = str(merge)[:idx]
        cell_end = str(merge)[idx+1:]
        print(f"cell begin: {cell_begin} \ncell end: {cell_end}")
        if (ws[cell_begin].value == tb_name):
            isfound = 1
            found_cnt += 1
            row_begin = coordinate_to_tuple(cell_begin)[0]
            col_begin = coordinate_to_tuple(cell_begin)[1]
            col_end = coordinate_to_tuple(cell_end)[1]
            print(f"col begin: {col_begin}\ncol end: {col_end}\nrow begin: {row_begin}")
            bor: Border = ws[setcell(col_begin, row_begin+1)].border.left.style
            row_end = row_begin + 1
            while(bor is not None):
                row_end += 1
                bor = ws[setcell(col_begin, row_end)].border.left.style
            row_end -= 1    
        
        else:
             pass
    if(isfound == 0):
        print(f"The table {tb_name} is not found") 
        return None
    elif(isfound ==1 and found_cnt > 1):
        print(f"More than 1 table \"{tb_name}\" present")
        return None
    else:
        print(f" col: {col_begin} {col_end} \nrow: {row_begin} {row_end}")
        return col_begin, col_end, row_begin, row_end
        
            
            # while ()
     
gettable(ws,tb_name) 
