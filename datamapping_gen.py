from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.utils import range_boundaries
from openpyxl.utils import column_index_from_string
#from openpyxl.utils import coordinate_from_string
from openpyxl.utils import coordinate_to_tuple
from openpyxl.worksheet.table import Table, TableStyleInfo
# from tkinter import *
from tkinter import messagebox
from tkinter import ttk
import getcolumn
from array import *
import tkinter as tk
from tkinter import *
from ttkthemes import ThemedTk, THEMES
from PIL import Image
from PIL import ImageTk, Image
from tkinter.font import Font as tkfont
from tkinter import filedialog
import gui_function as gui
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection
import os
import win32com.client
from pathlib import Path  # core library
from ploc_myTk import *

# excel_file = r"C:\Users\sytung\OneDrive - Synopsys, Inc\Desktop\py\Test.xlsx"
# pin_file = r"C:\Users\sytung\OneDrive - Synopsys, Inc\Desktop\py\DWORD_NET.txt"
def mynotif(texbox: TkTextbox,content : str):
    texbox.add_text(content + "\n")
    texbox.textbox.see("end")
def popup(notif: str):
    messagebox.showinfo("Notification", notif)

def show_error(error: str):
    messagebox.showerror("Error", error)  

def get_col_row_range(cell_begin, cell_end):
    row_begin = coordinate_to_tuple(cell_begin)[0]
    col_begin = coordinate_to_tuple(cell_begin)[1]
    row_end = coordinate_to_tuple(cell_end)[0]
    col_end = coordinate_to_tuple(cell_end)[1]
    return row_begin,row_end,col_begin,col_end


def getstring(string: str,c1: str, c2: str):
    cell = string
    idx1 = cell.find(c1)
    idx2 = cell.find(c2)
    if(idx1 == -1 or idx2 == -1):
        return None,None, None
    else:
        str_wo_c = cell[idx1+1:idx2]
        str_w_c = cell[idx1:idx2+1]
        str_cut = cell[:idx1]
        return str_wo_c,str_w_c, str_cut


def mapping_connections(wb,textbox:TkTextbox, input_params:dict,mapping_tb_out:dict):
    print("Generating Mapping table")
    mynotif(textbox,'Generating Mapping table...')
    try:
        sheet_list = wb.sheetnames
        out_mapws_name =  mapping_tb_out['sheet_name']
        if out_mapws_name in sheet_list:
                
                mapping_sheet = wb[out_mapws_name]
        else:
                # mynotif("")
                mynotif(textbox,"The \"" + out_mapws_name + "\" doesn't exist.")
                msg_ws = messagebox.askquestion('Create Sheet', 'The sheet \"' + out_mapws_name + '\" doesn\'t exist. Do you want to create it?', icon='question')
            
                if(msg_ws == 'yes'):
                    mapping_sheet = wb.create_sheet(out_mapws_name)
                    # mynotif("")
                    mynotif(textbox,'Creating the sheet...')
                else:
                    mynotif(textbox,"Aborted!")
                    # progress_bar(0)
                    return
    
        # mapping_sheet = wb_f.create_sheet("MAPPING")
        _input_params={
            'ch_sheet_name': input_params['ch_sheet_name'],
            'ch_begin_cell': input_params['ch_begin_cell'],
            'ch_end_cell': input_params['ch_end_cell'],
            'ch_num': input_params['ch_num'],
            'pwr_list': input_params['pwr_list'],
            'bus_char': input_params['bus_char'],
            'DieL_name': input_params['DieL_name'],
            'DieR_name': input_params['DieR_name'],
            'map_char': input_params['map_char'],
            'single_bus': input_params['signal_bus'][1],
            'multi_bus': input_params['signal_bus'][2],
            'ch_seq': input_params['ch_seq']
        }
        _mapping_tb_out = {
            'sheet':mapping_sheet,
            # 'tb_ch2ch_name':mapping_tb_out['tb_ch2ch_name'],
            'tb_ch2ch_loc':mapping_tb_out['tb_ch2ch_loc'],
            
            # 'tb_d2d_name': "DIE to DIE Mapping",
            
        }
        # try:
        #     with open(pin_file,'r') as pin:
        #         pin_list =  [line.rstrip() for line in pin]
        # except:
        #     print("The file " +pin_file + " doen't exist. Please re-check")
        powerlist = list(str(_input_params['pwr_list']).split(" "))
        mapchar = list(str(_input_params['map_char']).split(" "))
        ch_sheet_name = _input_params['ch_sheet_name']
        if ch_sheet_name in sheet_list:
                ch_sheet = wb[ch_sheet_name]
        else:
            msg_ws = messagebox.showerror('Open Sheet', 'The sheet: \"' + ch_sheet + '\" doesn\'t exist.')
            # mynotif("")
            mynotif(textbox,"The \"" + ch_sheet_name + "\" doesn't exist.")
        
            mynotif(textbox,"Aborted!!!")
            # progress_bar(0)
            return
        mynotif(textbox,'Generating Channel to Channel mapping table...')
        ch_row_col = get_col_row_range(_input_params['ch_begin_cell'], _input_params['ch_end_cell'])

        ch_row_begin = ch_row_col[0]
        ch_row_end = ch_row_col[1]
        ch_col_begin = ch_row_col[2]
        ch_col_end = ch_row_col[3]

        title_bg_fill = PatternFill(patternType='solid', fgColor='9e42f5')
        subtil_bg_fill = PatternFill(patternType='solid',fgColor='0e7bf0')
        row_ch2ch_begin = coordinate_to_tuple(_mapping_tb_out['tb_ch2ch_loc'])[0]
        col_ch2ch_begin = coordinate_to_tuple(_mapping_tb_out['tb_ch2ch_loc'])[1]
        
        row_d2d_begin = row_ch2ch_begin
        print(get_column_letter(col_ch2ch_begin) + str(row_ch2ch_begin)+":"+ get_column_letter(col_ch2ch_begin + 1) + str(row_ch2ch_begin))
        
        mapping_sheet[get_column_letter(col_ch2ch_begin) + str(row_ch2ch_begin)] = "CHANNEL - CHANNEL"
        mapping_sheet.merge_cells(get_column_letter(col_ch2ch_begin) + str(row_ch2ch_begin)+":"+get_column_letter(col_ch2ch_begin + 1) + str(row_ch2ch_begin))
        row_ch2ch_begin += 1
        mapping_sheet[get_column_letter(col_ch2ch_begin) + str(row_ch2ch_begin)] = "CHANNEL_1"
        mapping_sheet[get_column_letter(col_ch2ch_begin + 1) + str(row_ch2ch_begin)] = "CHANNEL_2"
        for c in range(col_ch2ch_begin, col_ch2ch_begin + 2):
            mapping_sheet[get_column_letter(c) + str(row_ch2ch_begin -1)].fill = title_bg_fill
            mapping_sheet[get_column_letter(c) + str(row_ch2ch_begin -1)].alignment = Alignment(horizontal='center')
            mapping_sheet[get_column_letter(c) + str(row_ch2ch_begin -1)].border = Border(left=Side(style='thin'),right=Side(style='thin'),top=Side(style='thin'),bottom=Side(style='thin'))
            mapping_sheet[get_column_letter(c) + str(row_ch2ch_begin)].fill = subtil_bg_fill
            mapping_sheet[get_column_letter(c) + str(row_ch2ch_begin)].alignment = Alignment(horizontal='center')
            mapping_sheet[get_column_letter(c) + str(row_ch2ch_begin)].border = Border(left=Side(style='thin'),right=Side(style='thin'),top=Side(style='thin'),bottom=Side(style='thin'))

        row_ch2ch_begin += 1
        ch2ch_tb_row_begin = row_ch2ch_begin
        
        ch1_col = col_ch2ch_begin
        ch2_col = col_ch2ch_begin + 1
        print("Generating Mapping for channel - channel")
        for col in range(ch_col_begin, ch_col_end + 1):
            
            col_l = get_column_letter(col)
            for row in range(ch_row_begin, ch_row_end + 1):
                    cell_val = ch_sheet[col_l + str(row)].value
                    if(cell_val != None):
                        if(cell_val not in powerlist):
                            mapping_sheet[get_column_letter(ch1_col) + str(row_ch2ch_begin)].value = cell_val
                            if(cell_val.find(mapchar[0]) != -1):
                                mapping_sheet[get_column_letter(ch2_col) + str(row_ch2ch_begin)].value = cell_val.replace(mapchar[0],mapchar[1])
                            elif(cell_val.find(mapchar[1]) != -1):
                                mapping_sheet[get_column_letter(ch2_col) + str(row_ch2ch_begin)].value = cell_val.replace(mapchar[1],mapchar[0])
                            else:
                                mapping_sheet[get_column_letter(ch2_col) + str(row_ch2ch_begin)].value = "NA"
                            row_ch2ch_begin += 1
        ##############################################
        mynotif(textbox,'Generating Die to Die mapping table...')
        ch2ch_tb_row_end = row_ch2ch_begin
        ch2ch_tb_col1 =  ch1_col
        ch2ch_tb_col2 = ch2_col              

        col_d2d1_begin = col_ch2ch_begin + 4
        col_d2d2_begin = col_ch2ch_begin + 5
    
        mapping_sheet[get_column_letter(col_d2d1_begin) + str(row_d2d_begin)] = "DIE to DIE COMMON"
        mapping_sheet.merge_cells(get_column_letter(col_d2d1_begin) + str(ch2ch_tb_row_begin -2)+":"+get_column_letter(col_d2d2_begin) + str(ch2ch_tb_row_begin -2))
        row_d2d_begin += 1
        mapping_sheet[get_column_letter(col_d2d1_begin) + str(row_d2d_begin)] = "DIE_LEFT/DIE_UP"
        mapping_sheet[get_column_letter(col_d2d2_begin) + str(row_d2d_begin)] = "DIE_RIGHT/DIE_DOWN"
        mapping_sheet.freeze_panes = 'A'+ str(row_d2d_begin + 1)
        for c in range(col_d2d1_begin, col_d2d2_begin + 1):
            mapping_sheet[get_column_letter(c) + str(row_d2d_begin -1)].fill = title_bg_fill
            mapping_sheet[get_column_letter(c) + str(row_d2d_begin -1)].alignment = Alignment(horizontal='center')
            mapping_sheet[get_column_letter(c) + str(row_d2d_begin -1)].border = Border(left=Side(style='thin'),right=Side(style='thin'),top=Side(style='thin'),bottom=Side(style='thin'))
            mapping_sheet[get_column_letter(c) + str(row_d2d_begin)].fill = subtil_bg_fill
            mapping_sheet[get_column_letter(c) + str(row_d2d_begin)].alignment = Alignment(horizontal='center')
            mapping_sheet[get_column_letter(c) + str(row_d2d_begin)].border = Border(left=Side(style='thin'),right=Side(style='thin'),top=Side(style='thin'),bottom=Side(style='thin'))

        row_d2d_begin += 1
        ch_num = int(_input_params['ch_num'])
        buschar = _input_params['bus_char']
        singlelist = _input_params['single_bus']
        multildict = _input_params['multi_bus']
        ch_seq = _input_params['ch_seq']
        ch_sequence = ["Right to Left","Left to Right", "Center to Left first","Center to Right first",  "Left Edge to Center first", "Right Edge to Center first"]
        ch_sequence_ew = ["Top to Bot","Bot to Top", "Center to Bot first","Center to Top first", "Bot Edge to Center first", "Top Edge to Center first"]
        # bit_num = int(_input_params['bit_num'])
        r_d2d = row_d2d_begin

        ch_cnt_down = ch_num -1
        print("Generating Mapping for DIE - DIE")
        # edge_gap = ch_num
        for ch_cnt in range(0, ch_num):
            # for col in range(ch2ch_tb_col_begin, ch2ch_tb_col_end + 1):
            ch_col_l = get_column_letter(ch2ch_tb_col1)
            d1_col_l = get_column_letter(col_d2d1_begin)
            d2_col_l = get_column_letter(col_d2d2_begin)
            for row in range(ch2ch_tb_row_begin, ch2ch_tb_row_end):
                cell_val = str(mapping_sheet[ch_col_l + str(row)].value)  
                if cell_val not in powerlist:
                    if cell_val in singlelist:
                        mapping_sheet[d1_col_l+str(r_d2d)].value = str(cell_val) + buschar[0] + str(ch_cnt) + buschar[1]
                        if(cell_val.find(mapchar[0]) != -1):
                            cell_val = cell_val.replace(mapchar[0],mapchar[0])
                            if ch_seq == ch_sequence[0] or ch_seq == ch_sequence[1] or ch_seq == ch_sequence_ew[0] or ch_seq == ch_sequence_ew[1]:
                                mapping_sheet[d2_col_l+str(r_d2d)].value = str(cell_val) + buschar[0] + str(ch_cnt_down) + buschar[1]
                            elif ch_seq == ch_sequence[2] or ch_seq == ch_sequence[3] or ch_seq == ch_sequence[4] or ch_seq == ch_sequence[5] or ch_seq == ch_sequence_ew[2] or ch_seq == ch_sequence_ew[3] or ch_seq == ch_sequence_ew[4] or ch_seq == ch_sequence_ew[5]:
                                mapping_sheet[d2_col_l+str(r_d2d)].value = str(cell_val) + buschar[0] + str(ch_cnt + int(ch_num/2)) + buschar[1]
                            # elif ch_seq == ch_sequence[4] or ch_seq == ch_sequence[5]:
                            #     mapping_sheet[d2_col_l+str(r_d2d)].value = str(cell_val) + buschar[0] + str(ch_cnt + edge_gap) + buschar[1]
                        elif(cell_val.find(mapchar[1]) != -1):
                            cell_val = cell_val.replace(mapchar[1], mapchar[0])
                            if ch_seq == ch_sequence[0] or ch_seq == ch_sequence[1] or ch_seq == ch_sequence_ew[0] or ch_seq == ch_sequence_ew[1]: 
                                mapping_sheet[d2_col_l+str(r_d2d)].value = str(cell_val) + buschar[0] + str(ch_cnt_down) + buschar[1]
                            elif ch_seq == ch_sequence[2] or ch_seq == ch_sequence[3] or ch_seq == ch_sequence[4] or ch_seq == ch_sequence[5] or ch_seq == ch_sequence_ew[2] or ch_seq == ch_sequence_ew[3] or ch_seq == ch_sequence_ew[4] or ch_seq == ch_sequence_ew[5]:
                                mapping_sheet[d2_col_l+str(r_d2d)].value = str(cell_val) + buschar[0] + str(ch_cnt + int(ch_num/2)) + buschar[1]
                            # elif ch_seq == ch_sequence[4] or ch_seq == ch_sequence[5]:
                            #     mapping_sheet[d2_col_l+str(r_d2d)].value = str(cell_val) + buschar[0] + str(ch_cnt + edge_gap) + buschar[1]
                    else:
                        index =  getstring(str(cell_val),buschar[0],buschar[1])
                        bit_num = multildict[index[2]]
                        mapping_sheet[d1_col_l+str(r_d2d)].value = str(cell_val).replace(index[1],'') + buschar[0] + str(ch_cnt * bit_num + int(index[0])) + buschar[1]
                        if(cell_val.find(mapchar[0]) != -1):
                            cell_val = cell_val.replace(mapchar[0],mapchar[1])
                            if ch_seq == ch_sequence[0] or ch_seq == ch_sequence[1] or ch_seq == ch_sequence_ew[0] or ch_seq == ch_sequence_ew[1]:
                                mapping_sheet[d2_col_l+str(r_d2d)].value = str(cell_val).replace(index[1],'') + buschar[0] + str(ch_cnt_down * bit_num + int(index[0])) + buschar[1]
                            elif ch_seq == ch_sequence[2] or ch_seq == ch_sequence[3] or ch_seq == ch_sequence[4] or ch_seq == ch_sequence[5] or ch_seq == ch_sequence_ew[2] or ch_seq == ch_sequence_ew[3] or ch_seq == ch_sequence_ew[4] or ch_seq == ch_sequence_ew[5]:
                                mapping_sheet[d2_col_l+str(r_d2d)].value = str(cell_val).replace(index[1],'') + buschar[0] + str((ch_cnt + int(ch_num/2)) * bit_num + int(index[0])) + buschar[1]
                            # elif ch_seq == ch_sequence[4] or ch_seq == ch_sequence[5]:
                            #     mapping_sheet[d2_col_l+str(r_d2d)].value = str(cell_val).replace(index[1],'') + buschar[0] + str((ch_cnt + edge_gap) * bit_num + int(index[0])) + buschar[1]
                        elif(cell_val.find(mapchar[1]) != -1):
                            cell_val = cell_val.replace(mapchar[1],mapchar[0])
                            if ch_seq == ch_sequence[0] or ch_seq == ch_sequence[1] or ch_seq == ch_sequence_ew[0] or ch_seq == ch_sequence_ew[1]:
                                mapping_sheet[d2_col_l+str(r_d2d)].value = str(cell_val).replace(index[1],'') + buschar[0] + str(ch_cnt_down * bit_num + int(index[0])) + buschar[1]
                            elif ch_seq == ch_sequence[2] or ch_seq == ch_sequence[3] or ch_seq == ch_sequence[4] or ch_seq == ch_sequence[5] or ch_seq == ch_sequence_ew[2] or ch_seq == ch_sequence_ew[3] or ch_seq == ch_sequence_ew[4] or ch_seq == ch_sequence_ew[5]:
                                mapping_sheet[d2_col_l+str(r_d2d)].value = str(cell_val).replace(index[1],'') + buschar[0] + str((ch_cnt + int(ch_num/2)) * bit_num + int(index[0])) + buschar[1]
                            # elif ch_seq == ch_sequence[4] or ch_seq == ch_sequence[5]:
                            #     mapping_sheet[d2_col_l+str(r_d2d)].value = str(cell_val).replace(index[1],'') + buschar[0] + str((ch_cnt + edge_gap) * bit_num + int(index[0])) + buschar[1]
                
                r_d2d += 1
            # edge_gap -= 1
            ch_cnt_down -= 1
            row_d2d_end = r_d2d
        print("Generating mapping for list Die")

        dieL_list = list(_input_params['DieL_name'].split())
        dieR_list = list(_input_params['DieR_name'].split())
        col_dL_begin = col_d2d1_begin + 4
        col_dR_begin = col_d2d2_begin + 4

        if(len(dieL_list) != len(dieR_list)):
            print("Error!!.Die Left and Die R is not equal. Please check")
            mynotif(textbox,"Error!!.Die Left and Die R is not equal. Please check")
        else:
            for die in range (0, len(dieL_list)):
                mynotif(textbox,f"Generating {dieL_list[die]} - {dieR_list[die]} Mapping table...")
                r_d2d = row_d2d_begin
                r_d = row_d2d_begin
                mapping_sheet[get_column_letter(col_dL_begin) + str(row_d2d_begin - 2)] = dieL_list[die] +"-" + dieR_list[die] + " MAPPING"
                mapping_sheet.merge_cells(get_column_letter(col_dL_begin) + str(r_d2d -2)+":"+get_column_letter(col_dR_begin) + str(r_d2d -2))
                
                for c in range(col_dL_begin, col_dR_begin + 1):
                    mapping_sheet[get_column_letter(c) + str(r_d2d -2)].fill = title_bg_fill
                    mapping_sheet[get_column_letter(c) + str(r_d2d -2)].alignment = Alignment(horizontal='center')
                    mapping_sheet[get_column_letter(c) + str(r_d2d -2)].border = Border(left=Side(style='thin'),right=Side(style='thin'),top=Side(style='thin'),bottom=Side(style='thin'))
                    mapping_sheet[get_column_letter(c) + str(r_d2d -1)].fill = subtil_bg_fill
                    mapping_sheet[get_column_letter(c) + str(r_d2d -1)].alignment = Alignment(horizontal='center')
                    mapping_sheet[get_column_letter(c) + str(r_d2d -1)].border = Border(left=Side(style='thin'),right=Side(style='thin'),top=Side(style='thin'),bottom=Side(style='thin'))

                mapping_sheet[get_column_letter(col_dL_begin) + str(row_d2d_begin - 1)] = dieL_list[die]
                mapping_sheet[get_column_letter(col_dR_begin) + str(row_d2d_begin - 1)] = dieR_list[die]
                for d in range(row_d2d_begin, row_d2d_end):
                    mapping_sheet[get_column_letter(col_dL_begin) + str(r_d)].value = dieL_list[die] +"_"+ mapping_sheet[get_column_letter(col_d2d1_begin)+str(r_d2d)].value
                    mapping_sheet[get_column_letter(col_dR_begin) + str(r_d)].value = dieR_list[die] +"_"+ mapping_sheet[get_column_letter(col_d2d2_begin)+str(r_d2d)].value
                    r_d +=1
                    r_d2d +=1
                col_dL_begin += 4
                col_dR_begin += 4
        iserr = False
        return wb, iserr 
    except Exception as e:
        mynotif(textbox, "An error occurred:: "+ e)
        iserr = True
        return wb, iserr



    # for row in range(0, len(pin_list)):

    #     mapping_sheet[get_column_letter(col_ch2ch_begin) + str(row_ch2ch_begin)] = pin_list[row]
    #     if(pin_list[row].find("TX") != -1):
    #         mapping_sheet[get_column_letter(col_ch2ch_begin + 1) + str(row_ch2ch_begin)] = pin_list[row].replace('TX','RX')
    #     elif(pin_list[row].find("RX") != -1):
    #         mapping_sheet[get_column_letter(col_ch2ch_begin + 1) + str(row_ch2ch_begin)] = pin_list[row].replace('RX','TX')
    #     else:
    #         mapping_sheet[get_column_letter(col_ch2ch_begin + 1) + str(row_ch2ch_begin)] = "NA"
    #     row_ch2ch_begin += 1

#     wb.save(excel_file)
# mapping_connections(wb_m, input_params, mapping_tb_out)