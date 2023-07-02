from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.utils import range_boundaries
from openpyxl.utils import column_index_from_string
#from openpyxl.utils import coordinate_from_string
from openpyxl.utils import coordinate_to_tuple
# from tkinter import *
from tkinter import messagebox
from tkinter import ttk
import getcolumn
from array import *
import tkinter as tk

from ttkthemes import ThemedTk, THEMES
from PIL import Image


# adv = 1

# label.pack(padx=40,pady=40)
# Create a style
root = ThemedTk()
# my_canvas=tk.Canvas(root)
root.set_theme("scidpurple")

root.title("PLOC TABLE GENERATOR")
root.geometry("1000x1000+30+100")

root.option_add("*tearOff", False) # This is always a good idea
# my_canvas.pack(side='left', fill='both', expand=1)
# root.configure(bg= "blue")

# Make the app responsive
# root.columnconfigure(index=0, weight=1)
# root.columnconfigure(index=1, weight=1)
# root.columnconfigure(index=2, weight=1)
# root.columnconfigure(index=3, weight=1)

# root.grid_columnconfigure(0, weight=1)
# root.grid_rowconfigure(0, weight=1)
# root.rowconfigure(index=0, weight=1)
# root.rowconfigure(index=1, weight=1)
# root.rowconfigure(index=2, weight=1)
# root.rowconfigure(index=3, weight=1)
# root.rowconfigure(index=4, weight=1)
# root.rowconfigure(index=5, weight=1)
# root.rowconfigure(index=6, weight=1)
# root.rowconfigure(index=7, weight=1)

# Create a style
# style = ttk.Style()

stfont= ("Franklin Gothic Medium", 10, 'underline', "italic")
# creating the theme with the
# style.theme_create('style_class',
  
#                    # getting the settings
#                    settings={
  
#                        # getting through the Labelframe
#                        # widget
#                        'TLabelframe': {
                           
#                            # configure the changes
#                            'configure': {
#                             #    'background': '#f5e6ff'
#                             'bordercolor': 'green',
#                            }
#                        },
  
#                        # getting through the Labelframe's 
#                        # label widget
#                        'TLabelframe.Label': {
#                            'configure': {
#                             #    'background': 'green',
#                                 'bordercolor': 'green',
#                                 'font' : stfont
#                            }
#                        }
#                    }
#                    )

# Import the tcl file
# root.tk.call("source", r"C:\Users\sytung\OneDrive - Synopsys, Inc\Desktop\py\forest-dark.tcl")
# root.tk.call("source", r"C:\Users\sytung\OneDrive - Synopsys, Inc\Desktop\py\forest-light.tcl")
# # Set the theme with the theme_use method
# style.theme_use("forest-dark")




# Create lists for the Comboboxes
theme_list = ["adapta", "aquativo", "arc", "black","blue", "breeze", "clearlooks", "elegance", "equilux", "itft1", "keramik", "keramik_alt", "kroc", "plastik", "radiance", "ubuntu", "scidblue", "scidgreen", "scidgrey", "scidmint", "scidpink", "scidpurple", "scidsand", "smog", "winxpblue", "yaru" ]
package_list = ["S-Organic", "A-CoWoS", "A-EMIB"]
foundry_list = ["TSMC-MapwSR", "TSMC-MapwoSR", "SS-MapwSR", "SS-MapwoSR", "GF-MapwSR", "GF-MapwSR"]

# Create control variables
a = tk.BooleanVar()
b = tk.BooleanVar(value=True)
c = tk.BooleanVar()
d = tk.IntVar(value=2)
# e = tk.StringVar(value=option_menu_list[1])
f = tk.BooleanVar()
g = tk.DoubleVar(value=75.0)
h = tk.BooleanVar()
tc_opt = tk.IntVar()

#Define a Function to enable the frame
def enable(children):
   for child in children:
      child.configure(state='enable')
def disable(children):  
    for child in children:
        child.configure(state='disable')
def entry_disable(*entries):
    for entry in entries:
        entry.config(state='disable')
def entry_enable(*entries):
    for entry in entries:
        entry.config(state='normal')
def entry_toggle():
    print("TUng day")
        # if(entry['state'] == 'disable'):
    if(tc_opt.get() == 1):
        foundry_combo.config(state='normal')
    elif(tc_opt.get() == 0):
        foundry_combo.config(state='disable')
def progress_bar(value):
    progress['value'] = value
    root.update_idletasks()

def choosetheme(event):
    for theme in theme_list:
        if (theme_combo.get() == theme):
            root.set_theme(theme)
def choosemode(event):   
    if(package_combo.get() == "S-Organic"):
       
       entry_disable(cor1_x1y1, cor1_x2y2, cor1_Xget, cor1_Yget,
                     cor2_x1y1, cor2_x2y2, cor2_Xget, cor2_Yget,
                     cor3_x1y1, cor3_x2y2, cor3_Xget, cor3_Yget,
                     cor4_x1y1, cor4_x2y2, cor4_Xget, cor4_Yget)
       entry_disable(c4_x1y1_i,c4_x2y2_i, c4_Xget_i, c4_Yget_i, c4_tb_name, c4_col,
                      u_col, u_tb_name, u_x1y1_i, u_x2y2_i, u_Xget_i, u_Yget_i)
       entry_enable(x1y1_i, x2y2_i, Xget_i, Yget_i)
       entry_enable(out_name_in, out_col_i)
       entry_disable(sheete_i, sheete_t)
       sheet_t['text']= "Bump sheet:"
        
       print(1)
    elif(package_combo.get() == "A-CoWoS"):
    #    enable(dmbump_frame.winfo_children())
        # cor1_x1y1.config(state='disable')
        entry_enable(cor1_x1y1, cor1_x2y2, cor1_Xget, cor1_Yget,
                     cor2_x1y1, cor2_x2y2, cor2_Xget, cor2_Yget,
                     cor3_x1y1, cor3_x2y2, cor3_Xget, cor3_Yget,
                     cor4_x1y1, cor4_x2y2, cor4_Xget, cor4_Yget)
        entry_enable(x1y1_i, x2y2_i, Xget_i, Yget_i)
        entry_disable(sheete_i, sheete_t)
        entry_enable(out_name_in, out_col_i)
        entry_disable(c4_x1y1_i,c4_x2y2_i, c4_Xget_i, c4_Yget_i, c4_tb_name, c4_col,
                      u_col, u_tb_name, u_x1y1_i, u_x2y2_i, u_Xget_i, u_Yget_i)
        sheet_t['text']= "Bump sheet:"
        print(0)
    else:
        entry_disable(cor1_x1y1, cor1_x2y2, cor1_Xget, cor1_Yget,
                     cor2_x1y1, cor2_x2y2, cor2_Xget, cor2_Yget,
                     cor3_x1y1, cor3_x2y2, cor3_Xget, cor3_Yget,
                     cor4_x1y1, cor4_x2y2, cor4_Xget, cor4_Yget)
        entry_disable(x1y1_i, x2y2_i, Xget_i, Yget_i)
        entry_disable(out_name_in, out_col_i)
        entry_enable(c4_x1y1_i,c4_x2y2_i, c4_Xget_i, c4_Yget_i, c4_tb_name, c4_col,
                      u_col, u_tb_name, u_x1y1_i, u_x2y2_i, u_Xget_i, u_Yget_i)
        entry_enable(sheete_i, sheete_t)
        sheet_t['text']= "uBump sheet:"

        popup("The EMIB package type have not developed yet, Please use S-Organic to gen 2 times (for C4 and uBump) instead!")

def mynotif(content):
        
        myLabel = ttk.Label(root,text=content)
        myLabel.grid(row=5, column=0, columnspan=2, padx=(20, 10), pady=(20, 10), sticky="nsew")
        
        # excel_path = "r"+ e1.get()
        # e1.delete(0,END)
# Create a Frame for the Checkbuttons
# style.configure("TLabelframe", bordercolor="red")
# pst = ttk.Style()
# pst.configure("TLabelframe", font= ('Arial', 15),
# background="red")
ploc_frame = ttk.LabelFrame(root, text="Ploc input config", padding=(20, 10))


ploc_frame.grid(row=1,column=0, columnspan=2, padx=(20, 10), pady=(20, 10), sticky="nsew")

# Create a Frame for input widgets
# widgets_frame = ttk.Frame(check_frame, padding=(0, 0, 0, 10))
# widgets_frame.grid(row=0, column=1, padx=10, pady=(30, 10), sticky="nsew", rowspan=3)
# widgets_frame.columnconfigure(index=0, weight=1)

# Entry
pfont= ("Rosewood Std Regular", 12, "bold", 'underline')
excel_t = ttk.Label(ploc_frame,text="PLOC path:",border=20,font=pfont, borderwidth=5)
excel_t.grid(row=0, column=0, columnspan=4, padx=5, pady=(0, 10), sticky="ew")


excel_i = ttk.Entry(ploc_frame, width=130)
excel_i.insert(0, r"C:\Users\sytung\OneDrive - Synopsys, Inc\Desktop\py\Bump_map.xlsx")
# excel_t.place(x=200, y=20)

excel_i.grid(row=0, column=1, columnspan=4, padx=5, pady=(0, 10), sticky="ew",ipady=10)

sheet_t = ttk.Label(ploc_frame,text="Sheet name:",border=20,font=pfont, borderwidth=3)
sheet_t.grid(row=1, column=0, padx=5, pady=(0, 10), sticky="ew")

sheet_i = ttk.Entry(ploc_frame, background="#217346", width=30)
sheet_i.insert(0, "EMIB_PUB_GPIO")
# sheet_i.place(x=150, y=40, height= 20)
sheet_i.grid(row=1, column=1, padx=5, pady=(0, 10), sticky='w',ipady=10)

sheete_t = ttk.Label(ploc_frame,text="C4 sheet:",border=20,font=pfont, borderwidth=3)

sheete_t.grid(row=1, column=2, padx=5, pady=(0, 10), sticky="e")
sheete_i = ttk.Entry(ploc_frame, background="#217346", width=30)
sheete_i.insert(0, "C4 sheet")

sheete_i.grid(row=1, column=3, padx=5, pady=(0, 10), sticky="w", ipady=10)

# Package type selection

pkg_t = ttk.Label(ploc_frame,text="Package type:",border=20,font=pfont, borderwidth=3)
pkg_t.grid(row=2, column=0, padx=5, pady=(0, 10), sticky="w", ipady=5)
package_combo = ttk.Combobox(ploc_frame, state="readonly", values=package_list, width=27)
package_combo.current(0)
# package_combo.place(x=130, y=105, height= 40)
package_combo.grid(row=2, column=1, padx=5, pady=(0, 10), ipady=5, sticky="w")
package_combo.bind('<<ComboboxSelected>>', choosemode)

xfont = ("System", 12, "bold", 'underline', 'italic')
theme_combo_t = ttk.Label(root,text="Choose theme:",border=20, font=xfont, background='#b434eb', borderwidth=3)

theme_combo_t.grid(row=0, column=0, padx=5, pady=(0, 10),ipady=2,sticky="se")
theme_combo = ttk.Combobox(root, state="readonly", values=theme_list, width=27)
theme_combo.current(0)
# package_combo.place(x=130, y=105, height= 40)
theme_combo.grid(row=0, column=1, padx=5, pady=(0, 10),ipady=2, sticky="sw")
theme_combo.bind('<<ComboboxSelected>>', choosetheme)



# foundry_t = ttk.Label(ploc_frame,text="Foundary:",border=20,font=pfont, borderwidth=3)
# foundry_t.grid(row=2, column=3, padx=5, pady=(0, 10), sticky="ew", ipady=10)
foundry_combo = ttk.Combobox(ploc_frame, state="readonly", values=foundry_list, width=27)
foundry_combo.current(0)
foundry_combo.grid(row=2, column=3,padx=5, pady=(0, 10), sticky="w", ipady=5)
# foundry_combo.bind('<<ComboboxSelected>>', choosemode)
# Checkbuttons
sr_opt = ttk.Checkbutton(ploc_frame, text="For TC", variable=tc_opt,command= entry_toggle)
# sr_opt.place(x=400, y=115, height=50 )
sr_opt.grid(row=2, column=2, padx=5, pady=10, sticky="e")
# Separator
# separator = ttk.Separator(root)
# separator.grid(row=1, column=0, padx=(20, 10), pady=10, sticky="ew")

bumpvisual_frame = ttk.LabelFrame(root, text="Bump map config", padding=(20, 10))
bumpvisual_frame.grid(row=2, column=0, padx=(20, 10), pady=(20, 10), sticky="nsew")

# x1y1_t = ttk.Label(bumpvisual_frame,text="Die start:",border=20, borderwidth=3)
# x1y1_t.grid(row=0, column=0, padx=5, pady=(0, 10), sticky="ew")
x1y1_i = ttk.Entry(bumpvisual_frame)
x1y1_i.insert(0, "C11")

x1y1_i.grid(row=1, column=0, padx=5, pady=(0, 10), sticky="ew")

# x2y2_t = ttk.Label(bumpvisual_frame,text="Die end:",border=20, borderwidth=3)
# x2y2_t.grid(row=0, column=1, padx=5, pady=(0, 10), sticky="ew")
x2y2_i = ttk.Entry(bumpvisual_frame)
x2y2_i.insert(0, "AP55")
x2y2_i.grid(row=1, column=1, padx=5, pady=(0, 10), sticky="ew")

# Xget_t = ttk.Label(bumpvisual_frame,text="Row contains X:",border=20, borderwidth=3)
# Xget_t.grid(row=3, column=0, padx=5, pady=(0, 10), sticky="ew")
Xget_i = ttk.Entry(bumpvisual_frame)
Xget_i.insert(0, "9")
Xget_i.grid(row=2, column=0, padx=5, pady=(0, 10), sticky="ew")

# Yget_t = ttk.Label(bumpvisual_frame,text="Column contains Y:",border=20, borderwidth=3)
# Yget_t.grid(row=3, column=1, padx=5, pady=(0, 10), sticky="ew")
Yget_i = ttk.Entry(bumpvisual_frame)
Yget_i.insert(0, "A")
Yget_i.grid(row=2, column=1, padx=5, pady=(0, 10), sticky="ew")

out_table_frame = ttk.LabelFrame(root, text="Output Bump table config", padding=(20, 10))
out_table_frame.grid(row=2, column=1, padx=(20, 10), pady=(20, 10), sticky="nsew")
out_name = ttk.Label(out_table_frame,text="Bump table name:")
out_name.grid(row=0, column=0, padx=5, pady=(0, 10), sticky="ew")
out_name_in = ttk.Entry(out_table_frame)
out_name_in.insert(0, "Name")
out_name_in.grid(row=0, column=1, padx=5, pady=(0, 10), sticky="ew")

out_col_t = ttk.Label(out_table_frame,text="Out table without SR:")
out_col_t.grid(row=1, column=0, padx=5, pady=(0, 10), sticky="ew")
out_col_i = ttk.Entry(out_table_frame)
out_col_i.insert(0, "O64")
out_col_i.grid(row=1, column=1, padx=5, pady=(0, 10), sticky="ew")

out_col_wsr_t = ttk.Label(out_table_frame,text="Out table with SR:")
out_col_wsr_t.grid(row=2, column=0, padx=5, pady=(0, 10), sticky="ew")
out_col_wsr_i = ttk.Entry(out_table_frame)
out_col_wsr_i.insert(0, "T64")
out_col_wsr_i.grid(row=2, column=1, padx=5, pady=(0, 10), sticky="ew")

# ---------------------------------------gui for EMIB-------------------------------------------------
separator1 = ttk.Separator(bumpvisual_frame)
separator1.grid(row=3, column=0, padx=(0, 10), pady=10, sticky="ew")
separator2 = ttk.Separator(bumpvisual_frame)
separator2.grid(row=3, column=1, padx=(0, 10), pady=10, sticky="ew")

emib_t = ttk.Label(bumpvisual_frame,text="EMIB:",border=20, borderwidth=3)
emib_t.grid(row=4, column=0, padx=5, pady=(0, 10), sticky="ew")
c4_x1y1_i = ttk.Entry(bumpvisual_frame)
c4_x1y1_i.insert(0, "C4 window top-left")
c4_x1y1_i.grid(row=5, column=0, padx=5, pady=(0, 10), sticky="ew")

# x2y2_t = ttk.Label(bumpvisual_frame,text="Die end:",border=20, borderwidth=3)
# x2y2_t.grid(row=0, column=1, padx=5, pady=(0, 10), sticky="ew")
c4_x2y2_i = ttk.Entry(bumpvisual_frame)
c4_x2y2_i.insert(0, "C4 window bot-right")
c4_x2y2_i.grid(row=5, column=1, padx=5, pady=(0, 10), sticky="ew")

# Xget_t = ttk.Label(bumpvisual_frame,text="Row contains X:",border=20, borderwidth=3)
# Xget_t.grid(row=3, column=0, padx=5, pady=(0, 10), sticky="ew")
c4_Xget_i = ttk.Entry(bumpvisual_frame)
c4_Xget_i.insert(0, "Row contains C4 X value")
c4_Xget_i.grid(row=6, column=0, padx=5, pady=(0, 10), sticky="ew")

# Yget_t = ttk.Label(bumpvisual_frame,text="Column contains Y:",border=20, borderwidth=3)
# Yget_t.grid(row=3, column=1, padx=5, pady=(0, 10), sticky="ew")
c4_Yget_i = ttk.Entry(bumpvisual_frame)
c4_Yget_i.insert(0, "Column contains C4 Y value")
c4_Yget_i.grid(row=6, column=1, padx=5, pady=(0, 10), sticky="ew")
# ------------
u_x1y1_i = ttk.Entry(bumpvisual_frame)
u_x1y1_i.insert(0, "uBump window top-left")
u_x1y1_i.grid(row=7, column=0, padx=5, pady=(0, 10), sticky="ew", ipadx=20)

# x2y2_t = ttk.Label(bumpvisual_frame,text="Die end:",border=20, borderwidth=3)
# x2y2_t.grid(row=0, column=1, padx=5, pady=(0, 10), sticky="ew")
u_x2y2_i = ttk.Entry(bumpvisual_frame)
u_x2y2_i.insert(0, "uBump window bot-right")
u_x2y2_i.grid(row=7, column=1, padx=5, pady=(0, 10), sticky="ew", ipadx=20)

# Xget_t = ttk.Label(bumpvisual_frame,text="Row contains X:",border=20, borderwidth=3)
# Xget_t.grid(row=3, column=0, padx=5, pady=(0, 10), sticky="ew")
u_Xget_i = ttk.Entry(bumpvisual_frame)
u_Xget_i.insert(0, "Row contains uBump X value")
u_Xget_i.grid(row=8, column=0, padx=5, pady=(0, 10), sticky="ew")

# Yget_t = ttk.Label(bumpvisual_frame,text="Column contains Y:",border=20, borderwidth=3)
# Yget_t.grid(row=3, column=1, padx=5, pady=(0, 10), sticky="ew")
u_Yget_i = ttk.Entry(bumpvisual_frame)
u_Yget_i.insert(0, "Column contains uBump Y value")
u_Yget_i.grid(row=8, column=1, padx=5, pady=(0, 10), sticky="ew")
# ------------------------------
separator1 = ttk.Separator(out_table_frame)
separator1.grid(row=3, column=0, padx=(0, 10), pady=10, sticky="ew")
separator2 = ttk.Separator(out_table_frame)
separator2.grid(row=3, column=1, padx=(0, 10), pady=10, sticky="ew")

emib_tb_t = ttk.Label(out_table_frame,text="EMIB:")
emib_tb_t.grid(row=4, column=0, padx=5, pady=(0, 10), sticky="ew")
c4_tb_name = ttk.Entry(out_table_frame)
c4_tb_name.insert(0, "C4 Name")
c4_tb_name.grid(row=5, column=0, padx=5, pady=(0, 10), sticky="ew")

# out_col_t = ttk.Label(out_table_frame,text="Out table location:")
# out_col_t.grid(row=1, column=0, padx=5, pady=(0, 10), sticky="ew")
c4_col = ttk.Entry(out_table_frame)
c4_col.insert(0, "C4 location")
c4_col.grid(row=5, column=1, padx=5, pady=(0, 10), sticky="ew")

u_tb_name = ttk.Entry(out_table_frame)
u_tb_name.insert(0, "uBump Name")
u_tb_name.grid(row=6, column=0, padx=5, pady=(0, 10), sticky="ew")

u_col = ttk.Entry(out_table_frame)
u_col.insert(0, "uBump location")
u_col.grid(row=6, column=1, padx=5, pady=(0, 10), sticky="ew")
# out_col.grid(row=0, column=0, padx=5, pady=(0, 10), sticky="ew")
# out_row = ttk.Entry(bumpvisual_frame)
# out_row.insert(0, "X axis value get")
# out_row.grid(row=0, column=1, padx=5, pady=(0, 10), sticky="ew")

#--------------------------------------------------------------------------------------------------------#
dmbump_frame = ttk.LabelFrame(root, text="Dummy bump config", padding=(20, 10))
dmbump_frame.grid(row=3, column=0, columnspan=2, padx=(20, 10), pady=(20, 10), sticky="nsew")

dmbump_cor1_frame = ttk.LabelFrame(dmbump_frame, text="Corner 1 config", padding=(20, 10))
dmbump_cor1_frame.grid(row=0, column=0,padx=(20, 10), pady=(20, 10), sticky="nsew")

cor1_x1y1 = ttk.Entry(dmbump_cor1_frame, width=33)
cor1_x1y1.insert(0, "window top-left")
cor1_x1y1.grid(row=0, column=0, padx=5, pady=(0, 10), sticky="ew")
cor1_x2y2 = ttk.Entry(dmbump_cor1_frame, width=32)
cor1_x2y2.insert(0, "window bot-right")
cor1_x2y2.grid(row=0, column=1, padx=5, pady=(0, 10), sticky="ew")

cor1_Xget = ttk.Entry(dmbump_cor1_frame)
cor1_Xget.insert(0, "Row contains X")
cor1_Xget.grid(row=1, column=0, padx=5, pady=(0, 10), sticky="ew")
cor1_Yget = ttk.Entry(dmbump_cor1_frame)
cor1_Yget.insert(0, "Column contains Y")
cor1_Yget.grid(row=1, column=1, padx=5, pady=(0, 10), sticky="ew")
#---------------------------------------------------------------------------------------------------------#

dmbump_cor2_frame = ttk.LabelFrame(dmbump_frame, text="Corner 2 config", padding=(20, 10))
dmbump_cor2_frame.grid(row=0, column=1, padx=(20, 10), pady=(20, 10), sticky="nsew")

cor2_x1y1 = ttk.Entry(dmbump_cor2_frame, width=25)
cor2_x1y1.insert(0, "window top-left")
cor2_x1y1.grid(row=0, column=0, padx=5, pady=(0, 10), sticky="ew")
cor2_x2y2 = ttk.Entry(dmbump_cor2_frame, width=25)
cor2_x2y2.insert(0, "window bot-right")
cor2_x2y2.grid(row=0, column=1, padx=5, pady=(0, 10), sticky="ew")

cor2_Xget = ttk.Entry(dmbump_cor2_frame)
cor2_Xget.insert(0, "Row contains X")
cor2_Xget.grid(row=1, column=0, padx=5, pady=(0, 10), sticky="ew")
cor2_Yget = ttk.Entry(dmbump_cor2_frame)
cor2_Yget.insert(0, "Column contains Y")
cor2_Yget.grid(row=1, column=1, padx=5, pady=(0, 10), sticky="ew")

#--------------------------------------------------------------------------------------------------------#
dmbump_cor3_frame = ttk.LabelFrame(dmbump_frame, text="Corner 3 config", padding=(20, 10))
dmbump_cor3_frame.grid(row=1, column=0, padx=(20, 10), pady=(20, 10), sticky="nsew")

cor3_x1y1 = ttk.Entry(dmbump_cor3_frame, width=33)
cor3_x1y1.insert(0, "window top-left")
cor3_x1y1.grid(row=0, column=0, padx=5, pady=(0, 10), sticky="ew")
cor3_x2y2 = ttk.Entry(dmbump_cor3_frame, width=32)
cor3_x2y2.insert(0, "window bot-right")
cor3_x2y2.grid(row=0, column=1, padx=5, pady=(0, 10), sticky="ew")

cor3_Xget = ttk.Entry(dmbump_cor3_frame)
cor3_Xget.insert(0, "Row contains X")
cor3_Xget.grid(row=1, column=0, padx=5, pady=(0, 10), sticky="ew")
cor3_Yget = ttk.Entry(dmbump_cor3_frame)
cor3_Yget.insert(0, "Column contains Y")
cor3_Yget.grid(row=1, column=1, padx=5, pady=(0, 10), sticky="ew")
#--------------------------------------------------------------------------------------------------------#
dmbump_cor4_frame = ttk.LabelFrame(dmbump_frame, text="Corner 4 config", padding=(20, 10))
dmbump_cor4_frame.grid(row=1, column=1, padx=(20, 10), pady=(20, 10), sticky="nsew")

cor4_x1y1 = ttk.Entry(dmbump_cor4_frame, width=25)
cor4_x1y1.insert(0, "window top-left")
cor4_x1y1.grid(row=0, column=0, padx=5, pady=(0, 10), sticky="ew")
cor4_x2y2 = ttk.Entry(dmbump_cor4_frame, width=25)
cor4_x2y2.insert(0, "window bot-right")
cor4_x2y2.grid(row=0, column=1, padx=5, pady=(0, 10), sticky="ew")

cor4_Xget = ttk.Entry(dmbump_cor4_frame)
cor4_Xget.insert(0, "Row contains X")
cor4_Xget.grid(row=1, column=0, padx=5, pady=(0, 10), sticky="ew")
cor4_Yget = ttk.Entry(dmbump_cor4_frame)
cor4_Yget.insert(0, "Column contains Y")
cor4_Yget.grid(row=1, column=1, padx=5, pady=(0, 10), sticky="ew")

#--------------------------------------------------------------------------------------------------------#

# Combobox
# combobox = ttk.Combobox(ploc_frame, values=combo_list)
# combobox.current(0)
# combobox.grid(row=2, column=0, padx=5, pady=10,  sticky="ew")


# ploc_frame = ttk.LabelFrame(root, text="Checkbuttons")
# ploc_frame.grid(row=0, column=0, padx=(20, 10), pady=(20, 10), sticky="nsew")
# # ploc_frame.pack()

# e1= Entry(ploc_frame, width=100, borderwidth=4,justify="left")
# # e1.pack()




#---------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------#
def popup(notif):
    messagebox.showinfo("Notification", notif)
#---------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------#
def get_path():
    # popup("Generating...")
    # button['state'] = tk.DISABLED
    mynotif("Processing the input parameter...")
    button['text']="Generating..."
    progress_bar(20)
    

    global excel_path 
 
    excel_path = excel_i.get()
    excel_sheet = sheet_i.get()

    corner1_w1 = cor1_x1y1.get()
    corner1_w2 = cor1_x2y2.get()
    corner1_Xget = cor1_Xget.get()
    corner1_Yget = cor1_Yget.get()

    corner2_w1 = cor2_x1y1.get()
    corner2_w2 = cor2_x2y2.get()
    corner2_Xget = cor2_Xget.get()
    corner2_Yget = cor2_Yget.get()

    corner3_w1 = cor3_x1y1.get()
    corner3_w2 = cor3_x2y2.get()
    corner3_Xget = cor3_Xget.get()
    corner3_Yget = cor3_Yget.get()

    corner4_w1 = cor4_x1y1.get()
    corner4_w2 = cor4_x2y2.get()
    corner4_Xget = cor4_Xget.get()
    corner4_Yget = cor4_Yget.get()

    die_x1y1 = x1y1_i.get()
    die_x2y2 = x2y2_i.get()
    die_x_get = Xget_i.get()
    die_y_get = Yget_i.get()

    die_params=[]
    die_params.append(die_x1y1)
    die_params.append(die_x2y2)
    die_params.append(die_x_get)
    die_params.append(die_y_get)

    out_table_params=[]
    out_table_params.append(out_name_in.get())
    out_table_params.append(out_col_i.get())

    dummy_params=[]
    dummy_params.append(corner1_w1)
    dummy_params.append(corner1_w2)
    dummy_params.append(corner1_Xget)
    dummy_params.append(corner1_Yget)
    dummy_params.append(corner2_w1)
    dummy_params.append(corner2_w2)
    dummy_params.append(corner2_Xget)
    dummy_params.append(corner2_Yget)
    dummy_params.append(corner3_w1)
    dummy_params.append(corner3_w2)
    dummy_params.append(corner3_Xget)
    dummy_params.append(corner3_Yget)
    dummy_params.append(corner4_w1)
    dummy_params.append(corner4_w2)
    dummy_params.append(corner4_Xget)
    dummy_params.append(corner4_Yget)
  
    print(dummy_params[0])
    # print(excel_path)

    # print(corner1_w1, corner1_w2, corner1_Xget,corner1_Yget)
    # print(corner2_w1, corner2_w2, corner2_Xget,corner2_Yget)
    # print(corner3_w1, corner3_w2, corner3_Xget,corner3_Yget)
    # print(corner4_w1, corner4_w2, corner4_Xget,corner4_Yget)

    package = package_combo.get()
    print(package)
    if (package == "A-CoWoS"):
        print(1)
        package_type = 1
    elif(package == "S-Organic"):
        # dmbump_frame.state = tk.DISABLED
        package_type = 0
    else:
        package_type = 0

    generate_bump_table(excel_path, excel_sheet, package_type, out_table_params, die_params, dummy_params )
    button['text']="Generate"
    



# nofi = ttk.Entry(root,)
def generate_bump_table(excel_path, excel_sheet, package_type, out_table_params, die_params, dummy_params ):

   

# Bump table config 
    table={
        "name": out_table_params[0],
        "location": out_table_params[1]
        
    }

    #---Bump map visual view parameter---#
    coordinate = {
        
        "window1": die_params[0], #Top Left of Bump map visual view
        "window2": die_params[1], #Bottom Right of Bump map visual view
        "xcoor": die_params[2], #This define row where Xaxis value can be got
        "ycoor": die_params[3] #This define row where Yaxis value can be got
    }

    #---Dummy Bump visual view parameter---#
    dummybump={
        "corner_1":{
            "window1": dummy_params[0],
            "window2": dummy_params[1],
            "xcoor": dummy_params[2],
            "ycoor": dummy_params[3]
            },
        "corner_2":{
          
            "window1": dummy_params[4],
            "window2": dummy_params[5],
            "xcoor": dummy_params[6],
            "ycoor": dummy_params[7]
        },
        "corner_3":{
         
            "window1": dummy_params[8],
            "window2": dummy_params[9],
            "xcoor": dummy_params[10],
            "ycoor": dummy_params[11]
        },
        "corner_4":{         
            "window1": dummy_params[12],
            "window2": dummy_params[13],
            "xcoor": dummy_params[14],
            "ycoor": dummy_params[15]
        }

    }


    

    # wb = load_workbook('Test.xlsx')
    # ws = wb.active


    wb = load_workbook(excel_path)

    print(wb)   
    # ws = wb.create_sheet('Tung')
    ws1 = wb[excel_sheet]


    #----- Create table from bump map-----------#
    tb_x = coordinate_to_tuple(table['location'])[1]
    tb_y = coordinate_to_tuple(table['location'])[0]

    r = tb_y + 2

    ws1[table['location']].value = table['name']
    # ws1.merge_cells(table['xcol'] + str(table['begin']) + ":" + table['bumpcol'] + str(table['begin']))
    # print(table['xcol'] + str(table['begin']) + ":" + table['bumpcol'] + str(table['begin']))
    ws1[get_column_letter(tb_x) + str(tb_y + 1)].value = "X"
    ws1[get_column_letter(tb_x + 1) + str(tb_y + 1)].value = "Y"
    ws1[get_column_letter(tb_x + 2)  + str(str(tb_y + 1))].value = "Bump name"

    #----- Create dummy bump at 4 corner 140x140u for advance package (CoWos)-----------#
    ymin = coordinate_to_tuple(coordinate['window1'])[0]
    xmin = coordinate_to_tuple(coordinate['window1'])[1]
    ymax = coordinate_to_tuple(coordinate['window2'])[0]
    xmax = coordinate_to_tuple(coordinate['window2'])[1]

    print(xmin,xmax)
    print(ymin,ymax)
    progress_bar(60)
    if (package_type == 1):
        dm_bump_coor= []
        dm_cnt=0
        mynotif("Generating Dummy bump...")
        for dm_bump in dummybump:
            bump = list(dummybump[dm_bump].values())
                
            ymin_dm = coordinate_to_tuple(bump[0])[0]
            xmin_dm = coordinate_to_tuple(bump[0])[1]
            ymax_dm = coordinate_to_tuple(bump[1])[0]
            xmax_dm = coordinate_to_tuple(bump[1])[1]
            xcoor_dm = str(bump[2])
            ycoor_dm = str(bump[3])

            print(xmin_dm,xmax_dm)
            print(ymin_dm,ymax_dm)

            for dummycol1 in range(xmin_dm, xmax_dm + 1):
                for dummyrow1 in range(ymin_dm, ymax_dm + 1):
                    col_dm = get_column_letter(dummycol1)
                    if (ws1[col_dm + str(dummyrow1)].value != None):
                        ws1[get_column_letter(tb_x + 2)+str(r)].value =  ws1[col_dm+ str(dummyrow1)].value
                        # print(col_l + " " + str(coordinate['xcoor']))
                        ws1[get_column_letter(tb_x)+str(r)].value = ws1[col_dm + xcoor_dm].value
                        # print(coordinate['ycoor'] + " " + str(dummyrow1)) 
                        ws1[get_column_letter(tb_x + 1)+str(r)].value = ws1[ycoor_dm + str(dummyrow1)].value
                        r = r + 1
                        coor = col_dm + str(dummyrow1)
                        dm_bump_coor.append(coor)
                        dm_cnt += 1

        #---------Create Die bump exclued dummy bump at 4 corner-----------#

        match = 0
        mynotif("Generating Die bump...")
        for col in range(xmin, xmax + 1):
            for row in range(ymin, ymax + 1):       
                col_l = get_column_letter(col)
                #print(col_l)
                i = 0 
                while(i < len(dm_bump_coor)):
                    xy = col_l + str(row)
                    if(xy ==  dm_bump_coor[i]):
                        match = 1
                    else:
                        match = 0
                    if(match == 1):
                        break
                    i += 1
                if (match == 0 and ws1[col_l + str(row)].value != None):
                    ws1[get_column_letter(tb_x + 2)+str(r)].value =  ws1[col_l+ str(row)].value
                    print(col_l + " " + str(coordinate['xcoor']))
                    ws1[get_column_letter(tb_x)+str(r)].value = ws1[col_l + str(coordinate['xcoor'])].value
                    print(coordinate['ycoor'] + " " + str(row)) 
                    ws1[get_column_letter(tb_x + 1)+str(r)].value = ws1[coordinate['ycoor'] + str(row)].value
                    r = r + 1
    else:
        mynotif("Generating Die bump...")
        for col in range(xmin, xmax + 1):
                for row in range(ymin , ymax + 1):       
                    col_l = get_column_letter(col)
                    #print(col_l)
                    if (ws1[col_l + str(row)].value != None):
                        ws1[get_column_letter(tb_x + 2)+str(r)].value =  ws1[col_l+ str(row)].value
                        print(col_l + " " + str(coordinate['xcoor']))
                        ws1[get_column_letter(tb_x)+str(r)].value = ws1[col_l + str(coordinate['xcoor'])].value
                        print(coordinate['ycoor'] + " " + str(row)) 
                        ws1[get_column_letter(tb_x + 1)+str(r)].value = ws1[coordinate['ycoor'] + str(row)].value
                        r = r + 1
    progress_bar(80)    
    wb.save(excel_path)
    progress_bar(100)
    mynotif("Generated")
    popup("PLOC generated successful!!!")
    mynotif("")
    # button['state'] = tk.NORMAL


                
# myButton = tk.Button(root,text="Button", command=get_path)
# myButton.pack()

entry_disable(cor1_x1y1, cor1_x2y2, cor1_Xget, cor1_Yget,
            cor2_x1y1, cor2_x2y2, cor2_Xget, cor2_Yget,
            cor3_x1y1, cor3_x2y2, cor3_Xget, cor3_Yget,
            cor4_x1y1, cor4_x2y2, cor4_Xget, cor4_Yget)
entry_disable(c4_x1y1_i,c4_x2y2_i, c4_Xget_i, c4_Yget_i, c4_tb_name, c4_col,
            u_col, u_tb_name, u_x1y1_i, u_x2y2_i, u_Xget_i, u_Yget_i)
entry_disable(sheete_i, sheete_t)
sheet_t['text']= "Bump sheet:"
mynotif("")
treeScroll = ttk.Scrollbar(root, orient = 'vertical')
treeScroll.grid(row=0, column=2, rowspan=6)
# root.configure(yscrollcomand)
progress = ttk.Progressbar(root, orient = 'horizontal',
              length = 100, mode = 'determinate')
progress.grid(row=6,  column=0,columnspan=2, padx=5, pady=30, sticky="nsew")

# Button
#Create style object
style = ttk.Style()
style.configure('TButton', font =
               ('calibri', 20, 'bold'),
                    borderwidth = '4')

#configure style
# sto.configure('W.TButton', font= ('System', 20, 'underline', 'bold'),foreground='#9900ff', border=50)

button = ttk.Button(root, text="Generate", command=get_path, style= 'TButton')
button.grid(row=7, column=0,columnspan=2, padx=5, pady=30, sticky="nsew")




root.mainloop()

