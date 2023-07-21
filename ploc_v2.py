from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.utils import range_boundaries
from openpyxl.utils import column_index_from_string
#from openpyxl.utils import coordinate_from_string
from openpyxl.utils import coordinate_to_tuple
from openpyxl.worksheet.table import Table, TableStyleInfo
# from tkinter import *
from tkinter import messagebox
from tkinter import ttk
import getcolumn
from array import *
import tkinter as tk
from tkinter import *
from ttkthemes import ThemedTk, THEMES
from PIL import Image
from PIL import ImageTk, Image
from tkinter.font import Font as tkfont
from tkinter import filedialog
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection
import os
import win32com.client
from pathlib import Path  

import tempfile
# adv = 1

# label.pack(padx=40,pady=40)
# Create a style
root = ThemedTk()
# my_canvas=tk.Canvas(root)


root.title("PLOC TABLE GENERATOR")
root.geometry("1000x1000+30+100")
root.resizable(width=False, height=False)
root.iconbitmap(r".\mylogo.ico")
root.option_add("*tearOff", False) # This is always a good idea
# theme_list = ["adapta", "aquativo", "arc", "black","blue", "breeze", "clearlooks", "elegance", "equilux", "itft1", "keramik", "keramik_alt", "kroc", "plastik", "radiance", "ubuntu", "scidblue", "scidgreen", "scidgrey", "scidmint", "scidpink", "scidpurple", "scidsand", "smog", "winxpblue", "yaru" ]

# temp_dir = tempfile.gettempdir()
try:
    temp_file =  os.path.join(tempfile.gettempdir(), ".ploctablebgen_params_saved.txt")
    print(temp_file)
    tmp_flag = 0
except:
    messagebox.showerror("Can not found the User Temp dir")
    tmp_flag = 1
img_path = r".\img\resize1000x1000"
# bg = ImageTk.PhotoImage(file=r".\img\mountain.png")
bgm = PhotoImage(file=img_path + r"\frog.png")
# bg2 = PhotoImage(file = r".\img\resize1000x1000\bee.png")
# bg3 = PhotoImage(file = r".\img\resize1000x1000\owl.png")
# bg4 = PhotoImage(file = r".\img\resize1000x1000\mountain.png")
# bg5 = PhotoImage(file = r".\img\resize1000x1000\whale.png")
# bg6 = PhotoImage(file = r".\img\resize1000x1000\penguin.png")
# bg1 = PhotoImage(file= r".\brain.png").subsample(2,2)
# bg2 = PhotoImage(file= r".\img\braincircuit.png").subsample(3,3)
# bg3 = PhotoImage(file= r".\img\gear.png").subsample(2,2)
# bg4 = PhotoImage(file= r".\img\internet.png").subsample(2,2)
# bg5 = PhotoImage(file= r".\img\rocket.png").zoom(2,2)
open_imag = PhotoImage(file = r".\open-folder.png")
img_list = ["owl.png", "mountain.png","whale2.png", "penguin.png","sunset1.png", "circuit1.png", "fight.png", "pug.png", "penguin.png", "whale2.png", "elephant_grey.png", "snowman.png", "bee4.png", "elephant.png", "bee2.png", "fox.png", "beach.png", "frog.png", "cow.png", "forest.png", "owlpink2.png", "dinosaurs.png", "sand1.png", "green.png", "pig.png", "discord1.png" ]

lable_bg_list = ["#F0F0F0","#EDEDED","#EBECEE","#F0F0F0","#F0F0F0","#FCFCFC","#EFF0F1","#EFF0F1","#EFF0F1","#EAECEF","#EFF0F1","#EFF0F1","#FECDD9","#EFF0F1","#EFF0F1","#EFF0F1","#EFF0F1","#EFF0F1","#EFF0F1","#EFF0F1","#EFF0F1", "#EFF0F1","#EFF0F1", "#EFF0F1","#EFF0F1","#E6EBEF"]

# Define Canvas
my_canvas = tk.Canvas(root, width=1200, height=800, bd=0, highlightthickness=0)
my_canvas.pack(fill="both", expand=True)

# Put the image on the canvas
bg_img = my_canvas.create_image(0,0, image=bgm, anchor="nw")

# Make the app responsive
# root.columnconfigure(index=0, weight=1)
# root.columnconfigure(index=1, weight=1)
# root.columnconfigure(index=2, weight=1)
# root.columnconfigure(index=3, weight=1)

# root.grid_columnconfigure(0, weight=1)
# root.grid_rowconfigure(0, weight=1)
# root.rowconfigure(index=0, weight=1)
# root.rowconfigure(index=1, weight=1)
# root.rowconfigure(index=2, weight=1)
# root.rowconfigure(index=3, weight=1)
# root.rowconfigure(index=4, weight=1)
# root.rowconfigure(index=5, weight=1)
# root.rowconfigure(index=6, weight=1)
# root.rowconfigure(index=7, weight=1)




stfont= ("Franklin Gothic Medium", 10, 'underline', "italic")
# Create lists for the Comboboxes
theme_list = ["adapta", "aquativo", "arc", "black","blue", "breeze", "clearlooks", "elegance", "equilux", "itft1", "keramik", "keramik_alt", "kroc", "plastik", "radiance", "ubuntu", "scidblue", "scidgreen", "scidgrey", "scidmint", "scidpink", "scidpurple", "scidsand", "smog", "winxpblue", "yaru" ]
colour_list = ["#09a5e8", "#292b33", "#1583eb", "#292a2b","#1a7cad", "#0664bd", "#8baac7", "#59564f", "#40454a", "#7aa7f5", "#7795b4", "#7795b4", "#ebab0c", "#0c99eb", "#eb830c", "#eb830c", "#0937ab", "#37ed80", "#707371", "#479403", "#d12a9f", "#9b34eb", "#787122", "#118cbd", "#505257", "#924d8b" ]
package_list = ["S-Organic", "A-CoWoS", "A-EMIB"]
foundry_list = ["TSMC-MapWSR", "TSMC-MapWoSR", "SS-MapWSR", "SS-MapWoSR", "GF-MapWSR", "GF-MapWSR"]
int_couple_number = ["2", "4", "6", "8", "10", "12", "14", "16"]

# Create control variables
a = tk.BooleanVar()
b = tk.BooleanVar(value=True)
c = tk.BooleanVar()
d = tk.IntVar(value=2)
# e = tk.StringVar(value=option_menu_list[1])
f = tk.BooleanVar()
g = tk.DoubleVar(value=75.0)
h = tk.BooleanVar()
tc_opt = tk.IntVar()
isIntp = tk.IntVar()

#Define a Function to enable the frame
def round_rectangle(x1, y1, x2, y2, radius=25, **kwargs):
        
    points = [x1+radius, y1,
              x1+radius, y1,
              x2-radius, y1,
              x2-radius, y1,
              x2, y1,
              x2, y1+radius,
              x2, y1+radius,
              x2, y2-radius,
              x2, y2-radius,
              x2, y2,
              x2-radius, y2,
              x2-radius, y2,
              x1+radius, y2,
              x1+radius, y2,
              x1, y2,
              x1, y2-radius,
              x1, y2-radius,
              x1, y1+radius,
              x1, y1+radius,
              x1, y1]

    return my_canvas.create_polygon(points, **kwargs, smooth=True)
def change_colour(index):
    listchange =[bum_visual_t,die_tb_out_t,die_dumy_t,int_size_t,int_s_loc_t,int_die_cnt_t,int_die_name_t,int_xo_t,int_yo_t]
    entry_list = [button,sheet_t, theme_combo_t,excel_t, pkg_t]
    #  sheet_t, theme_combo_t,excel_t, pkg_t
    for t in listchange:
        my_canvas.itemconfig(t, fill = colour_list[index])
    for l in entry_list:
        l.config(background = colour_list[index])
    
    text.configure(foreground=colour_list[index],bg=lable_bg_list[index], highlightbackground=colour_list[index])

    global bgm
    p = os.path.join(img_path, img_list[index])
    bgm = PhotoImage(file = p)
    my_canvas.itemconfigure(bg_img, image=bgm)

def enable(children):
   for child in children:
      child.configure(state='enable')

def disable(children):  
    for child in children:
        child.configure(state='disable')

def entry_disable(*entries):
    for entry in entries:
        entry.config(state='disable')

def entry_enable(*entries):
    for entry in entries:
        entry.config(state='normal')
def entry_toggle():
    
        # if(entry['state'] == 'disable'):
    if(tc_opt.get() == 1):
        srw_i.config(state='normal')
        entry_enable(out_name2_in)
        print("Gen table without sealring: ON")
        entry_enable(text)
        text_delete()
        mynotif("Gen table without sealring: ON. Please define sealring width at the next to entry")
        entry_disable(text)
    elif(tc_opt.get() == 0):
        srw_i.config(state='disable')
        entry_disable(out_name2_in)
        print("Gen table without sealring: OFF")
        entry_enable(text)
        text_delete()
        mynotif("Gen table without sealring: OFF")
        entry_disable(text)
def intp_toggle():
    if(isIntp.get() == 1):
        entry_enable(xwidth_i, yheight_i, Die1_xoffset_i, Die1_yoffset_i, Die2_xoffset_i, Die2_yoffset_i, intp_sheet, Die1_name, Die2_name, int_tb_loc, int_die_num_combo)
        print("Gen interposer Die table: ON")
        entry_enable(text)
        text_delete()
        mynotif("Gen interposer Die table: ON")
        entry_disable(text)
    elif(isIntp.get() == 0):
        entry_disable(xwidth_i, yheight_i, Die1_xoffset_i, Die1_yoffset_i, Die2_xoffset_i, Die2_yoffset_i, intp_sheet, Die1_name, Die2_name, int_tb_loc, int_die_num_combo)
        print("Gen interposer Die table: OFF")
        entry_enable(text)
        text_delete()
        mynotif("Gen interposer Die table: OFF")
        entry_disable(text)
def progress_bar(value):
    progress['value'] = value
    root.update_idletasks()

def choosetheme(event):
    root.set_theme(theme_combo.get())
    change_colour(theme_list.index(theme_combo.get())) 

           
def choosemode(event):  
    entry_enable(text) 
    if(package_combo.get() == "S-Organic"):
       
       entry_disable(cor1_x1y1, cor1_x2y2, cor1_Xget, cor1_Yget,
                     cor2_x1y1, cor2_x2y2, cor2_Xget, cor2_Yget,
                     cor3_x1y1, cor3_x2y2, cor3_Xget, cor3_Yget,
                     cor4_x1y1, cor4_x2y2, cor4_Xget, cor4_Yget)
      
       entry_enable(x1y1_i, x2y2_i, Xget_i, Yget_i)
       entry_enable(out_name_in, out_col_i)
       entry_disable(sheete_i, sheete_t)
       sheet_t['text']= "Bump sheet:"
        
       print("Package:" + package_combo.get())
       text_delete()
       mynotif("Package used:" + package_combo.get())
    elif(package_combo.get() == "A-CoWoS"):
        entry_enable(cor1_x1y1, cor1_x2y2, cor1_Xget, cor1_Yget,
                     cor2_x1y1, cor2_x2y2, cor2_Xget, cor2_Yget,
                     cor3_x1y1, cor3_x2y2, cor3_Xget, cor3_Yget,
                     cor4_x1y1, cor4_x2y2, cor4_Xget, cor4_Yget)
        entry_enable(x1y1_i, x2y2_i, Xget_i, Yget_i)
        entry_disable(sheete_i, sheete_t)
        entry_enable(out_name_in, out_col_i)
        
        sheet_t['text']= "Bump sheet:"
        print("Package:" + package_combo.get())
        text_delete()
        mynotif("Package used:" + package_combo.get())
    else:
        entry_disable(cor1_x1y1, cor1_x2y2, cor1_Xget, cor1_Yget,
                     cor2_x1y1, cor2_x2y2, cor2_Xget, cor2_Yget,
                     cor3_x1y1, cor3_x2y2, cor3_Xget, cor3_Yget,
                     cor4_x1y1, cor4_x2y2, cor4_Xget, cor4_Yget)
        entry_disable(x1y1_i, x2y2_i, Xget_i, Yget_i)
        entry_disable(out_name_in, out_col_i)
        
        entry_enable(sheete_i, sheete_t)
        sheet_t['text']= "uBump sheet:"

        popup("The EMIB package type have not developed yet, Please use S-Organic to gen 2 times (for C4 and uBump) instead!")
    entry_disable(text)
    
myLabel = ttk.Label(root,text="---")
myLabel_w =my_canvas.create_window(80,770,anchor="nw", window=myLabel)


def get_num_intdie(event):
    pass
def text_delete():
   text.delete("1.0","end")
def mynotif(content):
    text.insert(tk.END,content+"\n")
    text.see("end")
   
def process_notify(content): 
    root.update_idletasks()
    mynotif(content)
    root.update_idletasks()

def myguide(entries, content):
     entries.insert(tk.END, content)
def handle_click(event):
   pass

x1y1_guide = [
    "INFO: Die window begin cell\n\n ",
    "      * Example:   A0           "
]
x2y2_guide =  [
    "INFO: Die window end cell\n\n ",
    "      * Example:   CU100       "
]
Xget_guide = [
    "INFO: Row contains X axis value which is X location of Bump. \n",
    "      - Must be interger   \n\n",
    "      * Example:   8       "
]
Yget_guide = [
    "INFO: Row contains Y axis value which is Y location of Bump.\n", 
    "      - Must be Excel column format\n\n",
    "      * Example: CU " 
]
outtb_s_guide = [
    "INFO: Sheet to put Die table\n\n ",
    "      * Example: Bump coordination "
]
out_name_in_guide = [
    "INFO: This field to define the output table name\n\n    ",
    "      * Example: DieX "
]
out_name2_in_guide = [
    "INFO: This field to define the output table name for bump without sealring.\n",
    "      - This field will be used when TC option is turned on\n\n",
    "      * Example: DieX without sealring "
]
out_col_in_guide = [
    "INFO: This field to define the first output table location. \n",
    "      - The next tables placed away 2 column from previous table \n\n ",
    "      * Example: O64 "
]
out_col_wsr_i_guide = [
    "INFO: This field to define the first output table location. \n\n",
    "      * Example: O64 "
]
dummystart_guide = [
    "INFO: Dummy bump window begin cell\n\n ",
    "      * Example:   A0           "
]
dummyend_guide = [
    "INFO: Dummy bump window end cell\n\n ",
    "      * Example:   E3           "
]
dummy_Xget_guide = [
    "INFO: Row contains X axis value which is X location of Bump. \n",
    "      - Must be interger   \n\n",
    "      * Example:   8       "
]
dummy_Yget_guide = [
    "INFO: Row contains Y axis value which is Y location of Bump.\n",
    "      - Must be Excel column format\n\n",
    "      * Example: CU "
]
xwidth_i_guide = [
    "INFO: Width of Die/chip. \n\n",
    "      - This param used for Flip, Rotate die/chip to put on PKG  "
]
yheight_i_guide = [
    "INFO: Height of Die/chip. \n\n",
    "      - This param used for Flip, Rotate die/chip to put on PKG  "
]
Die1_xoffset_i_guide = [
    "INFO: List X Offset of Die Left/Up(Chip Left/Up). \n ",
    "      - This param used for Die/chip placement on PKG  "
]
Die1_yoffset_i_guide = [
    "INFO: List Y Offset of Die Left/Up(Chip Left/Up). \n ",
    "      - This param used for Die/chip placement on PKG  "
]
Die2_xoffset_i_guide = [
    "INFO: List X Offset of Die Right/Down(Chip Right/Down). \n",
    "      - This param used for Die/chip placement on PKG  "
]
Die2_yoffset_i_guide = [
    "INFO: List Y Offset of Die Right/Down(Chip Right/Down). \n",
    "      - This param used for Die/chip placement on PKG  "
]
intp_sheet_guide = [
    "INFO: Name of interposer sheet to put interposer Die table \n\n ",
    "      * Example: Packge_substrates "
]
Die1_name_guide = [
    "INFO: List Name of interposer Die Left/Up which is outcome of Die Flipped then Rotate -90\n ",
    "      - The dies name are separated by spaces.\n ",
    "NOTE: The Die name is mapping between Die Left/Up and Die Right/Down. \n",
    "For example:\n",
    "       Die Right list name: DIE5 DIE6 DIE7 DIE8, and\n",
    "       Die Left list name: DIE1 DIE2 DIE3 DIE4, and\n",
    "               (DIE1 <=> DIE5) \n",
    "               (DIE2 <=> DIE6) \n",
    "               (DIE3 <=> DIE7) \n",
    "               (DIE4 <=> DIE7)"
]
Die2_name_guide = [
    "INFO: List Name of interposer Die Right/Down which is outcome of Die Flipped then Rotate +90\n ",
    "      - The dies name are separated by spaces.\n ",
    "NOTE: The Die name is mapping between Die Left/Up and Die Right/Down. \n",
    "For example:\n",
    "       Die Left list name: DIE1 DIE2 DIE3 DIE4, and\n",
    "       Die Right list name: DIE5 DIE6 DIE7 DIE8, and\n",
    "               (DIE1 <=> DIE5) \n",
    "               (DIE2 <=> DIE6) \n",
    "               (DIE3 <=> DIE7) \n",
    "               (DIE4 <=> DIE7)"
]
int_tb_guide = [
    "INFO: This field to define the first output table cell. \n ",
    "       - The next tables placed away 1 column from previous table \n\n ",
    "       * Example: O64 "
]
srw_i_guide = [
    "INFO: This field to define the width of sealring.\n\n",
    "Note: Normally, TSMC is 21.6, SS/GF is 14.04 \n\n"
]
vssheet_gui = [
    "INFO: Name of bump visual sheet to generate Bump coordinate table\n\n ",
    "      * Example: N3P_CoWoS"
]

def guide(gui_list):
    entry_enable(text)
    text_delete()
    for gui in gui_list:
       myguide(text,gui)
    entry_disable(text)

        
xfont = ("System", 12, "bold", 'underline', 'italic')
theme_combo_t = ttk.Label(root,text="Choose theme:",border=20, font=xfont, background='#b434eb', borderwidth=3)
theme_combo_t_w = my_canvas.create_window(750, 15, window=theme_combo_t)

theme_combo = ttk.Combobox(root, state="readonly", values=theme_list, width=15)
theme_combo_w = my_canvas.create_window(870,15, window=theme_combo)
theme_combo.current(0)
theme_combo.bind('<<ComboboxSelected>>', choosetheme)

# -------------------------excelpath input--------------------------#
pfont= ("Rosewood Std Regular", 12, "bold", 'underline' )
excel_t = ttk.Label(root,text="PLOC path:",border=20,font=pfont, borderwidth=5)
excel_t_w = my_canvas.create_window(30,40, anchor="nw", window=excel_t)
excel_i = ttk.Entry(root, width=115)

excel_i_w = my_canvas.create_window(150,40, anchor="nw", window=excel_i)

# -------------------------excel sheet_name input--------------------------#
sheet_t = ttk.Label(root,text="Sheet name:",border=20,font=pfont, borderwidth=3)
sheet_t_w = my_canvas.create_window(30,80, anchor="nw", window=sheet_t)
sheet_i = ttk.Entry(root, background="#217346", width=20)

sheet_i_w = my_canvas.create_window(150,80, anchor="nw", window=sheet_i)
sheet_i.bind('<FocusIn>', lambda event: guide(vssheet_gui))
# sheet_i.bind('<FocusOut>', un_guide)
# -------------------------excel sheet_name input--------------------------#
sheete_t = ttk.Label(root,text="C4 sheet:",border=20,font=pfont, borderwidth=3)
sheete_t_w = my_canvas.create_window(300,80, anchor="nw", window=sheete_t)
sheete_i = ttk.Entry(root, background="#217346", width=20)

sheete_i_w = my_canvas.create_window(400,80, anchor="nw", window=sheete_i)

# -------------------------pkg type input--------------------------#
pkg_t = ttk.Label(root,text="Package type:",border=20,font=pfont, borderwidth=3)
pkg_t_w = my_canvas.create_window(30,120, anchor="nw", window=pkg_t)
package_combo = ttk.Combobox(root, state="readonly", values=package_list, width=17)
package_combo_w = my_canvas.create_window(150,120, anchor="nw", window=package_combo)

package_combo.bind('<<ComboboxSelected>>', choosemode)
# -------------------------sealring option input--------------------------#
sr_opt = ttk.Checkbutton(root, text="For TC", variable=tc_opt,command= entry_toggle)
sr_opt_w =my_canvas.create_window(300, 120, anchor="nw", window=sr_opt)




# -------------------------foundary selection --------------------------#

srw_i = ttk.Entry(root, width=20)
my_canvas.create_window(400, 120, anchor="nw", window=srw_i)

srw_i.bind('<FocusIn>', lambda event: guide(srw_i_guide))
# srw_i.bind('<FocusOut>', un_guide)

# Separator
separator = ttk.Separator(root)
separator_w = my_canvas.create_window(30, 130, anchor="nw", window=separator)


bum_visual_t = my_canvas.create_text(30, 200, text="Die bump map visual input:", anchor="nw",font=("Helvetica", 10, 'italic', 'underline', 'bold'), fill="#b434eb")



# ------------------------Die bump visual parameters input --------------------------#
x1y1_i = ttk.Entry(root, width=20)
x1y1_i_w = my_canvas.create_window(150, 230, anchor="nw", window=x1y1_i)
x1y1_i.bind('<FocusIn>', lambda event: guide(x1y1_guide))
# x1y1_i.bind('<FocusOut>', un_guide)


x2y2_i = ttk.Entry(root, width=20)
my_canvas.create_window(300, 230, anchor="nw", window=x2y2_i)

x2y2_i.bind('<FocusIn>', lambda event: guide(x2y2_guide))
# x2y2_i.bind('<FocusOut>', un_guide)

Xget_i = ttk.Entry(root, width=20)
Xget_i_w = my_canvas.create_window(150, 270, anchor="nw", window=Xget_i)

Xget_i.bind('<FocusIn>', lambda event: guide(Xget_guide))
# Xget_i.bind('<FocusOut>', un_guide)

Yget_i = ttk.Entry(root, width=20)
Yget_i_w = my_canvas.create_window(300, 270, anchor="nw", window=Yget_i)

Yget_i.bind('<FocusIn>', lambda event: guide(Yget_guide))
# Yget_i.bind('<FocusOut>', un_guide)

# ------------------------Output table configure --------------------------#
die_tb_out_t = my_canvas.create_text(500, 200, text="Die table out \nconfig:", anchor="nw",font=("Helvetica", 10, 'italic', 'underline', 'bold'), fill="#b434eb")
# my_canvas.create_text(680, 200, text="Sheet:", anchor="nw",font=("Helvetica", 10, 'italic', 'underline', 'bold'), fill="#b434eb")

out_tb_sheet = ttk.Entry(root)
out_tb_sheet_w = my_canvas.create_window(600, 195, anchor="nw", window=out_tb_sheet, width=275)

out_tb_sheet.bind('<FocusIn>', lambda event: guide(outtb_s_guide))
# out_tb_sheet.bind('<FocusOut>', un_guide)

out_name_in = ttk.Entry(root, width=20)
out_name_in_w = my_canvas.create_window(600, 230, anchor="nw", window=out_name_in)

out_name_in.bind('<FocusIn>', lambda event: guide(out_name_in_guide))
# out_name_in.bind('<FocusOut>', un_guide)
out_name2_in = ttk.Entry(root, width=20)
out_name2_in_w = my_canvas.create_window(750, 230, anchor="nw", window=out_name2_in)

out_name2_in.bind('<FocusIn>', lambda event: guide(out_name2_in_guide))
# out_name2_in.bind('<FocusOut>', un_guide)


out_col_i = ttk.Entry(root, width=20)
out_col_i_w = my_canvas.create_window(600, 270, anchor="nw", window=out_col_i)

out_col_i.bind('<FocusIn>', lambda event: guide(out_col_in_guide))
# out_col_i.bind('<FocusOut>', un_guide)



out_col_wsr_i = ttk.Entry(root)
out_col_wsr_w = my_canvas.create_window(750, 270, anchor="nw", window=out_col_wsr_i)

out_col_wsr_i.bind('<FocusIn>', lambda event: guide(out_col_wsr_i_guide))
# out_col_wsr_i.bind('<FocusOut>', un_guide)


#------------------------------------Dummybup at 4 corners for Advance package-----------------------------------------------------#

die_dumy_t = my_canvas.create_text(30, 310, text="Die Bummy bump input:", anchor="nw",font=("Helvetica", 10, 'italic', 'underline', 'bold'), fill="#b434eb")

my_canvas.create_text(245, 330, text="Corner 1 config", anchor="nw",font=("Helvetica", 10, 'italic', 'underline', 'bold'), fill="black")
cor1_x1y1 = ttk.Entry(root, width=20)
cor1_x1y1_w = my_canvas.create_window(150, 350, anchor="nw", window=cor1_x1y1)

cor1_x1y1.bind('<FocusIn>', lambda event: guide(dummystart_guide))
# cor1_x1y1.bind('<FocusOut>', un_guide)

cor1_x2y2 = ttk.Entry(root, width=20)
cor1_x2y2_w = my_canvas.create_window(300, 350, anchor="nw", window=cor1_x2y2)

cor1_x2y2.bind('<FocusIn>', lambda event: guide(dummyend_guide))
# cor1_x2y2.bind('<FocusOut>', un_guide)

cor1_Xget = ttk.Entry(root,width=20)
cor1_Xget_w = my_canvas.create_window(150, 380, anchor="nw", window=cor1_Xget)

cor1_Xget.bind('<FocusIn>', lambda event: guide(dummy_Xget_guide))
# cor1_Xget.bind('<FocusOut>', un_guide)

cor1_Yget = ttk.Entry(root, width=20)
cor1_Yget_w = my_canvas.create_window(300, 380, anchor="nw", window=cor1_Yget)

cor1_Yget.bind('<FocusIn>', lambda event: guide(dummy_Yget_guide))
# cor1_Yget.bind('<FocusOut>', un_guide)
#---------------------------------------------------------------------------------------------------------#

my_canvas.create_text(670, 330, text="Corner 2 config", anchor="nw",font=("Helvetica", 10, 'italic', 'underline', 'bold'), fill="black")

cor2_x1y1 = ttk.Entry(root, width=20)
cor2_x1y1_w = my_canvas.create_window(600, 350, anchor="nw", window=cor2_x1y1)

cor2_x1y1.bind('<FocusIn>', lambda event: guide(dummystart_guide))
# cor2_x1y1.bind('<FocusOut>', un_guide)

cor2_x2y2 = ttk.Entry(root, width=20)
cor2_x2y2_w = my_canvas.create_window(750, 350, anchor="nw", window=cor2_x2y2)

cor2_x2y2.bind('<FocusIn>', lambda event: guide(dummyend_guide))
# cor2_x2y2.bind('<FocusOut>', un_guide)

cor2_Xget = ttk.Entry(root, width=20)
cor2_Xget_w = my_canvas.create_window(600, 380, anchor="nw", window=cor2_Xget)

cor2_Xget.bind('<FocusIn>', lambda event: guide(dummy_Xget_guide))
# cor2_Xget.bind('<FocusOut>', un_guide)

cor2_Yget = ttk.Entry(root, width=20)
cor2_Yget_w = my_canvas.create_window(750, 380, anchor="nw", window=cor2_Yget)

cor2_Yget.bind('<FocusIn>', lambda event: guide(dummy_Yget_guide))
# cor2_Yget.bind('<FocusOut>', un_guide)

#--------------------------------------------------------------------------------------------------------#

my_canvas.create_text(245, 410, text="Corner 3 config", anchor="nw",font=("Helvetica", 10, 'italic', 'underline', 'bold'), fill="black")
cor3_x1y1 = ttk.Entry(root, width=20)
cor3_x1y1_w = my_canvas.create_window(150, 430, anchor="nw", window=cor3_x1y1)

cor3_x1y1.bind('<FocusIn>', lambda event: guide(dummystart_guide))
# cor3_x1y1.bind('<FocusOut>', un_guide)

cor3_x2y2 = ttk.Entry(root, width=20)
cor3_x2y2_w = my_canvas.create_window(300, 430, anchor="nw", window=cor3_x2y2)

cor3_x2y2.bind('<FocusIn>', lambda event: guide(dummyend_guide))
# cor3_x2y2.bind('<FocusOut>', un_guide)

cor3_Xget = ttk.Entry(root)
cor3_Xget_w = my_canvas.create_window(150, 460, anchor="nw", window=cor3_Xget)

cor3_Xget.bind('<FocusIn>', lambda event: guide(dummy_Xget_guide))
# cor3_Xget.bind('<FocusOut>', un_guide)

cor3_Yget = ttk.Entry(root)
cor3_Yget_w = my_canvas.create_window(300, 460, anchor="nw", window=cor3_Yget)

cor3_Yget.bind('<FocusIn>', lambda event: guide(dummy_Yget_guide))
# cor3_Yget.bind('<FocusOut>', un_guide)

#--------------------------------------------------------------------------------------------------------#

my_canvas.create_text(670, 410, text="Corner 4 config", anchor="nw",font=("Helvetica", 10, 'italic', 'underline', 'bold'), fill="black")
cor4_x1y1 = ttk.Entry(root, width=20)
cor4_x1y1_w = my_canvas.create_window(600, 430, anchor="nw", window=cor4_x1y1)

cor4_x1y1.bind('<FocusIn>', lambda event: guide(dummystart_guide))
# cor4_x1y1.bind('<FocusOut>', un_guide)

cor4_x2y2 = ttk.Entry(root, width=20)
cor4_x2y2_w = my_canvas.create_window(750, 430, anchor="nw", window=cor4_x2y2)

cor4_x2y2.bind('<FocusIn>', lambda event: guide(dummyend_guide))
# cor4_x2y2.bind('<FocusOut>', un_guide)

cor4_Xget = ttk.Entry(root, width=20)
cor4_Xget_w = my_canvas.create_window(600, 460, anchor="nw", window=cor4_Xget)

cor4_Xget.bind('<FocusIn>', lambda event: guide(dummy_Xget_guide))
# cor4_Xget.bind('<FocusOut>', un_guide)

cor4_Yget = ttk.Entry(root, width=20)
cor4_Yget_w = my_canvas.create_window(750, 460, anchor="nw", window=cor4_Yget)

cor4_Yget.bind('<FocusIn>', lambda event: guide(dummy_Yget_guide))
# cor4_Yget.bind('<FocusOut>', un_guide)

# ---------------------------------------INTERPOSER DIE-------------------------------------------------
interopser = ttk.Checkbutton(root, text="Interposer Die generator", variable=isIntp, onvalue=1, offvalue=0,command= intp_toggle)
interopser_w =my_canvas.create_window(30, 500, anchor="nw", window=interopser)

int_size_t = my_canvas.create_text(280, 540, text="Die/Chip size input:", anchor="nw",font=("Helvetica", 10, 'italic', 'underline', 'bold'), fill="#b434eb")
xwidth_i = ttk.Entry(root)
xwidth_i_w = my_canvas.create_window(150, 560, anchor="nw", window=xwidth_i, width= 170)

xwidth_i.bind('<FocusIn>', lambda event: guide(xwidth_i_guide))
# xwidth_i.bind('<FocusOut>', un_guide)

yheight_i = ttk.Entry(root, width=20)
yheight_w = my_canvas.create_window(340, 560, anchor="nw", window=yheight_i, width=170)

yheight_i.bind('<FocusIn>', lambda event: guide(yheight_i_guide))
# yheight_i.bind('<FocusOut>', un_guide)


int_s_loc_t = my_canvas.create_text(620, 540, text="OUT DIE sheet/location:", anchor="nw",font=("Helvetica", 10, 'italic', 'underline', 'bold'), fill="#b434eb")

intp_sheet = ttk.Entry(root)
intp_sheet_w = my_canvas.create_window(520, 560, anchor="nw", window=intp_sheet, width=170)

intp_sheet.bind('<FocusIn>', lambda event: guide(intp_sheet_guide))
# intp_sheet.bind('<FocusOut>', un_guide)

int_tb_loc = ttk.Entry(root)
int_tb_locc_w = my_canvas.create_window(710, 560, anchor="nw", window=int_tb_loc, width=170)

int_tb_loc.bind('<FocusIn>', lambda event: guide(int_tb_guide))
# int_tb_loc.bind('<FocusOut>', un_guide)

# my_canvas.create_text(30, 595, text="Die/Chip Offset:", anchor="nw",font=("Helvetica", 10, 'italic', 'underline', 'bold'), fill="#b434eb")
int_die_cnt_t = my_canvas.create_text(60, 600, text="Die count:", anchor="nw",font=("Helvetica", 10, 'italic', 'underline', 'bold'), fill="#b434eb")

int_die_num_combo = ttk.Combobox(root, state="readonly", values=int_couple_number, width=22)
int_die_num_combo_w = my_canvas.create_window(230,610, window=int_die_num_combo)

int_die_num_combo.bind('<<ComboboxSelected>>', get_num_intdie)

int_die_name_t = my_canvas.create_text(60, 635, text="Die name:", anchor="nw",font=("Helvetica", 10, 'italic', 'underline', 'bold'), fill="#b434eb")
Die1_name = ttk.Entry(root)
Die1_name_w = my_canvas.create_window(150, 630, anchor="nw", window=Die1_name, width= 360)

Die1_name.bind('<FocusIn>', lambda event: guide(Die1_name_guide))
# Die1_name.bind('<FocusOut>', un_guide)

Die2_name = ttk.Entry(root)
Die2_name_w = my_canvas.create_window(520, 630, anchor="nw", window=Die2_name, width=360)

Die2_name.bind('<FocusIn>', lambda event: guide(Die2_name_guide))
# Die2_name.bind('<FocusOut>', un_guide)

int_xo_t = my_canvas.create_text(60, 675, text="X offset:", anchor="nw",font=("Helvetica", 10, 'italic', 'underline', 'bold'), fill="#b434eb")
Die1_xoffset_i = ttk.Entry(root)
Die1_xoffset_w = my_canvas.create_window(150, 670, anchor="nw", window=Die1_xoffset_i, width=360)

Die1_xoffset_i.bind('<FocusIn>', lambda event: guide(Die1_xoffset_i_guide))
# Die1_xoffset_i.bind('<FocusOut>', un_guide)

int_yo_t = my_canvas.create_text(60, 715, text="Y offset:", anchor="nw",font=("Helvetica", 10, 'italic', 'underline', 'bold'), fill="#b434eb")
Die1_yoffset_i = ttk.Entry(root, width=20)
Die1_yoffset_w = my_canvas.create_window(150, 710, anchor="nw", window=Die1_yoffset_i, width=360)

Die1_yoffset_i.bind('<FocusIn>', lambda event: guide(Die1_yoffset_i_guide))
# Die1_yoffset_i.bind('<FocusOut>', un_guide)

Die2_xoffset_i = ttk.Entry(root)
Die2_xoffset_w = my_canvas.create_window(520, 670, anchor="nw", window=Die2_xoffset_i, width=360)

Die2_xoffset_i.bind('<FocusIn>', lambda event: guide(Die2_xoffset_i_guide))
# Die2_xoffset_i.bind('<FocusOut>', un_guide)

Die2_yoffset_i = ttk.Entry(root, width=20)
Die2_yoffset_w = my_canvas.create_window(520, 710, anchor="nw", window=Die2_yoffset_i, width=360)

Die2_yoffset_i.bind('<FocusIn>', lambda event: guide(Die2_yoffset_i_guide))
# Die2_yoffset_i.bind('<FocusOut>', un_guide)


separator1 = ttk.Separator(root)

separator2 = ttk.Separator(root)

# ------------------------------
separator1 = ttk.Separator(root)

separator2 = ttk.Separator(root)


#--------------------------------------------------------------------------------------------------------#

my_canvas.create_text(880,980, text= "Internal contact: sytung@synopsys.com" ,font=("Helvetica", 8, 'underline'), fill="grey")

def open_file():
	# global my_image
    root.filename = filedialog.askopenfilename(initialdir="./", title="Select A File", filetypes=(("Excel files", "*.xlsx"),("all files", "*.*")))
    excel_i.delete(0,END)
    print(root.filename) 
    excel_i.insert(0, root.filename)
	# my_image = ImageTk.PhotoImage(Image.open(root.filename))
	# my_image_label = Label(image=my_image).pack()

#---------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------#
def popup(notif):
    messagebox.showinfo("Notification", notif)

def show_error(error):
    messagebox.showerror("Error", error)
#---------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------#
def getstring(string: str,c1: str, c2: str):
	cell = string
	idx1 = cell.find(c1)
	cell_2 = cell[:idx1] + cell[idx1+1 :]
	idx2 = cell_2.find(c2)
	if(idx2 != -1):
		
		cell_tmpx = cell[idx1+1:idx2+1]
		if(cell_tmpx.find("+") != -1 or cell_tmpx.find("-") != -1 or cell_tmpx.find("*") != -1  or cell_tmpx.find("/") != -1):
			return 1
		else: return 0
	else:
		return 1


def get_params_and_generate():
    # popup("Generating...")
    # button['state'] = tk.DISABLED
    entry_enable(text)
    text_delete()
    
    mynotif("Processing the input parameter...")
    button['text']="Generating..."
    
    progress_bar(20)
    

    global excel_path 
 
    excel_path = excel_i.get()
    bump_visual_sheet = sheet_i.get()
    
    tc_sr={
        "isTC": tc_opt.get(),
        "sr_w": srw_i.get(),
        "sr_tb": out_name2_in.get(),
        "sr_tb_loc": out_col_wsr_i.get()
    }
    # isTC = tc_opt.get()
    # sr_w = srw_i.get()
    int_gen = isIntp.get()

    # bump_visual_params=[]
    # bump_visual_params.append(die_x1y1)
    # bump_visual_params.append(die_x2y2)
    # bump_visual_params.append(die_x_get)
    # bump_visual_params.append(die_y_get)

    
   
    die_table={
        "sheet": out_tb_sheet.get(),
        "name": out_name_in.get(),
        "location": out_col_i.get(),
        "name_wsr": out_name2_in.get(),
        "location_wsr": out_col_wsr_i.get(),
        
    }

    #---Bump map visual view parameter---#
    die_coor = {
        
        "window1": x1y1_i.get(), #Top Left of Bump map visual view
        "window2": x2y2_i.get(), #Bottom Right of Bump map visual view
        "xcoor": Xget_i.get(), #This define row where Xaxis value can be got
        "ycoor": Yget_i.get() #This define row where Yaxis value can be got
    }

    
# Die interposet prarams
    int_die_cnt = int(int_die_num_combo.get())
    die_params={
        "chip_width": xwidth_i.get(),
        "chip_height": yheight_i.get(),
        "die1_xoffset": Die1_xoffset_i.get(),
        "die1_yoffset": Die1_yoffset_i.get(),
        "die2_xoffset": Die2_xoffset_i.get(),
        "die2_yoffset": Die2_yoffset_i.get(),
    }

    int_die_tb={
        "sheet": intp_sheet.get(),
        "Die1_name": Die1_name.get(),
        "int_tb_location": int_tb_loc.get(),
        "Die2_name": Die2_name.get(),

    }
   
    package = package_combo.get()
    dummybump={
            "corner_1":{
                "window1": cor1_x1y1.get(),
                "window2": cor1_x2y2.get(),
                "xcoor": cor1_Xget.get(),
                "ycoor": cor1_Yget.get()
                },
            "corner_2":{
            
                "window1": cor2_x1y1.get(),
                "window2": cor2_x2y2.get(),
                "xcoor": cor2_Xget.get(),
                "ycoor": cor2_Yget.get()
            },
            "corner_3":{
            
                "window1": cor3_x1y1.get(),
                "window2": cor3_x2y2.get(),
                "xcoor": cor3_Xget.get(),
                "ycoor": cor3_Yget.get()
            },
            "corner_4":{         
                "window1": cor4_x1y1.get(),
                "window2": cor4_x2y2.get(),
                "xcoor": cor4_Xget.get(),
                "ycoor": cor4_Yget.get()
            }

        }
    global temp_file, tmp_flag
    if(tmp_flag == 0):
        with open(temp_file,'w') as params_saved:
            params_saved.writelines(excel_path+"\n")
            params_saved.writelines(bump_visual_sheet+"\n")
            params_saved.writelines(package+"\n")
            params_saved.writelines(str(tc_sr["isTC"])+"\n")
            params_saved.writelines(str(tc_sr["sr_w"])+"\n")
            params_saved.writelines(die_coor['window1'] +" "+ die_coor['window2'] +" "+ die_coor['xcoor']+" "+die_coor['ycoor'] + "\n")
            params_saved.writelines(cor1_x1y1.get() +" "+ cor1_x2y2.get() +" "+ cor1_Xget.get() +" "+ cor1_Yget.get() +" "+
                                    cor2_x1y1.get() +" "+ cor2_x2y2.get() +" "+ cor2_Xget.get() +" "+ cor2_Yget.get() +" "+
                                    cor3_x1y1.get() +" "+ cor3_x2y2.get() +" "+ cor3_Xget.get() +" "+ cor3_Yget.get() +" "+
                                    cor4_x1y1.get() +" "+ cor4_x2y2.get() +" "+ cor4_Xget.get() +" "+ cor4_Yget.get()
                                    + "\n")
            params_saved.writelines(die_table['sheet'] +"\n")
            params_saved.writelines(die_table['name'] + "\n")
            params_saved.writelines(die_table['name_wsr']+"\n")
            params_saved.writelines(die_table['location']+"\n")
        
            params_saved.writelines(str(int_gen) + "\n")
            params_saved.writelines(die_params["chip_width"] + " " + die_params['chip_height'] +"\n")
            params_saved.writelines(str(int_die_cnt) + "\n")
        

            params_saved.writelines(int_die_tb['Die1_name'] + "\n")
            params_saved.writelines(int_die_tb['Die2_name'] + "\n")
            
            params_saved.writelines(die_params['die1_xoffset'] + "\n")
            params_saved.writelines(die_params['die2_xoffset'] + "\n")
            params_saved.writelines(die_params['die1_yoffset'] + "\n")
            params_saved.writelines(die_params['die2_yoffset'] + "\n")
            params_saved.writelines(int_die_tb['sheet'] + "\n")
            params_saved.writelines(int_die_tb['int_tb_location'] + "\n")
            params_saved.writelines(theme_combo.get() + "\n")
        
    print("Package: " + package)
    if (package == "A-CoWoS"):
        text_delete()
        mynotif("Package type: A-CoWoS")
        package_type = 1
    elif(package == "S-Organic"):
        text_delete()
        mynotif("Package type: S-Organic")
        package_type = 0
        
    else:
        package_type = 0
       

  

    generate_bump_table(excel_path, bump_visual_sheet, package_type, die_table, die_coor, dummybump, die_params, int_die_tb, int_gen, int_die_cnt, tc_sr)
    button['text']="Generate"
    


def generate_bump_table(excel_path, bump_visual_sheet, package_type, die_table, die_coor, dummybump, die_params, int_die_tb, int_gen, int_die_cnt, tc_sr):


    
    root.update_idletasks()
    mynotif("Loading the ploc file...")
    root.update_idletasks()
    try:
        # wb_d = load_workbook(excel_path, data_only=True)
        print("Opening excel file...")
        mynotif("Opening excel file...")
        wb_f = load_workbook(excel_path)
        print(wb_f)   
    except:
        print("Wrong Ploc path or Ploc file is openning. Please recheck/close the PLOC file before generate :(")
        show_error("Wrong Ploc path or Ploc file is openning. Please recheck/close the PLOC file before generate :(")
        progress_bar(0)
        mynotif("Error")
        root.update_idletasks()
        return
    
    # ws = wb_f.create_sheet('Tung')
    try:
        sheet_list = wb_f.sheetnames
       #wsvisual_d = wb_d[bump_visual_sheet]  # use for further function
        wsvisual_f = wb_f[bump_visual_sheet]
      # wsdiebump_d = wb_d[die_table['sheet']] # use for further function
       
        if die_table['sheet'] in sheet_list:
            wsdiebump_f = wb_f[die_table['sheet']]
        else:
            msg_ws = messagebox.askquestion('Create Sheet', 'The ' + die_table['sheet'] + ' doesn\'t exist. Do you want to create it?',icon='question')
            # mynotif("")
            mynotif("The " + die_table['sheet'] + " doesn't exist.")
            if(msg_ws == 'yes'):
                # mynotif("")
                mynotif('Creating the sheet...')
                wsdiebump_f = wb_f.create_sheet(die_table['sheet'])
            else:
                # mynotif("")
                progress_bar(0)
                return
        if(int_gen == 1):
            if int_die_tb['sheet'] in sheet_list:
                wsintbump_f = wb_f[int_die_tb['sheet']]
            else:
                # mynotif("")
                mynotif("The " + int_die_tb['sheet'] + " doesn't exist.")
                msg_ws = messagebox.askquestion('Create Sheet', 'The ' + int_die_tb['sheet'] + ' doesn\'t exist. Do you want to create it?', icon='question')
            
                if(msg_ws == 'yes'):
                    wsintbump_f = wb_f.create_sheet(int_die_tb['sheet'])
                    # mynotif("")
                    mynotif('Creating the sheet...')
                else:
                    # mynotif("")
                    progress_bar(0)
                    return
            
      
        
       
       #wsintbump_d = wb_d[int_die_tb['sheet']] # use for further function
       
    except:
        print("Sheet " + bump_visual_sheet + " doesn't exist")
        show_error("Sheet " + bump_visual_sheet + " doesn't exist")
        progress_bar(0)
        mynotif("Error")
        root.update_idletasks()
        return
    
    

    #----- Create dummy bump at 4 corner 140x140u for advance package (CoWos)-----------#
    ymin = coordinate_to_tuple(die_coor['window1'])[0]
    xmin = coordinate_to_tuple(die_coor['window1'])[1]
    ymax = coordinate_to_tuple(die_coor['window2'])[0]
    xmax = coordinate_to_tuple(die_coor['window2'])[1]

    print(xmin,xmax)
    print(ymin,ymax)
    progress_bar(60)
    
    try:
        #----- Create table from bump map-----------#
        die_tb_x = coordinate_to_tuple(die_table['location'])[1]
        die_tb_y = coordinate_to_tuple(die_table['location'])[0]
        int_tb_x = coordinate_to_tuple(int_die_tb['int_tb_location'])[1]
        int_tb_y = coordinate_to_tuple(int_die_tb['int_tb_location'])[0]
        

        r_die = die_tb_y + 2
        r_int = int_tb_y + 2

        title_bg_fill = PatternFill(patternType='solid', fgColor='9e42f5')
        subtil_bg_fill = PatternFill(patternType='solid',fgColor='0e7bf0')
        wsdiebump_f[die_table['location']].value = die_table['name']
        
        wsdiebump_f.merge_cells(die_table['location'] + ":" + get_column_letter(die_tb_x + 2) + str(die_tb_y))
        for c1 in range(0,3):
            wsdiebump_f[get_column_letter(die_tb_x + c1) + str(die_tb_y)].fill = title_bg_fill
            wsdiebump_f[get_column_letter(die_tb_x + c1) + str(die_tb_y)].alignment = Alignment(horizontal='center')
            wsdiebump_f[get_column_letter(die_tb_x + c1) + str(die_tb_y)].border = Border(left=Side(style='thin'),right=Side(style='thin'),top=Side(style='thin'),bottom=Side(style='thin'))

        wsdiebump_f[get_column_letter(die_tb_x) + str(die_tb_y + 1)].value = "X"

        wsdiebump_f[get_column_letter(die_tb_x + 1) + str(die_tb_y + 1)].value = "Y"

        wsdiebump_f[get_column_letter(die_tb_x + 2)  + str(str(die_tb_y + 1))].value = "Bump name"
        for c2 in range(0,3):
            wsdiebump_f[get_column_letter(die_tb_x + c2) + str(die_tb_y + 1)].fill = subtil_bg_fill
            wsdiebump_f[get_column_letter(die_tb_x + c2) + str(die_tb_y + 1)].alignment = Alignment(horizontal='center')
            wsdiebump_f[get_column_letter(die_tb_x + c2) + str(die_tb_y + 1)].border = Border(left=Side(style='thin'),right=Side(style='thin'),top=Side(style='thin'),bottom=Side(style='thin'))
        if(int(tc_sr['isTC'] == 1)):
            wsdiebump_f[get_column_letter(die_tb_x + 4) + str(die_tb_y)].value = tc_sr['sr_tb']
            wsdiebump_f.merge_cells(get_column_letter(die_tb_x + 4) + str(die_tb_y) + ":" + get_column_letter(die_tb_x + 6) + str(die_tb_y))
            for c1 in range(0,3):
                wsdiebump_f[get_column_letter(die_tb_x + 4 + c1) + str(die_tb_y)].fill = title_bg_fill
                wsdiebump_f[get_column_letter(die_tb_x + 4 + c1) + str(die_tb_y)].alignment = Alignment(horizontal='center')
                wsdiebump_f[get_column_letter(die_tb_x + 4 + c1) + str(die_tb_y)].border = Border(left=Side(style='thin'),right=Side(style='thin'),top=Side(style='thin'),bottom=Side(style='thin'))
            wsdiebump_f[get_column_letter(die_tb_x + 4) + str(die_tb_y + 1)].value = "X"
            wsdiebump_f[get_column_letter(die_tb_x + 5) + str(die_tb_y + 1)].value = "Y"
            wsdiebump_f[get_column_letter(die_tb_x + 6)  + str(str(die_tb_y + 1))].value = "Bump name"
            for c1 in range(0,3):
                wsdiebump_f[get_column_letter(die_tb_x + 4 + c1) + str(die_tb_y + 1)].fill = subtil_bg_fill
                wsdiebump_f[get_column_letter(die_tb_x + 4 + c1) + str(die_tb_y + 1)].alignment = Alignment(horizontal='center')
                wsdiebump_f[get_column_letter(die_tb_x + 4 + c1) + str(die_tb_y + 1)].border = Border(left=Side(style='thin'),right=Side(style='thin'),top=Side(style='thin'),bottom=Side(style='thin'))

        if(int_gen == 1):
            die1_list = int_die_tb['Die1_name'].split()
            die2_list = int_die_tb['Die2_name'].split()
            die1_xoffset_list = die_params['die1_xoffset'].split()
            die1_yoffset_list = die_params['die1_yoffset'].split()
            die2_xoffset_list = die_params['die2_xoffset'].split()
            die2_yoffset_list = die_params['die2_yoffset'].split()
            if(len(die1_list) != int(int_die_cnt)/2 or len(die2_list) != int(int_die_cnt)/2 or len(die1_xoffset_list) != int(int_die_cnt)/2 or len(die2_xoffset_list) != int(int_die_cnt)/2 or len(die1_yoffset_list) != int(int_die_cnt)/2 or len(die2_yoffset_list) != int(int_die_cnt)/2):
                show_error('The input die parameters incorrect. Please re-check it')
                int_input_correct = 0
                mynotif('The input die parameters incorrect. Please re-check it')
                # mynotif("")
                progress_bar(0)
                return
            else:
                wsintbump_f[get_column_letter(int_tb_x ) + str(int_tb_y)].value = "Die Flipped by Y axis"
                wsintbump_f.merge_cells(get_column_letter(int_tb_x) + str(int_tb_y) + ":" + get_column_letter(int_tb_x + 2) + str(int_tb_y))
                for c1 in range(0,3):
                    wsintbump_f[get_column_letter(int_tb_x + c1) + str(int_tb_y)].fill = title_bg_fill
                    wsintbump_f[get_column_letter(int_tb_x + c1) + str(int_tb_y)].alignment = Alignment(horizontal='center',wrapText=True)
                    wsintbump_f[get_column_letter(int_tb_x + c1) + str(int_tb_y)].border = Border(left=Side(style='thin'),right=Side(style='thin'),top=Side(style='thin'),bottom=Side(style='thin'))
                wsintbump_f[get_column_letter(int_tb_x) + str(int_tb_y + 1)].value = "X"
                wsintbump_f[get_column_letter(int_tb_x + 1) + str(int_tb_y + 1)].value = "Y"
                wsintbump_f[get_column_letter(int_tb_x + 2)  + str(str(int_tb_y + 1))].value = "Bump name"
                wsintbump_f.freeze_panes = 'A' + str(int_tb_y + 1)
                for c2 in range(0,3):
                    wsintbump_f[get_column_letter(int_tb_x + c2) + str(int_tb_y + 1)].fill = subtil_bg_fill
                    wsintbump_f[get_column_letter(int_tb_x + c2) + str(int_tb_y + 1)].alignment = Alignment(horizontal='center')
                    wsintbump_f[get_column_letter(int_tb_x + c2) + str(int_tb_y + 1)].border = Border(left=Side(style='thin'),right=Side(style='thin'),top=Side(style='thin'),bottom=Side(style='thin'))
                tbidx = 4
                for tb in range(0,int(int_die_cnt/2)):

                    wsintbump_f[get_column_letter(int_tb_x + tbidx) + str(int_tb_y)].value =  str(die1_list[tb]) + " = Die Flipped, Rotate -90 + Offset" + "(" + str(die1_xoffset_list[tb]) + "," + str(die1_yoffset_list[tb]) + ")"
                    wsintbump_f.merge_cells(get_column_letter(int_tb_x + tbidx) + str(int_tb_y) + ":" + get_column_letter(int_tb_x + tbidx + 2) + str(int_tb_y))
                    
                    for c1 in range(0,3):
                        wsintbump_f[get_column_letter(int_tb_x + tbidx + c1) + str(int_tb_y)].fill = title_bg_fill
                        wsintbump_f[get_column_letter(int_tb_x + tbidx + c1) + str(int_tb_y)].alignment = Alignment(horizontal='center', wrapText=True)
                        wsintbump_f[get_column_letter(int_tb_x + tbidx + c1) + str(int_tb_y)].border = Border(left=Side(style='thin'),right=Side(style='thin'),top=Side(style='thin'),bottom=Side(style='thin'))
                    wsintbump_f[get_column_letter(int_tb_x + tbidx) + str(int_tb_y + 1)].value = "X"
                    wsintbump_f[get_column_letter(int_tb_x + tbidx + 1) + str(int_tb_y + 1)].value = "Y"
                    wsintbump_f[get_column_letter(int_tb_x + tbidx + 2)  + str(str(int_tb_y + 1))].value = "Bump name"
                    for c2 in range(0,3):
                        wsintbump_f[get_column_letter(int_tb_x + tbidx + c2) + str(int_tb_y + 1)].fill = subtil_bg_fill
                        wsintbump_f[get_column_letter(int_tb_x + tbidx + c2) + str(int_tb_y + 1)].alignment = Alignment(horizontal='center')
                        wsintbump_f[get_column_letter(int_tb_x + tbidx + c2) + str(int_tb_y + 1)].border = Border(left=Side(style='thin'),right=Side(style='thin'),top=Side(style='thin'),bottom=Side(style='thin'))

                    wsintbump_f[get_column_letter(int_tb_x + tbidx + 4) + str(int_tb_y)].value = str(die2_list[tb]) + " = Die Flipped, Rotate +90 + Offset" + "(" + str(die2_xoffset_list[tb]) + "," + str(die2_yoffset_list[tb]) + ")"
                    wsintbump_f.merge_cells(get_column_letter(int_tb_x + tbidx + 4) + str(int_tb_y) + ":" + get_column_letter(int_tb_x + tbidx + 6) + str(int_tb_y))
                    for c1 in range(0,3):
                        wsintbump_f[get_column_letter(int_tb_x + tbidx + c1 + 4) + str(int_tb_y)].fill = title_bg_fill
                        wsintbump_f[get_column_letter(int_tb_x + tbidx + c1 + 4) + str(int_tb_y)].alignment = Alignment(horizontal='center', wrapText=True)
                        wsintbump_f[get_column_letter(int_tb_x + tbidx + c1 + 4) + str(int_tb_y)].border = Border(left=Side(style='thin'),right=Side(style='thin'),top=Side(style='thin'),bottom=Side(style='thin'))
                    wsintbump_f[get_column_letter(int_tb_x + tbidx + 4) + str(int_tb_y + 1)].value = "X"
                    wsintbump_f[get_column_letter(int_tb_x + tbidx + 5) + str(int_tb_y + 1)].value = "Y"
                    wsintbump_f[get_column_letter(int_tb_x + tbidx + 6)  + str(str(int_tb_y + 1))].value = "Bump name"
                    for c2 in range(0,3):
                        wsintbump_f[get_column_letter(int_tb_x + tbidx + c2 + 4) + str(int_tb_y + 1)].fill = subtil_bg_fill
                        wsintbump_f[get_column_letter(int_tb_x + tbidx + c2 + 4) + str(int_tb_y + 1)].alignment = Alignment(horizontal='center')
                        wsintbump_f[get_column_letter(int_tb_x + tbidx + c2 + 4) + str(int_tb_y + 1)].border = Border(left=Side(style='thin'),right=Side(style='thin'),top=Side(style='thin'),bottom=Side(style='thin'))
                    tbidx += 8
                int_input_correct = 1
        else:
            pass
        # xwidth = float (ws_f[get_column_letter(xmax) + die_coor["xcoor"]].value)
        # minxval = float (ws_f[get_column_letter(xmin) + die_coor["xcoor"]].value)
        # ywidth = float (ws_f[die_coor["ycoor"] + str(ymin)].value)
        # minyval = float (ws_f[die_coor["ycoor"] + str(ymax)].value)
        # xwidth = ws_f[get_column_letter(xmax) + die_coor["xcoor"]].value
        # minxval = ws_f[get_column_letter(xmin) + die_coor["xcoor"]].value
        # ywidth = ws_f[die_coor["ycoor"] + str(ymin)].value
        # minyval = ws_f[die_coor["ycoor"] + str(ymax)].value
        if (package_type == 1):
            
            print("Generate for Advance Package")
            mynotif("Generate for Advance Package")
            dm_bump_coor= []
            dm_cnt=0
            # mynotif("")
            root.update_idletasks()
            mynotif("Generating Dummy bump...")
            root.update_idletasks()
            for dm_bump in dummybump:
                bump = list(dummybump[dm_bump].values())
                    
                ymin_dm = coordinate_to_tuple(bump[0])[0]
                xmin_dm = coordinate_to_tuple(bump[0])[1]
                ymax_dm = coordinate_to_tuple(bump[1])[0]
                xmax_dm = coordinate_to_tuple(bump[1])[1]
                xcoor_dm = str(bump[2])
                ycoor_dm = str(bump[3])

                print(xmin_dm,xmax_dm)
                print(ymin_dm,ymax_dm)
                print(dummybump)

                for dummycol1 in range(xmin_dm, xmax_dm + 1):
                    for dummyrow1 in range(ymin_dm, ymax_dm + 1):
                        col_dm = get_column_letter(dummycol1)
                        if (wsvisual_f[col_dm + str(dummyrow1)].value != None):
                            
                            print("Processing for Dummy bump at: " + col_dm + str(dummyrow1))
                            mynotif("Processing for Dummy bump at: " + col_dm + str(dummyrow1))
                            # Gen dummy bump table
                            wsdiebump_f[get_column_letter(die_tb_x)+str(r_die)].value = f"='{bump_visual_sheet}'!{col_dm + xcoor_dm}"
                            wsdiebump_f[get_column_letter(die_tb_x + 1)+str(r_die)].value = f"='{bump_visual_sheet}'!{ycoor_dm + str(dummyrow1)}"
                            wsdiebump_f[get_column_letter(die_tb_x + 2)+str(r_die)].value =  f"='{bump_visual_sheet}'!{col_dm+ str(dummyrow1)}"
                            if(int(tc_sr['isTC'] == 1)):
                                wsdiebump_f[get_column_letter(die_tb_x + 4)+str(r_die)].value = f"=('{bump_visual_sheet}'!{col_dm + xcoor_dm})-({tc_sr['sr_w']})"
                                wsdiebump_f[get_column_letter(die_tb_x + 5)+str(r_die)].value = f"=('{bump_visual_sheet}'!{ycoor_dm + str(dummyrow1)})-({tc_sr['sr_w']}) "
                                wsdiebump_f[get_column_letter(die_tb_x + 6)+str(r_die)].value =  f"='{bump_visual_sheet}'!{col_dm+ str(dummyrow1)}"
                            r_die += 1
                            coor = col_dm + str(dummyrow1)
                            dm_bump_coor.append(coor)
                            dm_cnt += 1

                            if(int_gen == 1 and int_input_correct == 1):

                                wsintbump_f[get_column_letter(int_tb_x)+str(r_int)].value = f"=({str(die_params['chip_width']).replace('=','')})-('{bump_visual_sheet}'!{str(col_dm + xcoor_dm)})" # Flip Y axis
                                wsintbump_f[get_column_letter(int_tb_x + 1)+str(r_int)].value = f"='{bump_visual_sheet}'!{ycoor_dm + str(dummyrow1)}" # Flip Y axis
                                wsintbump_f[get_column_letter(int_tb_x + 2)+str(r_int)].value =  f"='{bump_visual_sheet}'!{col_dm+ str(dummyrow1)}" #Flip Y axis
                                tbidx2 = 0
                                # r_current = 

                                for tb in range(0,int(int_die_cnt/2)):
                                # #----------------------------Flip bump map in y axis - Rotate -90 - Rotate +90---------------------------
                                    
                                    wsintbump_f[get_column_letter(int_tb_x + tbidx2 + 5)+str(r_int)].value = f"=({str(die_params['chip_width']).replace('=','')})-('{bump_visual_sheet}'!{str(col_dm + xcoor_dm)})+({str(die1_yoffset_list[tb])})" # Rotate -90
                                    wsintbump_f[get_column_letter(int_tb_x + tbidx2 + 9)+str(r_int)].value = f"=('{bump_visual_sheet}'!{str(col_dm + xcoor_dm)})+({str(die1_yoffset_list[tb]).replace('=','')})" # Rotate +90

                                    wsintbump_f[get_column_letter(int_tb_x + tbidx2 + 4)+str(r_int)].value = f"=({str(die_params['chip_height']).replace('=','')})-('{bump_visual_sheet}'!{ycoor_dm + str(dummyrow1)})+({str(die1_xoffset_list[tb])})" # Rotate -90
                                    wsintbump_f[get_column_letter(int_tb_x + tbidx2 + 8)+str(r_int)].value = f"=('{bump_visual_sheet}'!{ycoor_dm + str(dummyrow1)})+({str(die2_xoffset_list[tb]).replace('=','')})" # Rotate +90

                                
                                    if(wsvisual_f[col_dm+ str(dummyrow1)].value == "VSS"):
                                        wsintbump_f[get_column_letter(int_tb_x + tbidx2 + 6)+str(r_int)].value = f"='{bump_visual_sheet}'!{col_dm+ str(dummyrow1)}" # Rotate -90
                                        wsintbump_f[get_column_letter(int_tb_x + tbidx2 + 10)+str(r_int)].value = f"='{bump_visual_sheet}'!{col_dm+ str(dummyrow1)}" # Rotate +90
                                    else:
                                        wsintbump_f[get_column_letter(int_tb_x + tbidx2 + 6)+str(r_int)].value = f"=\"{die1_list[tb]}_\"&'{bump_visual_sheet}'!{col_dm+ str(dummyrow1)}" # Rotate -90
                                        wsintbump_f[get_column_letter(int_tb_x + tbidx2 + 10)+str(r_int)].value = f"=\"{die2_list[tb]}_\"&'{bump_visual_sheet}'!{col_dm+ str(dummyrow1)}" # Rotate +90
                                    tbidx2 += 8
                                
                                r_int += 1
                            
                            

            #---------Create Die bump exclued dummy bump at 4 corner-----------#

            match = 0
            # mynotif("")
            root.update_idletasks()
            mynotif("Generating Die bump...")
            root.update_idletasks()
            for col in range(xmin, xmax + 1):
                for row in range(ymin, ymax + 1):       
                    col_l = get_column_letter(col)
                    #print(col_l)
                    i = 0 
                    while(i < len(dm_bump_coor)):
                        xy = col_l + str(row)
                        if(xy ==  dm_bump_coor[i]):
                            match = 1
                        else:
                            match = 0
                        if(match == 1):
                            break
                        i += 1
                    if (match == 0 and wsvisual_f[col_l + str(row)].value != None):
                        print("Processing for Die bump at: " + col_l + str(row))
                        mynotif("Processing for Die bump at: " + col_l + str(row))
                        #  get the X value from Visual bump sheet
                        if (wsvisual_f[col_l + str(row)].value != None):
                        
                            #  get the X value from Visual bump sheet
                        
                            wsdiebump_f[get_column_letter(die_tb_x)+str(r_die)].value = f"='{bump_visual_sheet}'!{col_l + str(die_coor['xcoor'])}"
                        
                            # #  get the Y value from Visual bump sheet
                        
                            wsdiebump_f[get_column_letter(die_tb_x + 1)+str(r_die)].value = f"='{bump_visual_sheet}'!{die_coor['ycoor'] + str(row)}"
                            
                            #  get the Bump name from Visual bump sheet
                            wsdiebump_f[get_column_letter(die_tb_x + 2)+str(r_die)].value =  f"='{bump_visual_sheet}'!{col_l+ str(row)}"
                            if(int(tc_sr['isTC'] == 1)):
                               wsdiebump_f[get_column_letter(die_tb_x + 4)+str(r_die)].value = f"=('{bump_visual_sheet}'!{col_l + str(die_coor['xcoor'])})-({tc_sr['sr_w']})" 
                               wsdiebump_f[get_column_letter(die_tb_x + 5)+str(r_die)].value = f"=('{bump_visual_sheet}'!{die_coor['ycoor'] + str(row)})-({tc_sr['sr_w']})"
                               wsdiebump_f[get_column_letter(die_tb_x + 6)+str(r_die)].value =  f"='{bump_visual_sheet}'!{col_l+ str(row)}"
                            r_die += 1
                            
                            if(int_gen == 1 and int_input_correct == 1):

                                wsintbump_f[get_column_letter(int_tb_x )+str(r_int)].value = f"=({str(die_params['chip_width']).replace('=','')})-('{bump_visual_sheet}'!{col_l + str(die_coor['xcoor'])})" # Flip Y axis
                                wsintbump_f[get_column_letter(int_tb_x + 1)+str(r_int)].value = f"='{bump_visual_sheet}'!{die_coor['ycoor'] + str(row)}" # Flip Y axis
                                wsintbump_f[get_column_letter(int_tb_x + 2)+str(r_int)].value =  f"='{bump_visual_sheet}'!{col_l+ str(row)}" #Flip Y axis

                                tbidx2 = 0
                                # r_current = 

                                for tb in range(0,int(int_die_cnt/2)):
                                # #----------------------------Flip bump map in y axis - Rotate -90 - Rotate +90---------------------------
                                    
                                    wsintbump_f[get_column_letter(int_tb_x + tbidx2 + 5)+str(r_int)].value = f"=({str(die_params['chip_width']).replace('=','')})-('{bump_visual_sheet}'!{col_l + str(die_coor['xcoor'])})+({str(die1_yoffset_list[tb])})" # Rotate -90
                                    wsintbump_f[get_column_letter(int_tb_x + tbidx2 + 9)+str(r_int)].value = f"=('{bump_visual_sheet}'!{col_l + str(die_coor['xcoor'])})+({str(die2_yoffset_list[tb]).replace('=','')})" # Rotate +90
                                    
                                    wsintbump_f[get_column_letter(int_tb_x + tbidx2 + 4)+str(r_int)].value = f"=({str(die_params['chip_height']).replace('=','')})-('{bump_visual_sheet}'!{die_coor['ycoor']+str(row)})+({str(die1_xoffset_list[tb])})" # Rotate -90
                                    wsintbump_f[get_column_letter(int_tb_x + tbidx2 + 8)+str(r_int)].value = f"=('{bump_visual_sheet}'!{die_coor['ycoor'] + str(row)})+({str(die2_xoffset_list[tb]).replace('=','')})" # Rotate +90
                                
                                    if(wsvisual_f[col_l+ str(row)].value == "VSS"):
                                        wsintbump_f[get_column_letter(int_tb_x + tbidx2 + 6)+str(r_int)].value = f"='{bump_visual_sheet}'!{col_l+ str(row)}" # Rotate -90
                                        wsintbump_f[get_column_letter(int_tb_x + tbidx2 + 10)+str(r_int)].value = f"='{bump_visual_sheet}'!{col_l+ str(row)}" # Rotate +90
                                    else:
                                        wsintbump_f[get_column_letter(int_tb_x + tbidx2 + 6)+str(r_int)].value = f"=\"{die1_list[tb]}_\"&'{bump_visual_sheet}'!{col_l+ str(row)}" # Rotate -90
                                        wsintbump_f[get_column_letter(int_tb_x + tbidx2 + 10)+str(r_int)].value = f"=\"{die2_list[tb]}_\"&'{bump_visual_sheet}'!{col_l+ str(row)}" # Rotate +90
                                    tbidx2 += 8
                                
                                r_int += 1
                                
                                
        else:
            process_notify("Generating Die bump...")
            for col in range(xmin, xmax + 1):
                    for row in range(ymin , ymax + 1):       
                        col_l = get_column_letter(col)
                        #print(col_l)
                        if (wsvisual_f[col_l + str(row)].value != None):
                            print("Processing for Die bump at: " + col_l + str(row))
                            mynotif("Processing for Die bump at: " + col_l + str(row))
                            #  get the X value from Visual bump sheet
                        
                            wsdiebump_f[get_column_letter(die_tb_x)+str(r_die)].value = f"='{bump_visual_sheet}'!{col_l + str(die_coor['xcoor'])}"
                        
                            # #  get the Y value from Visual bump sheet
                        
                            wsdiebump_f[get_column_letter(die_tb_x + 1)+str(r_die)].value = f"='{bump_visual_sheet}'!{die_coor['ycoor'] + str(row)}"
                            
                            #  get the Bump name from Visual bump sheet
                            wsdiebump_f[get_column_letter(die_tb_x + 2)+str(r_die)].value =  f"='{bump_visual_sheet}'!{col_l+ str(row)}"
                            if(int(tc_sr['isTC'] == 1)):
                               wsdiebump_f[get_column_letter(die_tb_x + 4)+str(r_die)].value = f"=('{bump_visual_sheet}'!{col_l + str(die_coor['xcoor'])})-({tc_sr['sr_w']})" 
                               wsdiebump_f[get_column_letter(die_tb_x + 5)+str(r_die)].value = f"=('{bump_visual_sheet}'!{die_coor['ycoor'] + str(row)})-({tc_sr['sr_w']})"
                               wsdiebump_f[get_column_letter(die_tb_x + 6)+str(r_die)].value =  f"='{bump_visual_sheet}'!{col_l+ str(row)}"
                            r_die += 1
                            if(int_gen == 1 and int_input_correct == 1):
                                
                                wsintbump_f[get_column_letter(int_tb_x )+str(r_int)].value = f"=({str(die_params['chip_width']).replace('=','')})-('{bump_visual_sheet}'!{col_l + str(die_coor['xcoor'])})" # Flip Y axis
                                wsintbump_f[get_column_letter(int_tb_x + 1)+str(r_int)].value = f"='{bump_visual_sheet}'!{die_coor['ycoor'] + str(row)}" # Flip Y axis
                                wsintbump_f[get_column_letter(int_tb_x + 2)+str(r_int)].value =  f"='{bump_visual_sheet}'!{col_l+ str(row)}" #Flip Y axis

                                tbidx2 = 0
                                # r_current = 

                                for tb in range(0,int(int_die_cnt/2)):
                                # #----------------------------Flip bump map in y axis - Rotate -90 - Rotate +90---------------------------
                                    
                                    wsintbump_f[get_column_letter(int_tb_x + tbidx2 + 5)+str(r_int)].value = f"=({str(die_params['chip_width']).replace('=','')})-('{bump_visual_sheet}'!{col_l + str(die_coor['xcoor'])})+({str(die1_yoffset_list[tb])})" # Rotate -90
                                    wsintbump_f[get_column_letter(int_tb_x + tbidx2 + 9)+str(r_int)].value = f"=('{bump_visual_sheet}'!{col_l + str(die_coor['xcoor'])})+({str(die2_yoffset_list[tb]).replace('=','')})" # Rotate +90
                                
                                    
                                    wsintbump_f[get_column_letter(int_tb_x + tbidx2 + 4)+str(r_int)].value = f"=({str(die_params['chip_height']).replace('=','')})-('{bump_visual_sheet}'!{die_coor['ycoor']+str(row)})+({str(die1_xoffset_list[tb])})" # Rotate -90
                                    wsintbump_f[get_column_letter(int_tb_x + tbidx2 + 8)+str(r_int)].value = f"=('{bump_visual_sheet}'!{die_coor['ycoor'] + str(row)})+({str(die2_xoffset_list[tb]).replace('=','')})" # Rotate +90
                                
                                
                                    if(wsvisual_f[col_l+ str(row)].value == "VSS"):
                                        wsintbump_f[get_column_letter(int_tb_x + tbidx2 + 6)+str(r_int)].value = f"='{bump_visual_sheet}'!{col_l+ str(row)}" # Rotate -90
                                        wsintbump_f[get_column_letter(int_tb_x + tbidx2 + 10)+str(r_int)].value = f"='{bump_visual_sheet}'!{col_l+ str(row)}" # Rotate +90
                                    else:
                                        wsintbump_f[get_column_letter(int_tb_x + tbidx2 + 6)+str(r_int)].value = f"=\"{die1_list[tb]}_\"&'{bump_visual_sheet}'!{col_l+ str(row)}" # Rotate -90
                                        wsintbump_f[get_column_letter(int_tb_x + tbidx2 + 10)+str(r_int)].value = f"=\"{die2_list[tb]}_\"&'{bump_visual_sheet}'!{col_l+ str(row)}" # Rotate +90
                                    tbidx2 += 8
                                

                                r_int += 1
            # tab = Table(displayName="Table1", ref="O65:Q500")
            # ws_f.add_table(tab)
        
        progress_bar(80)   
        print("Saving excel...") 
        mynotif("Saving excel file...")
        wb_f.save(excel_path)
        progress_bar(100)
        mynotif("Successed!!!")
        print("Completed!!!")
        popup("PLOC generated successful!!!")
        entry_disable(text)
        
    except (ValueError):
        print ("Wrong input, Please check and regenerate")
        show_error("Wrong input, Please check and regenerate")
        progress_bar(0)
        mynotif("Error")
        root.update_idletasks()
        return
    except:
        print('Error!!!')
        
        show_error("There are an error in caculations, Please recheck and make sure the input is correct!")
        progress_bar(0)
        mynotif("Error")
        root.update_idletasks()
        return
            
            
    


                
# myButton = tk.Button(root,text="Button", command=get_path)
# myButton.pack()


treeScroll = ttk.Scrollbar(root, orient = 'vertical')



# root.configure(yscrollcomand)

progress = ttk.Progressbar(root, orient = 'horizontal',
              length = 100, mode = 'determinate')
progress_w = my_canvas.create_window(80,800, anchor="nw", window=progress, width= 800)


# Button
#Create style object
style = ttk.Style()

# #configure style
# style = ttk.Style()
# style.configure('TButton', font =
#                ('calibri', 20, 'bold'),
#                     borderwidth = '4',
#                     width = '80')
style.configure('TCheckbutton', font= ('System', 12, 'underline', 'bold'),
 foreground='black', border=50)
mediumFont = tkfont(
	family="System",
	size=12,
	weight="normal",
	slant="italic",
	underline=1,
	overstrike=0)
def hihi():
    button.configure(font=mediumFont, foreground='white', background='Green')
browse_btn = ttk.Button(root, text="Open File", image=open_imag, command=open_file)
browse_btn_w = my_canvas.create_window(865, 40, anchor="nw", window=browse_btn)
# button = tk.Button(root, text="Generate",font=("System", 14, 'underline', 'bold'), foreground='white', background='#9b34eb', command=get_path, width=40)
button = tk.Button(root, text="Generate",font = mediumFont, foreground='white', background='#9b34eb', command=get_params_and_generate, width=40)
# button = ttk.Button(root, text="Generate", command=get_path, width=80)

button_w = my_canvas.create_window(300, 860, anchor="nw", window=button)

text = tk.Text(my_canvas,width = 50, height = 100,bd=5,relief='groove', wrap='word', font=('arial',10), highlightthickness=2 ) #yscrollcommand=scroll_y.set
scroll_y = ttk.Scrollbar(text)
my_canvas.create_window(600,80, anchor='nw', window=text, height=100, width=280)

text.config(yscrollcommand=scroll_y.set)
scroll_y.pack(side=RIGHT, fill=Y)
scroll_y.config(command=text.yview)

# Get saved params
def get_saved_params():
    global temp_file
    try:
        with open(temp_file,'r') as params_saved:
            line1 = [line.rstrip() for line in params_saved]
            params = {
            'excel_path' : line1[0],
            'bump_visual_sheet' : line1[1],
            'package_type' : line1[2],
            'forTC' : line1[3],
            'sr_width' : line1[4],
            'die_visual' : line1[5],
            'die_dummy' : line1[6],
            'die_out_tb_sheet' : line1[7],
            'die_out_tb_name1': line1[8],
            'die_out_tb_name2': line1[9],
            'die_out_loc': line1[10],
            'is_interpos' : line1[11],
            'inter_size' : line1[12],
            'inter_diecount' : line1[13],
            'inter_dieL_name' : line1[14],
            'inter_dieR_name' : line1[15],
            'inter_xL_offset' : line1[16],
            'inter_xR_offset' : line1[17],
            'inter_yL_offset' : line1[18],
            'inter_yR_offset' : line1[19],
            'inter_out_tb_sheet' : line1[20],
            'inter_out_tb_loc' : line1[21],
            'theme': line1[22]
            }
        
        root.set_theme(params['theme'])
        theme_combo.current(theme_list.index(params['theme']))
        excel_i.insert(0,params['excel_path'])
        sheet_i.insert(0, params['bump_visual_sheet'])

        srw_i.insert(0, params['sr_width'])
        die_visual_list = params['die_visual'].split()
        x1y1_i.insert(0, die_visual_list[0])
        x2y2_i.insert(0, die_visual_list[1])
        Xget_i.insert(0, die_visual_list[2])
        Yget_i.insert(0, die_visual_list[3])

        dummy_list = params['die_dummy'].split()
        cor1_x1y1.insert(0, dummy_list[0])
        cor1_x2y2.insert(0, dummy_list[1])
        cor1_Xget.insert(0, dummy_list[2])
        cor1_Yget.insert(0, dummy_list[3])
        
        cor2_x1y1.insert(0, dummy_list[4])
        cor2_x2y2.insert(0, dummy_list[5])
        cor2_Xget.insert(0, dummy_list[6])
        cor2_Yget.insert(0, dummy_list[7])
        
        cor3_x1y1.insert(0, dummy_list[8])
        cor3_x2y2.insert(0, dummy_list[9])
        cor3_Xget.insert(0, dummy_list[10])
        cor3_Yget.insert(0, dummy_list[11])

        cor4_x1y1.insert(0, dummy_list[12])
        cor4_x2y2.insert(0, dummy_list[13])
        cor4_Xget.insert(0, dummy_list[14])
        cor4_Yget.insert(0, dummy_list[15])

        
        out_tb_sheet.insert(0, params['die_out_tb_sheet'])
        out_name_in.insert(0, params['die_out_tb_name1'])
        out_col_i.insert(0, params['die_out_loc'])
        out_name2_in.insert(0, params['die_out_tb_name2'])
        chip_size_list = params['inter_size'].split()
        xwidth_i.insert(0, chip_size_list[0])
        yheight_i.insert(0, chip_size_list[1])
        Die1_name.insert(0, params['inter_dieL_name'])
        Die2_name.insert(0, params['inter_dieR_name'])
        Die1_xoffset_i.insert(0, params['inter_xL_offset'])
        Die2_xoffset_i.insert(0, params['inter_xR_offset'])
        Die1_yoffset_i.insert(0, params['inter_yL_offset'])
        Die2_yoffset_i.insert(0, params['inter_yR_offset'])
        intp_sheet.insert(0, params['inter_out_tb_sheet'])
        int_tb_loc.insert(0, params['inter_out_tb_loc'])
        
        int_die_num_combo.current(int_couple_number.index(params['inter_diecount']))
        package_combo.current(0)
        change_colour(theme_list.index(params['theme']))
        # package_combo.current(package_list.index(params['package_type']))

    except:
        root.set_theme("scidpurple")
        theme_combo.current(theme_list.index("scidpurple"))
        change_colour(theme_list.index("scidpurple"))
        excel_i.insert(0, r"C:\Users\sytung\OneDrive - Synopsys, Inc\Desktop\py\Bump_CoWoS_S.xlsx")
        sheet_i.insert(0, "N3P_CoWoS")
        srw_i.insert(0, "21.6")
        x1y1_i.insert(0, "C11")
        x2y2_i.insert(0, "CW103")
        Xget_i.insert(0, "8")
        Yget_i.insert(0, "B")
        out_tb_sheet.insert(0, "Bump coordination")
        out_name_in.insert(0, "Die with sealring")
        out_name2_in.insert(0, "Die without sealring")
        out_col_i.insert(0, "S111")
        out_col_wsr_i.insert(0, "T64")
        cor1_x1y1.insert(0, "C11")
        cor1_x2y2.insert(0, "E13")
        cor1_Xget.insert(0, "9")
        cor1_Yget.insert(0, "B")
        cor2_x1y1.insert(0, "CU11")
        cor2_x2y2.insert(0, "CW13")
        cor2_Xget.insert(0, "9")
        cor2_Yget.insert(0, "B")
        cor3_x1y1.insert(0, "C101")
        cor3_x2y2.insert(0, "E103")
        cor3_Xget.insert(0, "9")
        cor3_Yget.insert(0, "B")
        cor4_x1y1.insert(0, "CU101")
        cor4_x2y2.insert(0, "CW103")
        cor4_Xget.insert(0, "9")
        cor4_Yget.insert(0, "B")
        xwidth_i.insert(0, "3938.352")
        yheight_i.insert(0, "2262.872")
        intp_sheet.insert(0, "Package_substrate")
        int_tb_loc.insert(0, "X111")
        Die1_name.insert(0, "DIE3")
        Die2_name.insert(0, "DIE7")
        Die1_xoffset_i.insert(0, "-4350.8")
        Die1_yoffset_i.insert(0, "16.2349999999999")
        Die2_xoffset_i.insert(0, "1571.96")
        Die2_yoffset_i.insert(0, "97.9849999999997")
        int_die_num_combo.current(0)
        package_combo.current(0)
    mynotif("\n\nINFO: This field is for showing the information or guidance")
get_saved_params()

entry_disable(cor1_x1y1, cor1_x2y2, cor1_Xget, cor1_Yget,
            cor2_x1y1, cor2_x2y2, cor2_Xget, cor2_Yget,
            cor3_x1y1, cor3_x2y2, cor3_Xget, cor3_Yget,
            cor4_x1y1, cor4_x2y2, cor4_Xget, cor4_Yget)

entry_disable(sheete_i, sheete_t)
entry_disable(srw_i, out_name2_in, out_col_wsr_i)
entry_disable(xwidth_i, yheight_i, Die1_xoffset_i, Die1_yoffset_i, Die2_xoffset_i, Die2_yoffset_i, intp_sheet, Die1_name, Die2_name, int_tb_loc, int_die_num_combo)
# my_canvas.itemconfigure(out_col_i_w, bac )
sheet_t['text']= "Bump sheet:"
mynotif("")
root.mainloop()

