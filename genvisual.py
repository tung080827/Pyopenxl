from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.utils import range_boundaries
from openpyxl.utils import column_index_from_string
#from openpyxl.utils import coordinate_from_string
from openpyxl.utils import coordinate_to_tuple
from openpyxl.worksheet.table import Table, TableStyleInfo
# from tkinter import *
from tkinter import messagebox
from tkinter import ttk
import getcolumn
from array import *
import tkinter as tk
from tkinter import *
from ttkthemes import ThemedTk, THEMES
from PIL import Image
from PIL import ImageTk, Image
from tkinter.font import Font as tkfont
from tkinter import filedialog
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection,Font,Fill,Color
from openpyxl.worksheet.worksheet import Worksheet
import gui_function as gui
from ploc_myTk import *
from tkinter.filedialog import asksaveasfile
from copy import copy, deepcopy


def cell(col,row):
    return get_column_letter(col)+str(row)
def getstring(string: str,c1: str, c2: str):
    cell = string
    idx1 = cell.find(c1)
    idx2 = cell.find(c2)
    if(idx1 == -1 or idx2 == -1):
        return "NA","NA"
    else:
        str_wo_c = cell[idx1+1:idx2]
        str_w_c = cell[idx1:idx2+1]
        return str_wo_c,str_w_c
def row_col(cell: str):
    row = coordinate_to_tuple(cell)[0]
    col = coordinate_to_tuple(cell)[1]
    return row, col
def gettable(ws: Worksheet,tb_name: str):
    isfound = 0
    found_cnt = 0
    for merge in ws.merged_cells.ranges:
        idx = str(merge).find(':')
        cell_begin = str(merge)[:idx]
        cell_end = str(merge)[idx+1:]
        # print(f"cell begin: {cell_begin} \ncell end: {cell_end}")
        print(ws[cell_begin].value)
        if (ws[cell_begin].value == tb_name):
            isfound = 1
            found_cnt += 1
            row_begin = coordinate_to_tuple(cell_begin)[0]
            col_begin = coordinate_to_tuple(cell_begin)[1]
            col_end = coordinate_to_tuple(cell_end)[1]
            # print(f"col begin: {col_begin}\ncol end: {col_end}\nrow begin: {row_begin}")
            bor: Border = ws[cell(col_begin, row_begin+1)].border.left.style
            row_end = row_begin + 1
            while(bor is not None):
                row_end += 1
                bor = ws[cell(col_begin, row_end)].border.left.style
            row_end -= 1 
             
        
        else:
             pass
    if(isfound == 0):
        print(f"The table {tb_name} is not found") 
        return None
    elif(isfound ==1 and found_cnt > 1):
        print(f"More than 1 table \"{tb_name}\" present")
        return None
    else:
        # print(f" col: {col_begin} {col_end} \nrow: {row_begin} {row_end}")
        return col_begin, col_end, row_begin, row_end    
def browse_file(entry: Tkentry):
    global prj_ls, summary_info
	# global my_image
    root.filename = filedialog.askopenfilename(title="Select A File", filetypes=(("Excel files", "*.xlsx"),("all files", "*.*")))
    # excel_i.delete(0,END)
    entry.add_new_content(root.filename)
    print(root.filename)
excel_path = r'C:\Users\sytung\Desktop\H140_Cayman_N3P_CoWos_test.xlsx'
table = "MMX1_REMOVED"
wb = load_workbook(excel_path, data_only= True)
ws = wb['BUMP_SQUARE_UNIFORM']
table_range = gettable(ws,table)
print(table_range)

Xcol = get_column_letter(table_range[0])
Ycol = get_column_letter(int(table_range[0])+1)
Bumpcol = get_column_letter(table_range[1])
begin = int(table_range[2])
end = int(table_range[3])
xcoor:list = []
ycoor: list = []
for row in range(begin+2,end+1):
    xcoor.append(float(ws[Xcol+str(row)].value))
    ycoor.append(float(ws[Ycol+ str(row)].value))
xcoor = list(dict.fromkeys(xcoor))
xcoor.sort()
ycoor = list(dict.fromkeys(ycoor))
ycoor.sort()
print(f"X: {xcoor} \n {len(xcoor)}")
print(f"Y: {ycoor} \n {len(ycoor)}")

vws = wb['Sheet2']
vrow = len(ycoor) + 10
vcolbegin = 6

for i in range(0,len(xcoor)):
    vws[get_column_letter(vcolbegin+i)+str(vrow)].value = xcoor[i]
for j in range(0,len(ycoor)):
    vws[get_column_letter(vcolbegin-1)+str(vrow-j-1)].value = ycoor[j]

for bump in range(begin+2,end+1):
    x = get_column_letter(vcolbegin + xcoor.index(ws[Xcol+str(bump)].value))
    yindex = ycoor.index(ws[Ycol+ str(bump)].value)
    y=  (vrow - 1) - yindex
    bump_name = ws[Bumpcol +str(bump)].value
    vws[x+str(y)].value = bump_name
    print(bump)
    # if bump == 63:
    #     print("vao day")

wb.save(excel_path)
