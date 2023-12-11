from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.utils import range_boundaries
from openpyxl.utils import column_index_from_string
#from openpyxl.utils import coordinate_from_string
from openpyxl.utils import coordinate_to_tuple
from openpyxl.worksheet.table import Table, TableStyleInfo
# from tkinter import *
from tkinter import messagebox
from tkinter import ttk
import getcolumn
from array import *
import tkinter as tk
from tkinter import *
from ttkthemes import ThemedTk, THEMES
from PIL import Image
from PIL import ImageTk, Image
from tkinter.font import Font as tkfont
from tkinter import filedialog
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection,Font,Fill,Color
from openpyxl.worksheet.worksheet import Worksheet
import gui_function as gui
from ploc_myTk import *
from tkinter.filedialog import asksaveasfile
from copy import copy, deepcopy


def cell(col,row):
    return get_column_letter(col)+str(row)
def getstring(string: str,c1: str, c2: str):
    cell = string
    idx1 = cell.find(c1)
    idx2 = cell.find(c2)
    if(idx1 == -1 or idx2 == -1):
        return "NA","NA"
    else:
        str_wo_c = cell[idx1+1:idx2]
        str_w_c = cell[idx1:idx2+1]
        return str_wo_c,str_w_c
def row_col(cell: str):
    row = coordinate_to_tuple(cell)[0]
    col = coordinate_to_tuple(cell)[1]
    return row, col
def gettable(ws: Worksheet,tb_name: str):
    isfound = 0
    found_cnt = 0
    for merge in ws.merged_cells.ranges:
        idx = str(merge).find(':')
        cell_begin = str(merge)[:idx]
        cell_end = str(merge)[idx+1:]
        # print(f"cell begin: {cell_begin} \ncell end: {cell_end}")
        print(ws[cell_begin].value)
        if (ws[cell_begin].value == tb_name):
            isfound = 1
            found_cnt += 1
            row_begin = coordinate_to_tuple(cell_begin)[0]
            col_begin = coordinate_to_tuple(cell_begin)[1]
            col_end = coordinate_to_tuple(cell_end)[1]
            # print(f"col begin: {col_begin}\ncol end: {col_end}\nrow begin: {row_begin}")
            bor: Border = ws[cell(col_begin, row_begin+1)].border.left.style
            row_end = row_begin + 1
            while(bor is not None):
                row_end += 1
                bor = ws[cell(col_begin, row_end)].border.left.style
            row_end -= 1 
             
        
        else:
             pass
    if(isfound == 0):
        print(f"The table {tb_name} is not found") 
        return None
    elif(isfound ==1 and found_cnt > 1):
        print(f"More than 1 table \"{tb_name}\" present")
        return None
    else:
        # print(f" col: {col_begin} {col_end} \nrow: {row_begin} {row_end}")
        return col_begin, col_end, row_begin, row_end    
def browse_file(entry: Tkentry):
    global prj_ls, summary_info
	# global my_image
    root.filename = filedialog.askopenfilename(title="Select A File", filetypes=(("Excel files", "*.xlsx"),("all files", "*.*")))
    # excel_i.delete(0,END)
    entry.add_new_content(root.filename)
    print(root.filename)

def get_cell(range: str):
    idx = str(range).find(':')
    cell_begin = str(range)[:idx]
    cell_end = str(range)[idx+1:]
    return cell_begin, cell_end

########### Input parammeters #########################################

excel_path = r'C:\Users\sytung\Desktop\Tung_temp.xlsx'

ref_map_range = "X26:AG80"
ref_ws_name = 'Sheet2'

out_map_loc = "K30"
out_ws_name = "Sheet3"
option = "rotate(-90)"

####################################################
wb = load_workbook(excel_path)
ref_ws = wb[ref_ws_name]
# out_ws =  wb[out_ws_name]
if out_ws_name in wb.sheetnames:            
            out_ws = wb[out_ws_name]
else:          
    # mynotif(textbox, "The " + wso_name + " doesn't exist.")
    msg_ws = messagebox.askquestion('Create Sheet', 'The sheet: ' + out_ws_name + ' doesn\'t exist. Do you want to create it?', icon='question')        
    if(msg_ws == 'yes'):
        out_ws = wb.create_sheet(out_ws_name)
        # mynotif(textbox, "Creating sheet...")
        # mynotif(textbox,'Creating the sheet...')
    # else:
        # mynotif(textbox, "Aborted!!!")
        # progress_bar(progressbar,0)
        # return       
 
ref_cell_begin = get_cell(ref_map_range)[0]
ref_cell_end = get_cell(ref_map_range)[1]

ref_row_begin = row_col(ref_cell_begin)[0]
ref_col_begin = row_col(ref_cell_begin)[1]
ref_row_end = row_col(ref_cell_end)[0]
ref_col_end = row_col(ref_cell_end)[1]

out_row_begin = row_col(out_map_loc)[0]
out_col_begin = row_col(out_map_loc)[1]
x_len = ref_col_end - ref_col_begin + 1
y_len = ref_row_end - ref_row_begin + 1
print(f"row begin: {ref_row_begin},  row end: {ref_row_end}")
print(f"col begin: {ref_col_begin},  col end: {ref_col_end}")
if option == "flip":
    out_r = out_row_begin
    out_c = out_col_begin + x_len
    for row in range(ref_row_begin, ref_row_end+ 1):
        for col in range(ref_col_begin, ref_col_end + 1):
            out_ws.cell(row=out_r, column=out_c).value = ref_ws.cell(row=row, column=col).value
            out_c -= 1
        out_c = out_col_begin + x_len
        out_r += 1
elif option == 'rotate(-90)':
    out_r = out_row_begin + y_len
    out_c = out_col_begin
    for row in range(ref_row_begin, ref_row_end+ 1):
        for col in range(ref_col_begin, ref_col_end + 1):
            out_ws.cell(row=out_r, column=out_c).value = ref_ws.cell(row=row, column=col).value
            out_r -= 1
        out_r = out_row_begin + y_len
        out_c += 1
elif option == 'rotate(+90)':
    out_r = out_row_begin
    out_c = out_col_begin + y_len
    for row in range(ref_row_begin, ref_row_end+ 1):
        for col in range(ref_col_begin, ref_col_end + 1):
            out_ws.cell(row=out_r, column=out_c).value = ref_ws.cell(row=row, column=col).value
            out_r += 1
        out_r = out_row_begin
        out_c -= 1
elif option == 'rotate(180)':
    out_r = out_row_begin + y_len
    out_c = out_col_begin + x_len
    for row in range(ref_row_begin, ref_row_end+ 1):
        for col in range(ref_col_begin, ref_col_end + 1):
            out_ws.cell(row=out_r, column=out_c).value = ref_ws.cell(row=row, column=col).value
            out_c -= 1
        out_c = out_col_begin + x_len
        out_r -= 1
wb.save(excel_path)