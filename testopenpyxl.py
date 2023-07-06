from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.utils import range_boundaries
from openpyxl.utils import column_index_from_string
#from openpyxl.utils import coordinate_from_string
from openpyxl.utils import coordinate_to_tuple
from openpyxl.worksheet.table import Table, TableStyleInfo
# from tkinter import *
from tkinter import messagebox
from tkinter import ttk
import getcolumn
from array import *
import tkinter as tk
from tkinter import *
from ttkthemes import ThemedTk, THEMES
from PIL import Image
from PIL import ImageTk, Image
from tkinter.font import Font
from tkinter import filedialog
import gui_function as gui

import os
import win32com.client
from pathlib import Path  # core library

# wb_f = load_workbook(r"C:\Users\sytung\OneDrive - Synopsys, Inc\Desktop\py\Test1.xlsx", data_only= False)
# wb_f.save(r"C:\Users\sytung\OneDrive - Synopsys, Inc\Desktop\py\Test1_temp.xlsx")
# wb_d = load_workbook(r"C:\Users\sytung\OneDrive - Synopsys, Inc\Desktop\py\Test1.xlsx",data_only=True)
# wb_d.save(r"C:\Users\sytung\OneDrive - Synopsys, Inc\Desktop\py\Test1_temp2.xlsx")
# wb_d.close()
# excel_file = os.path.join(r"C:\Users\sytung\OneDrive - Synopsys, Inc\Desktop\py","Test1.xlsx")
# excel = win32com.client.gencache.EnsureDispatch('Excel.Application')
# excel.DisplayAlerts = False # disabling prompts to overwrite existing file
# excel.Workbooks.Open(excel_file )

# excel.ActiveWorkbook.SaveAs("excel_file", FileFormat=51, ConflictResolution=2)
intpath = r"C:\Users\sytung\OneDrive - Synopsys, Inc\Desktop\py\INPUT"
outpath = r"C:\Users\sytung\OneDrive - Synopsys, Inc\Desktop\py\OUTPUT"
xl_file = "Test1.xlsx"
# excel.DisplayAlerts = True # enabling prompts
# excel.ActiveWorkbook.Close()
# wb_convert = load_workbook(r"C:\Users\sytung\OneDrive - Synopsys, Inc\Desktop\py\H137_UCIe_TC_Bump_coordination_official.xlsx",data_only=True)
# wb_convert.save("H137_UCIe_TC_Bump_coordination_valuecopy.xlsx")
# wb_convert.close()
# wb_d = load_workbook("H137_UCIe_TC_Bump_coordination_valuecopy.xlsx")
# wb_f = load_workbook(r"C:\Users\sytung\OneDrive - Synopsys, Inc\Desktop\py\H137_UCIe_TC_Bump_coordination_official.xlsx")



# ws_f = wb_f["Sheet1"]
# ws_d = wb_d["Sheet1"]

def refresh_excel(excelfile):
    excel_file = os.path.join(excelfile)
    excel = win32com.client.gencache.EnsureDispatch('Excel.Application')
    excel.DisplayAlerts = True # disabling prompts to overwrite existing file
    excel.Workbooks.Open(excel_file )
    excel.ActiveWorkbook.Save()
    excel.DisplayAlerts = True # enabling prompts
    excel.ActiveWorkbook.Close()

ball_table = {
    "tb_sheet": "BGA",
    "tb_begin_cell": "J3",
    "tb_end_cell": "M29"
}



def copy_table(cell):
    wb =  load_workbook(r"C:\Users\sytung\OneDrive - Synopsys, Inc\Desktop\py\Pyopenxl\Book1.xlsx")
    row_begin = coordinate_to_tuple(cell['tb_begin_cell'])[0]
    col_begin = coordinate_to_tuple(cell['tb_begin_cell'])[1]
    row_end = coordinate_to_tuple(cell['tb_end_cell'])[0]
    col_end = coordinate_to_tuple(cell['tb_end_cell'])[1]
    ws_create = wb.create_sheet(str(cell['tb_sheet']) + "TEMP")
    ws_create_name = str(cell['tb_sheet']) + "TEMP"
    tb_sheet = wb[cell['tb_sheet']]
    print(row_begin)
    print(col_begin)
    print(row_end)
    print(col_end)
    for i in range (row_begin, row_end + 1):
        for j in range (col_begin, col_end + 1):
            # print({str(tb_sheet[get_column_letter(j) + str(i)].value).replace('=','')})
            if(str(tb_sheet[get_column_letter(j) + str(i)].value).find("=") != -1):
                ws_create[get_column_letter(j) + str(i)].value = f"={cell['tb_sheet']}!({str(tb_sheet[get_column_letter(j) + str(i)].value).replace('=','')})"
            else:
                print(str(get_column_letter(j) + str(i)))
                ws_create[get_column_letter(j) + str(i)].value = f"={cell['tb_sheet']}!{get_column_letter(j) + str(i)}"

    wb.save(r"C:\Users\sytung\OneDrive - Synopsys, Inc\Desktop\py\Pyopenxl\Book1.xlsx")
    wb.close()
    return row_begin,col_begin,row_end,col_end,ws_create_name
ball_temp = copy_table(ball_table)
refresh_excel(r"C:\Users\sytung\OneDrive - Synopsys, Inc\Desktop\py\Pyopenxl\Book1.xlsx")
wb_d = load_workbook(r"C:\Users\sytung\OneDrive - Synopsys, Inc\Desktop\py\Pyopenxl\Book1.xlsx", data_only=True)
ws_d = wb_d[ball_temp[4]]
print(ws_d['M4'].value)
# wstemp = wb["BALL"]
# ws2 = wb.create_sheet('BALL_temp')
# print(ws_f['D4'].value)
# print(ws_d['D4'].value)
# if(ws_d['D4'].value == 79):
    
#     ws_f['E5'].value = f"=({str(ws_f['D4'].value).replace('=','')})+100"
#     print(ws_f['E5'].value)

# print(ws_d['D5'].value)
# print(ws_d['E5'].value)
# wb_f.save(r"C:\Users\sytung\OneDrive - Synopsys, Inc\Desktop\py\Test1.xlsx")