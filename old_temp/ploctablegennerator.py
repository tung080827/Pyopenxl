from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.utils import range_boundaries
from openpyxl.utils import column_index_from_string
#from openpyxl.utils import coordinate_from_string
from openpyxl.utils import coordinate_to_tuple
from openpyxl.worksheet.table import Table, TableStyleInfo
# from tkinter import *
from tkinter import messagebox
from tkinter import ttk
import getcolumn
from array import *
import tkinter as tk
from tkinter import *
from ttkthemes import ThemedTk, THEMES
from PIL import Image
from PIL import ImageTk, Image
from tkinter.font import Font
from tkinter import filedialog
# adv = 1

# label.pack(padx=40,pady=40)
# Create a style
root = ThemedTk()
# my_canvas=tk.Canvas(root)
root.set_theme("scidpurple")

root.title("PLOC TABLE GENERATOR")
root.geometry("1000x1000+30+100")
root.resizable(width=False, height=False)
root.iconbitmap(r".\mylogo.ico")
root.option_add("*tearOff", False) # This is always a good idea
# my_canvas.pack(side='left', fill='both', expand=1)
# root.configure(bg= "blue")
# root.resizable(width=False, height=False)

# Create lists for the Comboboxes

# Define Background Image
bg = ImageTk.PhotoImage(file=r".\bg3_1.png")
open_imag = PhotoImage(file = r".\open-folder.png")

# Define Canvas
my_canvas = tk.Canvas(root, width=800, height=800, bd=0, highlightthickness=0)
my_canvas.pack(fill="both", expand=True)

# Put the image on the canvas
my_canvas.create_image(0,0, image=bg, anchor="nw")
# Make the app responsive
# root.columnconfigure(index=0, weight=1)
# root.columnconfigure(index=1, weight=1)
# root.columnconfigure(index=2, weight=1)
# root.columnconfigure(index=3, weight=1)

# root.grid_columnconfigure(0, weight=1)
# root.grid_rowconfigure(0, weight=1)
# root.rowconfigure(index=0, weight=1)
# root.rowconfigure(index=1, weight=1)
# root.rowconfigure(index=2, weight=1)
# root.rowconfigure(index=3, weight=1)
# root.rowconfigure(index=4, weight=1)
# root.rowconfigure(index=5, weight=1)
# root.rowconfigure(index=6, weight=1)
# root.rowconfigure(index=7, weight=1)


# Create a style
# style = ttk.Style()

stfont= ("Franklin Gothic Medium", 10, 'underline', "italic")
# Create lists for the Comboboxes
theme_list = ["adapta", "aquativo", "arc", "black","blue", "breeze", "clearlooks", "elegance", "equilux", "itft1", "keramik", "keramik_alt", "kroc", "plastik", "radiance", "ubuntu", "scidblue", "scidgreen", "scidgrey", "scidmint", "scidpink", "scidpurple", "scidsand", "smog", "winxpblue", "yaru" ]
package_list = ["S-Organic", "A-CoWoS", "A-EMIB"]
foundry_list = ["TSMC-MapWSR", "TSMC-MapWoSR", "SS-MapWSR", "SS-MapWoSR", "GF-MapWSR", "GF-MapWSR"]

# Create control variables
a = tk.BooleanVar()
b = tk.BooleanVar(value=True)
c = tk.BooleanVar()
d = tk.IntVar(value=2)
# e = tk.StringVar(value=option_menu_list[1])
f = tk.BooleanVar()
g = tk.DoubleVar(value=75.0)
h = tk.BooleanVar()
tc_opt = tk.IntVar()

#Define a Function to enable the frame
def round_rectangle(x1, y1, x2, y2, radius=25, **kwargs):
        
    points = [x1+radius, y1,
              x1+radius, y1,
              x2-radius, y1,
              x2-radius, y1,
              x2, y1,
              x2, y1+radius,
              x2, y1+radius,
              x2, y2-radius,
              x2, y2-radius,
              x2, y2,
              x2-radius, y2,
              x2-radius, y2,
              x1+radius, y2,
              x1+radius, y2,
              x1, y2,
              x1, y2-radius,
              x1, y2-radius,
              x1, y1+radius,
              x1, y1+radius,
              x1, y1]

    return my_canvas.create_polygon(points, **kwargs, smooth=True)

def enable(children):
   for child in children:
      child.configure(state='enable')
def disable(children):  
    for child in children:
        child.configure(state='disable')
def entry_disable(*entries):
    for entry in entries:
        entry.config(state='disable')
def entry_enable(*entries):
    for entry in entries:
        entry.config(state='normal')
def entry_toggle():
    print("TUng day")
        # if(entry['state'] == 'disable'):
    if(tc_opt.get() == 1):
        foundry_combo.config(state='normal')
    elif(tc_opt.get() == 0):
        foundry_combo.config(state='disable')
def progress_bar(value):
    progress['value'] = value
    root.update_idletasks()

def choosetheme(event):
    for theme in theme_list:
        if (theme_combo.get() == theme):
            root.set_theme(theme)
def choosemode(event):   
    if(package_combo.get() == "S-Organic"):
       
       entry_disable(cor1_x1y1, cor1_x2y2, cor1_Xget, cor1_Yget,
                     cor2_x1y1, cor2_x2y2, cor2_Xget, cor2_Yget,
                     cor3_x1y1, cor3_x2y2, cor3_Xget, cor3_Yget,
                     cor4_x1y1, cor4_x2y2, cor4_Xget, cor4_Yget)
       entry_disable(c4_x1y1_i,c4_x2y2_i, c4_Xget_i, c4_Yget_i, c4_tb_name, c4_col,
                      u_col, u_tb_name, u_x1y1_i, u_x2y2_i, u_Xget_i, u_Yget_i)
       entry_enable(x1y1_i, x2y2_i, Xget_i, Yget_i)
       entry_enable(out_name_in, out_col_i)
       entry_disable(sheete_i, sheete_t)
       sheet_t['text']= "Bump sheet:"
        
       print(1)
    elif(package_combo.get() == "A-CoWoS"):
    #    enable(dmbump_frame.winfo_children())
        # cor1_x1y1.config(state='disable')
        entry_enable(cor1_x1y1, cor1_x2y2, cor1_Xget, cor1_Yget,
                     cor2_x1y1, cor2_x2y2, cor2_Xget, cor2_Yget,
                     cor3_x1y1, cor3_x2y2, cor3_Xget, cor3_Yget,
                     cor4_x1y1, cor4_x2y2, cor4_Xget, cor4_Yget)
        entry_enable(x1y1_i, x2y2_i, Xget_i, Yget_i)
        entry_disable(sheete_i, sheete_t)
        entry_enable(out_name_in, out_col_i)
        entry_disable(c4_x1y1_i,c4_x2y2_i, c4_Xget_i, c4_Yget_i, c4_tb_name, c4_col,
                      u_col, u_tb_name, u_x1y1_i, u_x2y2_i, u_Xget_i, u_Yget_i)
        sheet_t['text']= "Bump sheet:"
        print(0)
    else:
        entry_disable(cor1_x1y1, cor1_x2y2, cor1_Xget, cor1_Yget,
                     cor2_x1y1, cor2_x2y2, cor2_Xget, cor2_Yget,
                     cor3_x1y1, cor3_x2y2, cor3_Xget, cor3_Yget,
                     cor4_x1y1, cor4_x2y2, cor4_Xget, cor4_Yget)
        entry_disable(x1y1_i, x2y2_i, Xget_i, Yget_i)
        entry_disable(out_name_in, out_col_i)
        entry_enable(c4_x1y1_i,c4_x2y2_i, c4_Xget_i, c4_Yget_i, c4_tb_name, c4_col,
                      u_col, u_tb_name, u_x1y1_i, u_x2y2_i, u_Xget_i, u_Yget_i)
        entry_enable(sheete_i, sheete_t)
        sheet_t['text']= "uBump sheet:"

        popup("The EMIB package type have not developed yet, Please use S-Organic to gen 2 times (for C4 and uBump) instead!")

myLabel = ttk.Label(root,text="Info:")
myLabel_w =my_canvas.create_window(80,770,anchor="nw", window=myLabel)

frame = tk.Label(root, bg="#c9f2dc", font=("Courier New", 10), foreground="#f2a50a")
my_canvas.create_window(600, 80, window=frame, anchor="nw", width= 280, height=100)

def mynotif(content):
    if(content == ""):
        myLabel.configure(text="", anchor='w')
    else:
        myLabel.configure(text=content, anchor='w')
        # myLabel = ttk.Label(root,text=content)
        # myLabel_w =my_canvas.create_window(80,750,anchor="nw", window=myLabel)
        # myLabel.grid(row=5, column=0, columnspan=2, padx=(20, 10), pady=(20, 10), sticky="nsew")
# ------------------------------------------------------------------------------------------------------------------------------------------------

def myguide(entries, content):
    if(content == ""):
        entries.configure(text="")
       
    else:
        entries.configure(text=content)

def handle_click(event):
    x = c4_x1y1_i.get()
    c4_x1y1_i.delete(0,END)
    print(x)
def x1y1_guide(event):
     myguide(frame, "INFO:" + "Die window start cell\n\n - Example:   A0           ")
def un_guide(event):
     myguide(frame,"")

def x2y2_guide(event):
     myguide(frame, "INFO:" + "Die window end cell\n\n - Example:   CU100       ")
def Xget_guide(event):
     myguide(frame, "INFO:" + "Row contains X axis value     \n  which is X location of Bump.\n Must be interger           \n\n - Example: 8                       ")
def Yget_guide(event):
     myguide(frame, "INFO:" + "Row contains Y axis value  \nwhich is Y location of Bump.\n Must be Excel column format\n\n Example: CU ")
def out_name_in_guide(event):
    myguide(frame, "INFO:" + "This field to define the\n   output table name    ")
def out_name2_in_guide(event):
    myguide(frame, "INFO:" + "This field to define the \n  output table 2 name.\n Use for TC with 2 option\n with/without sealring ")
def out_col_in_guide(event):
    myguide(frame, "INFO:" + "This field to define the\n     output table location. \n\n - Example: O64 ")
def out_col_wsr_i_guide(event):
    myguide(frame, "INFO:" + "This field to define the\n   output table 2 location.\n. Use for TC with 2 option\n with/without sealring\n\n - Example: U64 ")  
def dummystart_guide(event):
    myguide(frame, "INFO:" + "Dummy bump window start cell.\n\n - Example:   A0                   ")  
def dummyend_guide(event):
    myguide(frame, "INFO:" + "Dummy bump window end cell.\n\n - Example:   E3                   ") 
def dummy_Xget_guide(event):
     myguide(frame, "INFO:" + "Row contains X axis values              \n  which is X location of dummy Bump.\n Must be interger                 \n\n- Example: 8                                      ")
def dummy_Yget_guide(event):
     myguide(frame, "INFO:" + "Row contains Y axis value     \nwhich is Y location of dummy Bump. \n Must be Excel column format\n\n- Example: CU               ")

        
xfont = ("System", 12, "bold", 'underline', 'italic')
theme_combo_t = ttk.Label(root,text="Choose theme:",border=20, font=xfont, background='#b434eb', borderwidth=3)
theme_combo_t_w = my_canvas.create_window(750, 15, window=theme_combo_t)

theme_combo = ttk.Combobox(root, state="readonly", values=theme_list, width=15)
theme_combo_w = my_canvas.create_window(870,15, window=theme_combo)
theme_combo.current(0)
theme_combo.bind('<<ComboboxSelected>>', choosetheme)

pfont= ("Rosewood Std Regular", 12, "bold", 'underline' )
excel_t = ttk.Label(root,text="PLOC path:",border=20,font=pfont, borderwidth=5)
# excel_t.configure(background='#b434eb', foreground="black")
excel_t_w = my_canvas.create_window(30,40, anchor="nw", window=excel_t)
# my_canvas.create_text(30,40, text="PLOC path:", anchor="nw",font=pfont, fill="#b434eb")
excel_i = ttk.Entry(root, width=115)
excel_i.insert(0, r"C:\Users\sytung\OneDrive - Synopsys, Inc\Desktop\py\Book1.xlsx")
excel_i_w = my_canvas.create_window(150,40, anchor="nw", window=excel_i)

sheet_t = ttk.Label(root,text="Sheet name:",border=20,font=pfont, borderwidth=3)
# sheet_t.configure(background='#b434eb', foreground="black")
sheet_t_w = my_canvas.create_window(30,80, anchor="nw", window=sheet_t)

sheet_i = ttk.Entry(root, background="#217346", width=20)
sheet_i.insert(0, "Sheet1")
sheet_i_w = my_canvas.create_window(150,80, anchor="nw", window=sheet_i)

sheete_t = ttk.Label(root,text="C4 sheet:",border=20,font=pfont, borderwidth=3)
# sheete_t.configure(background='#b434eb', foreground="black")
sheete_t_w = my_canvas.create_window(300,80, anchor="nw", window=sheete_t)

sheete_i = ttk.Entry(root, background="#217346", width=20)
sheete_i.insert(0, "C4 sheet")
sheete_i_w = my_canvas.create_window(400,80, anchor="nw", window=sheete_i)


pkg_t = ttk.Label(root,text="Package type:",border=20,font=pfont, borderwidth=3)
# pkg_t.configure(background='#b434eb', foreground="black")
pkg_t_w = my_canvas.create_window(30,120, anchor="nw", window=pkg_t)

package_combo = ttk.Combobox(root, state="readonly", values=package_list, width=17)
package_combo_w = my_canvas.create_window(150,120, anchor="nw", window=package_combo)
package_combo.current(0)
package_combo.bind('<<ComboboxSelected>>', choosemode)

sr_opt = ttk.Checkbutton(root, text="For TC", variable=tc_opt,command= entry_toggle)
sr_opt_w =my_canvas.create_window(300, 120, anchor="nw", window=sr_opt)




# foundry_t = ttk.Label(ploc_frame,text="Foundary:",border=20,font=pfont, borderwidth=3)
# foundry_t.grid(row=2, column=3, padx=5, pady=(0, 10), sticky="ew", ipady=10)
foundry_combo = ttk.Combobox(root, state="readonly", values=foundry_list, width=17)
foundry_combo_w = my_canvas.create_window(400, 120, anchor="nw", window=foundry_combo)
foundry_combo.current(0)

# foundry_combo.bind('<<ComboboxSelected>>', choosemode)


# Separator
separator = ttk.Separator(root)
separator_w = my_canvas.create_window(30, 130, anchor="nw", window=separator)
# separator.grid(row=1, column=0, padx=(20, 10), pady=10, sticky="ew")

# my_rectangle = round_rectangle(20, 200, 220,220, radius=10, fill ="#b434eb", outline = "#0b30e6", width = 2)
my_canvas.create_text(30, 200, text="Die bump map config", anchor="nw",font=("Helvetica", 10, 'italic', 'underline', 'bold'), fill="#b434eb")

# x1y1_t = ttk.Label(bumpvisual_frame,text="Die start:",border=20, borderwidth=3)
# x1y1_t.grid(row=0, column=0, padx=5, pady=(0, 10), sticky="ew")
x1y1_i = ttk.Entry(root, width=20)
x1y1_i_w = my_canvas.create_window(150, 230, anchor="nw", window=x1y1_i)
x1y1_i.insert(0, "C10")
x1y1_i.bind('<FocusIn>', x1y1_guide)
x1y1_i.bind('<FocusOut>', un_guide)


# x2y2_t = ttk.Label(bumpvisual_frame,text="Die end:",b1order=20, borderwidth=3)
# x2y2_t.grid(row=0, column=1, padx=5, pady=(0, 10), sticky="ew")
x2y2_i = ttk.Entry(root, width=20)
x2y2_i_w = my_canvas.create_window(300, 230, anchor="nw", window=x2y2_i)
x2y2_i.insert(0, "O29")
x2y2_i.bind('<FocusIn>', x2y2_guide)
x2y2_i.bind('<FocusOut>', un_guide)

# Xget_t = ttk.Label(bumpvisual_frame,text="Row contains X:",border=20, borderwidth=3)
# Xget_t.grid(row=3, column=0, padx=5, pady=(0, 10), sticky="ew")
Xget_i = ttk.Entry(root, width=20)
Xget_i_w = my_canvas.create_window(150, 270, anchor="nw", window=Xget_i)
Xget_i.insert(0, "30")
Xget_i.bind('<FocusIn>', Xget_guide)
Xget_i.bind('<FocusOut>', un_guide)

# Yget_t = ttk.Label(bumpvisual_frame,text="Column contains Y:",border=20, borderwidth=3)
# Yget_t.grid(row=3, column=1, padx=5, pady=(0, 10), sticky="ew")
Yget_i = ttk.Entry(root, width=20)
Yget_i_w = my_canvas.create_window(300, 270, anchor="nw", window=Yget_i)
Yget_i.insert(0, "B")
Yget_i.bind('<FocusIn>', Yget_guide)
Yget_i.bind('<FocusOut>', un_guide)
# emib_tb_t = ttk.Label(root,text="EMIB:")
my_canvas.create_text(50, 300, text="EMIB:", anchor="nw",font=("Helvetica", 10, 'italic', 'underline', 'bold'), fill="black")

c4_x1y1_i = ttk.Entry(root, width=20)
c4_x1y1_i_w = my_canvas.create_window(150, 320, anchor="nw", window=c4_x1y1_i)
c4_x1y1_i.insert(0, "C4 window top-left")
c4_x1y1_i.bind("<1>", handle_click)


# x2y2_t = ttk.Label(bumpvisual_frame,text="Die end:",border=20, borderwidth=3)
# x2y2_t.grid(row=0, column=1, padx=5, pady=(0, 10), sticky="ew")
c4_x2y2_i = ttk.Entry(root, width=20)
c4_x2y2_i_w = my_canvas.create_window(300, 320, anchor="nw", window=c4_x2y2_i)
c4_x2y2_i.insert(0, "C4 window bot-right")


# Xget_t = ttk.Label(bumpvisual_frame,text="Row contains X:",border=20, borderwidth=3)
# Xget_t.grid(row=3, column=0, padx=5, pady=(0, 10), sticky="ew")
c4_Xget_i = ttk.Entry(root, width=20)
c4_Xget_i_w = my_canvas.create_window(150, 360, anchor="nw", window=c4_Xget_i)
c4_Xget_i.insert(0, "Row contains C4 X value")


# Yget_t = ttk.Label(bumpvisual_frame,text="Column contains Y:",border=20, borderwidth=3)
# Yget_t.grid(row=3, column=1, padx=5, pady=(0, 10), sticky="ew")
c4_Yget_i = ttk.Entry(root, width=20)
c4_Yget_i_w = my_canvas.create_window(300, 360, anchor="nw", window=c4_Yget_i)
c4_Yget_i.insert(0, "Column contains C4 Y value")

# ------------
u_x1y1_i = ttk.Entry(root, width=20)
u_x1y1_i_w = my_canvas.create_window(150, 400, anchor="nw", window=u_x1y1_i)
u_x1y1_i.insert(0, "uBump window top-left")


# x2y2_t = ttk.Label(bumpvisual_frame,text="Die end:",border=20, borderwidth=3)
# x2y2_t.grid(row=0, column=1, padx=5, pady=(0, 10), sticky="ew")
u_x2y2_i = ttk.Entry(root)
u_x2y2_i_w = my_canvas.create_window(300, 400, anchor="nw", window=u_x2y2_i)
u_x2y2_i.insert(0, "uBump window bot-right")


# Xget_t = ttk.Label(bumpvisual_frame,text="Row contains X:",border=20, borderwidth=3)
# Xget_t.grid(row=3, column=0, padx=5, pady=(0, 10), sticky="ew")
u_Xget_i = ttk.Entry(root, width=20)
u_Xget_i_w = my_canvas.create_window(150, 440, anchor="nw", window=u_Xget_i)
u_Xget_i.insert(0, "Row contains uBump X value")


# Yget_t = ttk.Label(bumpvisual_frame,text="Column contains Y:",border=20, borderwidth=3)
# Yget_t.grid(row=3, column=1, padx=5, pady=(0, 10), sticky="ew")
u_Yget_i = ttk.Entry(root, width=20)
u_Yget_i_w = my_canvas.create_window(300, 440, anchor="nw", window=u_Yget_i)
u_Yget_i.insert(0, "Column contains uBump Y value")


# out_table_frame = ttk.LabelFrame(root, text="Output Bump table config", padding=(20, 10))
# my_rectangle = round_rectangle(480, 200, 980,480, radius=20, fill ="", outline = "#0b30e6", width = 2 )

my_canvas.create_text(500, 200, text="Output table config", anchor="nw",font=("Helvetica", 10, 'italic', 'underline', 'bold'), fill="#b434eb")
# out_name = ttk.Label(root,text="Bump table name:")

out_name_in = ttk.Entry(root, width=20)
out_name_in_w = my_canvas.create_window(600, 230, anchor="nw", window=out_name_in)
out_name_in.insert(0, "Name")
out_name_in.bind('<FocusIn>', out_name_in_guide)
out_name_in.bind('<FocusOut>', un_guide)
out_name2_in = ttk.Entry(root, width=20)
out_name2_in_w = my_canvas.create_window(750, 230, anchor="nw", window=out_name2_in)
out_name2_in.insert(0, "Name WO SR")
out_name2_in.bind('<FocusIn>', out_name2_in_guide)
out_name2_in.bind('<FocusOut>', un_guide)

# out_col_t = ttk.Label(root,text="Out table without SR:")

out_col_i = ttk.Entry(root, width=20)
out_col_i_w = my_canvas.create_window(600, 270, anchor="nw", window=out_col_i)
out_col_i.insert(0, "H33")
out_col_i.bind('<FocusIn>', out_col_in_guide)
out_col_i.bind('<FocusOut>', un_guide)

# out_col_wsr_t = ttk.Label(root,text="Out table with SR:")

out_col_wsr_i = ttk.Entry(root)
out_col_wsr_w = my_canvas.create_window(750, 270, anchor="nw", window=out_col_wsr_i)
out_col_wsr_i.insert(0, "T64")
out_col_wsr_i.bind('<FocusIn>', out_col_wsr_i_guide)
out_col_wsr_i.bind('<FocusOut>', un_guide)

# ---------------------------------------gui for EMIB-------------------------------------------------
separator1 = ttk.Separator(root)

separator2 = ttk.Separator(root)


# emib_t = ttk.Label(root,text="EMIB:",border=20, borderwidth=3)
my_canvas.create_text(550, 300, text="EMIB:", anchor="nw",font=("Helvetica", 10, 'italic', 'underline', 'bold'), fill="black")

c4_tb_name = ttk.Entry(root)
c4_tb_name_w = my_canvas.create_window(600, 320, anchor="nw", window=c4_tb_name)
c4_tb_name.insert(0, "C4 Name")


# out_col_t = ttk.Label(out_table_frame,text="Out table location:")
# out_col_t.grid(row=1, column=0, padx=5, pady=(0, 10), sticky="ew")
c4_col = ttk.Entry(root, width=20)
c4_col_w = my_canvas.create_window(750, 320, anchor="nw", window=c4_col)
c4_col.insert(0, "C4 location")


u_tb_name = ttk.Entry(root, width=20)
u_tb_name_w = my_canvas.create_window(600, 360, anchor="nw", window=u_tb_name)
u_tb_name.insert(0, "uBump Name")


u_col = ttk.Entry(root, width=20)
u_col_w = my_canvas.create_window(750, 360, anchor="nw", window=u_col)
u_col.insert(0, "uBump location")


# ------------------------------
separator1 = ttk.Separator(root)

separator2 = ttk.Separator(root)




# out_col.grid(row=0, column=0, padx=5, pady=(0, 10), sticky="ew")
# out_row = ttk.Entry(bumpvisual_frame)
# out_row.insert(0, "X axis value get")
# out_row.grid(row=0, column=1, padx=5, pady=(0, 10), sticky="ew")

#--------------------------------------------------------------------------------------------------------#
# my_rectangle = round_rectangle(20, 500, 980,760, radius=20, fill ="", outline = "#0b30e6", width = 2 )
# dmbump_frame = ttk.LabelFrame(root, text="Dummy bump config", padding=(20, 10))
my_canvas.create_text(30, 500, text="Bummy bump config", anchor="nw",font=("Helvetica", 10, 'italic', 'underline', 'bold'), fill="#b434eb")

# dmbump_cor1_frame = ttk.LabelFrame(dmbump_frame, text="Corner 1 config", padding=(20, 10))

my_canvas.create_text(60, 540, text="Corner 1 config", anchor="nw",font=("Helvetica", 10, 'italic', 'underline', 'bold'), fill="black")
cor1_x1y1 = ttk.Entry(root, width=20)
cor1_x1y1_w = my_canvas.create_window(150, 560, anchor="nw", window=cor1_x1y1)
cor1_x1y1.insert(0, "window top-left")
cor1_x1y1.bind('<FocusIn>', dummystart_guide)
cor1_x1y1.bind('<FocusOut>', un_guide)

cor1_x2y2 = ttk.Entry(root, width=20)
cor1_x2y2_w = my_canvas.create_window(300, 560, anchor="nw", window=cor1_x2y2)
cor1_x2y2.insert(0, "window bot-right")
cor1_x2y2.bind('<FocusIn>', dummyend_guide)
cor1_x2y2.bind('<FocusOut>', un_guide)

cor1_Xget = ttk.Entry(root,width=20)
cor1_Xget_w = my_canvas.create_window(150, 600, anchor="nw", window=cor1_Xget)
cor1_Xget.insert(0, "Row contains X")
cor1_Xget.bind('<FocusIn>', dummy_Xget_guide)
cor1_Xget.bind('<FocusOut>', un_guide)

cor1_Yget = ttk.Entry(root, width=20)
cor1_Yget_w = my_canvas.create_window(300, 600, anchor="nw", window=cor1_Yget)
cor1_Yget.insert(0, "Column contains Y")
cor1_Yget.bind('<FocusIn>', dummy_Yget_guide)
cor1_Yget.bind('<FocusOut>', un_guide)
#---------------------------------------------------------------------------------------------------------#

# dmbump_cor2_frame = ttk.LabelFrame(dmbump_frame, text="Corner 2 config", padding=(20, 10))
my_canvas.create_text(560, 540, text="Corner 2 config", anchor="nw",font=("Helvetica", 10, 'italic', 'underline', 'bold'), fill="black")

cor2_x1y1 = ttk.Entry(root, width=20)
cor2_x1y1_w = my_canvas.create_window(600, 560, anchor="nw", window=cor2_x1y1)
cor2_x1y1.insert(0, "window top-left")
cor2_x1y1.bind('<FocusIn>', dummystart_guide)
cor2_x1y1.bind('<FocusOut>', un_guide)

cor2_x2y2 = ttk.Entry(root, width=20)
cor2_x2y2_w = my_canvas.create_window(750, 560, anchor="nw", window=cor2_x2y2)
cor2_x2y2.insert(0, "window bot-right")
cor2_x2y2.bind('<FocusIn>', dummyend_guide)
cor2_x2y2.bind('<FocusOut>', un_guide)

cor2_Xget = ttk.Entry(root, width=20)
cor2_Xget_w = my_canvas.create_window(600, 600, anchor="nw", window=cor2_Xget)
cor2_Xget.insert(0, "Row contains X")
cor2_Xget.bind('<FocusIn>', dummy_Xget_guide)
cor2_Xget.bind('<FocusOut>', un_guide)

cor2_Yget = ttk.Entry(root, width=20)
cor2_Yget_w = my_canvas.create_window(750, 600, anchor="nw", window=cor2_Yget)
cor2_Yget.insert(0, "Column contains Y")
cor2_Yget.bind('<FocusIn>', dummy_Yget_guide)
cor2_Yget.bind('<FocusOut>', un_guide)

#--------------------------------------------------------------------------------------------------------#
# dmbump_cor3_frame = ttk.LabelFrame(dmbump_frame, text="Corner 3 config", padding=(20, 10))

my_canvas.create_text(60, 660, text="Corner 3 config", anchor="nw",font=("Helvetica", 10, 'italic', 'underline', 'bold'), fill="black")
cor3_x1y1 = ttk.Entry(root, width=20)
cor3_x1y1_w = my_canvas.create_window(150, 680, anchor="nw", window=cor3_x1y1)
cor3_x1y1.insert(0, "window top-left")
cor3_x1y1.bind('<FocusIn>', dummystart_guide)
cor3_x1y1.bind('<FocusOut>', un_guide)

cor3_x2y2 = ttk.Entry(root, width=20)
cor3_x2y2_w = my_canvas.create_window(300, 680, anchor="nw", window=cor3_x2y2)
cor3_x2y2.insert(0, "window bot-right")
cor3_x2y2.bind('<FocusIn>', dummyend_guide)
cor3_x2y2.bind('<FocusOut>', un_guide)

cor3_Xget = ttk.Entry(root)
cor3_Xget_w = my_canvas.create_window(150, 720, anchor="nw", window=cor3_Xget)
cor3_Xget.insert(0, "Row contains X")
cor3_Xget.bind('<FocusIn>', dummy_Xget_guide)
cor3_Xget.bind('<FocusOut>', un_guide)

cor3_Yget = ttk.Entry(root)
cor3_Yget_w = my_canvas.create_window(300, 720, anchor="nw", window=cor3_Yget)
cor3_Yget.insert(0, "Column contains Y")
cor3_Yget.bind('<FocusIn>', dummy_Yget_guide)
cor3_Yget.bind('<FocusOut>', un_guide)

#--------------------------------------------------------------------------------------------------------#
# dmbump_cor4_frame = ttk.LabelFrame(dmbump_frame, text="Corner 4 config", padding=(20, 10))

my_canvas.create_text(560, 660, text="Corner 4 config", anchor="nw",font=("Helvetica", 10, 'italic', 'underline', 'bold'), fill="black")
cor4_x1y1 = ttk.Entry(root, width=20)
cor4_x1y1_w = my_canvas.create_window(600, 680, anchor="nw", window=cor4_x1y1)
cor4_x1y1.insert(0, "window top-left")
cor4_x1y1.bind('<FocusIn>', dummystart_guide)
cor4_x1y1.bind('<FocusOut>', un_guide)

cor4_x2y2 = ttk.Entry(root, width=20)
cor4_x2y2_w = my_canvas.create_window(750, 680, anchor="nw", window=cor4_x2y2)
cor4_x2y2.insert(0, "window bot-right")
cor4_x2y2.bind('<FocusIn>', dummyend_guide)
cor4_x2y2.bind('<FocusOut>', un_guide)

cor4_Xget = ttk.Entry(root, width=20)
cor4_Xget_w = my_canvas.create_window(600, 720, anchor="nw", window=cor4_Xget)
cor4_Xget.insert(0, "Row contains X")
cor4_Xget.bind('<FocusIn>', dummy_Xget_guide)
cor4_Xget.bind('<FocusOut>', un_guide)

cor4_Yget = ttk.Entry(root, width=20)
cor4_Yget_w = my_canvas.create_window(750, 720, anchor="nw", window=cor4_Yget)
cor4_Yget.insert(0, "Column contains Y")
cor4_Yget.bind('<FocusIn>', dummy_Yget_guide)
cor4_Yget.bind('<FocusOut>', un_guide)

#--------------------------------------------------------------------------------------------------------#

my_canvas.create_text(880,980, text= "Internal contact: sytung@synopsys.com" ,font=("Helvetica", 8, 'underline'), fill="grey")
def open():
	# global my_image
    root.filename = filedialog.askopenfilename(initialdir="./", title="Select A File", filetypes=(("excel files", "*.xlsx"),("all files", "*.*")))
    excel_i.delete(0,END)
    print(root.filename) 
    excel_i.insert(0, root.filename)

	# my_image = ImageTk.PhotoImage(Image.open(root.filename))
	# my_image_label = Label(image=my_image).pack()

#---------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------#
def popup(notif):
    messagebox.showinfo("Notification", notif)

def show_error(error):
    messagebox.showerror("showerror", error)
#---------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------#
def get_path():
    # popup("Generating...")
    # button['state'] = tk.DISABLED

    mynotif("Processing the input parameter...")
    button['text']="Generating..."
    
    progress_bar(20)
    

    global excel_path 
 
    excel_path = excel_i.get()
    excel_sheet = sheet_i.get()

    corner1_w1 = cor1_x1y1.get()
    corner1_w2 = cor1_x2y2.get()
    corner1_Xget = cor1_Xget.get()
    corner1_Yget = cor1_Yget.get()

    corner2_w1 = cor2_x1y1.get()
    corner2_w2 = cor2_x2y2.get()
    corner2_Xget = cor2_Xget.get()
    corner2_Yget = cor2_Yget.get()

    corner3_w1 = cor3_x1y1.get()
    corner3_w2 = cor3_x2y2.get()
    corner3_Xget = cor3_Xget.get()
    corner3_Yget = cor3_Yget.get()

    corner4_w1 = cor4_x1y1.get()
    corner4_w2 = cor4_x2y2.get()
    corner4_Xget = cor4_Xget.get()
    corner4_Yget = cor4_Yget.get()

    die_x1y1 = x1y1_i.get()
    die_x2y2 = x2y2_i.get()
    die_x_get = Xget_i.get()
    die_y_get = Yget_i.get()

    die_params=[]
    die_params.append(die_x1y1)
    die_params.append(die_x2y2)
    die_params.append(die_x_get)
    die_params.append(die_y_get)

    out_table_params=[]
    out_table_params.append(out_name_in.get())
    out_table_params.append(out_col_i.get())
    out_table_params.append(out_name2_in.get())
    out_table_params.append(out_col_wsr_i.get())

    dummy_params=[]
    dummy_params.append(corner1_w1)
    dummy_params.append(corner1_w2)
    dummy_params.append(corner1_Xget)
    dummy_params.append(corner1_Yget)
    dummy_params.append(corner2_w1)
    dummy_params.append(corner2_w2)
    dummy_params.append(corner2_Xget)
    dummy_params.append(corner2_Yget)
    dummy_params.append(corner3_w1)
    dummy_params.append(corner3_w2)
    dummy_params.append(corner3_Xget)
    dummy_params.append(corner3_Yget)
    dummy_params.append(corner4_w1)
    dummy_params.append(corner4_w2)
    dummy_params.append(corner4_Xget)
    dummy_params.append(corner4_Yget)
  
    print(dummy_params[0])
    # print(excel_path)

    # print(corner1_w1, corner1_w2, corner1_Xget,corner1_Yget)
    # print(corner2_w1, corner2_w2, corner2_Xget,corner2_Yget)
    # print(corner3_w1, corner3_w2, corner3_Xget,corner3_Yget)
    # print(corner4_w1, corner4_w2, corner4_Xget,corner4_Yget)

    package = package_combo.get()
    print(package)
    if (package == "A-CoWoS"):
        print(1)
        package_type = 1
    elif(package == "S-Organic"):
        # dmbump_frame.state = tk.DISABLED
        package_type = 0
    else:
        package_type = 0

    isTC = tc_opt.get()

    if(isTC == 1 and foundry_combo.get() == "TSMC-MapWSR"):
        opt_sr = 1
    elif(isTC == 1 and foundry_combo.get() == "TSMC-MapWoSR"):
        opt_sr = 2
    elif(isTC == 3 and foundry_combo.get() == "SS-MapWSR"):
        opt_sr = 4
    elif(isTC == 1 and foundry_combo.get() == "SS-MapWoSR"):
        opt_sr = 5
    elif(isTC == 1 and foundry_combo.get() == "GF-MapWoSR"):
        opt_sr = 6
    elif(isTC == 1 and foundry_combo.get() == "GF-MapWoSR"):
        opt_sr = 7
    else:
        opt_sr = 0
    generate_bump_table(excel_path, excel_sheet, package_type, out_table_params, die_params, dummy_params, opt_sr )
    button['text']="Generate"
    

# nofi = ttk.Entry(root,)
def generate_bump_table(excel_path, excel_sheet, package_type, out_table_params, die_params, dummy_params, opt_sr ):

   
    die_size={
        "width": 5080.68,
        "height": 2778.84
    }
# Bump table config 
    table={
        "name": out_table_params[0],
        "location": out_table_params[1],
        "name_wsr": out_table_params[2],
        "location_wsr": out_table_params[3],
        
    }

    #---Bump map visual view parameter---#
    coordinate = {
        
        "window1": die_params[0], #Top Left of Bump map visual view
        "window2": die_params[1], #Bottom Right of Bump map visual view
        "xcoor": die_params[2], #This define row where Xaxis value can be got
        "ycoor": die_params[3] #This define row where Yaxis value can be got
    }

    #---Dummy Bump visual view parameter---#
    dummybump={
        "corner_1":{
            "window1": dummy_params[0],
            "window2": dummy_params[1],
            "xcoor": dummy_params[2],
            "ycoor": dummy_params[3]
            },
        "corner_2":{
          
            "window1": dummy_params[4],
            "window2": dummy_params[5],
            "xcoor": dummy_params[6],
            "ycoor": dummy_params[7]
        },
        "corner_3":{
         
            "window1": dummy_params[8],
            "window2": dummy_params[9],
            "xcoor": dummy_params[10],
            "ycoor": dummy_params[11]
        },
        "corner_4":{         
            "window1": dummy_params[12],
            "window2": dummy_params[13],
            "xcoor": dummy_params[14],
            "ycoor": dummy_params[15]
        }

    }


    mynotif("")
    root.update_idletasks()
    mynotif("Loading the ploc file...")
    root.update_idletasks()
    try:
        wb = load_workbook(excel_path)
        print(wb)   
    except:
        print("Wrong Ploc path or Ploc file is openning. Please recheck/close the PLOC file before generate :(")
        show_error("Wrong Ploc path or Ploc file is openning. Please recheck/close the PLOC file before generate :(")
        progress_bar(0)
        mynotif("Error")
        root.update_idletasks()
    
    # ws = wb.create_sheet('Tung')
    try:
       ws1 = wb[excel_sheet] 
    except:
        print("Sheet name doesn't exist")
        show_error("Sheet name doesn't exist")
        progress_bar(0)
        mynotif("Error")
        root.update_idletasks()
    
    

    #----- Create dummy bump at 4 corner 140x140u for advance package (CoWos)-----------#
    ymin = coordinate_to_tuple(coordinate['window1'])[0]
    xmin = coordinate_to_tuple(coordinate['window1'])[1]
    ymax = coordinate_to_tuple(coordinate['window2'])[0]
    xmax = coordinate_to_tuple(coordinate['window2'])[1]

    print(xmin,xmax)
    print(ymin,ymax)
    progress_bar(60)
    if(opt_sr == 0):
        try:
          #----- Create table from bump map-----------#
            tb_x = coordinate_to_tuple(table['location'])[1]
            tb_y = coordinate_to_tuple(table['location'])[0]

            r = tb_y + 2
            ws1[table['location']].value = table['name']
            ws1.merge_cells(table['location'] + ":" + get_column_letter(tb_x + 2) + str(tb_y))
          
            ws1[get_column_letter(tb_x) + str(tb_y + 1)].value = "X"
            ws1[get_column_letter(tb_x + 1) + str(tb_y + 1)].value = "Y"
            ws1[get_column_letter(tb_x + 2)  + str(str(tb_y + 1))].value = "Bump name"

            # xwidth = float (ws1[get_column_letter(xmax) + coordinate["xcoor"]].value)
            # minxval = float (ws1[get_column_letter(xmin) + coordinate["xcoor"]].value)
            # ywidth = float (ws1[coordinate["ycoor"] + str(ymin)].value)
            # minyval = float (ws1[coordinate["ycoor"] + str(ymax)].value)
            xwidth = ws1[get_column_letter(xmax) + coordinate["xcoor"]].value
            minxval = ws1[get_column_letter(xmin) + coordinate["xcoor"]].value
            ywidth = ws1[coordinate["ycoor"] + str(ymin)].value
            minyval = ws1[coordinate["ycoor"] + str(ymax)].value
            if (package_type == 1):
                dm_bump_coor= []
                dm_cnt=0
                mynotif("")
                root.update_idletasks()
                mynotif("Generating Dummy bump...")
                root.update_idletasks()
                for dm_bump in dummybump:
                    bump = list(dummybump[dm_bump].values())
                        
                    ymin_dm = coordinate_to_tuple(bump[0])[0]
                    xmin_dm = coordinate_to_tuple(bump[0])[1]
                    ymax_dm = coordinate_to_tuple(bump[1])[0]
                    xmax_dm = coordinate_to_tuple(bump[1])[1]
                    xcoor_dm = str(bump[2])
                    ycoor_dm = str(bump[3])

                    print(xmin_dm,xmax_dm)
                    print(ymin_dm,ymax_dm)

                    for dummycol1 in range(xmin_dm, xmax_dm + 1):
                        for dummyrow1 in range(ymin_dm, ymax_dm + 1):
                            col_dm = get_column_letter(dummycol1)
                            if (ws1[col_dm + str(dummyrow1)].value != None):
                                ws1[get_column_letter(tb_x + 2)+str(r)].value =  ws1[col_dm+ str(dummyrow1)].value
                                # print(col_l + " " + str(coordinate['xcoor']))
                                ws1[get_column_letter(tb_x)+str(r)].value = ws1[col_dm + xcoor_dm].value
                                # print(coordinate['ycoor'] + " " + str(dummyrow1)) 
                                ws1[get_column_letter(tb_x + 1)+str(r)].value = ws1[ycoor_dm + str(dummyrow1)].value
                                r = r + 1
                                coor = col_dm + str(dummyrow1)
                                dm_bump_coor.append(coor)
                                dm_cnt += 1

                #---------Create Die bump exclued dummy bump at 4 corner-----------#

                match = 0
                mynotif("")
                root.update_idletasks()
                mynotif("Generating Die bump...")
                root.update_idletasks()
                for col in range(xmin, xmax + 1):
                    for row in range(ymin, ymax + 1):       
                        col_l = get_column_letter(col)
                        #print(col_l)
                        i = 0 
                        while(i < len(dm_bump_coor)):
                            xy = col_l + str(row)
                            if(xy ==  dm_bump_coor[i]):
                                match = 1
                            else:
                                match = 0
                            if(match == 1):
                                break
                            i += 1
                        if (match == 0 and ws1[col_l + str(row)].value != None):
                            ws1[get_column_letter(tb_x + 2)+str(r)].value =  ws1[col_l+ str(row)].value
                            print(col_l + " " + str(coordinate['xcoor']))
                            ws1[get_column_letter(tb_x)+str(r)].value = ws1[col_l + str(coordinate['xcoor'])].value
                            print(coordinate['ycoor'] + " " + str(row)) 
                            ws1[get_column_letter(tb_x + 1)+str(r)].value = ws1[coordinate['ycoor'] + str(row)].value
                            r = r + 1
            else:
                mynotif("")
                root.update_idletasks()
                mynotif("Generating Die bump...")
                root.update_idletasks()
                for col in range(xmin, xmax + 1):
                        for row in range(ymin , ymax + 1):       
                            col_l = get_column_letter(col)
                            #print(col_l)
                            if (ws1[col_l + str(row)].value != None):
                                ws1[get_column_letter(tb_x + 2)+str(r)].value =  ws1[col_l+ str(row)].value
                                print(col_l + " " + str(coordinate['xcoor']))
                                ws1[get_column_letter(tb_x)+str(r)].value = ws1[col_l + str(coordinate['xcoor'])].value
                                print(coordinate['ycoor'] + " " + str(row)) 
                                ws1[get_column_letter(tb_x + 1)+str(r)].value = ws1[coordinate['ycoor'] + str(row)].value
                                #----------------------------flip bump map y axis---------------------------
                                # ws1[get_column_letter(tb_x + 7)+str(r)].value =  ws1[col_l+ str(row)].value
                                # print(col_l + " " + str(coordinate['xcoor']))
                                # ws1[get_column_letter(tb_x + 5)+str(r)].value = xwidth - float (ws1[col_l + str(coordinate['xcoor'])].value) + minxval
                                # print(coordinate['ycoor'] + " " + str(row)) 
                                # ws1[get_column_letter(tb_x + 6)+str(r)].value = ws1[coordinate['ycoor'] + str(row)].value

                                ws1[get_column_letter(tb_x + 7)+str(r)].value =  ws1[col_l+ str(row)].value
                                print(col_l + " " + str(coordinate['xcoor']))
                                ws1[get_column_letter(tb_x + 5)+str(r)].value = f"=({str(xwidth).replace('=','')})-({str(ws1[col_l + str(coordinate['xcoor'])].value).replace('=','')})+({str(minxval).replace('=','')})"
                                print(ws1[get_column_letter(tb_x + 5)+str(r)].value)
                                print(coordinate['ycoor'] + " " + str(row)) 
                                ws1[get_column_letter(tb_x + 6)+str(r)].value = ws1[coordinate['ycoor'] + str(row)].value

                                #----------------------------rotate -90 bump map after flip---------------------------
                                # ws1[get_column_letter(tb_x + 12)+str(r)].value =  ws1[col_l+ str(row)].value
                                # print(col_l + " " + str(coordinate['xcoor']))
                                # ws1[get_column_letter(tb_x + 10)+str(r)].value = ywidth - ws1[coordinate['ycoor'] + str(row)].value + minyval
                                # print(coordinate['ycoor'] + " " + str(row)) 
                                # ws1[get_column_letter(tb_x + 11)+str(r)].value = xwidth - float (ws1[col_l + str(coordinate['xcoor'])].value) + minxval

                                ws1[get_column_letter(tb_x + 12)+str(r)].value =  ws1[col_l+ str(row)].value
                                print(col_l + " " + str(coordinate['xcoor']))
                                ws1[get_column_letter(tb_x + 10)+str(r)].value = f"=({str(ywidth).replace('=','')})-({str(ws1[coordinate['ycoor']+str(row)].value).replace('=','')})+({str(minyval).replace('=','')})"
                                print(coordinate['ycoor'] + " " + str(row)) 
                                ws1[get_column_letter(tb_x + 11)+str(r)].value = f"=({str(xwidth).replace('=','')})-({str(ws1[col_l + str(coordinate['xcoor'])].value).replace('=','')})+({str(minxval).replace('=','')})"

                                #---------------------------rotate 90 bump map after flip -----------------------------
                                ws1[get_column_letter(tb_x + 17)+str(r)].value =  ws1[col_l+ str(row)].value
                                print(col_l + " " + str(coordinate['xcoor']))
                                ws1[get_column_letter(tb_x + 15)+str(r)].value = ws1[coordinate['ycoor'] + str(row)].value
                                print(coordinate['ycoor'] + " " + str(row)) 
                                ws1[get_column_letter(tb_x + 16)+str(r)].value = f"=({str(ws1[col_l + str(coordinate['xcoor'])].value).replace('=','')})"

                                r = r + 1
                # tab = Table(displayName="Table1", ref="O65:Q500")
                # ws1.add_table(tab)
            
            progress_bar(80)    
            wb.save(excel_path)
            progress_bar(100)
            mynotif("Generated")
            popup("PLOC generated successful!!!")
            mynotif("")
        except (ValueError):
            print ("loi roi")
            show_error("Wrong input, Please check and regenerate")
            progress_bar(0)
            mynotif("Error")
            root.update_idletasks()
        except:
            print('Loi quan que` gi` za^y')
            
            show_error("Wrong input, Please check and regenerate")
            progress_bar(0)
            mynotif("Error")
            root.update_idletasks()
            
            
    elif(opt_sr == 1):

        tb_x = coordinate_to_tuple(table['location'])[1]
        tb_y = coordinate_to_tuple(table['location'])[0]
        tb_x2 = coordinate_to_tuple(table['location_wsr'])[1]
        tb_y2 = coordinate_to_tuple(table['location_wsr'])[0]
        r = tb_y + 2
        r2 = tb_y2 + 2

        ws1[table['location']].value = table['name']
        ws1[table['location_wsr']].value = table['name_wsr']
        # ws1.merge_cells(table['xcol'] + str(table['begin']) + ":" + table['bumpcol'] + str(table['begin']))
        # print(table['xcol'] + str(table['begin']) + ":" + table['bumpcol'] + str(table['begin']))
        ws1[get_column_letter(tb_x) + str(tb_y + 1)].value = "X"
        ws1[get_column_letter(tb_x2) + str(tb_y2 + 1)].value = "X"
        ws1[get_column_letter(tb_x + 1) + str(tb_y + 1)].value = "Y"
        ws1[get_column_letter(tb_x2 + 1) + str(tb_y2 + 1)].value = "Y"
        ws1[get_column_letter(tb_x + 2)  + str(str(tb_y + 1))].value = "Bump name"
        ws1[get_column_letter(tb_x2 + 2)  + str(str(tb_y2 + 1))].value = "Bump name"
        if (package_type == 1):
            dm_bump_coor= []
            dm_cnt=0
            mynotif("")
            root.update_idletasks()
            mynotif("Generating Dummy bump...")
            root.update_idletasks()
            for dm_bump in dummybump:
                bump = list(dummybump[dm_bump].values())
                    
                ymin_dm = coordinate_to_tuple(bump[0])[0]
                xmin_dm = coordinate_to_tuple(bump[0])[1]
                ymax_dm = coordinate_to_tuple(bump[1])[0]
                xmax_dm = coordinate_to_tuple(bump[1])[1]
                xcoor_dm = str(bump[2])
                ycoor_dm = str(bump[3])

                print(xmin_dm,xmax_dm)
                print(ymin_dm,ymax_dm)

                for dummycol1 in range(xmin_dm, xmax_dm + 1):
                    for dummyrow1 in range(ymin_dm, ymax_dm + 1):
                        col_dm = get_column_letter(dummycol1)
                        if (ws1[col_dm + str(dummyrow1)].value != None):
                            ws1[get_column_letter(tb_x + 2)+str(r)].value =  ws1[col_dm+ str(dummyrow1)].value
                            ws1[get_column_letter(tb_x2 + 2)+str(r2)].value =  ws1[col_dm+ str(dummyrow1)].value 
                      
                            ws1[get_column_letter(tb_x)+str(r)].value = ws1[col_dm + xcoor_dm].value
                           
                            ws1[get_column_letter(tb_x + 1)+str(r)].value = ws1[ycoor_dm + str(dummyrow1)].value
                            
                            r = r + 1
                            r2 = r2 + 1
                            coor = col_dm + str(dummyrow1)
                            dm_bump_coor.append(coor)
                            dm_cnt += 1

            #---------Create Die bump exclued dummy bump at 4 corner-----------#

            match = 0
            mynotif("")
            root.update_idletasks()
            mynotif("Generating Die bump...")
            root.update_idletasks()
            for col in range(xmin, xmax + 1):
                for row in range(ymin, ymax + 1):       
                    col_l = get_column_letter(col)
                    #print(col_l)
                    i = 0 
                    while(i < len(dm_bump_coor)):
                        xy = col_l + str(row)
                        if(xy ==  dm_bump_coor[i]):
                            match = 1
                        else:
                            match = 0
                        if(match == 1):
                            break
                        i += 1
                    if (match == 0 and ws1[col_l + str(row)].value != None):
                        ws1[get_column_letter(tb_x + 2)+str(r)].value =  ws1[col_l+ str(row)].value
                        ws1[get_column_letter(tb_x2 + 2)+str(r2)].value =  ws1[col_l+ str(row)].value
                        print(col_l + " " + str(coordinate['xcoor']))
                        ws1[get_column_letter(tb_x)+str(r)].value = ws1[col_l + str(coordinate['xcoor'])].value
                       
                        print(coordinate['ycoor'] + " " + str(row)) 
                        ws1[get_column_letter(tb_x + 1)+str(r)].value = ws1[coordinate['ycoor'] + str(row)].value
                        
                        r = r + 1
                        r2 = r2 + 1
        else:
            mynotif("")
            root.update_idletasks()
            mynotif("Generating Die bump...")
            root.update_idletasks()
            for col in range(xmin, xmax + 1):
                    for row in range(ymin , ymax + 1):       
                        col_l = get_column_letter(col)
                        #print(col_l)
                        if (ws1[col_l + str(row)].value != None):
                            ws1[get_column_letter(tb_x + 2)+str(r)].value =  ws1[col_l+ str(row)].value
                            ws1[get_column_letter(tb_x2 + 2)+str(r2)].value =  ws1[col_l+ str(row)].value
                            print(col_l + " " + str(coordinate['xcoor']))
                            ws1[get_column_letter(tb_x)+str(r)].value = ws1[col_l + str(coordinate['xcoor'])].value
                            
                            print(coordinate['ycoor'] + " " + str(row)) 
                            ws1[get_column_letter(tb_x + 1)+str(r)].value = ws1[coordinate['ycoor'] + str(row)].value
                           
                            r = r + 1
                            r2 = r2 + 1
    
    # button['state'] = tk.NORMAL


                
# myButton = tk.Button(root,text="Button", command=get_path)
# myButton.pack()

entry_disable(cor1_x1y1, cor1_x2y2, cor1_Xget, cor1_Yget,
            cor2_x1y1, cor2_x2y2, cor2_Xget, cor2_Yget,
            cor3_x1y1, cor3_x2y2, cor3_Xget, cor3_Yget,
            cor4_x1y1, cor4_x2y2, cor4_Xget, cor4_Yget)
entry_disable(c4_x1y1_i,c4_x2y2_i, c4_Xget_i, c4_Yget_i, c4_tb_name, c4_col,
            u_col, u_tb_name, u_x1y1_i, u_x2y2_i, u_Xget_i, u_Yget_i)
entry_disable(sheete_i, sheete_t)
entry_disable(sr_opt, foundry_combo, out_name2_in, out_col_wsr_i)
sheet_t['text']= "Bump sheet:"
mynotif("")
treeScroll = ttk.Scrollbar(root, orient = 'vertical')



# root.configure(yscrollcomand)

progress = ttk.Progressbar(root, orient = 'horizontal',
              length = 100, mode = 'determinate')
progress_w = my_canvas.create_window(80,800, anchor="nw", window=progress, width= 800)


# Button
#Create style object
# sto = ttk.Style()

# #configure style
# style = ttk.Style()
# style.configure('TButton', font =
#                ('calibri', 20, 'bold'),
#                     borderwidth = '4',
#                     width = '80')
# sto.configure('W.TButton', font= ('System', 10, 'underline', 'bold'),
#  foreground='#9900ff', border=50)
mediumFont = Font(
	family="System",
	size=16,
	weight="normal",
	slant="italic",
	underline=1,
	overstrike=0)
def hihi():
    button.configure(font=mediumFont, foreground='white', background='Green')
browse_btn = ttk.Button(root, text="Open File", image=open_imag, command=open)
browse_btn_w = my_canvas.create_window(865, 40, anchor="nw", window=browse_btn)
# button = tk.Button(root, text="Generate",font=("System", 14, 'underline', 'bold'), foreground='white', background='#9b34eb', command=get_path, width=40)
button = tk.Button(root, text="Generate",font = mediumFont, foreground='white', background='#9b34eb', command=get_path, width=40)
# button = ttk.Button(root, text="Generate", command=get_path, width=80)

button_w = my_canvas.create_window(300, 860, anchor="nw", window=button)





root.mainloop()

