from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.utils import range_boundaries
from openpyxl.utils import column_index_from_string
from openpyxl.utils import coordinate_to_tuple
from openpyxl.worksheet.table import Table, TableStyleInfo
# from tkinter import *

from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection
import os
# import win32com.client
from pathlib import Path  
import time
import tempfile
from ploc_myTk import *


ini_text =[
    "This tool is used for PLOC table generation from Bump map visual view.\n"

    
]

excel_g = [
    "\n\nINFO: This field for puting excel file\n\n ",
   
]
v_sheet_g = [
    "INFO: Name of bump visual sheet to generate Bump/Ball coordinate table\n\n ",
    "      * Example: N3P_CoWoS"
]
v_start_g = [
    "INFO: Die window begin cell\n\n ",
    "      * Example:   A0           "
]
v_end_g =  [
    "INFO: Die window end cell\n\n ",
    "      * Example:   CU100       "
]
v_X_g = [
    "INFO: Row contains X axis value which is X location of Bump. \n",
    "- Must be interger   \n\n",
    "      * Example:   8       "
]
v_Y_g = [
    "INFO: Row contains Y axis value which is Y location of Bump.\n", 
    "- Must be Excel column format\n\n",
    "      * Example: CU " 
]
die_tb_o_sheet_g = [
    "INFO: Sheet to put Die table\n\n ",
    "      * Example: Bump coordination "
]
die_tb_o_name_g = [
    "INFO: This field to define the output table name\n\n    ",
    "      * Example: DieX "
]
die_tbsr_o_name_g = [
    "INFO: This field to define the output table name for bump without sealring.\n",
    "- This field will be used when TC option is turned on\n\n",
    "      * Example: DieX without sealring "
]
die_tb_o_loc_g = [
    "INFO: This field to define the first output table location. \n",
    "- The next tables placed away 2 column from previous table \n\n ",
    "      * Example: O64 "
]
die_tbsr_o_loc_g = [
    "INFO: This field to define the first output table location. \n\n",
    "      * Example: O64 "
]
dummystart_g = [
    "INFO: Dummy bump window begin cell\n ",
    "- This param used for A-CoWoS package type only\n"
    "      * Example:   A0           "
]
dummyend_g = [
    "INFO: Dummy bump window end cell\n ",
    "- This param used for A-CoWoS package type only\n"
    "      * Example:   E3           "
]
dummy_X_g = [
    "INFO: Row contains X axis value which is X location of Bump. \n",
    "- This param used for A-CoWoS package type only\n"
    "- Must be interger   \n\n",
    "      * Example:   8       "
]
dummy_Y_g = [
    "INFO: Row contains Y axis value which is Y location of Bump.\n",
    "- This param used for A-CoWoS package type only"
    "- Must be Excel column format\n\n",
    "      * Example: CU "
]
chip_size_x_g = [
    "INFO: Width of Die/chip. \n",
    "- This param used for Flip, Rotate die/chip to put on PKG  "
]
chip_size_y_g = [
    "INFO: Height of Die/chip. \n",
    "- This param used for Flip, Rotate die/chip to put on PKG  "
]
int_chip_L_xoffset_g = [
    "INFO: List X Offset of Die Left/Up(Chip Left/Up). \n ",
    "- This param used for Die/chip placement on PKG \n ",
    "- The offset values are separated by spaces.\n ",
]
int_chip_L_yoffset_g = [
    "INFO: List Y Offset of Die Left/Up(Chip Left/Up). \n ",
    "- This param used for Die/chip placement on PKG \n",
    "- The offset values are separated by spaces.\n ",
]
int_chip_R_xoffset_g = [
    "INFO: List X Offset of Die Right/Down(Chip Right/Down). \n",
    "- This param used for Die/chip placement on PKG \n ",
    "- The offset values are separated by spaces.\n ",
]
int_chip_R_yoffset_g = [
    "INFO: List Y Offset of Die Right/Down(Chip Right/Down). \n",
    "- This param used for Die/chip placement on PKG \n ",
    "- The offset values are separated by spaces.\n ",
]
int_sheet_g = [
    "INFO: Name of interposer sheet to put interposer Die table \n\n ",
    "      * Example: Packge_substrates "
]
chipL_name_g = [
    "INFO: List Name of interposer Die Left/Up which is outcome of Die Flipped then Rotate -90 degree\n ",
    "- The dies name are separated by spaces.\n ",
    "NOTE: The Die name is mapping between Die Left/Up and Die Right/Down. \n",
    "  For example:\n",
    "   + Die Right list name: DIE5 DIE6 DIE7 DIE8, and\n",
    "   + Die Left list name: DIE1 DIE2 DIE3 DIE4 then:\n",
    "               DIE1 <=> DIE5 \n",
    "               DIE2 <=> DIE6 \n",
    "               DIE3 <=> DIE7 \n",
    "               DIE4 <=> DIE8"
]
chipR_name_g = [
    "INFO: List Name of interposer Die Right/Down which is outcome of Die Flipped then Rotate +90 degree\n ",
    "- The dies name are separated by spaces.\n ",
    "NOTE: The Die name is mapping between Die Left/Up and Die Right/Down. \n",
    "   For example:\n",
    "    + Die Left list name: DIE1 DIE2 DIE3 DIE4, and\n",
    "    + Die Right list name: DIE5 DIE6 DIE7 DIE8, then:\n",
    "               DIE1 <=> DIE5 \n",
    "               DIE2 <=> DIE6 \n",
    "               DIE3 <=> DIE7 \n",
    "               DIE4 <=> DIE8"
]
int_tb_o_loc_g = [
    "INFO: This field to define the first output table cell. \n ",
    "- The next tables placed away 1 column from previous table \n\n ",
    "       * Example: O64 "
]
srw_g = [
    "INFO: This field to define the width of sealring.\n\n",
    "Normally:\n"
            "+ TSMC: 21.6\n"
            "+ SS/GF is 14.04 \n"
]
shrink_g = [
    "INFO: This field to define the die shrink value.\n\n",
    "       * Example: 0.98 "
]

# style =Style(theme='darkly')
# root = ThemedTk()
root = ttk.Window(themename='united')
root.title("PLOC TABLE GENERATOR")
root.geometry("1000x1000")
root.resizable(width=True, height=True)
root.iconbitmap(r"./mylogo.ico")
root.option_add("*tearOff", False) # This is always a good idea
img_path = r"./img/resize1000x1000"

bgm = PhotoImage(file=img_path + r"/owlpurple.png")

open_imag = PhotoImage(file = r"./open-folder.png")
# img_list = ["owl.png", "mountain.png","car.png", "penguin.png","sunset1.png", "flower3.png", "kid.png", "pug.png", "cat.png", "whale2.png", "elephant_grey.png", "snowman.png", "bee4.png", "elephant.png", "bee2.png", "fox.png", "beach.png", "frog.png", "cow.png", "forest.png", "owlpink2.png", "girl.png", "sand1.png", "baby2.png", "pig.png", "discord1.png" ]

# text_fg_ls = ['#78C2AD', '#923BFF','#FF5B14', '#2943FF', '#FFFDFF','#6A8BFF','#CBFF61' ]
# theme_list = ["adapta", "aquativo", "arc", "black","blue", "breeze", "clearlooks", "elegance", "equilux", "itft1", "keramik", "keramik_alt", "kroc", "plastik", "radiance", "ubuntu", "scidblue", "scidgreen", "scidgrey", "scidmint", "scidpink", "scidpurple", "scidsand", "smog", "winxpblue", "yaru" ]
colour_ls = ["#09a5e8", "#292b33", "#1583eb", "#292a2b","#1a7cad", "#0664bd", "#8baac7", "#59564f", "#40454a", "#7aa7f5", "#7795b4", "#7795b4", "#ebab0c", "#0c99eb", "#eb830c", "#eb830c", "#0937ab", "#37ed80", "#707371", "#479403", "#d12a9f", "#9b34eb", "#787122", "#118cbd", "#505257", "#924d8b" ]
package_ls = ["S-Organic", "A-CoWoS", "A-EMIB"]
foundry_ls = ["TSMC-MapWSR", "TSMC-MapWoSR", "SS-MapWSR", "SS-MapWoSR", "GF-MapWSR", "GF-MapWSR"]
int_couple_ls = ["2", "4", "6", "8", "10", "12", "14", "16"]
theme_ls = ['minty','pulse','united','morph','darkly','cyborg','superhero']
textentry_fg_ls = ['black', 'black','black', 'black', 'white','white','white' ]
# text_bg_color_ls = ['#78C2AD', '#B4A7D6', '#EB6536', '#378DFC', '#375A7F', '#2A9FD6', '#4C9BE8']
text_bg_color_ls = ['#78C2AD', '#B4A7D6', '#FF9378', '#48C1F8', '#2F2F2F', '#373737', '#414D59'] 
text_fg_ls = ['black', 'black','black', 'black', 'white','white','white' ]
img_ls = ["frog3.png", "owlpurple.png","fox.png", "car.png", "kid.png", "snowman.png", "cow.png"]

try:
    temp_file =  os.path.join(tempfile.gettempdir(), ".ploctablebgen_params_saved.txt")
    print(temp_file)
    tmp_flag = 0
except:
    messagebox.showerror("Can not found the User Temp dir")
    tmp_flag = 1

def disable_entries(*entries:Tkentry):
    for entry in entries:
        entry.disable()
def enable_entries(*entries:Tkentry):
    for entry in entries:
        entry.enable()
def choosetheme(combo_list: dict[str,TkCombobox], entry_list: dict[str, Tkentry], text_list: dict[str, CanvasText], checkbutton_list : dict[str, TKcheckbtn]):
    # root.set_theme(combo.combobox.get())
    root.style.theme_use(combo_list['theme_cb'].get_value())
    idx = theme_ls.index(combo_list['theme_cb'].get_value())
    enable_fg = text_fg_ls[idx]
    for entry in entry_list:
      entry_list[entry].entry.config(foreground=textentry_fg_ls[idx])
    for text in text_list:
        text_list[text].change_color(text_fg_ls[idx])
        text_list[text].set_bg_color(text_bg_color_ls[idx])
    resize_image_bg(my_canvas, combo_list['theme_cb'])
    sr_opt_val = checkbutton_list['sr_opt'].get_state()
    int_opt_val = checkbutton_list['int_opt'].get_state()
    if(sr_opt_val == 1):
        for entry in entry_list:
            if str(entry).find('sr_w') != -1 :
                entry_list[entry].set_fg(enable_fg)
    else:
        for entry in entry_list:
            if str(entry).find('sr_w') != -1 :
                entry_list[entry].set_fg('#A9A9A9')
    if(int_opt_val == 1):
        for entry in entry_list:
            if str(entry).find('int_') != -1 or str(entry).find('chip_') != -1:
                entry_list[entry].set_fg(enable_fg)
    else:
        for entry in entry_list:
            if str(entry).find('int_') != -1 or str(entry).find('chip_') != -1 :
                entry_list[entry].set_fg('#A9A9A9')
    choosemode(combo_list,entry_list)
                
    
def resize_image_bg(myCanvas: Canvas, combo: TkCombobox):
    global bgm
    width = myCanvas.winfo_width()
    height = myCanvas.winfo_height()
    idx = theme_ls.index(combo.get_value())
    p = os.path.join(img_path, img_ls[idx])
    if(width>=height):
        size = height
    else:
        size = width
    img= (Image.open(p))

    # #Resize the Image using resize method
    resized_image= img.resize((size,size), Image.LANCZOS)
    bgm= ImageTk.PhotoImage(resized_image)
    myCanvas.itemconfigure(bg_img, image=bgm)
    myCanvas.moveto(bg_img, (width - size)/2, (height - size)/2)
    

def set_theme(combo : TkCombobox, theme_name : str, themelist: list, entry_list: dict[str, Tkentry], text_list: dict[str, CanvasText]):
    root.style.theme_use(theme_name)
    combo.set_current(themelist.index(theme_name))
    for entry in entry_list:
        entry_list[entry].entry.config(foreground=textentry_fg_ls[theme_ls.index(combo.get_value())])
    for text in text_list:
        text_list[text].change_color(text_fg_ls[theme_ls.index(combo.get_value())])
        text_list[text].set_bg_color(text_bg_color_ls[theme_ls.index(combo.get_value())])

def choosemode(combo_list: dict[str,TkCombobox], entry_list : dict[str, Tkentry]):
    pkg = combo_list['pkg_cb'].get_value()
    idx = theme_ls.index(combo_list['theme_cb'].get_value())
    enable_fg = text_fg_ls[idx]
    if(pkg == "S-Organic"):
        for entry in entry_list:
            if str(entry).find("dm_") != -1:
                disable_entries(entry_list[entry])
                entry_list[entry].set_fg('#A9A9A9')
    elif(pkg == "A-CoWoS"): 
        for entry in entry_list:
            if str(entry).find("dm_") != -1:
                enable_entries(entry_list[entry])
                entry_list[entry].set_fg(enable_fg)
    elif(pkg == "A-EMIB"):
        crying = icons.Emoji.get('CRYING FACE')
        messagebox.showinfo("Notification", f"This feature is not developed yet {crying}.\nPlease use S-Organic option and run 2 time (C4 and uBump) for this package type")
        combo_list['pkg_cb'].set_current(0)
        for entry in entry_list:
            if str(entry).find("dm_") != -1:
                disable_entries(entry_list[entry])
                entry_list[entry].set_fg('#A9A9A9')

def on_vertical(event):
    my_canvas.yview_scroll(-1 * event.delta, 'units')

def on_horizontal(event):
    my_canvas.xview_scroll(-1 * event.delta, 'units')

def entry_responsive(entry_list :dict[str, Tkentry], w: int, h : int):
    for entry in entry_list:
        entry_list[entry].moveto(w,h)
        entry_list[entry].change_width_height(w,h)
        entry_list[entry].change_textsize(w,h)
def textbox_responsive(textbox_list : dict[str, TkTextbox], w: int, h : int):
    for box in textbox_list:
        textbox_list[box].moveto(w,h)
        textbox_list[box].change_width_height(w,h)
def text_reponsive(text_list : dict[str, CanvasText] ,w : int, h : int):
    for text in text_list:
        text_list[text].moveto(w,h)
        text_list[text].set_size(w,h)
def button_responsive(btn_list : dict[str, Tkbutton],w : int,h : int):
    for btn in btn_list:
        btn_list[btn].moveto(w,h)
        btn_list[btn].change_width_height(w,h)
def progressbar_responsive(pgbar_list : dict[str, Tkprogressbar],w : int,h : int):
    for pgbar in pgbar_list:
        pgbar_list[pgbar].moveto(w,h)
        pgbar_list[pgbar].change_width_height(w,h)
def checkbtn_responsive(checkbtn_list : dict[str, TKcheckbtn], w: int, h : int) :
    for chkbtn in checkbtn_list:
        checkbtn_list[chkbtn].moveto(w,h)
        # chkbtn.change_width_height(h)
def cobobox_responsive(cb_list : dict[str, TkCombobox], w: int, h : int):
    for cb in cb_list:
        cb_list[cb].moveto(w,h)
        cb_list[cb].change_width_height(w,h)
def on_window_resize(entry_list :dict[str, Tkentry],text_list : dict[str, CanvasText],textbox_list : dict[str, TkTextbox], btn_list : dict[str, Tkbutton],
                      pgbar_list : dict[str, Tkprogressbar], chkbtn_list : dict[str, TKcheckbtn],combo_list : dict[str, TkCombobox]):
    global bgm
    if my_canvas.winfo_width() < 500:
        root.geometry(f"500x{my_canvas.winfo_height()}")
    if my_canvas.winfo_height()<820:
        root.geometry(f"{my_canvas.winfo_width()}x820")
    width = my_canvas.winfo_width()
    height = my_canvas.winfo_height()
    # print(f"Window resized to {width}x{height}")    
    entry_responsive(entry_list,w=width,h=height)
    text_reponsive(text_list, w=width,h=height)
    textbox_responsive(textbox_list,w=width, h=height)
    button_responsive(btn_list,w=width,h=height)
    progressbar_responsive(pgbar_list,w=width, h=height)
    checkbtn_responsive(chkbtn_list, w=width, h=height)
    cobobox_responsive(combo_list, w=width, h=height)
    resize_image_bg(my_canvas, combo_list['theme_cb'])

def toggle(checkbtn : TKcheckbtn,combo_list: dict[str, TkCombobox], entry_list : dict[str, Tkentry],textbox_list : dict[str, TkTextbox], content : list[str]):
    print(f"toggled, state: {checkbtn.get_state()}")
    idx = theme_ls.index(combo_list['theme_cb'].get_value())
    enable_fg = text_fg_ls[idx]
    if(content[len(content)-1] == 'not_yet'):
        crying = icons.Emoji.get('CRYING FACE')
        messagebox.showinfo("Notification", f"This feature is not developed yet {crying}")
    else:
        if(checkbtn.get_state() == 1):        
            print(content[0])
            for entry in entry_list:
                for c in range(2,len(content)):
                    if str(entry).find(content[c]) != -1 :
                        enable_entries(entry_list[entry])
                        entry_list[entry].set_fg(enable_fg)  
            textbox_list['text_box'].add_new_text(content[0] + "\n")

        elif(checkbtn.get_state() == 0):            
            print(content[0])
            for entry in entry_list:                
                for c in range(2,len(content)):
                    if str(entry).find(content[c]) != -1 :
                        disable_entries(entry_list[entry])
                        entry_list[entry].set_fg('#A9A9A9')                        
            textbox_list['text_box'].add_new_text(content[1] + "\n")
                
def progress_bar(bar: Tkprogressbar,value):
    bar.update(value)
    root.update_idletasks()
def open_file():
    root.filename = filedialog.askopenfilename(initialdir="./", title="Select A File", filetypes=(("Excel files", "*.xlsx"),("all files", "*.*")))
    p_excel.del_content()
    print(root.filename) 
    p_excel.add_content(root.filename)

def mynotif(texbox: TkTextbox,content : str):
    texbox.add_text(content + "\n")
    texbox.textbox.see("end")
def popup(notif: str):
    messagebox.showinfo("Notification", notif)

def show_error(error: str):
    messagebox.showerror("Error", error)
def get_input(entry_list : dict[str, Tkentry], checkbtn_list: dict[str, TKcheckbtn], combo_list: dict[str, TkCombobox]):
    inputparams : dict[str,str] = {}
    for entry in entry_list:
        inputparams.__setitem__(entry,entry_list[entry].get())
    for chkbtn in checkbtn_list:
        inputparams.__setitem__(chkbtn,checkbtn_list[chkbtn].get_state())
    for combo in combo_list:
        inputparams.__setitem__(combo,combo_list[combo].get_value())
    return inputparams
def get_saved_params(entry_list: dict[str, Tkentry], checkbtn_list: dict[str, TKcheckbtn], combo_list: dict[str, TkCombobox], text_list: dict[str, CanvasText], int_couple_list: list, package_list: list, theme_list: list):
    global temp_file
    try:
        with open(temp_file,'r') as params_saved:
            line1 = [line.rstrip() for line in params_saved]
            params = {
            'excel_path' : line1[0],
            'bump_visual_sheet' : line1[1],
            'package_type' : line1[2],
            'forTC' : line1[3],
            'sr_width' : line1[4],
            'die_visual' : line1[5],
            'die_dummy' : line1[6],
            'die_out_tb_sheet' : line1[7],
            'die_wsr_out_tb_name': line1[8],
            'die_wosr_out_tb_name': line1[9],
            'die_out_loc': line1[10],
            'is_interpos' : line1[11],
            'inter_size' : line1[12],
            'inter_diecount' : line1[13],
            'inter_dieL_name' : line1[14],
            'inter_dieR_name' : line1[15],
            'inter_xL_offset' : line1[16],
            'inter_xR_offset' : line1[17],
            'inter_yL_offset' : line1[18],
            'inter_yR_offset' : line1[19],
            'inter_out_tb_sheet' : line1[20],
            'inter_out_tb_loc' : line1[21],
            'theme': line1[22]
            }
        
        set_theme(combo=combo_list['theme_cb'],theme_name= params['theme'],themelist=theme_list, entry_list=entry_list,text_list=text_list )
 
        entry_list['p_excel'].add_new_content(params['excel_path'])
        entry_list['v_sheet'].add_new_content(params['bump_visual_sheet'])

        entry_list['sr_w'].add_new_content(params['sr_width'])
        die_visual_list = params['die_visual'].split()
        entry_list['v_start'].add_new_content(die_visual_list[0])
        entry_list['v_end'].add_new_content(die_visual_list[1])
        entry_list['v_X'].add_new_content(die_visual_list[2])
        entry_list['v_Y'].add_new_content(die_visual_list[3])

        dummy_list = params['die_dummy'].split()
        entry_list['dm_v_cor1_start'].add_new_content(dummy_list[0])
        entry_list['dm_v_cor1_end'].add_new_content(dummy_list[1])
        entry_list['dm_v_cor1_X'].add_new_content(dummy_list[2])
        entry_list['dm_v_cor1_Y'].add_new_content(dummy_list[3])
     
        entry_list['dm_v_cor2_start'].add_new_content(dummy_list[4])
        entry_list['dm_v_cor2_end'].add_new_content(dummy_list[5])
        entry_list['dm_v_cor2_X'].add_new_content(dummy_list[6])
        entry_list['dm_v_cor2_Y'].add_new_content(dummy_list[7])
        
        entry_list['dm_v_cor3_start'].add_new_content(dummy_list[8])
        entry_list['dm_v_cor3_end'].add_new_content(dummy_list[9])
        entry_list['dm_v_cor3_X'].add_new_content(dummy_list[10])
        entry_list['dm_v_cor3_Y'].add_new_content(dummy_list[11])

        entry_list['dm_v_cor4_start'].add_new_content(dummy_list[12])
        entry_list['dm_v_cor4_end'].add_new_content(dummy_list[13])
        entry_list['dm_v_cor4_X'].add_new_content(dummy_list[14])
        entry_list['dm_v_cor4_Y'].add_new_content(dummy_list[15])

        
        entry_list['die_tb_o_sheet'].add_new_content(params['die_out_tb_sheet'])
        entry_list['die_tb_o_name'].add_new_content(params['die_wsr_out_tb_name'])
        entry_list['die_tb_o_loc'].add_new_content(params['die_out_loc'])
        entry_list['die_tbsr_o_name'].add_new_content(params['die_wosr_out_tb_name'])
        chip_size_list = params['inter_size'].split()
        entry_list['chip_size_x'].add_new_content(chip_size_list[0])
        entry_list['chip_size_y'].add_new_content(chip_size_list[1])
        entry_list['int_chip_L_name'].add_new_content(params['inter_dieL_name'])
        entry_list['int_chip_R_name'].add_new_content(params['inter_dieR_name'])
        entry_list['int_chip_L_xoffset'].add_new_content(params['inter_xL_offset'])
        entry_list['int_chip_R_xoffset'].add_new_content(params['inter_xR_offset'])
        entry_list['int_chip_L_yoffset'].add_new_content(params['inter_yL_offset'])
        entry_list['int_chip_R_yoffset'].add_new_content(params['inter_yR_offset'])
        entry_list['int_sheet'].add_new_content(params['inter_out_tb_sheet'])
        entry_list['int_tb_o_loc'].add_new_content(params['inter_out_tb_loc'])
        
        combo_list['chip_cnt_cb'].set_current(int_couple_list.index(params['inter_diecount']))
        
        combo_list['pkg_cb'].set_current(package_list.index(params['package_type']))
        # change_colour(theme_ls.index(params['theme']))
        # package_combo.current(package_ls.index(params['package_type']))

    except:
        print('Can not get the old value. Use temple params')
        
        set_theme(combo=combo_list['theme_cb'],theme_name= 'pulse',themelist=theme_list, entry_list=entry_list,text_list=text_list )
 
        entry_list['p_excel'].add_new_content( r"C:\Users\sytung\OneDrive - Synopsys, Inc\Desktop\py\Bump_CoWoS_S.xlsx")
        entry_list['v_sheet'].add_new_content('N3P_CoWoS')

        entry_list['sr_w'].add_new_content('21.6')
    
        entry_list['v_start'].add_new_content('C11')
        entry_list['v_end'].add_new_content('CW103')
        entry_list['v_X'].add_new_content('8')
        entry_list['v_Y'].add_new_content('B')

        entry_list['dm_v_cor1_start'].add_new_content('C11')
        entry_list['dm_v_cor1_end'].add_new_content('E13')
        entry_list['dm_v_cor1_X'].add_new_content('9')
        entry_list['dm_v_cor1_Y'].add_new_content('B')
     
        entry_list['dm_v_cor2_start'].add_new_content('CU11')
        entry_list['dm_v_cor2_end'].add_new_content('CW13')
        entry_list['dm_v_cor2_X'].add_new_content('9')
        entry_list['dm_v_cor2_Y'].add_new_content('B')
        
        entry_list['dm_v_cor3_start'].add_new_content('C101')
        entry_list['dm_v_cor3_end'].add_new_content('E103')
        entry_list['dm_v_cor3_X'].add_new_content('9')
        entry_list['dm_v_cor3_Y'].add_new_content('B')

        entry_list['dm_v_cor4_start'].add_new_content('CU101')
        entry_list['dm_v_cor4_end'].add_new_content('CW103')
        entry_list['dm_v_cor4_X'].add_new_content('9')
        entry_list['dm_v_cor4_Y'].add_new_content('B')

        
        entry_list['die_tb_o_sheet'].add_new_content('Bump coordination')
        entry_list['die_tb_o_name'].add_new_content('Die with sealring')
        entry_list['die_tb_o_loc'].add_new_content('S111')
        entry_list['die_tbsr_o_name'].add_new_content('Die without sealring')
      
        entry_list['chip_size_x'].add_new_content('3938.352')
        entry_list['chip_size_y'].add_new_content('2262.872')
        entry_list['int_chip_L_name'].add_new_content('DIE3')
        entry_list['int_chip_R_name'].add_new_content('DIE7')
        entry_list['int_chip_L_xoffset'].add_new_content('-4350.8')
        entry_list['int_chip_R_xoffset'].add_new_content('1571.96')
        entry_list['int_chip_L_yoffset'].add_new_content('16.235')
        entry_list['int_chip_R_yoffset'].add_new_content('97.985')
        entry_list['int_sheet'].add_new_content('Package_substrate')
        entry_list['int_tb_o_loc'].add_new_content('X111')
        
        combo_list['chip_cnt_cb'].set_current(0)
        
        combo_list['pkg_cb'].set_current(0)
    # mynotif("\n\nINFO: This field is for showing the information or guidance")

# Define Canvas
my_canvas = tk.Canvas(root, bd=0, highlightthickness=0,relief='groove',scrollregion=(0,0,800,1200))
my_canvas.pack(fill="both", expand=True)
my_canvas.bind_all('<Shift-MouseWheel>', on_vertical)
my_canvas.bind("<Configure>",lambda event: on_window_resize(entry_ls,text_ls, textbox_ls, btn_ls, pgbar_ls, chkbtn_ls,combo_ls ))
bg_img = my_canvas.create_image(0,0,image=bgm,anchor='nw')

text_box = TkTextbox(canvas=my_canvas,x=600,y=80,w=280, h=100)
text_box.add_new_text(ini_text)
text_box.textbox.config(foreground='#F75726', font=tkfont(family="Courier", size="10"))

sr_opt = TKcheckbtn(win=root,canvas=my_canvas,x=280, y=125,text= "wo sealring?", anchor='sw')
sealring_noti = ["- Generate table without sealring: ON \n- Please define sealring width at the next to entry", "- Generate table without sealring: OFF", "sr_w","die_tbsr_o_name"]
sr_opt.checkbtn.config(command= lambda : toggle(sr_opt, combo_ls, entry_ls, textbox_ls, sealring_noti))

shrink_opt = TKcheckbtn(win=root,canvas=my_canvas,x=280, y=160,text= "Shrink?", anchor='sw')
shrink_noti = ["- Generate table for shrink die based on normal die visual bump map: ON \n- Please define sealring width at the next to entry", "- Generate table for shrink die based on normal die visual bump map: OFF", "not_yet"]
shrink_opt.checkbtn.config(command= lambda : toggle(shrink_opt,combo_ls, entry_ls, textbox_ls, shrink_noti))

int_opt = TKcheckbtn(win=root,canvas=my_canvas,x=30,y=500, text="Interposer Die generator?")
interposer_noti = ["-Gen interposer Die table: ON", "-Gen interposer Die table: OFF", "int_", "chip_"]
int_opt.checkbtn.config(command= lambda : toggle(int_opt, combo_ls, entry_ls, textbox_ls, interposer_noti))

sr_w = Tkentry(canvas=my_canvas,x=400,y=120,w=90,TkTextbox=text_box,guide_text=srw_g)
shrink_e = Tkentry(canvas=my_canvas,x=400,y=155,w=90,TkTextbox=text_box,guide_text=shrink_g)
p_excel = Tkentry(canvas=my_canvas,x=150,y=40,w=700,TkTextbox=text_box,guide_text=excel_g, justify='left')
v_sheet = Tkentry(canvas=my_canvas,x=150,y=80,w=340, TkTextbox=text_box,guide_text=v_sheet_g)
v_start = Tkentry(canvas=my_canvas,x=150,y=230,w=140,TkTextbox=text_box,guide_text=v_start_g)
v_end = Tkentry(canvas=my_canvas,x=300,y=230,w=140,TkTextbox=text_box,guide_text=v_end_g)
v_X = Tkentry(canvas=my_canvas,x=150,y=270,w=140,TkTextbox=text_box,guide_text=v_X_g)
v_Y = Tkentry(canvas=my_canvas,x=300,y=270,w=140,TkTextbox=text_box,guide_text=v_Y_g)
die_tb_o_sheet = Tkentry(canvas=my_canvas,x=600,y=195,w=290,TkTextbox=text_box,guide_text=die_tb_o_sheet_g)

die_tb_o_name = Tkentry(canvas=my_canvas,x=600,y=230,w=140,TkTextbox=text_box,guide_text=die_tb_o_name_g)
die_tb_o_loc = Tkentry(canvas=my_canvas,x=600,y=270,w=140,TkTextbox=text_box,guide_text=die_tb_o_loc_g)

die_tbsr_o_name = Tkentry(canvas=my_canvas,x=750,y=230,w=140,TkTextbox=text_box,guide_text=die_tbsr_o_name_g)
die_tbsr_o_loc = Tkentry(canvas=my_canvas,x=750,y=270,w=140,TkTextbox=text_box,guide_text=die_tbsr_o_loc_g)
dm_v_cor1_start = Tkentry(canvas=my_canvas,x=150,y=350,w=140,TkTextbox=text_box,guide_text=dummystart_g)
dm_v_cor1_end = Tkentry(canvas=my_canvas,x=300,y=350,w=140,TkTextbox=text_box,guide_text=dummyend_g)
dm_v_cor1_X = Tkentry(canvas=my_canvas,x=150,y=380,w=140,TkTextbox=text_box,guide_text=dummy_X_g)
dm_v_cor1_Y = Tkentry(canvas=my_canvas,x=300,y=380,w=140,TkTextbox=text_box,guide_text=dummy_Y_g)

dm_v_cor2_start = Tkentry(canvas=my_canvas,x=600,y=350,w=140,TkTextbox=text_box,guide_text=dummystart_g)
dm_v_cor2_end = Tkentry(canvas=my_canvas,x=750,y=350,w=140,TkTextbox=text_box,guide_text=dummyend_g)
dm_v_cor2_X = Tkentry(canvas=my_canvas,x=600,y=380,w=140,TkTextbox=text_box,guide_text=dummy_X_g)
dm_v_cor2_Y = Tkentry(canvas=my_canvas,x=750,y=380,w=140,TkTextbox=text_box,guide_text=dummy_Y_g)

dm_v_cor3_start = Tkentry(canvas=my_canvas,x=150,y=430,w=140,TkTextbox=text_box,guide_text=dummystart_g)
dm_v_cor3_end = Tkentry(canvas=my_canvas,x=300,y=430,w=140,TkTextbox=text_box,guide_text=dummyend_g)
dm_v_cor3_X = Tkentry(canvas=my_canvas,x=150,y=460,w=140,TkTextbox=text_box,guide_text=dummy_X_g)
dm_v_cor3_Y = Tkentry(canvas=my_canvas,x=300,y=460,w=140,TkTextbox=text_box,guide_text=dummy_Y_g)

dm_v_cor4_start = Tkentry(canvas=my_canvas,x=600,y=430,w=140,TkTextbox=text_box,guide_text=dummystart_g)
dm_v_cor4_end = Tkentry(canvas=my_canvas,x=750,y=430,w=140,TkTextbox=text_box,guide_text=dummyend_g)
dm_v_cor4_X = Tkentry(canvas=my_canvas,x=600,y=460,w=140,TkTextbox=text_box,guide_text=dummy_X_g)
dm_v_cor4_Y = Tkentry(canvas=my_canvas,x=750,y=460,w=140,TkTextbox=text_box,guide_text=dummy_Y_g)

chip_size_x = Tkentry(canvas=my_canvas,x=150,y=560,w=170,TkTextbox=text_box,guide_text=chip_size_x_g)
chip_size_y = Tkentry(canvas=my_canvas,x=340,y=560,w=170,TkTextbox=text_box,guide_text=chip_size_y_g)
int_sheet = Tkentry(canvas=my_canvas,x=530,y=560,w=170,TkTextbox=text_box,guide_text=int_sheet_g)
int_tb_o_loc = Tkentry(canvas=my_canvas,x=720,y=560,w=170,TkTextbox=text_box,guide_text=int_tb_o_loc_g)
int_chip_L_name = Tkentry(canvas=my_canvas,x=150,y=630,w=360,TkTextbox=text_box,guide_text=chipL_name_g)
int_chip_R_name = Tkentry(canvas=my_canvas,x=530,y=630,w=360,TkTextbox=text_box,guide_text=chipR_name_g)
int_chip_L_xoffset = Tkentry(canvas=my_canvas,x=150,y=670,w=360,TkTextbox=text_box,guide_text=int_chip_L_xoffset_g)
int_chip_R_xoffset = Tkentry(canvas=my_canvas,x=530,y=670,w=360,TkTextbox=text_box,guide_text=int_chip_R_xoffset_g)
int_chip_L_yoffset = Tkentry(canvas=my_canvas,x=150,y=710,w=360,TkTextbox=text_box,guide_text=int_chip_L_yoffset_g)
int_chip_R_yoffset = Tkentry(canvas=my_canvas,x=530,y=710,w=360,TkTextbox=text_box,guide_text=int_chip_R_yoffset_g)

p_excel_text = CanvasText(canvas=my_canvas,x=30,y=40,text="PLOC file:",font=tkfont(family="Helvetica", size=11, slant='italic', underline=True, weight='bold'), fill='black', bgx=110, bgy= 32, isbg=True)
v_sheet_text = CanvasText(canvas=my_canvas,x=30,y=80,text="Bump sheet:",font=tkfont(family="Helvetica", size=11, slant='italic', underline=True, weight='bold'), fill='black', bgx=110, bgy= 32, isbg=True)
pkg_text = CanvasText(canvas=my_canvas,x=30,y=120,text="Package type:",font=tkfont(family="Helvetica", size=11, slant='italic', underline=True, weight='bold'), fill='black', bgx=110, bgy= 32, isbg=True)
v_text = CanvasText(canvas=my_canvas,x=30,y=200,text="Die bump map visual input:", bgx=220, bgy= 32, isbg=True)
die_tb_o_text = CanvasText(canvas=my_canvas,x=480,y=200,text="Die table out \nconfig:", bgx=100, bgy= 56, isbg=True)
dm_v_text = CanvasText(canvas=my_canvas,x=30,y=310,text="Die Dummy bump input:", bgx=200, bgy= 32, isbg=True)
# dm_cor1_text = CanvasText(canvas=my_canvas,x=245,y=330,text="Corner 1 config:",font=("Helvetica", 10, 'italic', True, 'bold'), fill='black')
dm_cor1_text = CanvasText(canvas=my_canvas,x=dm_v_cor1_end.x -50,y=330,text="Corner 1 config:",font=tkfont(family="Helvetica", size=10, slant='italic', underline=True, weight='bold'), fill='black')
dm_cor2_text = CanvasText(canvas=my_canvas,x=670,y=330,text="Corner 2 config:",font=tkfont(family="Helvetica", size=10, slant='italic', underline=True, weight='bold'), fill='black')
dm_cor3_text = CanvasText(canvas=my_canvas,x=245,y=410,text="Corner 3 config:",font=tkfont(family="Helvetica", size=10, slant='italic', underline=True, weight='bold'), fill='black')
dm_cor4_text = CanvasText(canvas=my_canvas,x=670,y=410,text="Corner 4 config:",font=tkfont(family="Helvetica", size=10, slant='italic', underline=True, weight='bold'), fill='black')
chip_size_text = CanvasText(canvas=my_canvas,x=280,y=540,text="Die/Chip size input:", bgx=160, bgy= 22, isbg=True, bg_xo=5, bg_yo=4)
int_o_text = CanvasText(canvas=my_canvas,x=620,y=540,text="OUT DIE sheet/location:", bgx=180, bgy= 22, isbg=True, bg_xo=5, bg_yo=4)
chip_cnt_text = CanvasText(canvas=my_canvas,x=30,y=600,text="Chip count:", bgx=110, bgy= 32, isbg=True)
chip_ls_name_text = CanvasText(canvas=my_canvas,x=30,y=635,text="List name:", bgx=110, bgy= 32, isbg=True)
chip_ls_xoffset_text = CanvasText(canvas=my_canvas,x=30,y=675,text="List X offset:",bgx=110, bgy= 32, isbg=True)
chip_ls_yoffset_text = CanvasText(canvas=my_canvas,x=30,y=715,text="List Y offset:", bgx=110, bgy= 32, isbg=True)
source_text =  CanvasText(canvas=my_canvas,x=700,y=980,text="Internal contact: sytung@synopsys.com", font=tkfont(family="Helvetica", size=8, slant='italic', underline=True, weight='normal'), fill='grey')

theme_cb = TkCombobox(canvas=my_canvas,x=750, y=10,values=theme_ls)
theme_cb.combobox.bind('<<ComboboxSelected>>', lambda event: choosetheme(combo_ls, entry_ls, text_ls, chkbtn_ls))

pkg_cb = TkCombobox(canvas=my_canvas,x=150, y=120,values=package_ls, w=100)
pkg_cb.combobox.bind('<<ComboboxSelected>>',lambda event: choosemode(combo_ls, entry_ls))
chip_cnt_cb = TkCombobox(canvas=my_canvas,x=150, y=600,values=int_couple_ls,w=100)

pg_bar = Tkprogressbar(my_canvas,x=80,y=800,w=800)





def delete_item():
    x1y1_i.delete_item()  
def change():
    # root.style.theme_use('superhero')
    get_input()
    # x1y1_i.change_width(500)
    # print(x1y1_i.get_width())





mediumFont = tkfont(
	family="System",
	size=12,
	weight="normal",
	slant="italic",
	underline=1,
	overstrike=0)

# button = tk.Button(root, text="Generate",font = mediumFont, foreground='black', background='green', command = change_width) #get_params_and_generate
# button_w = my_canvas.create_window(300, 860, anchor="nw", window=button,width=300)
browse_btn = Tkbutton(canvas=my_canvas,x=852, y=40, w=40, h=25)
# browse_btn_w = my_canvas.create_window(865, 40, anchor="nw", window=browse_btn)
btn_gen = Tkbutton(canvas=my_canvas,x=350,y=860,w=300, text="GENERATE")

browse_btn.button.config(image=open_imag, command=open_file)


entry_ls = {
    'p_excel' : p_excel,
    'v_sheet': v_sheet,
    'sr_w' : sr_w,
    'v_start' : v_start,
    'v_end' : v_end,
    'v_X' : v_X,
    'v_Y' : v_Y,
    'die_tb_o_sheet' : die_tb_o_sheet,
    'die_tb_o_name' : die_tb_o_name,
    'die_tb_o_loc' : die_tb_o_loc,
    'die_tbsr_o_name' : die_tbsr_o_name,
    'die_tbsr_o_loc' : die_tbsr_o_loc,
    'dm_v_cor1_start' :dm_v_cor1_start,
    'dm_v_cor1_end' : dm_v_cor1_end,
    'dm_v_cor1_X' : dm_v_cor1_X,
    'dm_v_cor1_Y' : dm_v_cor1_Y,
    'dm_v_cor2_start' : dm_v_cor2_start,
    'dm_v_cor2_end' : dm_v_cor2_end,
    'dm_v_cor2_X' : dm_v_cor2_X,
    'dm_v_cor2_Y' :dm_v_cor2_Y,
    'dm_v_cor3_start' : dm_v_cor3_start,
    'dm_v_cor3_end' : dm_v_cor3_end,
    'dm_v_cor3_X' :dm_v_cor3_X,
    'dm_v_cor3_Y' : dm_v_cor3_Y,
    'dm_v_cor4_start' : dm_v_cor4_start,
    'dm_v_cor4_end' : dm_v_cor4_end,
    'dm_v_cor4_X' : dm_v_cor4_X,
    'dm_v_cor4_Y' : dm_v_cor4_Y,
    'chip_size_x' :chip_size_x,
    'chip_size_y' :chip_size_y,
    'int_sheet' : int_sheet,
    'int_tb_o_loc' : int_tb_o_loc,
    'int_chip_L_name' : int_chip_L_name,
    'int_chip_R_name' :int_chip_R_name,
    'int_chip_L_xoffset' : int_chip_L_xoffset,
    'int_chip_R_xoffset' : int_chip_R_xoffset,
    'int_chip_L_yoffset' : int_chip_L_yoffset,
    'int_chip_R_yoffset' : int_chip_R_yoffset,
    'shrink_e': shrink_e
}
text_ls = {
    'p_excel_text':	p_excel_text,	
    'v_sheet_text':	v_sheet_text,	
    'pkg_text':	pkg_text,	
    'v_text':	v_text,	
    'die_tb_o_text':	die_tb_o_text,	
    'dm_v_text':	dm_v_text,	
    'dm_cor1_text':	dm_cor1_text,	
    'dm_cor2_text':	dm_cor2_text,	
    'dm_cor3_text':	dm_cor3_text,	
    'dm_cor4_text':	dm_cor4_text,	
    'chip_size_text':	chip_size_text,	
    'int_o_text':	int_o_text,	
    'chip_cnt_text':	chip_cnt_text,	
    'chip_ls_name_text':	chip_ls_name_text,	
    'chip_ls_xoffset_text':	chip_ls_xoffset_text,	
    'chip_ls_yoffset_text':	chip_ls_yoffset_text,	
    'source_text':	source_text,	

}
  
textbox_ls = {
    'text_box': text_box
    }
btn_ls ={
    'btn_gen' : btn_gen,
    'browse_btn' : browse_btn
    }
pgbar_ls = {
    'pg_bar':pg_bar 
    }
chkbtn_ls = {
    'int_opt' : int_opt,
    'sr_opt' : sr_opt,
    'shrink_opt': shrink_opt
    }
combo_ls = {
    'chip_cnt_cb' : chip_cnt_cb,
    'pkg_cb' : pkg_cb,
    'theme_cb' : theme_cb
    }

# button_gen.button.config(command=)

######################################################## BACK END ##################################################################################################
def save_current_input(inputted_params : dict[str:str]):
    global temp_file, tmp_flag
    if(tmp_flag == 0):
        with open(temp_file,'w') as params_saved:
            params_saved.writelines(inputted_params['p_excel']+"\n")
            params_saved.writelines(inputted_params['v_sheet']+"\n")
            params_saved.writelines(inputted_params['pkg_cb']+"\n")
            params_saved.writelines(str(inputted_params['sr_opt'])+"\n")
            params_saved.writelines(str(inputted_params['sr_w'])+"\n")
            params_saved.writelines(f"{inputted_params['v_start']} {inputted_params['v_end']} {inputted_params['v_X']} {inputted_params['v_Y']} \n")
            params_saved.writelines(inputted_params['dm_v_cor1_start'] +" "+ inputted_params['dm_v_cor1_end'] +" "+ inputted_params['dm_v_cor1_X'] +" "+ inputted_params['dm_v_cor1_Y'] +" "+
                                    inputted_params['dm_v_cor2_start'] +" "+ inputted_params['dm_v_cor2_end'] +" "+ inputted_params['dm_v_cor2_X'] +" "+ inputted_params['dm_v_cor2_Y'] +" "+
                                    inputted_params['dm_v_cor3_start'] +" "+ inputted_params['dm_v_cor3_end'] +" "+ inputted_params['dm_v_cor3_X'] +" "+ inputted_params['dm_v_cor3_Y'] +" "+
                                    inputted_params['dm_v_cor4_start'] +" "+ inputted_params['dm_v_cor4_end'] +" "+ inputted_params['dm_v_cor4_X'] +" "+ inputted_params['dm_v_cor4_Y']
                                    + "\n")
            params_saved.writelines(inputted_params['die_tb_o_sheet'] +"\n")
            params_saved.writelines(inputted_params['die_tb_o_name'] + "\n")
            params_saved.writelines(inputted_params['die_tbsr_o_name']+"\n")
            params_saved.writelines(inputted_params['die_tb_o_loc']+"\n")
        
            params_saved.writelines(str(inputted_params['int_opt']) + "\n")
            params_saved.writelines(inputted_params['chip_size_x'] + " " + inputted_params['chip_size_y'] +"\n")
            params_saved.writelines(str(inputted_params['chip_cnt_cb']) + "\n")
        

            params_saved.writelines(inputted_params['int_chip_L_name'] + "\n")
            params_saved.writelines(inputted_params['int_chip_R_name'] + "\n")
            
            params_saved.writelines(inputted_params['int_chip_L_xoffset'] + "\n")
            params_saved.writelines(inputted_params['int_chip_R_xoffset'] + "\n")
            params_saved.writelines(inputted_params['int_chip_L_yoffset'] + "\n")
            params_saved.writelines(inputted_params['int_chip_R_yoffset'] + "\n")
            params_saved.writelines(inputted_params['int_sheet'] + "\n")
            params_saved.writelines(inputted_params['int_tb_o_loc'] + "\n")
            params_saved.writelines(inputted_params['theme_cb'] + "\n")

def process_and_generate(entry_list: dict[str, Tkentry], checkbtn_list: dict[str, TKcheckbtn], combo_list: dict[str, TkCombobox], textbox :TkTextbox, progressbar: Tkprogressbar, button : Tkbutton):
    
    button.set_text("Generating...")
    button.state(state='disable')
    input_params = get_input(entry_list, checkbtn_list, combo_list)
    print(input_params)
    print(f"input pexcel:  {input_params['p_excel']}")
    print(f"input isTC:  {input_params['sr_opt']} \n input isint: {input_params['int_opt']}")
    save_current_input(input_params)
# ------------------------params--------------
    excel_path = input_params['p_excel']
    bump_visual_sheet = input_params['v_sheet']
    
    tc_sr={
        "isTC": input_params['sr_opt'],
        "sr_w": input_params['sr_w'],
        "sr_tb_name": input_params['die_tbsr_o_name'],
        "sr_tb_loc": input_params['die_tbsr_o_loc']
    }

   
   
    die_table={
        "sheet": input_params['die_tb_o_sheet'],
        "name": input_params['die_tb_o_name'],
        "location": input_params['die_tb_o_loc'],
        "wsr_name": input_params['die_tbsr_o_name'],
        "wsr_location": input_params['die_tbsr_o_name'],
        
    }

    #---Bump map visual view parameter---#
    die_coor = {
        
        "cell_begin": input_params['v_start'], #Top Left of Bump map visual view
        "cell_end": input_params['v_end'], #Bottom Right of Bump map visual view
        "xcoor": input_params['v_X'], #This define row where Xaxis value can be got
        "ycoor": input_params['v_Y'] #This define column where Yaxis value can be got
    }

    
# Die interposet prarams

    int_gen = input_params['int_opt']
    int_die_cnt = int(input_params['chip_cnt_cb'])
    int_die_params={
        "chip_width": input_params['chip_size_x'],
        "chip_height": input_params['chip_size_y'],
        "die1_xoffset":input_params['int_chip_L_xoffset'],
        "die1_yoffset": input_params['int_chip_L_yoffset'],
        "die2_xoffset": input_params['int_chip_R_xoffset'],
        "die2_yoffset": input_params['int_chip_R_yoffset'],
    }

    int_die_tb={
        "sheet": input_params['int_sheet'],
        "Die1_name": input_params['int_chip_L_name'],
        "Die2_name": input_params['int_chip_R_name'],
        "int_tb_location": input_params['int_tb_o_loc'],
    }

    dummybump={
        "corner_1":{
            "cell_begin": input_params['dm_v_cor1_start'],
            "cell_end": input_params['dm_v_cor1_end'],
            "xcoor": input_params['dm_v_cor1_X'],
            "ycoor": input_params['dm_v_cor1_Y']
            },
        "corner_2":{            
            "cell_begin": input_params['dm_v_cor2_start'],
            "cell_end": input_params['dm_v_cor2_end'],
            "xcoor": input_params['dm_v_cor2_X'],
            "ycoor": input_params['dm_v_cor2_Y']
        },
        "corner_3":{
            "cell_begin": input_params['dm_v_cor3_start'],
            "cell_end": input_params['dm_v_cor3_end'],
            "xcoor": input_params['dm_v_cor3_X'],
            "ycoor": input_params['dm_v_cor3_Y']
        },
        "corner_4":{         
            "cell_begin": input_params['dm_v_cor4_start'],
            "cell_end": input_params['dm_v_cor4_end'],
            "xcoor": input_params['dm_v_cor4_X'],
            "ycoor": input_params['dm_v_cor4_Y']
        }

        }
    package = input_params['pkg_cb']
    print("Package: " + package)
    if (package == "A-CoWoS"):
        mynotif(textbox,"Package type: A-CoWoS")
        package_type = 1
    elif(package == "S-Organic"):
        mynotif(textbox,"Package type: S-Organic")
        package_type = 0
        
    else:
        package_type = 0
       

    root.update_idletasks()
    mynotif(textbox, "Loading the ploc file...")
    root.update_idletasks()
    try:
        # wb_d = load_workbook(excel_path, data_only=True)
        print("Opening excel file...")
        mynotif(textbox, "Opening excel file...")
        wb_f = load_workbook(excel_path)
        print(wb_f)   
    except:
        print("Wrong Ploc path or Ploc file is openning. Please recheck/close the PLOC file before generate :(")
        show_error("Wrong Ploc path or Ploc file is openning. Please recheck/close the PLOC file before generate :(")
        progress_bar(progressbar, 0)
        mynotif(textbox, "Error")
        button.state(state='normal')
        button.set_text("GENERATE")
        root.update_idletasks()
        return
    
    # ws = wb_f.create_sheet('Tung')
    try:
        sheet_list = wb_f.sheetnames
       #wsvisual_d = wb_d[bump_visual_sheet]  # use for further function
        wsvisual_f = wb_f[bump_visual_sheet]
      # wsdiebump_d = wb_d[die_table['sheet']] # use for further function
       
        if die_table['sheet'] in sheet_list:
            wsdiebump_f = wb_f[die_table['sheet']]
        else:
            msg_ws = messagebox.askquestion('Create Sheet', 'The ' + die_table['sheet'] + ' doesn\'t exist. Do you want to create it?',icon='question')
            mynotif(textbox, "The " + die_table['sheet'] + " doesn't exist.")
            if(msg_ws == 'yes'):
                mynotif(textbox, 'Creating the sheet...')
                wsdiebump_f = wb_f.create_sheet(die_table['sheet'])
            else:
                progress_bar(progressbar, 0)
                return
        if(int_gen == 1):
            if int_die_tb['sheet'] in sheet_list:
                wsintbump_f = wb_f[int_die_tb['sheet']]
            else:
                mynotif(textbox, "The " + int_die_tb['sheet'] + " doesn't exist.")
                msg_ws = messagebox.askquestion('Create Sheet', 'The ' + int_die_tb['sheet'] + ' doesn\'t exist. Do you want to create it?', icon='question')
            
                if(msg_ws == 'yes'):
                    wsintbump_f = wb_f.create_sheet(int_die_tb['sheet'])
                    mynotif(textbox, 'Creating the sheet...')
                else:
                    progress_bar(progressbar, 0)
                    return

       #wsintbump_d = wb_d[int_die_tb['sheet']] # use for further function
       
    except:
        print("Sheet " + bump_visual_sheet + " doesn't exist")
        show_error("Sheet " + bump_visual_sheet + " doesn't exist")
        progress_bar(progressbar, 0)
        mynotif(textbox, "Error")
        button.state(state='normal')
        button.set_text("GENERATE")
        root.update_idletasks()
        return
    #----- Create dummy bump at 4 corner 140x140u for advance package (CoWos)-----------#
    ymin = coordinate_to_tuple(die_coor['cell_begin'])[0]
    xmin = coordinate_to_tuple(die_coor['cell_begin'])[1]
    ymax = coordinate_to_tuple(die_coor['cell_end'])[0]
    xmax = coordinate_to_tuple(die_coor['cell_end'])[1]

    print(xmin,xmax)
    print(ymin,ymax)
    progress_bar(progressbar, 60)    
    try:
        #----- Create table from bump map-----------#
        die_tb_x = coordinate_to_tuple(die_table['location'])[1]
        die_tb_y = coordinate_to_tuple(die_table['location'])[0]
        int_tb_x = coordinate_to_tuple(int_die_tb['int_tb_location'])[1]
        int_tb_y = coordinate_to_tuple(int_die_tb['int_tb_location'])[0]

        r_die = die_tb_y + 2
        r_int = int_tb_y + 2

        title_bg_fill = PatternFill(patternType='solid', fgColor='9e42f5')
        subtil_bg_fill = PatternFill(patternType='solid',fgColor='0e7bf0')
        wsdiebump_f[die_table['location']].value = die_table['name']
        
        wsdiebump_f.merge_cells(die_table['location'] + ":" + get_column_letter(die_tb_x + 2) + str(die_tb_y))
        for c1 in range(0,3):
            wsdiebump_f[get_column_letter(die_tb_x + c1) + str(die_tb_y)].fill = title_bg_fill
            wsdiebump_f[get_column_letter(die_tb_x + c1) + str(die_tb_y)].alignment = Alignment(horizontal='center')
            wsdiebump_f[get_column_letter(die_tb_x + c1) + str(die_tb_y)].border = Border(left=Side(style='thin'),right=Side(style='thin'),top=Side(style='thin'),bottom=Side(style='thin'))

        wsdiebump_f[get_column_letter(die_tb_x) + str(die_tb_y + 1)].value = "X"

        wsdiebump_f[get_column_letter(die_tb_x + 1) + str(die_tb_y + 1)].value = "Y"

        wsdiebump_f[get_column_letter(die_tb_x + 2)  + str(str(die_tb_y + 1))].value = "Bump name"
        for c2 in range(0,3):
            wsdiebump_f[get_column_letter(die_tb_x + c2) + str(die_tb_y + 1)].fill = subtil_bg_fill
            wsdiebump_f[get_column_letter(die_tb_x + c2) + str(die_tb_y + 1)].alignment = Alignment(horizontal='center')
            wsdiebump_f[get_column_letter(die_tb_x + c2) + str(die_tb_y + 1)].border = Border(left=Side(style='thin'),right=Side(style='thin'),top=Side(style='thin'),bottom=Side(style='thin'))
        if(int(tc_sr['isTC'] == 1)):
            wsdiebump_f[get_column_letter(die_tb_x + 4) + str(die_tb_y)].value = die_table['wsr_name']
            wsdiebump_f.merge_cells(get_column_letter(die_tb_x + 4) + str(die_tb_y) + ":" + get_column_letter(die_tb_x + 6) + str(die_tb_y))
            for c1 in range(0,3):
                wsdiebump_f[get_column_letter(die_tb_x + 4 + c1) + str(die_tb_y)].fill = title_bg_fill
                wsdiebump_f[get_column_letter(die_tb_x + 4 + c1) + str(die_tb_y)].alignment = Alignment(horizontal='center')
                wsdiebump_f[get_column_letter(die_tb_x + 4 + c1) + str(die_tb_y)].border = Border(left=Side(style='thin'),right=Side(style='thin'),top=Side(style='thin'),bottom=Side(style='thin'))
            wsdiebump_f[get_column_letter(die_tb_x + 4) + str(die_tb_y + 1)].value = "X"
            wsdiebump_f[get_column_letter(die_tb_x + 5) + str(die_tb_y + 1)].value = "Y"
            wsdiebump_f[get_column_letter(die_tb_x + 6)  + str(str(die_tb_y + 1))].value = "Bump name"
            for c1 in range(0,3):
                wsdiebump_f[get_column_letter(die_tb_x + 4 + c1) + str(die_tb_y + 1)].fill = subtil_bg_fill
                wsdiebump_f[get_column_letter(die_tb_x + 4 + c1) + str(die_tb_y + 1)].alignment = Alignment(horizontal='center')
                wsdiebump_f[get_column_letter(die_tb_x + 4 + c1) + str(die_tb_y + 1)].border = Border(left=Side(style='thin'),right=Side(style='thin'),top=Side(style='thin'),bottom=Side(style='thin'))

        if(int_gen == 1):
            die1_list = int_die_tb['Die1_name'].split()
            die2_list = int_die_tb['Die2_name'].split()
            die1_xoffset_list = int_die_params['die1_xoffset'].split()
            die1_yoffset_list = int_die_params['die1_yoffset'].split()
            die2_xoffset_list = int_die_params['die2_xoffset'].split()
            die2_yoffset_list = int_die_params['die2_yoffset'].split()
            if(len(die1_list) != int(int_die_cnt)/2 or len(die2_list) != int(int_die_cnt)/2 or len(die1_xoffset_list) != int(int_die_cnt)/2 or len(die2_xoffset_list) != int(int_die_cnt)/2 or len(die1_yoffset_list) != int(int_die_cnt)/2 or len(die2_yoffset_list) != int(int_die_cnt)/2):
                show_error('The input die parameters incorrect. Please re-check it')
                int_input_correct = 0
                mynotif(textbox, 'The input die parameters incorrect. Please re-check it')
                # mynotif(textbox, "")
                progress_bar(progressbar, 0)
                return
            else:
                wsintbump_f[get_column_letter(int_tb_x ) + str(int_tb_y)].value = "Die Flipped by Y axis"
                wsintbump_f.merge_cells(get_column_letter(int_tb_x) + str(int_tb_y) + ":" + get_column_letter(int_tb_x + 2) + str(int_tb_y))
                for c1 in range(0,3):
                    wsintbump_f[get_column_letter(int_tb_x + c1) + str(int_tb_y)].fill = title_bg_fill
                    wsintbump_f[get_column_letter(int_tb_x + c1) + str(int_tb_y)].alignment = Alignment(horizontal='center',wrapText=True)
                    wsintbump_f[get_column_letter(int_tb_x + c1) + str(int_tb_y)].border = Border(left=Side(style='thin'),right=Side(style='thin'),top=Side(style='thin'),bottom=Side(style='thin'))
                wsintbump_f[get_column_letter(int_tb_x) + str(int_tb_y + 1)].value = "X"
                wsintbump_f[get_column_letter(int_tb_x + 1) + str(int_tb_y + 1)].value = "Y"
                wsintbump_f[get_column_letter(int_tb_x + 2)  + str(str(int_tb_y + 1))].value = "Bump name"
                wsintbump_f.freeze_panes = 'A' + str(int_tb_y + 1)
                for c2 in range(0,3):
                    wsintbump_f[get_column_letter(int_tb_x + c2) + str(int_tb_y + 1)].fill = subtil_bg_fill
                    wsintbump_f[get_column_letter(int_tb_x + c2) + str(int_tb_y + 1)].alignment = Alignment(horizontal='center')
                    wsintbump_f[get_column_letter(int_tb_x + c2) + str(int_tb_y + 1)].border = Border(left=Side(style='thin'),right=Side(style='thin'),top=Side(style='thin'),bottom=Side(style='thin'))
                tbidx = 4
                for tb in range(0,int(int_die_cnt/2)):

                    wsintbump_f[get_column_letter(int_tb_x + tbidx) + str(int_tb_y)].value =  str(die1_list[tb]) + " = Die Flipped, Rotate -90 + Offset" + "(" + str(die1_xoffset_list[tb]) + "," + str(die1_yoffset_list[tb]) + ")"
                    wsintbump_f.merge_cells(get_column_letter(int_tb_x + tbidx) + str(int_tb_y) + ":" + get_column_letter(int_tb_x + tbidx + 2) + str(int_tb_y))
                    
                    for c1 in range(0,3):
                        wsintbump_f[get_column_letter(int_tb_x + tbidx + c1) + str(int_tb_y)].fill = title_bg_fill
                        wsintbump_f[get_column_letter(int_tb_x + tbidx + c1) + str(int_tb_y)].alignment = Alignment(horizontal='center', wrapText=True)
                        wsintbump_f[get_column_letter(int_tb_x + tbidx + c1) + str(int_tb_y)].border = Border(left=Side(style='thin'),right=Side(style='thin'),top=Side(style='thin'),bottom=Side(style='thin'))
                    wsintbump_f[get_column_letter(int_tb_x + tbidx) + str(int_tb_y + 1)].value = "X"
                    wsintbump_f[get_column_letter(int_tb_x + tbidx + 1) + str(int_tb_y + 1)].value = "Y"
                    wsintbump_f[get_column_letter(int_tb_x + tbidx + 2)  + str(str(int_tb_y + 1))].value = "Bump name"
                    for c2 in range(0,3):
                        wsintbump_f[get_column_letter(int_tb_x + tbidx + c2) + str(int_tb_y + 1)].fill = subtil_bg_fill
                        wsintbump_f[get_column_letter(int_tb_x + tbidx + c2) + str(int_tb_y + 1)].alignment = Alignment(horizontal='center')
                        wsintbump_f[get_column_letter(int_tb_x + tbidx + c2) + str(int_tb_y + 1)].border = Border(left=Side(style='thin'),right=Side(style='thin'),top=Side(style='thin'),bottom=Side(style='thin'))

                    wsintbump_f[get_column_letter(int_tb_x + tbidx + 4) + str(int_tb_y)].value = str(die2_list[tb]) + " = Die Flipped, Rotate +90 + Offset" + "(" + str(die2_xoffset_list[tb]) + "," + str(die2_yoffset_list[tb]) + ")"
                    wsintbump_f.merge_cells(get_column_letter(int_tb_x + tbidx + 4) + str(int_tb_y) + ":" + get_column_letter(int_tb_x + tbidx + 6) + str(int_tb_y))
                    for c1 in range(0,3):
                        wsintbump_f[get_column_letter(int_tb_x + tbidx + c1 + 4) + str(int_tb_y)].fill = title_bg_fill
                        wsintbump_f[get_column_letter(int_tb_x + tbidx + c1 + 4) + str(int_tb_y)].alignment = Alignment(horizontal='center', wrapText=True)
                        wsintbump_f[get_column_letter(int_tb_x + tbidx + c1 + 4) + str(int_tb_y)].border = Border(left=Side(style='thin'),right=Side(style='thin'),top=Side(style='thin'),bottom=Side(style='thin'))
                    wsintbump_f[get_column_letter(int_tb_x + tbidx + 4) + str(int_tb_y + 1)].value = "X"
                    wsintbump_f[get_column_letter(int_tb_x + tbidx + 5) + str(int_tb_y + 1)].value = "Y"
                    wsintbump_f[get_column_letter(int_tb_x + tbidx + 6)  + str(str(int_tb_y + 1))].value = "Bump name"
                    for c2 in range(0,3):
                        wsintbump_f[get_column_letter(int_tb_x + tbidx + c2 + 4) + str(int_tb_y + 1)].fill = subtil_bg_fill
                        wsintbump_f[get_column_letter(int_tb_x + tbidx + c2 + 4) + str(int_tb_y + 1)].alignment = Alignment(horizontal='center')
                        wsintbump_f[get_column_letter(int_tb_x + tbidx + c2 + 4) + str(int_tb_y + 1)].border = Border(left=Side(style='thin'),right=Side(style='thin'),top=Side(style='thin'),bottom=Side(style='thin'))
                    tbidx += 8
                int_input_correct = 1
        else:
            pass
  
        if (package_type == 1):
            
            print("Generate for Advance Package")
            mynotif(textbox, "Generate for Advance Package")
            dm_bump_coor= []
            dm_cnt=0
            # mynotif(textbox, "")
            root.update_idletasks()
            mynotif(textbox, "Generating Dummy bump...")
            root.update_idletasks()
            for dm_bump in dummybump:
                bump = list(dummybump[dm_bump].values())
                    
                ymin_dm = coordinate_to_tuple(bump[0])[0]
                xmin_dm = coordinate_to_tuple(bump[0])[1]
                ymax_dm = coordinate_to_tuple(bump[1])[0]
                xmax_dm = coordinate_to_tuple(bump[1])[1]
                xcoor_dm = str(bump[2])
                ycoor_dm = str(bump[3])

                print(xmin_dm,xmax_dm)
                print(ymin_dm,ymax_dm)
                print(dummybump)

                for dummycol1 in range(xmin_dm, xmax_dm + 1):
                    for dummyrow1 in range(ymin_dm, ymax_dm + 1):
                        col_dm = get_column_letter(dummycol1)
                        if (wsvisual_f[col_dm + str(dummyrow1)].value != None):
                            
                            print("Processing for Dummy bump at: " + col_dm + str(dummyrow1))
                            mynotif(textbox, "Processing for Dummy bump at: " + col_dm + str(dummyrow1))
                            # Gen dummy bump table
                            wsdiebump_f[get_column_letter(die_tb_x)+str(r_die)].value = f"='{bump_visual_sheet}'!{col_dm + xcoor_dm}"
                            wsdiebump_f[get_column_letter(die_tb_x + 1)+str(r_die)].value = f"='{bump_visual_sheet}'!{ycoor_dm + str(dummyrow1)}"
                            wsdiebump_f[get_column_letter(die_tb_x + 2)+str(r_die)].value =  f"='{bump_visual_sheet}'!{col_dm+ str(dummyrow1)}"
                            if(int(tc_sr['isTC'] == 1)):
                                wsdiebump_f[get_column_letter(die_tb_x + 4)+str(r_die)].value = f"=('{bump_visual_sheet}'!{col_dm + xcoor_dm})-({tc_sr['sr_w']})"
                                wsdiebump_f[get_column_letter(die_tb_x + 5)+str(r_die)].value = f"=('{bump_visual_sheet}'!{ycoor_dm + str(dummyrow1)})-({tc_sr['sr_w']}) "
                                wsdiebump_f[get_column_letter(die_tb_x + 6)+str(r_die)].value =  f"='{bump_visual_sheet}'!{col_dm+ str(dummyrow1)}"
                            r_die += 1
                            coor = col_dm + str(dummyrow1)
                            dm_bump_coor.append(coor)
                            dm_cnt += 1

                            if(int_gen == 1 and int_input_correct == 1):

                                wsintbump_f[get_column_letter(int_tb_x)+str(r_int)].value = f"=({str(int_die_params['chip_width']).replace('=','')})-('{bump_visual_sheet}'!{str(col_dm + xcoor_dm)})" # Flip Y axis
                                wsintbump_f[get_column_letter(int_tb_x + 1)+str(r_int)].value = f"='{bump_visual_sheet}'!{ycoor_dm + str(dummyrow1)}" # Flip Y axis
                                wsintbump_f[get_column_letter(int_tb_x + 2)+str(r_int)].value =  f"='{bump_visual_sheet}'!{col_dm+ str(dummyrow1)}" #Flip Y axis
                                tbidx2 = 0

                                for tb in range(0,int(int_die_cnt/2)):
                                # #----------------------------Flip bump map in y axis - Rotate -90 - Rotate +90---------------------------
                                    
                                    wsintbump_f[get_column_letter(int_tb_x + tbidx2 + 5)+str(r_int)].value = f"=({str(int_die_params['chip_width']).replace('=','')})-('{bump_visual_sheet}'!{str(col_dm + xcoor_dm)})+({str(die1_yoffset_list[tb])})" # Rotate -90
                                    wsintbump_f[get_column_letter(int_tb_x + tbidx2 + 9)+str(r_int)].value = f"=('{bump_visual_sheet}'!{str(col_dm + xcoor_dm)})+({str(die1_yoffset_list[tb]).replace('=','')})" # Rotate +90

                                    wsintbump_f[get_column_letter(int_tb_x + tbidx2 + 4)+str(r_int)].value = f"=({str(int_die_params['chip_height']).replace('=','')})-('{bump_visual_sheet}'!{ycoor_dm + str(dummyrow1)})+({str(die1_xoffset_list[tb])})" # Rotate -90
                                    wsintbump_f[get_column_letter(int_tb_x + tbidx2 + 8)+str(r_int)].value = f"=('{bump_visual_sheet}'!{ycoor_dm + str(dummyrow1)})+({str(die2_xoffset_list[tb]).replace('=','')})" # Rotate +90

                                
                                    if(wsvisual_f[col_dm+ str(dummyrow1)].value == "VSS"):
                                        wsintbump_f[get_column_letter(int_tb_x + tbidx2 + 6)+str(r_int)].value = f"='{bump_visual_sheet}'!{col_dm+ str(dummyrow1)}" # Rotate -90
                                        wsintbump_f[get_column_letter(int_tb_x + tbidx2 + 10)+str(r_int)].value = f"='{bump_visual_sheet}'!{col_dm+ str(dummyrow1)}" # Rotate +90
                                    else:
                                        wsintbump_f[get_column_letter(int_tb_x + tbidx2 + 6)+str(r_int)].value = f"=\"{die1_list[tb]}_\"&'{bump_visual_sheet}'!{col_dm+ str(dummyrow1)}" # Rotate -90
                                        wsintbump_f[get_column_letter(int_tb_x + tbidx2 + 10)+str(r_int)].value = f"=\"{die2_list[tb]}_\"&'{bump_visual_sheet}'!{col_dm+ str(dummyrow1)}" # Rotate +90
                                    tbidx2 += 8
                                
                                r_int += 1

            #---------Create Die bump exclued dummy bump at 4 corner-----------#

            match = 0
            # mynotif(textbox, "")
            root.update_idletasks()
            mynotif(textbox, "Generating Die bump...")
            root.update_idletasks()
            for col in range(xmin, xmax + 1):
                for row in range(ymin, ymax + 1):       
                    col_l = get_column_letter(col)
                    #print(col_l)
                    i = 0 
                    while(i < len(dm_bump_coor)):
                        xy = col_l + str(row)
                        if(xy ==  dm_bump_coor[i]):
                            match = 1
                        else:
                            match = 0
                        if(match == 1):
                            break
                        i += 1
                    if (match == 0 and wsvisual_f[col_l + str(row)].value != None):
                        print("Processing for Die bump at: " + col_l + str(row))
                        mynotif(textbox, "Processing for Die bump at: " + col_l + str(row))
                        #  get the X value from Visual bump sheet
                        if (wsvisual_f[col_l + str(row)].value != None):
                        
                            #  get the X value from Visual bump sheet
                        
                            wsdiebump_f[get_column_letter(die_tb_x)+str(r_die)].value = f"='{bump_visual_sheet}'!{col_l + str(die_coor['xcoor'])}"
                        
                            # #  get the Y value from Visual bump sheet
                        
                            wsdiebump_f[get_column_letter(die_tb_x + 1)+str(r_die)].value = f"='{bump_visual_sheet}'!{die_coor['ycoor'] + str(row)}"
                            
                            #  get the Bump name from Visual bump sheet
                            wsdiebump_f[get_column_letter(die_tb_x + 2)+str(r_die)].value =  f"='{bump_visual_sheet}'!{col_l+ str(row)}"
                            if(int(tc_sr['isTC'] == 1)):
                               wsdiebump_f[get_column_letter(die_tb_x + 4)+str(r_die)].value = f"=('{bump_visual_sheet}'!{col_l + str(die_coor['xcoor'])})-({tc_sr['sr_w']})" 
                               wsdiebump_f[get_column_letter(die_tb_x + 5)+str(r_die)].value = f"=('{bump_visual_sheet}'!{die_coor['ycoor'] + str(row)})-({tc_sr['sr_w']})"
                               wsdiebump_f[get_column_letter(die_tb_x + 6)+str(r_die)].value =  f"='{bump_visual_sheet}'!{col_l+ str(row)}"
                            r_die += 1
                            
                            if(int_gen == 1 and int_input_correct == 1):

                                wsintbump_f[get_column_letter(int_tb_x )+str(r_int)].value = f"=({str(int_die_params['chip_width']).replace('=','')})-('{bump_visual_sheet}'!{col_l + str(die_coor['xcoor'])})" # Flip Y axis
                                wsintbump_f[get_column_letter(int_tb_x + 1)+str(r_int)].value = f"='{bump_visual_sheet}'!{die_coor['ycoor'] + str(row)}" # Flip Y axis
                                wsintbump_f[get_column_letter(int_tb_x + 2)+str(r_int)].value =  f"='{bump_visual_sheet}'!{col_l+ str(row)}" #Flip Y axis

                                tbidx2 = 0
                                # r_current = 

                                for tb in range(0,int(int_die_cnt/2)):
                                # #----------------------------Flip bump map in y axis - Rotate -90 - Rotate +90---------------------------
                                    
                                    wsintbump_f[get_column_letter(int_tb_x + tbidx2 + 5)+str(r_int)].value = f"=({str(int_die_params['chip_width']).replace('=','')})-('{bump_visual_sheet}'!{col_l + str(die_coor['xcoor'])})+({str(die1_yoffset_list[tb])})" # Rotate -90
                                    wsintbump_f[get_column_letter(int_tb_x + tbidx2 + 9)+str(r_int)].value = f"=('{bump_visual_sheet}'!{col_l + str(die_coor['xcoor'])})+({str(die2_yoffset_list[tb]).replace('=','')})" # Rotate +90
                                    
                                    wsintbump_f[get_column_letter(int_tb_x + tbidx2 + 4)+str(r_int)].value = f"=({str(int_die_params['chip_height']).replace('=','')})-('{bump_visual_sheet}'!{die_coor['ycoor']+str(row)})+({str(die1_xoffset_list[tb])})" # Rotate -90
                                    wsintbump_f[get_column_letter(int_tb_x + tbidx2 + 8)+str(r_int)].value = f"=('{bump_visual_sheet}'!{die_coor['ycoor'] + str(row)})+({str(die2_xoffset_list[tb]).replace('=','')})" # Rotate +90
                                
                                    if(wsvisual_f[col_l+ str(row)].value == "VSS"):
                                        wsintbump_f[get_column_letter(int_tb_x + tbidx2 + 6)+str(r_int)].value = f"='{bump_visual_sheet}'!{col_l+ str(row)}" # Rotate -90
                                        wsintbump_f[get_column_letter(int_tb_x + tbidx2 + 10)+str(r_int)].value = f"='{bump_visual_sheet}'!{col_l+ str(row)}" # Rotate +90
                                    else:
                                        wsintbump_f[get_column_letter(int_tb_x + tbidx2 + 6)+str(r_int)].value = f"=\"{die1_list[tb]}_\"&'{bump_visual_sheet}'!{col_l+ str(row)}" # Rotate -90
                                        wsintbump_f[get_column_letter(int_tb_x + tbidx2 + 10)+str(r_int)].value = f"=\"{die2_list[tb]}_\"&'{bump_visual_sheet}'!{col_l+ str(row)}" # Rotate +90
                                    tbidx2 += 8
                                
                                r_int += 1
                                
                                
        else:
            mynotif(textbox, "Generating Die bump...")
            for col in range(xmin, xmax + 1):
                    for row in range(ymin , ymax + 1):       
                        col_l = get_column_letter(col)
                        #print(col_l)
                        if (wsvisual_f[col_l + str(row)].value != None):
                            print("Processing for Die bump at: " + col_l + str(row))
                            mynotif(textbox, "Processing for Die bump at: " + col_l + str(row))
                            #  get the X value from Visual bump sheet
                        
                            wsdiebump_f[get_column_letter(die_tb_x)+str(r_die)].value = f"='{bump_visual_sheet}'!{col_l + str(die_coor['xcoor'])}"
                        
                            # #  get the Y value from Visual bump sheet
                        
                            wsdiebump_f[get_column_letter(die_tb_x + 1)+str(r_die)].value = f"='{bump_visual_sheet}'!{die_coor['ycoor'] + str(row)}"
                            
                            #  get the Bump name from Visual bump sheet
                            wsdiebump_f[get_column_letter(die_tb_x + 2)+str(r_die)].value =  f"='{bump_visual_sheet}'!{col_l+ str(row)}"
                            if(int(tc_sr['isTC'] == 1)):
                               wsdiebump_f[get_column_letter(die_tb_x + 4)+str(r_die)].value = f"=('{bump_visual_sheet}'!{col_l + str(die_coor['xcoor'])})-({tc_sr['sr_w']})" 
                               wsdiebump_f[get_column_letter(die_tb_x + 5)+str(r_die)].value = f"=('{bump_visual_sheet}'!{die_coor['ycoor'] + str(row)})-({tc_sr['sr_w']})"
                               wsdiebump_f[get_column_letter(die_tb_x + 6)+str(r_die)].value =  f"='{bump_visual_sheet}'!{col_l+ str(row)}"
                            r_die += 1
                            if(int_gen == 1 and int_input_correct == 1):
                                
                                wsintbump_f[get_column_letter(int_tb_x )+str(r_int)].value = f"=({str(int_die_params['chip_width']).replace('=','')})-('{bump_visual_sheet}'!{col_l + str(die_coor['xcoor'])})" # Flip Y axis
                                wsintbump_f[get_column_letter(int_tb_x + 1)+str(r_int)].value = f"='{bump_visual_sheet}'!{die_coor['ycoor'] + str(row)}" # Flip Y axis
                                wsintbump_f[get_column_letter(int_tb_x + 2)+str(r_int)].value =  f"='{bump_visual_sheet}'!{col_l+ str(row)}" #Flip Y axis

                                tbidx2 = 0
                                # r_current = 

                                for tb in range(0,int(int_die_cnt/2)):
                                # #----------------------------Flip bump map in y axis - Rotate -90 - Rotate +90---------------------------
                                    
                                    wsintbump_f[get_column_letter(int_tb_x + tbidx2 + 5)+str(r_int)].value = f"=({str(int_die_params['chip_width']).replace('=','')})-('{bump_visual_sheet}'!{col_l + str(die_coor['xcoor'])})+({str(die1_yoffset_list[tb])})" # Rotate -90
                                    wsintbump_f[get_column_letter(int_tb_x + tbidx2 + 9)+str(r_int)].value = f"=('{bump_visual_sheet}'!{col_l + str(die_coor['xcoor'])})+({str(die2_yoffset_list[tb]).replace('=','')})" # Rotate +90
                                
                                    
                                    wsintbump_f[get_column_letter(int_tb_x + tbidx2 + 4)+str(r_int)].value = f"=({str(int_die_params['chip_height']).replace('=','')})-('{bump_visual_sheet}'!{die_coor['ycoor']+str(row)})+({str(die1_xoffset_list[tb])})" # Rotate -90
                                    wsintbump_f[get_column_letter(int_tb_x + tbidx2 + 8)+str(r_int)].value = f"=('{bump_visual_sheet}'!{die_coor['ycoor'] + str(row)})+({str(die2_xoffset_list[tb]).replace('=','')})" # Rotate +90
                                
                                
                                    if(wsvisual_f[col_l+ str(row)].value == "VSS"):
                                        wsintbump_f[get_column_letter(int_tb_x + tbidx2 + 6)+str(r_int)].value = f"='{bump_visual_sheet}'!{col_l+ str(row)}" # Rotate -90
                                        wsintbump_f[get_column_letter(int_tb_x + tbidx2 + 10)+str(r_int)].value = f"='{bump_visual_sheet}'!{col_l+ str(row)}" # Rotate +90
                                    else:
                                        wsintbump_f[get_column_letter(int_tb_x + tbidx2 + 6)+str(r_int)].value = f"=\"{die1_list[tb]}_\"&'{bump_visual_sheet}'!{col_l+ str(row)}" # Rotate -90
                                        wsintbump_f[get_column_letter(int_tb_x + tbidx2 + 10)+str(r_int)].value = f"=\"{die2_list[tb]}_\"&'{bump_visual_sheet}'!{col_l+ str(row)}" # Rotate +90
                                    tbidx2 += 8
                                

                                r_int += 1
            # tab = Table(displayName="Table1", ref="O65:Q500")
            # ws_f.add_table(tab)
        
        progress_bar(progressbar, 80)   
        print("Saving excel...") 
        mynotif(textbox, "Saving excel file...")
        wb_f.save(excel_path)
        progress_bar(progressbar, 100)
        mynotif(textbox, f"Successed!!!")
        print("Completed!!!")
        popup("PLOC generated successful!!!")
       
        
    except (ValueError):
        print ("Wrong input, Please check and regenerate")
        show_error("Wrong input, Please check and regenerate")
        progress_bar(progressbar, 0)
        mynotif(textbox, "Error")
        root.update_idletasks()
        button.state(state='normal')
        button.set_text("GENERATE")
        return
    except:
        print('Error!!!')
        
        show_error("There are an error in caculations, Please recheck and make sure the input is correct!")
        progress_bar(progressbar, 0)
        mynotif(textbox, "Error")
        root.update_idletasks()
        button.state(state='normal')
        button.set_text("GENERATE")
        return
    button.state(state='normal')
    button.set_text("GENERATE")
    
#######################################BEGIN FUNCT################################################################
get_saved_params(entry_ls, chkbtn_ls, combo_ls, text_ls, int_couple_ls, package_ls, theme_ls)
disable_entries(die_tbsr_o_name,die_tbsr_o_loc,dm_v_cor1_start,dm_v_cor1_end,dm_v_cor1_X,dm_v_cor1_Y,
                  dm_v_cor2_start,dm_v_cor2_end,dm_v_cor2_X,dm_v_cor2_Y,dm_v_cor3_start,dm_v_cor3_end,dm_v_cor3_X,dm_v_cor3_Y,
                  dm_v_cor4_start,dm_v_cor4_end,dm_v_cor4_X,dm_v_cor4_Y,chip_size_x,chip_size_y,
                  int_sheet,int_tb_o_loc,int_chip_L_name,int_chip_R_name,int_chip_L_xoffset,int_chip_R_xoffset,
                  int_chip_L_yoffset,int_chip_R_yoffset,sr_w,shrink_e )
sealring_begin = [" ", " ", "sr_w"]
toggle(sr_opt, combo_ls, entry_ls,textbox_ls, sealring_begin)

interposer_begin = [" ", " ", "int_", "chip_"]
toggle(int_opt, combo_ls, entry_ls, textbox_ls, interposer_begin)
choosemode(combo_ls, entry_ls)

btn_gen.button.config(command= lambda: process_and_generate(entry_ls,chkbtn_ls,combo_ls,text_box,pg_bar,btn_gen))
root.mainloop()