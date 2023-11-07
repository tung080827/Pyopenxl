from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.utils import range_boundaries
from openpyxl.utils import column_index_from_string
#from openpyxl.utils import coordinate_from_string
from openpyxl.utils import coordinate_to_tuple
from openpyxl.worksheet.table import Table, TableStyleInfo
# from tkinter import *
from tkinter import messagebox
from tkinter import ttk
import getcolumn
from array import *
import tkinter as tk
from tkinter import *
from ttkthemes import ThemedTk, THEMES
from PIL import Image
from PIL import ImageTk, Image
from tkinter.font import Font as tkfont
from tkinter import filedialog
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection,Font,Fill,Color
from openpyxl.worksheet.worksheet import Worksheet
import gui_function as gui
from ploc_myTk import *
from tkinter.filedialog import asksaveasfile
from copy import copy, deepcopy



def getstring(string: str,c1: str, c2: str):
    cell = string
    idx1 = cell.find(c1)
    idx2 = cell.find(c2)
    if(idx1 == -1 or idx2 == -1):
        return "NA","NA"
    else:
        str_wo_c = cell[idx1+1:idx2]
        str_w_c = cell[idx1:idx2+1]
        return str_wo_c,str_w_c
def row_col(cell: str):
    row = coordinate_to_tuple(cell)[0]
    col = coordinate_to_tuple(cell)[1]
    return row, col


# wb = load_workbook(r"C:\Users\sytung\Desktop\UCIe_IP_Bump_coordination_Standard.xlsx")
# ws = wb['UCIe_A_CoWoS_MMPL2']
# tb_name = 'MMX4'
summary_title_list =['Project', 'Technode','UCIe_type', 'Package_type', 'Stack_Interposer','Package_substrate', 'Data_bit_num', 'Bump_pitch_tc', 'Bump_pitch_Prod', 'Bump_UBM', 'Bump_PAD', 'Bumpcolla', 'Sheet', 'SI', 'PI' ]

prj_ls:list = ['None']
summary_info = None

# print(merge_lsit)
def cell(col,row):
    return get_column_letter(col)+str(row)
def getbumpvisualwindow(ws: Worksheet,tb_name: str):
    isfound = 0
    found_cnt = 0
    for merge in ws.merged_cells.ranges:
        idx = str(merge).find(':')
        cell_begin = str(merge)[:idx]
        cell_end = str(merge)[idx+1:]
        # print(f"cell begin: {cell_begin} \ncell end: {cell_end}")
        if (ws[cell_begin].value == tb_name):
            isfound = 1
            found_cnt += 1
            row_begin = row_col(cell_begin)[0]
            col_begin = row_col(cell_begin)[1]            
            col_end = row_col(cell_end)[1]
            # print(f"col begin: {col_begin}\ncol end: {col_end}\nrow begin: {row_begin}")
            borbegin: Border = ws[cell(col_begin, row_begin+1)].border.left.style
            row_end = row_begin + 1
            bor_cnt = 0
            row_end = row_begin
            row_b: int
            row_e: int = row_end
            for col in range(col_begin, col_end + 1):
                row = row_begin
                find_cell = 0
                while True:
                    bor = ws[cell(col, row)].border.left.style
                    if(find_cell == 0 and (bor is not None)):
                        # print(f"{row_e} : {row}")
                        if col == col_begin:
                            row_b = row
                            row_e = row
                        else:
                            if row < row_b:
                                row_b = row
                                # print(row_b)
                            
                        find_cell = 1
                    if(find_cell == 1 and (bor is None)):
                        if row > row_e:
                                row_e = row
                                # print(row_e)
                        break
                    row += 1
            if row_end < row_e:
                row_end = row_e
            if (row_begin < row_b):
                row_begin = row_b
            print(row_begin,row_end)
            merge_loc = merge.coord 

        else:
             pass
    if(isfound == 0):
        print(f"The table {tb_name} is not found") 
        return None
    elif(isfound ==1 and found_cnt > 1):
        print(f"More than 1 table \"{tb_name}\" present")
        return None
    else:
        # print(f" col: {col_begin} {col_end} \nrow: {row_begin} {row_end}")
        return col_begin, col_end, row_begin, row_end - 1, merge_loc

def gettable(ws: Worksheet,tb_name: str):
    isfound = 0
    found_cnt = 0
    for merge in ws.merged_cells.ranges:
        idx = str(merge).find(':')
        cell_begin = str(merge)[:idx]
        cell_end = str(merge)[idx+1:]
        # print(f"cell begin: {cell_begin} \ncell end: {cell_end}")
        if (ws[cell_begin].value == tb_name):
            isfound = 1
            found_cnt += 1
            row_begin = coordinate_to_tuple(cell_begin)[0]
            col_begin = coordinate_to_tuple(cell_begin)[1]
            col_end = coordinate_to_tuple(cell_end)[1]
            # print(f"col begin: {col_begin}\ncol end: {col_end}\nrow begin: {row_begin}")
            bor: Border = ws[cell(col_begin, row_begin+1)].border.left.style
            row_end = row_begin + 1
            while(bor is not None):
                row_end += 1
                bor = ws[cell(col_begin, row_end)].border.left.style
            row_end -= 1 
             
        
        else:
             pass
    if(isfound == 0):
        print(f"The table {tb_name} is not found") 
        return None
    elif(isfound ==1 and found_cnt > 1):
        print(f"More than 1 table \"{tb_name}\" present")
        return None
    else:
        # print(f" col: {col_begin} {col_end} \nrow: {row_begin} {row_end}")
        return col_begin, col_end, row_begin, row_end    
def getsummary_info(excel:str):
    wb = load_workbook(excel, data_only=True)
# getbumpvisualwindow(ws,tb_name) 
    summary_tb = gettable(wb['Summary'],'Project Summary')
    summary_col: dict = {}

    for title in range(summary_tb[0], summary_tb[1] + 1):
        summary_col.__setitem__(summary_title_list[title - summary_tb[0]],title)

    print(summary_col)
    prj_ls: list =[]
    summary_prj_row: dict = {}
    print(summary_col['Project'])
    for row in range(summary_tb[2]+3, summary_tb[3]+1):
        prj_ls.append(wb['Summary'][cell(summary_col['Project'],row)].value)
        summary_prj_row.__setitem__(wb['Summary'][cell(summary_col['Project'],row)].value,row)
    print(prj_ls, summary_prj_row)
    return summary_tb, summary_col, summary_prj_row, prj_ls,wb
def get_prj_sheet():
    global summary_info
    v_bump_list:list = []

    prj = proj_cb.get_value()
    # sheet = summary_info[4]['Summary'][cell(summary_info[1]['Sheet'],summary_info[2][prj])].value
    sheet = summary_info[4]['Summary'][cell(summary_info[1]['Sheet'],summary_info[2][prj])].value
    
    # print(sheet)
    v_refsheet_e.add_new_content(sheet)
    for merge in summary_info[4][sheet].merged_cells.ranges:
        idx = str(merge).find(':')
        cell_begin = str(merge)[:idx]
        cell_end = str(merge)[idx+1:]
        tb_name = summary_info[4][sheet][cell_begin].value

        if  tb_name not in summary_info[3]:
             v_bump_list.append(tb_name)
    v_bump_e.add_new_content(v_bump_list)
    print(v_bump_list)


def browse_file(entry: Tkentry):
    global prj_ls, summary_info
	# global my_image
    root.filename = filedialog.askopenfilename(title="Select A File", filetypes=(("Excel files", "*.xlsx"),("all files", "*.*")))
    # excel_i.delete(0,END)
    entry.add_new_content(root.filename)
    print(root.filename)
    summary_info = getsummary_info(root.filename)
    prj_ls = summary_info[3]
    proj_cb.combobox.config(values=prj_ls)
    proj_cb.set_current(0)
    # v_refsheet_e.add_new_content(summary_info[4]['Summary'][cell(summary_info[1]['Sheet'],summary_info[2][proj_cb.get_value()])].value)
    get_prj_sheet()

def save_file():
    prj = proj_cb.get_value()
    if(prj == 'None'):
        messagebox.showerror("Error!!!", 'The inputs is not correcct, please re-check')
    else:
        f = filedialog.asksaveasfile(initialfile = f'{prj}_Bump_coordination.xlsx', title="Save file",
        defaultextension=".xlsx",filetypes=[("All Files","*.*"),("Excel","*.xlsx")])
        p_out_excel_e.add_new_content(f.name)


root = ttk.Window(themename='united')
root.title("PLOC DATA CHANNEL VISUAL GENERATOR")
root.geometry("800x800")
root.resizable(width=True, height=True)
root.iconbitmap(r"./mylogo.ico")
root.option_add("*tearOff", False) # This is always a good idea
open_imag = PhotoImage(file = r".\open-folder.png")

img_path = r".\img\resize1000x1000"

bgm = PhotoImage(file=img_path + r"\frog.png")

my_canvas = tk.Canvas(root, bd=0, highlightthickness=0,relief='groove',scrollregion=(0,0,800,1200))
my_canvas.pack(fill="both", expand=True)
# my_canvas.bind_all('<Shift-MouseWheel>', on_vertical)
# my_canvas.bind("<Configure>",lambda event: on_window_resize(entry_ls,text_ls, textbox_ls, btn_ls, pgbar_ls, chkbtn_ls,combo_ls ))
bg_img = my_canvas.create_image(0,0,image=bgm,anchor='nw')
g = ['Day la cho d bo guiline']
p_excel_e = Tkentry(canvas=my_canvas,x=150,y=50,w=550,guide_text=g, win_defaultx=800, win_defaulty=800, justify='left')
v_refsheet_e = Tkentry(canvas=my_canvas,x=430,y=100,w=270,guide_text=g,win_defaultx=800, win_defaulty=800, justify='center')
v_bump_e = Tkentry(canvas=my_canvas,x=430,y=150,w=270,guide_text=g,win_defaultx=800, win_defaulty=800, justify='center')

p_out_excel_e = Tkentry(canvas=my_canvas,x=150,y=600,w=550,guide_text=g, win_defaultx=800, win_defaulty=800, justify='left')
# pwr_list_e = Tkentry(canvas=my_canvas,x=150,y=370,w=290,guide_text=g,win_defaultx=800, win_defaulty=800, justify='center')
# bus_char_e = Tkentry(canvas=my_canvas,x=300,y=330,w=140,guide_text=g,win_defaultx=800, win_defaulty=800, justify='center')
# ref_cell_start_e = Tkentry(canvas=my_canvas,x=150,y=290,w=140,guide_text=g,win_defaultx=800, win_defaulty=800, justify='center')
# ref_cell_end_e = Tkentry(canvas=my_canvas,x=150,y=330,w=140,guide_text=g,win_defaultx=800, win_defaulty=800, justify='center')


p_excel_t = CanvasText(canvas=my_canvas,x=30,y=55,win_defaultx=800, win_defaulty=800,text="PLOC file:",font=tkfont(family="Helvetica", size=11, slant='italic', underline=True, weight='bold'), fill='black', bgx=110, bgy= 24, bg_xo=6, bg_yo=6, isbg=True)
proj_t = CanvasText(canvas=my_canvas,x=30,y=100,win_defaultx=800, win_defaulty=800,text="Project:",font=tkfont(family="Helvetica", size=11, slant='italic', underline=True, weight='bold'), fill='black', bgx=110, bgy= 24,bg_xo=6, bg_yo=6, isbg=True)
v_refsheet_t = CanvasText(canvas=my_canvas,x=350,y=100,win_defaultx=800, win_defaulty=800,text="Sheet:",font=tkfont(family="Helvetica", size=11, slant='italic', underline=True, weight='bold'), fill='black', bgx=70, bgy= 24,bg_xo=6, bg_yo=6, isbg=True)
v_refsheet_t = CanvasText(canvas=my_canvas,x=350,y=100,win_defaultx=800, win_defaulty=800,text="Sheet:",font=tkfont(family="Helvetica", size=11, slant='italic', underline=True, weight='bold'), fill='black', bgx=70, bgy= 24,bg_xo=6, bg_yo=6, isbg=True)
p_out_excel_t = CanvasText(canvas=my_canvas,x=350,y=150,win_defaultx=800, win_defaulty=800,text="Bump conf:",font=tkfont(family="Helvetica", size=11, slant='italic', underline=True, weight='bold'), fill='black', bgx=110, bgy= 24, bg_xo=6, bg_yo=6, isbg=True)

proj_cb = TkCombobox(canvas=my_canvas,x=150, y=100,win_defaultx=800, win_defaulty=800, values=prj_ls)
proj_cb.combobox.bind('<<ComboboxSelected>>',lambda event: get_prj_sheet())
# sheet_cb = TkCombobox(canvas=my_canvas,x=300, y=290,win_defaultx=800, win_defaulty=800, values=prj_ls)


browse_btn = Tkbutton(canvas=my_canvas,x=705, y=50, w=40, h=25, win_defaultx=800, win_defaulty=800,)
browse_btn.button.config(image=open_imag, command= lambda: browse_file(p_excel_e))

save_btn = Tkbutton(canvas=my_canvas,x=705, y=600, w=40, h=25, win_defaultx=800, win_defaulty=800,)
save_btn.button.config(image=open_imag, command= lambda: save_file())
# theme_cb = TkCombobox(canvas=my_canvas,x=675, y=15,w=70,win_defaultx=800, win_defaulty=800,values=theme_ls)

# theme_cb.combobox.bind('<<ComboboxSelected>>', lambda event: choosetheme(combo_ls, entry_ls, text_ls, chkbtn_ls))
shrink_ckbtn = TKcheckbtn(win=root,canvas=my_canvas,x=30, y=410,win_defaultx=800, win_defaulty=800,text= "Shrink?", anchor='sw')

gen_btn = Tkbutton(canvas=my_canvas,x=250,y=680,w=300,win_defaultx=800, win_defaulty=800, text="GENERATE")

entry_ls : dict = {
    'p_excel' : p_excel_e,
    'v_refsheet': v_refsheet_e,
    'p_out_excel' : p_out_excel_e,
    'v_bump': v_bump_e

}
combo_ls ={
    'proj' : proj_cb,
}
chkbtn_ls ={
    'shrink' : shrink_ckbtn,
}
################################################BACK END################################################

def get_input(entry_list : dict[str, Tkentry], checkbtn_list: dict[str, TKcheckbtn], combo_list: dict[str, TkCombobox]):
    inputparams : dict[str,str] = {}
    for entry in entry_list:
        inputparams.__setitem__(entry,entry_list[entry].get())
    for chkbtn in checkbtn_list:
        inputparams.__setitem__(chkbtn,checkbtn_list[chkbtn].get_state())
    for combo in combo_list:
        inputparams.__setitem__(combo,combo_list[combo].get_value())
    print(inputparams)
    return inputparams
def getbumpmap_params(ws: Worksheet, tb_name: str):
    tb_col_begin,tb_col_end,tb_row_begin,tb_row_end = gettable(ws,tb_name)
    for tb_row in range(tb_row_begin, tb_row_end + 1):
        cell = ws.cell(column=tb_col_begin, row=tb_row)
        if(str(cell.value).lower().replace(' ','').find('xoffset')) != -1:
            xoffset = ws.cell(column=tb_col_begin + 1, row=tb_row).value
            # print(xoffset)
        elif(str(cell.value).lower().replace(' ','').find('yoffset')) != -1:
            yoffset = ws.cell(column=tb_col_begin + 1, row=tb_row).value
            # print(yoffset)
        elif(str(cell.value).lower().replace(' ','').find('xpitch')) != -1:
            xpitch = ws.cell(column=tb_col_begin + 1, row=tb_row).value
            # print(xpitch)
        elif(str(cell.value).lower().replace(' ','').find('ypitch')) != -1:
            ypitch = ws.cell(column=tb_col_begin + 1, row=tb_row).value
        elif(str(cell.value).lower().replace(' ','').find('shrinkfactor') != -1):
            shrink_factor = float(ws.cell(column=tb_col_begin + 1, row=tb_row).value)
        elif(str(cell.value).lower().replace(' ','').find('precision') != -1):
            precision = float(ws.cell(column=tb_col_begin + 1, row=tb_row).value)
            # print(ypitch)
    return xpitch, ypitch,xoffset, yoffset, shrink_factor, precision
def process_and_generate(entry_list : dict[str, Tkentry], checkbtn_list: dict[str, TKcheckbtn], combo_list: dict[str, TkCombobox]):
    global summary_info
    inwb = summary_info[4]
    input_params = get_input(entry_ls, chkbtn_ls, combo_ls)
    inexcel = input_params['p_excel']
    ref_sheet = input_params['v_refsheet']
    outexcel = input_params['p_out_excel']
    project = input_params['proj']
    shrink = int(input_params['shrink'])
    vbump_ls = list(str(input_params['v_bump']).split(' '))
    outwb = Workbook()
    in_v_ws:Worksheet = inwb[ref_sheet]
    out_v_ws: Worksheet = outwb.create_sheet('Bump Coordination')
    Xp, Yp, Xoffset, Yoffset,shrink_factor, precision = getbumpmap_params(inwb['Summary'], project)
    print(Xp, Yp, Xoffset, Yoffset, shrink_factor)
    roundecimal = str(precision)[::-1].find('.')
    if shrink == 1:
        Xp = Xp/shrink_factor
        Yp = Yp/shrink_factor
        Xoffset = Xoffset/shrink_factor
        Yoffset = Yoffset/shrink_factor
    title_bg_fill = PatternFill(patternType='solid', fgColor='9e42f5')
    subtil_bg_fill = PatternFill(patternType='solid',fgColor='0e7bf0')
    border_all = Border(left=Side(style='thin'),right=Side(style='thin'),top=Side(style='thin'),bottom=Side(style='thin'))
    for vbump in vbump_ls:
        col_begin, col_end, row_begin, row_end,merge_loc = getbumpvisualwindow(ws=in_v_ws, tb_name=vbump)
        out_v_ws.merge_cells(merge_loc)
        idx = str(merge_loc).find(':')
        bumpmatrix_name_cell = str(merge_loc)[:idx]
        out_v_ws[bumpmatrix_name_cell].value = vbump
        out_v_ws[bumpmatrix_name_cell].fill = title_bg_fill     
        out_v_ws[bumpmatrix_name_cell].alignment = Alignment(horizontal='center')
        #### Create Y coordinate
        row_Y = row_end
        Ycnt = 0
        while row_Y >= row_begin:
            outYcell = out_v_ws.cell(column=col_begin-1, row=row_Y)
            outYcell.value = round(Yoffset + (Yp/2)*Ycnt,roundecimal)
            Ycnt +=1
            row_Y -=1
        #### Create X coordinate        
        Xcnt = 0
        for col_X in range(col_begin, col_end +1):
            outXcell = out_v_ws.cell(column=col_X, row=row_end + 1)
            outXcell.value = round(Xoffset + (Xp/2)*Xcnt, roundecimal)
            Xcnt +=1
     
        ############## Create bump visual map
        for col in range (col_begin, col_end + 1):
            # for row in range (row_begin, row_end + 1):
            row = row_end
            while row >= row_begin:
                incell = in_v_ws.cell(column=col, row=row)
                outcell = out_v_ws.cell(column=col, row=row)
                outcell.border = copy(incell.border)
                outcell.font = copy(incell.font)
                outcell.alignment = copy(incell.alignment)
                outcell.fill = copy(incell.fill)
                if incell.value != None:
                    outcell.value = incell.value
                row -= 1
        #########Generate Coordinate table##########
        tb_begin_row = row_end + 10
        tb_begin_col = col_begin        
        
        out_v_ws.cell(row=tb_begin_row, column=tb_begin_col).value = f"{vbump}_coordinate"
        out_v_ws.merge_cells(f"{cell(tb_begin_col, tb_begin_row)}:{cell(tb_begin_col+2, tb_begin_row)}")
        out_v_ws.cell(row=tb_begin_row, column=tb_begin_col).fill = title_bg_fill
        out_v_ws.cell(row=tb_begin_row, column=tb_begin_col).border = border_all
        out_v_ws.cell(row=tb_begin_row, column=tb_begin_col).alignment = Alignment(horizontal='center')
        tb_header = ["X", "Y", "Bump"]
        hd_idx = 0
        for header_col in range(tb_begin_col, tb_begin_col+3):
            out_v_ws.cell(row=tb_begin_row+1, column=header_col).value = tb_header[hd_idx]
            out_v_ws.cell(row=tb_begin_row+1, column=header_col).fill = subtil_bg_fill
            out_v_ws.cell(row=tb_begin_row+1, column=header_col).border = border_all
            out_v_ws.cell(row=tb_begin_row+1, column=header_col).alignment = Alignment(horizontal='center')
            hd_idx+=1
        r = tb_begin_row + 2
        for col in range(col_begin, col_end+1):
            for row in range(row_begin, row_end+1):
                celloutX = out_v_ws.cell(column=tb_begin_col, row=r)
                celloutY = out_v_ws.cell(column=tb_begin_col+1, row=r)
                celloutBump = out_v_ws.cell(column=tb_begin_col+2, row=r)
                cellX = out_v_ws.cell(column=col, row=row_end+1)
                cellY = out_v_ws.cell(column=col_begin-1, row=row)
                cellBump = out_v_ws.cell(column=col, row=row)
                if(cellBump.value != None and cellBump.value != '' ):
                    celloutX.value = cellX.value
                    celloutY.value = cellY.value
                    celloutBump.value = cellBump.value
                    celloutX.border = border_all
                    celloutY.border = border_all
                    celloutBump.border = border_all
                    celloutX.alignment = Alignment(horizontal='center')
                    celloutY.alignment = Alignment(horizontal='center')
                    celloutBump.alignment = Alignment(shrinkToFit=True,horizontal='center')
                    r+=1

    
    outwb.save('Tungtest.xlsx')
    print("Completed!!")

gen_btn.button.config(command= lambda: process_and_generate(entry_ls, chkbtn_ls, combo_ls))
root.mainloop()