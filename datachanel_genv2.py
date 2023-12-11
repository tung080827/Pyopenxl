from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.utils import range_boundaries
from openpyxl.utils import column_index_from_string
#from openpyxl.utils import coordinate_from_string
from openpyxl.utils import coordinate_to_tuple
from openpyxl.worksheet.table import Table, TableStyleInfo
# from tkinter import *
from tkinter import messagebox
from tkinter import ttk
import getcolumn
from array import *
import tkinter as tk
from tkinter import *
# from ttkthemes import ThemedTk, THEMES
from PIL import Image
from PIL import ImageTk, Image
from tkinter.font import Font as tkfont
from tkinter import filedialog
import gui_function as gui
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection
import os
import win32com.client
from pathlib import Path  # core library
# excel_file = r"C:\Users\sytung\OneDrive - Synopsys, Inc\Desktop\py\Test.xlsx"
from datamapping_gen import * 
import tempfile

import random

root = ttk.Window(themename='united')
root.title("PLOC DATA CHANNEL VISUAL GENERATOR")
root.geometry("800x800")
root.resizable(width=True, height=True)
root.iconbitmap(r"./mylogo.ico")
root.option_add("*tearOff", False) # This is always a good idea
# bg = ImageTk.PhotoImage(file=r".\bg3_1.png")
try:
    temp_file =  os.path.join(tempfile.gettempdir(), ".datachannelgen_params_saved.txt")
    print(temp_file)
    tmp_flag = 0
except:
    messagebox.showerror("Can not find the User Temp dir")
    tmp_flag = 1

open_imag = PhotoImage(file = r".\open-folder.png")

img_path = r".\img\resize1000x1000"

bgm = PhotoImage(file=img_path + r"\frog.png")

# img_list = ["owl.png", "mountain.png","car.png", "penguin.png","sunset1.png", "flower3.png", "kid.png", "pug.png", "cat.png", "whale2.png", "elephant_grey.png", "snowman.png", "bee4.png", "elephant.png", "bee2.png", "fox.png", "beach.png", "frog.png", "cow.png", "forest.png", "owlpink2.png", "girl.png", "sand1.png", "baby2.png", "pig.png", "discord1.png" ]

# lable_bg_list = ["#F0F0F0","#EDEDED","#EBECEE","#F0F0F0","#F0F0F0","#FCFCFC","#EFF0F1","#EFF0F1","#EFF0F1","#EAECEF","#EFF0F1","#EFF0F1","#FECDD9","#EFF0F1","#EFF0F1","#EFF0F1","#EFF0F1","#EFF0F1","#EFF0F1","#EFF0F1","#EFF0F1", "#EFF0F1","#EFF0F1", "#EFF0F1","#EFF0F1","#E6EBEF"]

stfont= ("Franklin Gothic Medium", 10, 'underline', "italic")
# Create lists for the Comboboxes
# theme_list = ["adapta", "aquativo", "arc", "black","blue", "breeze", "clearlooks", "elegance", "equilux", "itft1", "keramik", "keramik_alt", "kroc", "plastik", "radiance", "ubuntu", "scidblue", "scidgreen", "scidgrey", "scidmint", "scidpink", "scidpurple", "scidsand", "smog", "winxpblue", "yaru" ]
# colour_list = ["#09a5e8", "#292b33", "#1583eb", "#292a2b","#1a7cad", "#0664bd", "#8baac7", "#59564f", "#40454a", "#7aa7f5", "#7795b4", "#7795b4", "#ebab0c", "#0c99eb", "#eb830c", "#eb830c", "#0937ab", "#37ed80", "#707371", "#479403", "#d12a9f", "#9b34eb", "#787122", "#118cbd", "#505257", "#924d8b" ]

theme_ls = ['minty','pulse','united','morph','darkly','cyborg','superhero']
textentry_fg_ls = ['black', 'black','black', 'black', 'white','white','white' ]
# text_bg_color_ls = ['#78C2AD', '#B4A7D6', '#EB6536', '#378DFC', '#375A7F', '#2A9FD6', '#4C9BE8']
text_bg_color_ls = ['#78C2AD', '#B4A7D6', '#FF9378', '#48C1F8', '#2F2F2F', '#373737', '#414D59'] 
text_fg_ls = ['black', 'black','black', 'black', 'white','white','white' ]
img_ls = ["frog3.png", "owlpurple.png","fox.png", "car.png", "kid.png", "snowman.png", "cow.png"]

# ch_number_ls = [1,2,3,4,5,6,7,8,9, 10, 11, 12, 13, 14, 15, 16,17,18,19,20,21,22,23,24,25,26,27,28,29,30,31,32]
ch_number_ls = [2,4,6,8,10,12,14,16,18,20,22,24,26,28,30,32]
ch_sequence_ls = ["Right to Left","Left to Right", "Center to Left first","Center to Right first", "Left Edge to Center first", "Right Edge to Center first"]
orient_ls = ["NS", "EW"]

# Create control variables

ismapgen = tk.IntVar()



# ------------------------------------------------------------------------------------------------------------------------------------------------

ref_cell_start_g = [
    "INFO: Reference Visual Channel window end cell\n ",
    "* Example:   A0           "
]
map_tb_sheet_g = [
    "INFO: Sheet to put Channels Mapping table\n\n ",
    "* Example: Mapping Data Channel "
]
map_tb_loc_g = [
    "INFO: This field to define the output mapping table location.  \n\n",
    "* Example: O64 "
]
ref_cell_end_g = [
    "INFO: Reference Visual Channel window end cell.  \n\n",
    "* Example: CU100  "
]
out_name_in_guide = [
    "INFO: This field to define the output mapping table name.  \n\n",
    "* Example: Mapping  "
]
ch_o_loc_e = [
    "INFO: This field to define the output Bump visual location.  \n\n",
    "* Example: D10  "
]
ch_o_sheet_g = [
    "INFO: This field to define Sheet to put output Bump Channels visual .  \n\n",
    "* Example: Bump coordination  "
]
pwr_list_g = [
    "INFO: Power list.  \n",
    "- The power names are separated by spaces.\n ",
    "* Example: VDD VCCAON VCCIO VSS  "
]

v_refsheet_g = [
    "INFO: This field to define Sheet of reference channel bump visual.  \n\n",
    "* Example: Bump coordination  "
]
excel_g = [
    "\n\nINFO: This field for puting excel file\n\n ",   
]
die_L_list_g = [
    "INFO: List Name of interposer Die Left/Up which is outcome of Die Flipped then Rotate -90 degree\n ",
    "- The dies name are separated by spaces.\n ",
    "NOTE: The Die name is mapping between Die Left/Up and Die Right/Down. \n",
    "  For example:\n",
    "   + Die Right list name: DIE5 DIE6 DIE7 DIE8, and\n",
    "   + Die Left list name: DIE1 DIE2 DIE3 DIE4 then:\n",
    "               DIE1 <=> DIE5 \n",
    "               DIE2 <=> DIE6 \n",
    "               DIE3 <=> DIE7 \n",
    "               DIE4 <=> DIE8"
]
die_R_list_g = [
    "INFO: List Name of interposer Die Right/Down which is outcome of Die Flipped then Rotate +90 degree\n ",
    "- The dies name are separated by spaces.\n ",
    "NOTE: The Die name is mapping between Die Left/Up and Die Right/Down. \n",
    "   For example:\n",
    "    + Die Left list name: DIE1 DIE2 DIE3 DIE4, and\n",
    "    + Die Right list name: DIE5 DIE6 DIE7 DIE8, then:\n",
    "               DIE1 <=> DIE5 \n",
    "               DIE2 <=> DIE6 \n",
    "               DIE3 <=> DIE7 \n",
    "               DIE4 <=> DIE8"
]
map_char_g = [
    "INFO: This field to define mapping character between 2 Die.\n",
    "- The characters are separated by spaces.\n ",
    "* Example: TX RX  "
]
bus_char_g = [
    "INFO: This field to define Signal Bus character between 2 Die.\n",
    "- The characters are not seprated.\n ",
    "* Example: []  "
]
num_ch_g = [
    "INFO: This field to define number of channel that would to be generated.\n",

]
ch_combo_g = [
    "INFO: Please choose the output Channel number here.\n",

]
def on_vertical(event):
    my_canvas.yview_scroll(-1 * event.delta, 'units')

def on_horizontal(event):
    my_canvas.xview_scroll(-1 * event.delta, 'units')
def entry_responsive(entry_list :dict[str, Tkentry], w: int, h : int):
    for entry in entry_list:
        entry_list[entry].moveto(w,h)
        entry_list[entry].change_width_height(w,h)
        entry_list[entry].change_textsize(w,h)
def textbox_responsive(textbox_list : dict[str, TkTextbox], w: int, h : int):
    for box in textbox_list:
        textbox_list[box].moveto(w,h)
        textbox_list[box].change_width_height(w,h)
def text_reponsive(text_list : dict[str, CanvasText] ,w : int, h : int):
    for text in text_list:
        text_list[text].moveto(w,h)
        text_list[text].set_size(w,h)
def button_responsive(btn_list : dict[str, Tkbutton],w : int,h : int):
    for btn in btn_list:
        btn_list[btn].moveto(w,h)
        btn_list[btn].change_width_height(w,h)
def progressbar_responsive(pgbar_list : dict[str, Tkprogressbar],w : int,h : int):
    for pgbar in pgbar_list:
        pgbar_list[pgbar].moveto(w,h)
        pgbar_list[pgbar].change_width_height(w,h)
def checkbtn_responsive(checkbtn_list : dict[str, TKcheckbtn], w: int, h : int) :
    for chkbtn in checkbtn_list:
        checkbtn_list[chkbtn].moveto(w,h)
        # chkbtn.change_width_height(h)
def cobobox_responsive(cb_list : dict[str, TkCombobox], w: int, h : int):
    for cb in cb_list:
        cb_list[cb].moveto(w,h)
        cb_list[cb].change_width_height(w,h)
def resize_image_bg(myCanvas: Canvas, combo: TkCombobox):
    global bgm
    width = myCanvas.winfo_width()
    height = myCanvas.winfo_height()
    idx = theme_ls.index(combo.get_value())
    p = os.path.join(img_path, img_ls[idx])
    if(width>=height):
        size = height
    else:
        size = width
    img= (Image.open(p))

    # #Resize the Image using resize method
    resized_image= img.resize((size,size), Image.LANCZOS)
    bgm= ImageTk.PhotoImage(resized_image)
    myCanvas.itemconfigure(bg_img, image=bgm)
    myCanvas.moveto(bg_img, (width - size)/2, (height - size)/2)
def on_window_resize(entry_list :dict[str, Tkentry],text_list : dict[str, CanvasText],textbox_list : dict[str, TkTextbox], btn_list : dict[str, Tkbutton],
                      pgbar_list : dict[str, Tkprogressbar], chkbtn_list : dict[str, TKcheckbtn],combo_list : dict[str, TkCombobox]):
    global bgm
    if my_canvas.winfo_width() < 500:
        root.geometry(f"500x{my_canvas.winfo_height()}")
    if my_canvas.winfo_height()<700:
        root.geometry(f"{my_canvas.winfo_width()}x700")
    width = my_canvas.winfo_width()
    height = my_canvas.winfo_height()
    # print(f"Window resized to {width}x{height}")    
    entry_responsive(entry_list,w=width,h=height)
    text_reponsive(text_list, w=width,h=height)
    textbox_responsive(textbox_list,w=width, h=height)
    button_responsive(btn_list,w=width,h=height)
    progressbar_responsive(pgbar_list,w=width, h=height)
    checkbtn_responsive(chkbtn_list, w=width, h=height)
    cobobox_responsive(combo_list, w=width, h=height)
    resize_image_bg(my_canvas, combo_list['theme'])

def disable_entries(*entries:Tkentry):
    for entry in entries:
        entry.disable()
def enable_entries(*entries:Tkentry):
    for entry in entries:
        entry.enable()

def toggle(checkbtn : TKcheckbtn,combo_list: dict[str, TkCombobox], entry_list : dict[str, Tkentry],textbox_list : dict[str, TkTextbox], content : list[str]):
    print(f"toggled, state: {checkbtn.get_state()}")
    idx = theme_ls.index(combo_list['theme'].get_value())
    enable_fg = text_fg_ls[idx]
    if(content[len(content)-1] == 'not_yet'):
        crying = icons.Emoji.get('CRYING FACE')
        messagebox.showinfo("Notification", f"This feature is not developed yet {crying}")
    else:
        if(checkbtn.get_state() == 1):        
            print(content[0])
            for entry in entry_list:
                for c in range(2,len(content)):
                    if str(entry).find(content[c]) != -1 :
                        enable_entries(entry_list[entry])
                        entry_list[entry].set_fg(enable_fg)  
            textbox_list['text_box'].add_new_text(content[0] + "\n")

        elif(checkbtn.get_state() == 0):            
            print(content[0])
            for entry in entry_list:                
                for c in range(2,len(content)):
                    if str(entry).find(content[c]) != -1 :
                        disable_entries(entry_list[entry])
                        entry_list[entry].set_fg('#A9A9A9')                        
            textbox_list['text_box'].add_new_text(content[1] + "\n")

def browse_file(entry: Tkentry):
	# global my_image
    root.filename = filedialog.askopenfilename(initialdir="./", title="Select A File", filetypes=(("Excel files", "*.xlsx"),("all files", "*.*")))
    # excel_i.delete(0,END)
    entry.add_new_content(root.filename)
    print(root.filename) 
def get_input(entry_list : dict[str, Tkentry], checkbtn_list: dict[str, TKcheckbtn], combo_list: dict[str, TkCombobox]):
    inputparams : dict[str,str] = {}
    for entry in entry_list:
        inputparams.__setitem__(entry,entry_list[entry].get())
    for chkbtn in checkbtn_list:
        inputparams.__setitem__(chkbtn,checkbtn_list[chkbtn].get_state())
    for combo in combo_list:
        inputparams.__setitem__(combo,combo_list[combo].get_value())
    print(inputparams)
    return inputparams
    # excel_i.insert(0, root.filename)


def resize_image_bg(myCanvas: Canvas, combo: TkCombobox):
    global bgm
    width = myCanvas.winfo_width()
    height = myCanvas.winfo_height()
    idx = theme_ls.index(combo.get_value())
    p = os.path.join(img_path, img_ls[idx])
    if(width>=height):
        size = height
    else:
        size = width
    img= (Image.open(p))

    # #Resize the Image using resize method
    resized_image= img.resize((size,size), Image.LANCZOS)
    bgm= ImageTk.PhotoImage(resized_image)
    myCanvas.itemconfigure(bg_img, image=bgm)
    myCanvas.moveto(bg_img, (width - size)/2, (height - size)/2)
    

def set_theme(combo : TkCombobox, theme_name : str, themelist: list, entry_list: dict[str, Tkentry], text_list: dict[str, CanvasText]):
    root.style.theme_use(theme_name)
    combo.set_current(themelist.index(theme_name))
    idx = theme_ls.index(combo.get_value())
    for entry in entry_list:
        entry_list[entry].entry.config(foreground=textentry_fg_ls[idx])
    for text in text_list:
        text_list[text].change_color(text_fg_ls[idx])
        text_list[text].set_bg_color(text_bg_color_ls[idx])

def choosetheme(combo_list: dict[str,TkCombobox], entry_list: dict[str, Tkentry], text_list: dict[str, CanvasText], checkbutton_list : dict[str, TKcheckbtn]):
    # root.set_theme(combo.combobox.get())
    root.style.theme_use(combo_list['theme'].get_value())
    idx = theme_ls.index(combo_list['theme'].get_value())
    enable_fg = text_fg_ls[idx]
    for entry in entry_list:
      entry_list[entry].entry.config(foreground=textentry_fg_ls[idx])
    for text in text_list:
        text_list[text].change_color(text_fg_ls[idx])
        text_list[text].set_bg_color(text_bg_color_ls[idx])
    resize_image_bg(my_canvas, combo_list['theme'])
 
    map_opt_val = checkbutton_list['gen_map'].get_state()
    ls = ['map_', 'die_']
    if(map_opt_val == 1):
        for entry in entry_list:
            for e in ls:
                if str(entry).find(e) != -1 :
                    entry_list[entry].set_fg(enable_fg)
    else:
        for entry in entry_list:
            for e in ls:
                if str(entry).find(e) != -1 :
                    entry_list[entry].set_fg('#A9A9A9')
def choose_sequence(combo_list: dict[str,TkCombobox], entry_list: dict[str, Tkentry], text_list: dict[str, CanvasText], checkbutton_list : dict[str, TKcheckbtn]):
    orient = combo_list['orient'].get_value()
    global ch_sequence_ls
    if orient == "NS":
        ch_sequence_ls = ["Right to Left","Left to Right", "Center to Left first","Center to Right first", "Left Edge to Center first", "Right Edge to Center first"]

    elif orient == "EW":
        ch_sequence_ls = ["Top to Bot","Bot to Top", "Center to Bot first","Center to Top first", "Bot Edge to Center first", "Top Edge to Center first"]
    combo_list['ch_seq'].combobox.config(values=ch_sequence_ls)
    combo_list['ch_seq'].combobox.current(0)
def progress_bar(bar: Tkprogressbar,value):
    bar.update(value)
    root.update_idletasks()
xfont = ("System", 12, "bold", 'underline', 'italic')
# Define Canvas
my_canvas = tk.Canvas(root, bd=0, highlightthickness=0,relief='groove',scrollregion=(0,0,800,1200))
my_canvas.pack(fill="both", expand=True)
my_canvas.bind_all('<Shift-MouseWheel>', on_vertical)
my_canvas.bind("<Configure>",lambda event: on_window_resize(entry_ls,text_ls, textbox_ls, btn_ls, pgbar_ls, chkbtn_ls,combo_ls ))
bg_img = my_canvas.create_image(0,0,image=bgm,anchor='nw')

text_box = TkTextbox(canvas=my_canvas,x=400,y=80,w=300, h=100, win_defaultx=800, win_defaulty=800,)
text_box.add_new_text("This tool used for Ploc data channel genetation based on standard data channel.\nAnd generate Mapping table")
text_box.textbox.config(foreground='#F75726', font=tkfont(family="Courier", size="10"))

p_excel_e = Tkentry(canvas=my_canvas,x=150,y=50,w=550,guide_text=excel_g, win_defaultx=800, win_defaulty=800, justify='left')
v_refsheet_e = Tkentry(canvas=my_canvas,x=150,y=250,w=140,guide_text=v_refsheet_g,win_defaultx=800, win_defaulty=800, justify='center')
pwr_list_e = Tkentry(canvas=my_canvas,x=150,y=370,w=290,guide_text=pwr_list_g,win_defaultx=800, win_defaulty=800, justify='center')
bus_char_e = Tkentry(canvas=my_canvas,x=300,y=330,w=140,guide_text=bus_char_g,win_defaultx=800, win_defaulty=800, justify='center')
ref_cell_start_e = Tkentry(canvas=my_canvas,x=150,y=290,w=140,guide_text=ref_cell_start_g,win_defaultx=800, win_defaulty=800, justify='center')
ref_cell_end_e = Tkentry(canvas=my_canvas,x=150,y=330,w=140,guide_text=ref_cell_end_g,win_defaultx=800, win_defaulty=800, justify='center')
ch_o_sheet_e = Tkentry(canvas=my_canvas,x=500,y=250,w=140,guide_text=ch_o_sheet_g,win_defaultx=800, win_defaulty=800, justify='center')
ch_o_loc_e = Tkentry(canvas=my_canvas,x=500,y=290,w=140,guide_text=ch_o_loc_e,win_defaultx=800, win_defaulty=800, justify='center')
die_L_list_e = Tkentry(canvas=my_canvas,x=150,y=505,w=290,guide_text=die_L_list_g,win_defaultx=800, win_defaulty=800, justify='center')
die_R_list_e = Tkentry(canvas=my_canvas,x=150,y=575,w=290,guide_text=die_R_list_g,win_defaultx=800, win_defaulty=800, justify='center')
map_tb_sheet_e = Tkentry(canvas=my_canvas,x=500,y=505,w=140,guide_text=map_tb_sheet_g,win_defaultx=800, win_defaulty=800, justify='center')
map_tb_loc_e = Tkentry(canvas=my_canvas,x=500,y=575,w=140,guide_text=map_tb_loc_g,win_defaultx=800, win_defaulty=800, justify='center')
map_char_e = Tkentry(canvas=my_canvas,x=260,y=435,w=180,guide_text=map_char_g,win_defaultx=800, win_defaulty=800, justify='center')


theme_cb = TkCombobox(canvas=my_canvas,x=675, y=15,w=70,win_defaultx=800, win_defaulty=800,values=theme_ls)
theme_cb.combobox.bind('<<ComboboxSelected>>', lambda event: choosetheme(combo_ls, entry_ls, text_ls, chkbtn_ls))
ch_cb = TkCombobox(canvas=my_canvas,x=300, y=250,win_defaultx=800, win_defaulty=800,values=ch_number_ls, is_bind=1, guidetext=ch_combo_g)
ch_seq_cb = TkCombobox(canvas=my_canvas,x=300, y=290,win_defaultx=800, win_defaulty=800,values=ch_sequence_ls)
orient_cb = TkCombobox(canvas=my_canvas,x=300, y=220,win_defaultx=800, win_defaulty=800,values=orient_ls)
orient_cb.combobox.bind('<<ComboboxSelected>>', lambda event: choose_sequence(combo_ls, entry_ls, text_ls, chkbtn_ls))

gen_map_ckbtn = TKcheckbtn(win=root,canvas=my_canvas,x=30, y=410,win_defaultx=800, win_defaulty=800,text= "Gen Mapping table?", anchor='sw')
gen_map_noti = ["- Generate Die mapping table: ON ", "- Generate Die mapping table: OFF", "map_","die_"]
gen_map_ckbtn.checkbtn.config(command= lambda : toggle(gen_map_ckbtn, combo_ls, entry_ls, textbox_ls, gen_map_noti))




p_excel_t = CanvasText(canvas=my_canvas,x=30,y=55,win_defaultx=800, win_defaulty=800,text="PLOC file:",font=tkfont(family="Helvetica", size=11, slant='italic', underline=True, weight='bold'), fill='black', bgx=110, bgy= 24, bg_xo=6, bg_yo=6, isbg=True)
theme_t = CanvasText(canvas=my_canvas,x=570,y=20,win_defaultx=800, win_defaulty=800,text="Choose theme:",font=tkfont(family="Helvetica", size=8, slant='italic', underline=True, weight='bold'), fill='black', bgx=100, bgy= 24,bg_xo=6, bg_yo=6, isbg=True)
inout_cfg_t = CanvasText(canvas=my_canvas,x=30,y=180,win_defaultx=800, win_defaulty=800,text="Input/Output Config:",font=tkfont(family="Helvetica", size=11, slant='italic', underline=True, weight='bold'), fill='black', bgx=180, bgy= 24,bg_xo=6, bg_yo=6, isbg=True)
in_cfg_t = CanvasText(canvas=my_canvas,x=150,y=220,win_defaultx=800, win_defaulty=800,text="Input:",font=tkfont(family="Helvetica", size=9, slant='italic', underline=True, weight='bold'), fill='black', bgx=90, bgy= 24,bg_xo=6, bg_yo=6, isbg=True)
out_cfg_t = CanvasText(canvas=my_canvas,x=500,y=220,win_defaultx=800, win_defaulty=800,text="Output:",font=tkfont(family="Helvetica", size=9, slant='italic', underline=True, weight='bold'), fill='black', bgx=90, bgy= 24,bg_xo=6, bg_yo=6, isbg=True)
map_char_list_t = CanvasText(canvas=my_canvas,x=150,y=440,win_defaultx=800, win_defaulty=800,text="Map char list:",font=tkfont(family="Helvetica", size=9, slant='italic', underline=True, weight='bold'), fill='black', bgx=110, bgy= 24,bg_xo=6, bg_yo=6, isbg=True)
die_L_list_t = CanvasText(canvas=my_canvas,x=150,y=480,win_defaultx=800, win_defaulty=800,text="Die L name list:",font=tkfont(family="Helvetica", size=9, slant='italic', underline=True, weight='bold'), fill='black', bgx=110, bgy= 24,bg_xo=6, bg_yo=6, isbg=True)
die_R_list_t = CanvasText(canvas=my_canvas,x=150,y=550,win_defaultx=800, win_defaulty=800,text="Die R name list:",font=tkfont(family="Helvetica", size=9, slant='italic', underline=True, weight='bold'), fill='black', bgx=110, bgy= 24,bg_xo=6, bg_yo=6, isbg=True)
map_tb_sheet_t = CanvasText(canvas=my_canvas,x=500,y=480,win_defaultx=800, win_defaulty=800,text="Mapping sheet:",font=tkfont(family="Helvetica", size=9, slant='italic', underline=True, weight='bold'), fill='black', bgx=110, bgy= 24,bg_xo=6, bg_yo=6, isbg=True)
map_tb_loc_t = CanvasText(canvas=my_canvas,x=500,y=550,win_defaultx=800, win_defaulty=800,text="Mapping location:",font=tkfont(family="Helvetica", size=9, slant='italic', underline=True, weight='bold'), fill='black', bgx=120, bgy= 24,bg_xo=6, bg_yo=6, isbg=True)
# in_cfg_t = CanvasText(canvas=my_canvas,x=150,y=230,text="Input:",font=tkfont(family="Helvetica", size=11, slant='italic', underline=True, weight='bold'), fill='black', bgx=110, bgy= 24, isbg=True)
pg_bar = Tkprogressbar(my_canvas,x=80,y=630,w=600, win_defaultx=800, win_defaulty=800,)
browse_btn = Tkbutton(canvas=my_canvas,x=705, y=50, w=40, h=25, win_defaultx=800, win_defaulty=800,)
browse_btn.button.config(image=open_imag, command= lambda: browse_file(entry_ls['pxcel']))

gen_btn = Tkbutton(canvas=my_canvas,x=250,y=680,w=300,win_defaultx=800, win_defaulty=800, text="GENERATE")



entry_ls={
    'pxcel': p_excel_e,
    'v_refsheet': v_refsheet_e,
    'pwr_list': pwr_list_e,
    'bus_char': bus_char_e,
    'ref_cell_start':ref_cell_start_e,
    'ref_cell_end':ref_cell_end_e,
    'ch_o_sheet':ch_o_sheet_e,
    'ch_o_loc':ch_o_loc_e,
    'die_L_list':die_L_list_e,
    'die_R_list':die_R_list_e,
    'map_tb_sheet':map_tb_sheet_e,
    'map_tb_loc': map_tb_loc_e,
    'map_char': map_char_e
}
text_ls={
    'p_excel': p_excel_t,
    'theme':theme_t,
    'inout_cfg' : inout_cfg_t,
    'in_cfg' : in_cfg_t,
    'out_cfg' : out_cfg_t,
    'die_L_list' : die_L_list_t,
    'die_R_list' : die_R_list_t,
    'map_tb_sheet' :map_tb_sheet_t,
    'map_tb_loc' : map_tb_loc_t,
    'map_char_list': map_char_list_t
}
textbox_ls={
    'text_box' : text_box
}
btn_ls={
    'browse' : browse_btn,
    'gen' : gen_btn
} 
pgbar_ls={
    'pg' : pg_bar
}
chkbtn_ls={
    'gen_map' : gen_map_ckbtn
}    
combo_ls={
    'theme' :theme_cb,
    'ch' :ch_cb,
    'ch_seq' : ch_seq_cb,
    'orient' : orient_cb
}





# Button
#Create style object
# style = ttk.Style()

# style.configure('TCheckbutton', font= ('System', 12, 'underline', 'bold'),
#  foreground='black', border=50)
mediumFont = tkfont(
	family="System",
	size=16,
	weight="normal",
	slant="italic",
	underline=1,
	overstrike=0)



def get_indexsubstring(string:str,sub:str):
    count_er=0
    start_index=0
    idx:list=[]
    for i in range(len(string)):
        j = string.find(sub,start_index)
        if(j!=-1):
            start_index = j+1
            count_er+=1
            idx.append(j)
        print("Total occurrences are: ", count_er)
        print("index: ",idx)
    return idx
def get_lastsub(string:str, singlechar:str):
    idx_ls = get_indexsubstring(string, singlechar)
    return string[idx_ls[len(idx_ls)-1]:]

# power_list = ['VDD', 'VCCIO', 'VCCAON', 'VSS']
# buschar ="[]"
def get_bus(bumplist: list, power_list:list, buschar:str):
    buscharls = list(buschar)
    netdict : dict ={}
    for pwr in power_list:
        bumplist.remove(pwr)
    for net in bumplist:
        s = getstring(net,buscharls[0], buscharls[1])
        if s[0] == None:
            if net not in power_list:
                netdict.__setitem__(net,1)
    keysList = list(netdict.keys())
    for key in keysList:
        bumplist.remove(key)
    print(bumplist)

    while bumplist:
        s = getstring(bumplist[0],buscharls[0], buscharls[1])
        cnt = 0
        templist = []
        for net in bumplist:
            s1 = getstring(net,buscharls[0], buscharls[1])
            if s1[2] == s[2]:
            # if str(net).find(s[2]) != -1:
            # if str(s[2]).find(net) != -1:
                cnt +=1
                templist.append(net)
        netdict.__setitem__(s[2],cnt)
        for ls in templist:
            bumplist.remove(ls)
    print(netdict)
    return netdict

def generate_color():
    color = '{:02x}{:02x}{:02x}'.format(*map(lambda x: random.randint(50, 255), range(3)))
    return color

def process(textbox: TkTextbox, entry_list: dict[str, Tkentry], checkbtn_list: dict[str, TKcheckbtn], combo_list: dict[str, TkCombobox], text_list: dict[str, CanvasText],theme_list: list, progressbar: Tkprogressbar, button:Tkbutton):
    button.set_text("Generating...")
    button.state(state='disable')
    mynotif(textbox, "Getting params...")
    progress_bar(progressbar, 10)
    try:
        input_got = get_input(entry_list, checkbtn_list, combo_list)
        input_params ={
            "excel_file": input_got['pxcel'],
            "ch_sheet": input_got['v_refsheet'],
            "ch_cell_start": input_got['ref_cell_start'],
            "ch_cell_end": input_got['ref_cell_end'],

            "ch_cnt": input_got['ch'],
            "pwr_list":input_got['pwr_list'],
            "bus_char":input_got['bus_char'],
            "ch_seq":input_got['ch_seq'],
            'DieL_name': input_got['die_L_list'],
            'DieR_name': input_got['die_R_list'],
            'gen_map': input_got['gen_map'],
            'map_char': input_got['map_char'],
            'orient': input_got['orient']
        }

        output_params ={
            "ch_o_sheet": input_got['ch_o_sheet'],
            "ch_o_loc": input_got['ch_o_loc']
        }

        mapping_tb_out = {
            'sheet_name': input_got['map_tb_sheet'],
            'tb_ch2ch_loc':input_got['map_tb_loc'],
    }
        global tmp_flag, temp_file
        if(tmp_flag == 0):
            with open(temp_file,'w') as params_saved:
                params_saved.writelines(input_params['excel_file'] +"\n")
                params_saved.writelines(input_params['ch_sheet'] +"\n")
                params_saved.writelines(input_params['ch_cnt'] +"\n")
                params_saved.writelines(input_params['ch_seq'] +"\n")
                params_saved.writelines(input_params['pwr_list'] +"\n")
                params_saved.writelines(input_params['ch_cell_start'] +"\n")
                params_saved.writelines(input_params['ch_cell_end'] +"\n")
                params_saved.writelines(output_params['ch_o_sheet'] +"\n")
                params_saved.writelines(output_params['ch_o_loc'] +"\n")
                params_saved.writelines(input_params['DieL_name'] + "\n")
                params_saved.writelines(input_params['DieR_name'] + "\n")
                params_saved.writelines(mapping_tb_out['sheet_name'] + "\n")
                params_saved.writelines(mapping_tb_out['tb_ch2ch_loc'] + "\n")
                params_saved.writelines(combo_list['theme'].get_value() + "\n")
                params_saved.writelines(str(checkbtn_list['gen_map'].get_state()) + "\n")
                params_saved.writelines(input_params['bus_char'] + "\n")
                params_saved.writelines(input_params['map_char'] + "\n")

        progress_bar(progressbar,20)
        gen_datachanel(textbox, progressbar,input_params,output_params, mapping_tb_out, button)
    except:
        messagebox.showerror("Error", "Some things wrong. Please re-check")
        progress_bar(progressbar,0)
        mynotif(textbox, "Error")
        button.set_text("GENERATE")
        button.state(state='normal')
def Right2left(textbox: TkTextbox, progressbar: Tkprogressbar,params):
    ch_cnt = params['ch_cnt']   
    col_begin = params['col_begin']
    col_end = params['col_end']
    row_begin = params['row_begin']
    row_end = params['row_end']
    wsi_f = params['wsi_f']
    wso_f = params['wso_f']
    c = params['out_col']
    r = params['out_row']
    ch_begin = params['ch_begin']
    ch_end = params['ch_end']
    pwrlist = params['pwr_list']
    buschar = list(params['bus_char'])
    orient = params['orient']

    signal_bus:dict = getbus_name_size(params)
    signalname = list(signal_bus.keys())
    singlebus:list =[]
    singlebus_color: dict ={}
    multibus: dict ={}
    multibus_color: dict = {}
    for sig in signalname:
        if signal_bus[sig]==1:
            singlebus.append(sig)
            singlebus_color.__setitem__(sig, generate_color())
    for s in signalname:
        if signal_bus[s] != 1:
            multibus.__setitem__(s,signal_bus[s])
            multibus_color.__setitem__(s,generate_color())

    pwr_color: dict={}
    powerlist = list(str(params['pwr_list']).split(" "))
    for pwr in powerlist:
        pwr_color.__setitem__(pwr, generate_color())
    if orient == "NS":
        while(ch_end>=ch_begin):   
            for col in range(col_begin, col_end + 1):
                for row in range(row_begin, row_end + 1):       
                    col_l = get_column_letter(col)
                    cell_val = wsi_f[col_l + str(row)].value
    
                    if (cell_val != None):
                        wso_f[get_column_letter(c)+str(r)].alignment = Alignment(shrinkToFit=True, horizontal='center')
                        wso_f[get_column_letter(c)+str(r)].border = Border(left=Side(style='thin'),right=Side(style='thin'),top=Side(style='thin'),bottom=Side(style='thin'))
                        if cell_val in pwrlist:
                                wso_f[get_column_letter(c)+str(r)].value = cell_val
                                # wso_f[get_column_letter(c)+str(r)].fill = PatternFill(patternType='solid', fgColor=pwr_color[cell_val])
                        else:
                            index =  getstring(str(cell_val),buschar[0],buschar[1])
                            if cell_val in singlebus:
                                if(index[1]!= None):
                                    wso_f[get_column_letter(c)+str(r)].value = f"{str(cell_val).replace(index[1],'')}[{str(ch_end)}]"                                
                                else:
                                    wso_f[get_column_letter(c)+str(r)].value = str(cell_val) + buschar[0] + str(ch_end) + buschar[1]
                                # wso_f[get_column_letter(c)+str(r)].fill = PatternFill(patternType='solid', fgColor=singlebus_color[cell_val])
                            else:
                                bit_cnt = multibus[index[2]]
                                wso_f[get_column_letter(c)+str(r)].value = str(cell_val).replace(index[1],'') + buschar[0] + str(ch_end * bit_cnt + int(index[0])) + buschar[1]
                                # wso_f[get_column_letter(c)+str(r)].fill = PatternFill(patternType='solid', fgColor=multibus_color[index[2]])
                        
                    r += 1
                    print("Processing at: "+col_l + str(row) )
                    
                    mynotif(textbox,"Processing at: "+col_l + str(row))
                c += 1
                r = params['out_row']
            ch_end -= 1
        return c, singlebus, multibus
    elif orient == "EW":
        ew_ch = ch_begin
        while(ch_end>=ch_begin): 
            for row in range(row_begin, row_end + 1):
                for col in range(col_begin, col_end + 1):       
                    col_l = get_column_letter(col)
                    cell_val = wsi_f[col_l + str(row)].value
    
                    if (cell_val != None):
                        wso_f[get_column_letter(c)+str(r)].alignment = Alignment(shrinkToFit=True, horizontal='center')
                        wso_f[get_column_letter(c)+str(r)].border = Border(left=Side(style='thin'),right=Side(style='thin'),top=Side(style='thin'),bottom=Side(style='thin'))
                        if cell_val in pwrlist:
                                wso_f[get_column_letter(c)+str(r)].value = cell_val
                                # wso_f[get_column_letter(c)+str(r)].fill = PatternFill(patternType='solid', fgColor=pwr_color[cell_val])
                        else:
                            index =  getstring(str(cell_val),buschar[0],buschar[1])
                            if cell_val in singlebus:
                                if(index[1]!= None):
                                    wso_f[get_column_letter(c)+str(r)].value = f"{str(cell_val).replace(index[1],'')}[{str(ew_ch)}]"                                
                                else:
                                    wso_f[get_column_letter(c)+str(r)].value = str(cell_val) + buschar[0] + str(ew_ch) + buschar[1]
                                # wso_f[get_column_letter(c)+str(r)].fill = PatternFill(patternType='solid', fgColor=singlebus_color[cell_val])
                            else:
                                bit_cnt = multibus[index[2]]
                                wso_f[get_column_letter(c)+str(r)].value = str(cell_val).replace(index[1],'') + buschar[0] + str(ew_ch * bit_cnt + int(index[0])) + buschar[1]
                                # wso_f[get_column_letter(c)+str(r)].fill = PatternFill(patternType='solid', fgColor=multibus_color[index[2]])
                        
                    c += 1
                    print("Processing at: "+col_l + str(row) )
                    
                    mynotif(textbox,"Processing at: "+col_l + str(row))
                c = params['out_col']
                r += 1

            ch_end -= 1
            ew_ch += 1
        return r, singlebus, multibus
def Left2Right(textbox: TkTextbox, progressbar: Tkprogressbar,params: dict):
    ch_cnt = params['ch_cnt']   
    col_begin = params['col_begin']
    col_end = params['col_end']
    row_begin = params['row_begin']
    row_end = params['row_end']
    wsi_f = params['wsi_f']
    wso_f = params['wso_f']
    c = params['out_col']
    r = params['out_row']
    ch_begin = params['ch_begin']
    ch_end = params['ch_end']
    pwrlist = params['pwr_list']
    buschar = list(params['bus_char'])
    orient = params['orient']
    signal_bus:dict = getbus_name_size(params)
    signalname = list(signal_bus.keys())
    singlebus:list =[]
    singlebus_color: dict ={}
    multibus: dict ={}
    multibus_color: dict = {}
    for sig in signalname:
        if signal_bus[sig]==1:
            singlebus.append(sig)
            singlebus_color.__setitem__(sig, generate_color())
    for s in signalname:
        if signal_bus[s] != 1:
            multibus.__setitem__(s,signal_bus[s])
            multibus_color.__setitem__(s,generate_color())

    pwr_color: dict={}
    powerlist = list(str(params['pwr_list']).split(" "))
    for pwr in powerlist:
        pwr_color.__setitem__(pwr, generate_color())
    if orient == "NS":
        for cnt in range(ch_begin,ch_end + 1):             
            for col in range(col_begin, col_end + 1):
                for row in range(row_begin, row_end + 1):       
                    col_l = get_column_letter(col)
                    cell_val = wsi_f[col_l + str(row)].value    
                    if (cell_val != None):
                        wso_f[get_column_letter(c)+str(r)].alignment = Alignment(shrinkToFit=True, horizontal='center')
                        wso_f[get_column_letter(c)+str(r)].border = Border(left=Side(style='thin'),right=Side(style='thin'),top=Side(style='thin'),bottom=Side(style='thin'))
                        if cell_val in pwrlist:
                            wso_f[get_column_letter(c)+str(r)].value = cell_val
                            # wso_f[get_column_letter(c)+str(r)].fill = PatternFill(patternType='solid', fgColor=pwr_color[cell_val])
                        else:
                            index =  getstring(str(cell_val),buschar[0],buschar[1])
                            if cell_val in singlebus:
                                if(index[1]!= None):
                                    wso_f[get_column_letter(c)+str(r)].value = f"{str(cell_val).replace(index[1],'')}[{str(cnt)}]"
                                else:
                                    wso_f[get_column_letter(c)+str(r)].value = str(cell_val) + buschar[0] + str(cnt) + buschar[1]
                                # wso_f[get_column_letter(c)+str(r)].fill = PatternFill(patternType='solid', fgColor=singlebus_color[cell_val])
                            else:
                                bit_cnt = multibus[index[2]]
                                wso_f[get_column_letter(c)+str(r)].value = str(cell_val).replace(index[1],'') + buschar[0] + str(cnt * bit_cnt + int(index[0])) + buschar[1]
                                # wso_f[get_column_letter(c)+str(r)].fill = PatternFill(patternType='solid', fgColor=multibus_color[index[2]])
                        
                    r += 1
                    print("Processing at: "+col_l + str(row) )
                    mynotif(textbox,"Processing at: "+col_l + str(row))
                c += 1
                r = params['out_row']
        return c, singlebus, multibus
    elif orient == "EW":
        ew_ch = ch_end
        for cnt in range(ch_begin,ch_end + 1):          
            for row in range(row_begin, row_end + 1):
                for col in range(col_begin, col_end + 1):       
                    col_l = get_column_letter(col)
                    cell_val = wsi_f[col_l + str(row)].value
                    if (cell_val != None):
                        wso_f[get_column_letter(c)+str(r)].alignment = Alignment(shrinkToFit=True, horizontal='center')
                        wso_f[get_column_letter(c)+str(r)].border = Border(left=Side(style='thin'),right=Side(style='thin'),top=Side(style='thin'),bottom=Side(style='thin'))
                        if cell_val in pwrlist:
                            wso_f[get_column_letter(c)+str(r)].value = cell_val
                            # wso_f[get_column_letter(c)+str(r)].fill = PatternFill(patternType='solid', fgColor=pwr_color[cell_val])
                        else:
                            index =  getstring(str(cell_val),buschar[0],buschar[1])
                            if cell_val in singlebus:
                                if(index[1]!= None):
                                    wso_f[get_column_letter(c)+str(r)].value = f"{str(cell_val).replace(index[1],'')}[{str(ew_ch)}]"
                                else:
                                    wso_f[get_column_letter(c)+str(r)].value = str(cell_val) + buschar[0] + str(ew_ch) + buschar[1]
                                # wso_f[get_column_letter(c)+str(r)].fill = PatternFill(patternType='solid', fgColor=singlebus_color[cell_val])
                            else:
                                bit_cnt = multibus[index[2]]
                                wso_f[get_column_letter(c)+str(r)].value = str(cell_val).replace(index[1],'') + buschar[0] + str(ew_ch * bit_cnt + int(index[0])) + buschar[1]
                                # wso_f[get_column_letter(c)+str(r)].fill = PatternFill(patternType='solid', fgColor=multibus_color[index[2]])
                        
                    c += 1
                    print("Processing at: "+col_l + str(row) )
                    mynotif(textbox,"Processing at: "+col_l + str(row))
                c = params['out_col']
                r += 1
            ew_ch -= 1
        return r, singlebus, multibus

def getbus_name_size(params: dict):
 
    col_begin = params['col_begin']
    col_end = params['col_end']
    row_begin = params['row_begin']
    row_end = params['row_end']
    wsi_f = params['wsi_f']
    bus_char = params['bus_char']
    powerlist = list(str(params['pwr_list']).split(" "))
    
    bump_name:list =[]  
    for col in range(col_begin, col_end + 1):
        for row in range(row_begin, row_end + 1):       
            col_l = get_column_letter(col)
            cell_val = wsi_f[col_l + str(row)].value
            if (cell_val != None):    
                bump_name.append(cell_val)
    bump_name = list(dict.fromkeys(bump_name)) #remove duplicates

    return get_bus(bump_name,powerlist,bus_char)

def gen_datachanel(textbox: TkTextbox, progressbar: Tkprogressbar,input_params, output_params, mapping_tb_out, button: Tkbutton):

    excel_file = input_params["excel_file"]

    print("Opening excel file...")

    mynotif(textbox, "Opening excel file...")
    try:
        wb_f = load_workbook(excel_file,data_only=True)
    except:
        messagebox.showerror("Error", "The PLOC file is openning or not exist. Please close/check it :(")
        progress_bar(progressbar,0)
        mynotif(textbox, "Error")
        button.set_text("GENERATE")
        button.state(state='normal')
        return
    progress_bar(progressbar,50)
    print("Generating data channel..")
  
    mynotif(textbox, "Generating data channel..")
    
    wsi_name = input_params['ch_sheet']
    wso_name = output_params['ch_o_sheet']
    try:
        sheet_list = wb_f.sheetnames

        if wsi_name in sheet_list:
            wsi_f = wb_f[wsi_name]
        else:
            msg_ws = messagebox.showerror('Create Sheet', 'The sheet: ' + wsi_name + ' doesn\'t exist.')            
            mynotif(textbox, "The " + wsi_name + " doesn't exist.")        
            mynotif(textbox, "Error!!!")
            progress_bar(progressbar,0)
            return
        
        if wso_name in sheet_list:            
            wso_f = wb_f[wso_name]
        else:          
            mynotif(textbox, "The " + wso_name + " doesn't exist.")
            msg_ws = messagebox.askquestion('Create Sheet', 'The sheet: ' + wso_name + ' doesn\'t exist. Do you want to create it?', icon='question')        
            if(msg_ws == 'yes'):
                wso_f = wb_f.create_sheet(wso_name)
                mynotif(textbox, "Creating sheet...")
                mynotif(textbox,'Creating the sheet...')
            else:
                mynotif(textbox, "Aborted!!!")
                progress_bar(progressbar,0)
                return       
    except:
        print("Sheet: " + wsi_name + " doesn't exist")
        show_error("Sheet: " + wsi_name + " doesn't exist")
        progress_bar(progressbar,0)
        mynotif(textbox, "Error")
        root.update_idletasks()
        button.set_text("GENERATE")
        button.state(state='normal')
        return    
    row_begin = coordinate_to_tuple(input_params['ch_cell_start'])[0]
    col_begin = coordinate_to_tuple(input_params['ch_cell_start'])[1]
    row_end = coordinate_to_tuple(input_params['ch_cell_end'])[0]
    col_end = coordinate_to_tuple(input_params['ch_cell_end'])[1]
    if(row_begin > row_end or col_begin > col_end):
            print("Data channel params input is wrong. Please re-check")          
            mynotif(textbox, "Data channel params input is wrong. Please re-check")
            return

    out_col_begin = coordinate_to_tuple(output_params['ch_o_loc'])[1]
    out_row_begin = coordinate_to_tuple(output_params['ch_o_loc'])[0]

    c = out_col_begin
    r = out_row_begin
    ch_cnt = int(input_params["ch_cnt"]) - 1
    pwr = input_params["pwr_list"]
    buschar = input_params["bus_char"]
    ch_seq = input_params["ch_seq"]
    orient = input_params["orient"]

    params = {
        'ch_cnt': ch_cnt,
        'col_begin': col_begin,
        'col_end': col_end,
        'row_begin': row_begin,
        'row_end': row_end,
        'wsi_f': wsi_f,
        'wso_f': wso_f,
        'pwr_list': pwr,
        'bus_char': buschar,
        'out_col': c,
        'out_row': r,
        'ch_begin': 0,
        'ch_end': ch_cnt,
        
        'signal_bus': None,
        'ch_seq': ch_seq,
        'orient' : orient
    }
    ch = 0
    params['ch_begin'] = 0
    params['ch_end'] = ch_cnt
    # Left2Right(textbox, progressbar,params)
    # # Right2left(textbox, progressbar,params)
    if(ch_seq == "Right to Left" or ch_seq == "Top to Bot"):
        if((ch_cnt + 1)%2 != 0): # This feature used for odd number channels generation. But it is not permited as this time
            msg = messagebox.askquestion('Number channels choose', 'The number chanels is not even. Do you want to continue?', icon='question')
            if(msg == 'yes'):
                params['ch_begin'] = 0
                params['ch_end'] = ch_cnt
                params['signal_bus'] = Right2left(textbox, progressbar, params) 
            else:
                mynotif(textbox, "Aborted!!")
                print("Aborted!!")
                progress_bar(progressbar,0)
                return
        else:
            params['ch_begin'] = 0
            params['ch_end'] = ch_cnt
            params['signal_bus'] = Right2left(textbox, progressbar,params)
    elif(ch_seq == "Left to Right" or ch_seq == "Bot to Top"):
        if((ch_cnt + 1)%2 != 0):
            msg = messagebox.askquestion('Number channels choose', 'The number chanels is not even. Do you want to continue?', icon='question')
            if(msg == 'yes'):
                params['ch_begin'] = 0
                params['ch_end'] = ch_cnt
                params['signal_bus'] = Left2Right(textbox, progressbar,params)
            else:
                mynotif(textbox, "Aborted!!")
                print("Aborted!!")
                progress_bar(progressbar,0)
                return
        else:
            params['ch_begin'] = 0
            params['ch_end'] = ch_cnt
            params['signal_bus'] = Left2Right(textbox, progressbar,params)
    elif(ch_seq == "Center to Left first" or ch_seq == "Center to Bot first"):
        if((ch_cnt + 1)%2 != 0): # This feature used for odd number channels generation. But it is not permited as this time
            # msg = messagebox.askquestion('Number channels choose', 'The number chanel is not even. Do you want to continue?', icon='question')
            # if(msg == 'yes'):
            #     msg2 = messagebox.askquestion('Number channels', '\"Yes\" means Number of Left Channels is more than Right Channels\n \"No\" means Number of Right Channels is more than Left Channels', icon='question')
            #     if (msg2 == 'yes'):
            #        center_nu=int((ch_cnt + 1)/2)
            #        params['ch_begin'] = 0
            #        params['ch_end'] = center_nu
            #     #    if NS
            #     #    current_col = Right2left(textbox, progressbar, params)[0]
            #     #    params['out_col'] = current_col
            #        current_row = Right2left(textbox, progressbar, params)[0]
            #        params['out_row'] = current_row
            #        params['ch_begin'] = center_nu + 1
            #        params['ch_end'] = ch_cnt
            #        params['signal_bus'] = Left2Right(textbox, progressbar,params)
            #     else:
            #        center_nu=int((ch_cnt + 1)/2 -1)
            #        params['ch_begin'] = 0
            #        params['ch_end'] = center_nu
            #     #    if NS
            #     #    current_col = Right2left(textbox, progressbar,params)[0]
            #     #    params['out_col'] = current_col
            #        current_row = Right2left(textbox, progressbar, params)[0]
            #        params['out_row'] = current_row
            #        params['ch_begin'] = center_nu + 1
            #        params['ch_end'] = ch_cnt
            #        params['signal_bus'] = Left2Right(textbox, progressbar,params)    
            # else:
                mynotif(textbox, "Aborted!!")
                print("Aborted!!")
                progress_bar(progressbar,0)
                return
        else:
            center_nu=int((ch_cnt + 1)/2 - 1)
            
            if params['orient'] == "NS":
                params['ch_begin'] = 0
                params['ch_end'] = center_nu
                current_col = Right2left(textbox, progressbar,params)[0]
                params['out_col'] = current_col
                params['ch_begin'] = center_nu + 1
                params['ch_end'] = ch_cnt
                params['signal_bus'] = Left2Right(textbox, progressbar,params)
            elif params['orient'] == "EW":
                params['ch_begin'] = center_nu + 1
                params['ch_end'] = ch_cnt
                current_row = Left2Right(textbox, progressbar, params)[0]
                params['out_row'] = current_row
                params['ch_begin'] = 0
                params['ch_end'] = center_nu
                params['signal_bus'] = Right2left(textbox, progressbar,params)
    elif(ch_seq == "Center to Right first" or ch_seq == "Center to Top first"):
            center_nu=int((ch_cnt + 1)/2 - 1)

            if params['orient'] == "NS":
                params['ch_begin'] = center_nu + 1
                params['ch_end'] = ch_cnt
                current_col = Right2left(textbox, progressbar,params)[0]
                params['out_col'] = current_col
                params['ch_begin'] = 0
                params['ch_end'] = center_nu
                params['signal_bus'] = Left2Right(textbox, progressbar,params)
            elif params['orient'] == "EW":
                params['ch_begin'] = 0
                params['ch_end'] = center_nu
                current_row = Left2Right(textbox, progressbar, params)[0]
                params['out_row'] = current_row
                params['ch_begin'] = center_nu + 1
                params['ch_end'] = ch_cnt
                params['signal_bus'] = Right2left(textbox, progressbar,params)
    elif (ch_seq == "Left Edge to Center first" or ch_seq == "Bot Edge to Center first"):
            center_nu=int((ch_cnt + 1)/2 - 1)
            if params['orient'] == "NS":
                params['ch_begin'] = 0
                params['ch_end'] = center_nu
                current_col = Left2Right(textbox, progressbar,params)[0]
                params['out_col'] = current_col
                params['ch_begin'] = center_nu + 1
                params['ch_end'] = ch_cnt
                params['signal_bus'] = Right2left(textbox, progressbar,params)
            if params['orient'] == "EW":
                params['ch_begin'] = center_nu + 1
                params['ch_end'] = ch_cnt
                current_row = Right2left(textbox, progressbar,params)[0]
                params['out_row'] = current_row
                params['ch_begin'] = 0
                params['ch_end'] = center_nu
                params['signal_bus'] = Left2Right(textbox, progressbar,params)
    elif (ch_seq == "Right Edge to Center first" or ch_seq == "Top Edge to Center first"):
            center_nu=int((ch_cnt + 1)/2 - 1)
            if params['orient'] == "NS":
                params['ch_begin'] = center_nu + 1
                params['ch_end'] = ch_cnt
                current_col = Left2Right(textbox, progressbar,params)[0]
                params['out_col'] = current_col
                params['ch_begin'] = 0
                params['ch_end'] = center_nu
                params['signal_bus'] = Right2left(textbox, progressbar,params)
            if params['orient'] == "EW":
                params['ch_begin'] = 0
                params['ch_end'] = center_nu
                current_row = Right2left(textbox, progressbar,params)[0]
                params['out_row'] = current_row
                params['ch_begin'] = center_nu + 1
                params['ch_end'] = ch_cnt
                params['signal_bus'] = Left2Right(textbox, progressbar,params)
    mapgen = int(input_params['gen_map'])
    if(mapgen == 1):
        mapping_input = {
            'ch_sheet_name': input_params['ch_sheet'],
            'ch_begin_cell': input_params['ch_cell_start'],
            'ch_end_cell': input_params['ch_cell_end'],
            'ch_num': input_params['ch_cnt'],
            'pwr_list': input_params['pwr_list'],
            'bus_char': input_params['bus_char'],
            'DieL_name': input_params['DieL_name'],
            'DieR_name': input_params['DieR_name'],
            'map_char': input_params['map_char'],
            'signal_bus': params['signal_bus'],
            'ch_seq': params['ch_seq']
        }
        mapping_output = {
            'sheet_name': mapping_tb_out['sheet_name'],
            'tb_ch2ch_loc': mapping_tb_out['tb_ch2ch_loc'],            
        }
        mapg = mapping_connections(wb_f,textbox, mapping_input, mapping_output)
        wb_f = mapg[0]
        is_error = mapg[1]
        if is_error == True:
          messagebox.showinfo("Error", "An error occurred!!!")
        
    print("Saving excel file...")  
    mynotif(textbox, "Saving excel file...")
    progress_bar(progressbar,80)
    wb_f.save(excel_file)
    progress_bar(progressbar,100)  
    mynotif(textbox, "Successful!!!")
    messagebox.showinfo("Notification", "Data channel has been generated successful!!!")
    button.set_text("GENERATE")
    button.state(state='normal')

def get_saved_params(entry_list: dict[str, Tkentry], combo_list: dict[str, TkCombobox], text_list: dict[str, CanvasText], theme_list: list, checkbutton_list: dict[str, TKcheckbtn]):
    global temp_file
    try:
        with open(temp_file,'r') as params_saved:
            line1 = [line.rstrip() for line in params_saved]
            params = {
                'excel_file': line1[0],
                'sheet': line1[1],
                'ch_combo': line1[2],
                'ch_seq_combo': line1[3],
                'pwr_list': line1[4],
                'ch_cell_start': line1[5],
                'ch_cell_end': line1[6],
                'ch_o_sheet': line1[7],
                'ch_o_loc': line1[8],
                'DieL_name': line1[9],
                'DieR_name': line1[10],
                'map_sheet_name': line1[11],
                'map_tb_ch2ch_loc': line1[12],
                'theme': line1[13],
                'gen_map': line1[14],
                'bus_char': line1[15],
                'map_char': line1[16]
            }

        set_theme(combo=combo_list['theme'], theme_name=params['theme'], themelist=theme_list, entry_list=entry_list, text_list=text_list)
        entry_list['pxcel'].add_new_content(params['excel_file'])
        entry_list['v_refsheet'].add_new_content(params['sheet'])
        
        # combo_list['ch'].set_current(params['ch_combo'])
        # combo_list['ch_seq'].set_current(params['ch_seq_combo'])
        entry_list['pwr_list'].add_new_content(params['pwr_list'])
        entry_list['ref_cell_start'].add_new_content(params['ch_cell_start'])
        entry_list['ref_cell_end'].add_new_content(params['ch_cell_end'])
        entry_list['ch_o_sheet'].add_new_content(params['ch_o_sheet'])
        entry_list['ch_o_loc'].add_new_content(params['ch_o_loc'])
        entry_list['die_L_list'].add_new_content(params['DieL_name'])
        entry_list['die_R_list'].add_new_content(params['DieR_name'])
        entry_list['map_tb_sheet'].add_new_content(params['map_sheet_name'])
        entry_list['map_tb_loc'].add_new_content(params['map_tb_ch2ch_loc'])
        entry_list['bus_char'].add_new_content(params['bus_char'])
        entry_list['map_char'].add_new_content(params['map_char'])
        # checkbutton_list['gen_map'].checkbtn.config(state=)

    except:
        set_theme(combo=combo_list['theme'], theme_name='pulse', themelist=theme_list, entry_list=entry_list, text_list=text_list)
        entry_list['pxcel'].add_new_content(r"C:\Users\sytung\OneDrive - Synopsys, Inc\Desktop\py\Test3.xlsx")
        entry_list['v_refsheet'].add_new_content('DWORD')
        # combo_list['ch'].set_current(params['ch_combo'])
        # combo_list['ch_seq'].set_current(params['ch_seq_combo'])
        entry_list['pwr_list'].add_new_content('VDD VCCIO VCCAON VSS')
        entry_list['ref_cell_start'].add_new_content('S16')
        entry_list['ref_cell_end'].add_new_content('AD30')
        entry_list['ch_o_sheet'].add_new_content('Data_Channelx')
        entry_list['ch_o_loc'].add_new_content('D10')
        entry_list['die_L_list'].add_new_content("DIE1 DIE2 DIE3 DIE4")
        entry_list['die_R_list'].add_new_content("DIE5 DIE6 DIE7 DIE8")
        entry_list['map_tb_sheet'].add_new_content('MAPPINGD2D')
        entry_list['map_tb_loc'].add_new_content('D10')
        entry_list['bus_char'].add_new_content('<>')
        entry_list['map_char'].add_new_content('TX RX')
        # checkbutton_list['gen_map'].checkbtn.config(state=)

get_saved_params(entry_ls, combo_ls,text_ls,theme_ls,chkbtn_ls)
gen_btn.button.config(command= lambda: process(textbox=text_box, entry_list=entry_ls,checkbtn_list=chkbtn_ls, combo_list=combo_ls,progressbar=pg_bar,text_list=text_ls, theme_list=theme_ls,button=gen_btn))

# for ent in entry_list:
#     ent.configure(justify='center')
# entry_disable(die_R_list, die_L_list, map_tb_sheet, map_loc_i)
root.mainloop()


    


