from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.utils import range_boundaries
from openpyxl.utils import column_index_from_string
#from openpyxl.utils import coordinate_from_string
from openpyxl.utils import coordinate_to_tuple
from openpyxl.worksheet.table import Table, TableStyleInfo
# from tkinter import *
from tkinter import messagebox
from tkinter import ttk
import getcolumn
from array import *
import tkinter as tk
from tkinter import *
from ttkthemes import ThemedTk, THEMES
from PIL import Image
from PIL import ImageTk, Image
from tkinter.font import Font as tkfont
from tkinter import filedialog
import gui_function as gui
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection
import os
import win32com.client
from pathlib import Path  # core library
from time import sleep


# excel_file = r"C:\Users\sytung\OneDrive - Synopsys, Inc\Desktop\py\Test_adp.xlsx"

root = ThemedTk()
# my_canvas=tk.Canvas(root)
root.set_theme("scidpurple")

root.title("ADP Netlist Generator")
root.geometry("1000x1000+30+100")
root.resizable(width=False, height=False)
root.iconbitmap(r".\mylogo.ico")
root.option_add("*tearOff", False) # This is always a good idea

bg = PhotoImage(file=r".\brain.png").subsample(2,2)
# bg = ImageTk.PhotoImage(file=r".\img\internet.png")
open_imag = PhotoImage(file = r".\open-folder.png")

# Define Canvas
my_canvas = tk.Canvas(root, width=1200, height=800, bd=0, highlightthickness=0)
my_canvas.pack(fill="both", expand=True)

# Put the image on the canvas
my_canvas.create_image(0,0, image=bg, anchor="nw")
# Make the app responsive
# root.columnconfigure(index=0, weight=1)
# root.columnconfigure(index=1, weight=1)
# root.columnconfigure(index=2, weight=1)
# root.columnconfigure(index=3, weight=1)

# root.grid_columnconfigure(0, weight=1)
# root.grid_rowconfigure(0, weight=1)
# root.rowconfigure(index=0, weight=1)
# root.rowconfigure(index=1, weight=1)
# root.rowconfigure(index=2, weight=1)
# root.rowconfigure(index=3, weight=1)
# root.rowconfigure(index=4, weight=1)
# root.rowconfigure(index=5, weight=1)
# root.rowconfigure(index=6, weight=1)
# root.rowconfigure(index=7, weight=1)




stfont= ("Franklin Gothic Medium", 10, 'underline', "italic")
# Create lists for the Comboboxes
theme_list = ["adapta", "aquativo", "arc", "black","blue", "breeze", "clearlooks", "elegance", "equilux", "itft1", "keramik", "keramik_alt", "kroc", "plastik", "radiance", "ubuntu", "scidblue", "scidgreen", "scidgrey", "scidmint", "scidpink", "scidpurple", "scidsand", "smog", "winxpblue", "yaru" ]
colour_list = ["#09a5e8", "#292b33", "#1583eb", "#292a2b","#1a7cad", "#0664bd", "#8baac7", "#063f75", "#40454a", "#7aa7f5", "#1c4894", "#1c4894", "#ebab0c", "#0c99eb", "#eb830c", "#eb830c", "#0937ab", "#37ed80", "#707371", "#479403", "#d12a9f", "#9b34eb", "#787122", "#118cbd", "#a3945f", "#621ba8" ]


# Create control variables
a = tk.BooleanVar()
b = tk.BooleanVar(value=True)
c = tk.BooleanVar()
d = tk.IntVar(value=2)
# e = tk.StringVar(value=option_menu_list[1])
f = tk.BooleanVar()
g = tk.DoubleVar(value=75.0)
h = tk.BooleanVar()
tc_opt = tk.IntVar()
isIntp = tk.IntVar()

#Define a Function to enable the frame
def round_rectangle(x1, y1, x2, y2, radius=25, **kwargs):
        
    points = [x1+radius, y1,
              x1+radius, y1,
              x2-radius, y1,
              x2-radius, y1,
              x2, y1,
              x2, y1+radius,
              x2, y1+radius,
              x2, y2-radius,
              x2, y2-radius,
              x2, y2,
              x2-radius, y2,
              x2-radius, y2,
              x1+radius, y2,
              x1+radius, y2,
              x1, y2,
              x1, y2-radius,
              x1, y2-radius,
              x1, y1+radius,
              x1, y1+radius,
              x1, y1]

    return my_canvas.create_polygon(points, **kwargs, smooth=True)
def change_colour(index):
    listchange =[ball_in_t,mapingtb_t,adptb_t,die_tb_t,die_l_t,die_r_t,die_name_t,die_begin_t,die_end_t, adp_tb_out_t]
    entry_list = [button,theme_combo_t,excel_t]
    #  sheet_t, theme_combo_t,excel_t, pkg_t
    for t in listchange:
        my_canvas.itemconfig(t, fill = colour_list[index])
    for l in entry_list:
        l.config(background = colour_list[index])
def enable(children):
   for child in children:
      child.configure(state='enable')
def disable(children):  
    for child in children:
        child.configure(state='disable')
def entry_disable(*entries):
    for entry in entries:
        entry.config(state='disable')
def entry_enable(*entries):
    for entry in entries:
        entry.config(state='normal')

def intp_toggle():
    
        # if(entry['state'] == 'disable'):
    if(isIntp.get() == 1):
        entry_enable(DieL_begincell_i, DieL_endcell_i, DieR_begincell_i, DieR_endcell_i, dietb_sheet, DieL_name, DieR_name)
        print("Gen interposer Die table: ON")
    elif(isIntp.get() == 0):
        entry_disable(DieL_begincell_i, DieL_endcell_i, DieR_begincell_i, DieR_endcell_i, dietb_sheet, DieL_name, DieR_name)
        print("Gen interposer Die table: OFF")
def progress_bar(value):
    progress['value'] = value
    root.update_idletasks()

def choosetheme(event):
    # for theme in theme_list:
        # if (theme_combo.get() == theme):
    root.set_theme(theme_combo.get())
    change_colour(theme_list.index(theme_combo.get()))
           

myLabel = ttk.Label(root,text="----")
myLabel_w =my_canvas.create_window(80,770,anchor="nw", window=myLabel)

frame = tk.Label(root, bg="#c9f2dc", font=("Courier New", 10), foreground="#f2a50a")
my_canvas.create_window(300, 80, window=frame, anchor="nw", width= 400, height=100)

def popup(notif):
    messagebox.showinfo("Notification", notif)
def progress_bar(value):
    progress['value'] = value
    root.update_idletasks()
def get_num_intdie(event):
    pass
def mynotif(content):
    if(content == ""):
        myLabel.configure(text="", anchor='w')
    else:
        myLabel.configure(text=content, anchor='w')
        # myLabel = ttk.Label(root,text=content)
        # myLabel_w =my_canvas.create_window(80,750,anchor="nw", window=myLabel)
        # myLabel.grid(row=5, column=0, columnspan=2, padx=(20, 10), pady=(20, 10), sticky="nsew")
def process_notify(content):    
        mynotif("")
        root.update_idletasks()
        mynotif(content)
        root.update_idletasks()
# ------------------------------------------------------------------------------------------------------------------------------------------------

def myguide(entries, content):
    if(content == ""):
        entries.configure(text="")
       
    else:
        entries.configure(text=content)

def handle_click(event):
   pass
    
def ball_sheet_guide(event):
     myguide(frame, "INFO:" + "Ball sheet name    \n\n - Example:   BGA           ")
def un_guide(event):
     myguide(frame,"")

def ball_begin_cell_guide(event):
     myguide(frame, "INFO:" + "Ball table begin cell\n\n - Example:   CU100       ")
def mapping_sheet_guide(event):
     myguide(frame, "INFO:" + "Mapping sheet name          \n\n - Example:   DIE_Mapping           ")
def mapping_begin_cell_guide(event):
     myguide(frame, "INFO:" + "Mapping table begin cell\n\n - Example:   CU100       ")
def ball_end_cell_guide(event):
    myguide(frame, "INFO:" + "Ball table end cell\n\n - Example:   CU100       ")
def adp_tb_guide(event):
    myguide(frame, "INFO:" + "This field to define the\n   ADP Netlist table name    ")
def adp_tb_loc_guide(event):
    myguide(frame, "INFO:" + "This field to define the \n  ADP table location.\n  Ex: A10 ")
def mapping_end_cell_in_guide(event):
    myguide(frame, "INFO:" + "Mapping table begin cell\n\n - Example:   CU100     ")

def DieL_begincell_i_guide(event):
     myguide(frame, "INFO:" + "List of Left Die table begin cell\n\n - Example:   A10  H10 G10       ")  
def DieL_endcell_i_guide(event):
     myguide(frame, "INFO:" + "List of Left Die table end cell\n\n - Example:   C998 K999 N100       ")
def DieR_begincell_i_guide(event):
     myguide(frame, "INFO:" + "List of Right Die table begin cell\n\n - Example:   A10  H10 G10       ")
def DieR_endcell_i_guide(event):
     myguide(frame, "INFO:" + "List of Right Die table end cell\n\n - Example:   A10  H10 G10     ")
def dietb_sheet_guide(event):
     myguide(frame, "INFO:" + "Die tables sheet name         \n\n - Example:    Die_table           ")
def DieL_name_guide(event):
     myguide(frame, "INFO:" + "Name list of left Die.\n (Die Flipped + Rotate -90)\n- Die name MUST NOT contain spaces character\n- The dies name are separated by spaces     ")
def DieR_name_guide(event):
     myguide(frame, "INFO:" + "Name list of right Die.\n (Die Flipped + Rotate +90)\n- Die name MUST NOT contain spaces character\n- The dies name are separated by spaces     ")
def int_tb_guide(event):
    myguide(frame, "INFO:" + "This field to define the\n  first output table cell. \n The next tables placed away \n1 column from previous table \n\n - Example: O64 ")
def srw_i_guide(event):
    myguide(frame, "INFO:" + "This field to define the\n  width of sealring . \n - TSMC: 21.6, SS/GF: 14.04 \n\n")
        
xfont = ("System", 12, "bold", 'underline', 'italic')
theme_combo_t = ttk.Label(root,text="Choose theme:",border=20, font=xfont, background='#b434eb', borderwidth=3)
theme_combo_t_w = my_canvas.create_window(750, 15, window=theme_combo_t)

theme_combo = ttk.Combobox(root, state="readonly", values=theme_list, width=15)
theme_combo_w = my_canvas.create_window(870,15, window=theme_combo)

theme_combo.bind('<<ComboboxSelected>>', choosetheme)

# -------------------------excelpath input--------------------------#
pfont= ("Rosewood Std Regular", 12, "bold", 'underline' )
excel_t = ttk.Label(root,text="PLOC path:",border=20,font=pfont, borderwidth=5)
excel_t_w = my_canvas.create_window(30,40, anchor="nw", window=excel_t)
excel_i = ttk.Entry(root, width=115)

excel_i_w = my_canvas.create_window(150,40, anchor="nw", window=excel_i)

# Separator
separator = ttk.Separator(root)
separator_w = my_canvas.create_window(30, 130, anchor="nw", window=separator)


ball_in_t = my_canvas.create_text(30, 200, text="Ball Table inputs", anchor="nw",font=("Helvetica", 10, 'italic', 'underline', 'bold'), fill="#b434eb")


# ------------------------Die bump visual parameters input --------------------------#
ball_sheet_i = ttk.Entry(root, width=20)
my_canvas.create_window(150, 200, anchor="nw", window=ball_sheet_i, width=225)

ball_sheet_i.bind('<FocusIn>', ball_sheet_guide)
ball_sheet_i.bind('<FocusOut>', un_guide)


ball_begin_cell_i = ttk.Entry(root, width=20)
my_canvas.create_window(400, 200, anchor="nw", window=ball_begin_cell_i, width=225)

ball_begin_cell_i.bind('<FocusIn>', ball_begin_cell_guide)
ball_begin_cell_i.bind('<FocusOut>', un_guide)

ball_end_cell = ttk.Entry(root)
ball_end_cell_w = my_canvas.create_window(650, 200, anchor="nw", window=ball_end_cell, width=225)

ball_end_cell.bind('<FocusIn>', ball_end_cell_guide)
ball_end_cell.bind('<FocusOut>', un_guide)

mapingtb_t = my_canvas.create_text(30, 260, text="Mapping table\n inputs", anchor="nw",font=("Helvetica", 10, 'italic', 'underline', 'bold'), fill="#b434eb")
mapping_sheet_i = ttk.Entry(root, width=20)
mapping_sheet_i_w = my_canvas.create_window(150, 260, anchor="nw", window=mapping_sheet_i, width=225)

mapping_sheet_i.bind('<FocusIn>', mapping_sheet_guide)
mapping_sheet_i.bind('<FocusOut>', un_guide)

mapping_begin_cell_i = ttk.Entry(root, width=20)
mapping_begin_cell_i_w = my_canvas.create_window(400, 260, anchor="nw", window=mapping_begin_cell_i, width=225)

mapping_begin_cell_i.bind('<FocusIn>', mapping_begin_cell_guide)
mapping_begin_cell_i.bind('<FocusOut>', un_guide)

mapping_end_cell_i = ttk.Entry(root, width=20)
mapping_end_cell_i_w = my_canvas.create_window(650, 260, anchor="nw", window=mapping_end_cell_i, width=225)

mapping_end_cell_i.bind('<FocusIn>', mapping_end_cell_in_guide)
mapping_end_cell_i.bind('<FocusOut>', un_guide)

adptb_t = my_canvas.create_text(30, 310, text="Die to gen ADP inputs", anchor="nw",font=("Helvetica", 10, 'italic', 'underline', 'bold'), fill="#b434eb")


# ---------------------------------------DIE INPUTS-------------------------------------------------

die_tb_t = my_canvas.create_text(60, 340, text="DIE table\nsheet:", anchor="nw",font=("Helvetica", 10, 'underline', 'bold'), fill="#003feb")

dietb_sheet = ttk.Entry(root)
dietb_sheet_w = my_canvas.create_window(150, 340, anchor="nw", window=dietb_sheet, width=360)

dietb_sheet.bind('<FocusIn>', dietb_sheet_guide)
dietb_sheet.bind('<FocusOut>', un_guide)



# my_canvas.create_text(30, 595, text="Die/Chip Offset:", anchor="nw",font=("Helvetica", 10, 'italic', 'underline', 'bold'), fill="#b434eb")
die_l_t = my_canvas.create_text(300, 380, text="Left Die:", anchor="nw",font=("Helvetica", 10, 'italic', 'underline', 'bold'), fill="#003feb")
die_r_t = my_canvas.create_text(680, 380, text="Right Die:", anchor="nw",font=("Helvetica", 10, 'italic', 'underline', 'bold'), fill="#003feb")


die_name_t = my_canvas.create_text(60, 410, text="Die name:", anchor="nw",font=("Helvetica", 10, 'underline', 'bold'), fill="#003feb")
DieL_name = ttk.Entry(root)
DieL_name_w = my_canvas.create_window(150, 410, anchor="nw", window=DieL_name, width= 360)

DieL_name.bind('<FocusIn>', DieL_name_guide)
DieL_name.bind('<FocusOut>', un_guide)



DieR_name = ttk.Entry(root)
DieR_name_w = my_canvas.create_window(520, 410, anchor="nw", window=DieR_name, width=360)

DieR_name.bind('<FocusIn>', DieR_name_guide)
DieR_name.bind('<FocusOut>', un_guide)



die_begin_t = my_canvas.create_text(60, 450, text="Begin cell:", anchor="nw",font=("Helvetica", 10, 'underline', 'bold'), fill="#003feb")
DieL_begincell_i = ttk.Entry(root)
DieL_begincell_w = my_canvas.create_window(150, 450, anchor="nw", window=DieL_begincell_i, width=360)

DieL_begincell_i.bind('<FocusIn>', DieL_begincell_i_guide)
DieL_begincell_i.bind('<FocusOut>', un_guide)

die_end_t = my_canvas.create_text(60, 490, text="End cell:", anchor="nw",font=("Helvetica", 10, 'underline', 'bold'), fill="#003feb")
DieL_endcell_i = ttk.Entry(root, width=20)
DieL_endcell_w = my_canvas.create_window(150, 490, anchor="nw", window=DieL_endcell_i, width=360)

DieL_endcell_i.bind('<FocusIn>', DieL_endcell_i_guide)
DieL_endcell_i.bind('<FocusOut>', un_guide)

DieR_begincell_i = ttk.Entry(root)
DieR_begincell_w = my_canvas.create_window(520, 450, anchor="nw", window=DieR_begincell_i, width=360)

DieR_begincell_i.bind('<FocusIn>', DieR_begincell_i_guide)
DieR_begincell_i.bind('<FocusOut>', un_guide)

DieR_endcell_i = ttk.Entry(root, width=20)
DieR_endcell_w = my_canvas.create_window(520, 490, anchor="nw", window=DieR_endcell_i, width=360)

DieR_endcell_i.bind('<FocusIn>', DieR_endcell_i_guide)
DieR_endcell_i.bind('<FocusOut>', un_guide)

adp_tb_out_t = my_canvas.create_text(30, 550, text="ADP Table config\nsheet/location:", anchor="nw",font=("Helvetica", 10, 'italic', 'underline', 'bold'), fill="#b434eb")

adp_tb = ttk.Entry(root, width=20)
adp_tb_w = my_canvas.create_window(150, 555, anchor="nw", window=adp_tb, width=360)

adp_tb.bind('<FocusIn>', adp_tb_guide)
adp_tb.bind('<FocusOut>', un_guide)

adp_tb_loc = ttk.Entry(root, width=20)
adp_tb_loc_w = my_canvas.create_window(520, 555, anchor="nw", window=adp_tb_loc, width=360)

adp_tb_loc.bind('<FocusIn>', adp_tb_loc_guide)
adp_tb_loc.bind('<FocusOut>', un_guide)



separator1 = ttk.Separator(root)

separator2 = ttk.Separator(root)


# ------------------------------
separator1 = ttk.Separator(root)

separator2 = ttk.Separator(root)


#--------------------------------------------------------------------------------------------------------#

my_canvas.create_text(880,980, text= "Internal contact: sytung@synopsys.com" ,font=("Helvetica", 8, 'underline'), fill="grey")

def browse_file():
	# global my_image
    root.filename = filedialog.askopenfilename(initialdir="./", title="Select A File", filetypes=(("Excel files", "*.xlsx"),("all files", "*.*")))
    excel_i.delete(0,END)
    print(root.filename) 
    excel_i.insert(0, root.filename)
	# my_image = ImageTk.PhotoImage(Image.open(root.filename))
	# my_image_label = Label(image=my_image).pack()

#---------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------#
def popup(notif):
    messagebox.showinfo("Notification", notif)

def show_error(error):
    messagebox.showerror("Error", error)




def get_col_row_range(cell_begin, cell_end):
    row_begin = coordinate_to_tuple(cell_begin)[0]
    col_begin = coordinate_to_tuple(cell_begin)[1]
    row_end = coordinate_to_tuple(cell_end)[0]
    col_end = coordinate_to_tuple(cell_end)[1]
    return row_begin,row_end,col_begin,col_end
def get_config():

    # die_params = {
    #     "Die_sheet":"Package_substrate",
    #     "Die_L_begin_cell":"T19",
    #     "Die_L_end_cell": "V791",
    #     "Die_L_name": "DIE3",
    #     "Die_R_begin_cell":"X19",
    #     "Die_R_end_cell": "Z791",
    #     "Die_R_name": "DIE7",
    # }
    die_params= {
        "die_sheet" : str(dietb_sheet.get()),
        "diel_list" : list((DieL_name.get()).split()),
        "diel_begin_list" : list((DieL_begincell_i.get()).split()),
        "diel_end_list" : list((DieL_endcell_i.get()).split()),
        "dier_list" : list((DieR_name.get()).split()),
        "dier_begin_list" : list((DieR_begincell_i.get()).split()),
        "dier_end_list" : list((DieR_endcell_i.get()).split()),
        "Pad_stack": "60x60"
    }
 
    input_params = {
        "excel_file": str(excel_i.get()),

        "mapping_sheet": str(mapping_sheet_i.get()),
        "mapping_begin_cell": str(mapping_begin_cell_i.get()),
        "mapping_end_cell":str(mapping_end_cell_i.get()),
        
        "ball_tb_sheet": str(ball_sheet_i.get()),
        "ball_tb_begin_cell": str(ball_begin_cell_i.get()),
        "ball_tb_end_cell": str(ball_end_cell.get()),
  
    }

    # adp table out put config 
    out_put = {
        "sheet": str(adp_tb.get()),
        "tb_loc": str(adp_tb_loc.get())
       
    }
    with open(".adp_params_saved.txt",'w') as params_saved:
            params_saved.writelines(theme_combo.get() +"\n")
            params_saved.writelines(input_params['excel_file'] +"\n")
            params_saved.writelines(input_params['ball_tb_sheet'] +"\n")
            params_saved.writelines(input_params['ball_tb_begin_cell'] +"\n")
            params_saved.writelines(input_params['ball_tb_end_cell'] +"\n")
            params_saved.writelines(input_params['mapping_sheet'] +"\n")
            params_saved.writelines(input_params['mapping_begin_cell'] +"\n")
            params_saved.writelines(input_params['mapping_end_cell'] +"\n")
            params_saved.writelines(die_params['die_sheet'] +"\n")
            
            
            params_saved.writelines(' '.join(die_params['diel_list']) +"\n")
            params_saved.writelines(' '.join(die_params['dier_list']) +"\n")
            params_saved.writelines(' '.join(die_params['diel_begin_list']) +"\n")
            params_saved.writelines(' '.join(die_params['diel_end_list']) +"\n")
            params_saved.writelines(' '.join(die_params['dier_begin_list']) +"\n")
            params_saved.writelines(' '.join(die_params['dier_end_list']) +"\n")
            params_saved.writelines(out_put['sheet'] +"\n")
            params_saved.writelines(out_put['tb_loc'] +"\n")
          


    return die_params,input_params,out_put

def refresh_excel(excelfile):
    excel_file = os.path.join(excelfile)
    excel = win32com.client.gencache.EnsureDispatch('Excel.Application')
    excel.DisplayAlerts = False # disabling prompts to overwrite existing file
    excel.Workbooks.Open(excel_file )
    excel.ActiveWorkbook.Save()
    excel.DisplayAlerts = True # enabling prompts
    excel.ActiveWorkbook.Close()
    print('Waiting save done')
    sleep(20)
def copy_table(cell):
    try:
        wb_tempsheet = load_workbook(cell['excel_file'])
    except:
        print("Ploc file does not exist or opening. Please close/recheck it")
        return

    row_begin = get_col_row_range(cell['ball_tb_begin_cell'],cell['ball_tb_end_cell'])[0]
    row_end =  get_col_row_range(cell['ball_tb_begin_cell'],cell['ball_tb_end_cell'])[1]
    col_begin = get_col_row_range(cell['ball_tb_begin_cell'],cell['ball_tb_end_cell'])[2]
    col_end =  get_col_row_range(cell['ball_tb_begin_cell'],cell['ball_tb_end_cell'])[3]
    sheet_ls = wb_tempsheet.sheetnames
    
    wstmp_create_name = str(cell['ball_tb_sheet'])

    if wstmp_create_name in sheet_ls:
        msg_ws = messagebox.askquestion('Create Sheet', 'The ' + wstmp_create_name + ' already exist. Do you want to overwrite it?',icon='question')
        # sh_index = sheet_ls.index(wstmp_create_name) 
        tmp_s = wb_tempsheet.get_sheet_by_name(wstmp_create_name)
        print("The " + wstmp_create_name + " exist.")
        if(msg_ws == 'yes'):
            wb_tempsheet.remove_sheet(tmp_s)
                       
        else:
            print("Stop the process")
            return
    wstmp_create = wb_tempsheet.create_sheet(wstmp_create_name) 
    cell_list = wb_tempsheet.sheetnames
    if cell['ball_tb_sheet'] in cell_list:
        source_sheet = wb_tempsheet[cell['ball_tb_sheet']]
        print(row_begin)
        print(col_begin)
        print(row_end)
        print(col_end)
        for i in range (row_begin, row_end + 1):
            for j in range (col_begin, col_end + 1):
                # if(str(source_sheet[get_column_letter(j) + str(i)].value).find("=") != -1):
                    # wstmp_create[get_column_letter(j) + str(i)].value = f"={cell['ball_tb_sheet']}!{str(source_sheet[get_column_letter(j) + str(i)].value).replace('=','')}"
                # else:
                wstmp_create[get_column_letter(j) + str(i)].value = f"={cell['ball_tb_sheet']}!{str(get_column_letter(j) + str(i))}"
        wb_tempsheet.save(cell['excel_file'])
        wb_tempsheet.close()
        print('Waiting save done')
        sleep(10)
        
    return row_begin,col_begin,row_end,col_end,wstmp_create_name 
 

def gen(adp, die, mapping, mapping_prefix, ball, last_ball):
        for title in range(adp['begincol'], adp['begincol'] + 10):
            adp['sheet'][get_column_letter(title) + str(adp['beginrow'])].border = Border(left=Side(style='thin'),right=Side(style='thin'),top=Side(style='thin'),bottom=Side(style='thin'))
            adp['sheet'][get_column_letter(title) + str(adp['beginrow'])].fill = PatternFill(patternType='solid', fgColor='9e42f5')
        adp['sheet'][adp['Pin_number'] + str(adp['beginrow'])].value = "Pin_number"
        adp['sheet'][adp['Pad_stack'] + str(adp['beginrow'])].value = "Pad_stack"
        adp['sheet'][adp['X'] + str(adp['beginrow'])].value = "X_coord"
        adp['sheet'][adp['Y'] + str(adp['beginrow'])].value = "Y_coord"
        adp['sheet'][adp['rotation'] + str(adp['beginrow'])].value = "Rotation"
        adp['sheet'][adp['Pad_use'] + str(adp['beginrow'])].value = "Pad_use"
        adp['sheet'][adp['pkg_name'] + str(adp['beginrow'])].value = "Net_name(Package)"
        adp['sheet'][adp['die_name'] + str(adp['beginrow'])].value = "Net_name(Die)"
        adp['sheet'][adp['resdef'] + str(adp['beginrow'])].value = "ResDef"
        adp['sheet'][adp['bga_pin'] + str(adp['beginrow'])].value = "Package_pin"
        adp['sheet'].freeze_panes = 'A'+ str(adp['beginrow'] + 1)
        
        
        for i in range(die['row_min'] + 2, die['row_max'] + 1):
            adp['sheet'][adp['Pin_number'] + str(adp['r'])].value = adp['pin_num'] 
            adp['sheet'][adp['Pad_stack'] + str(adp['r'])].value = die['name'] + "_" + die['Pad_stack'] 
            adp['sheet'][adp['rotation'] + str(adp['r'])].value = 0
            print("Processing for: " + die['sheet_d'][die['net_name'] + str(i)].value)
            #------------------------------Die2Die Connections--------------------------------------------
            if(str(die['sheet_d'][die['net_name'] + str(i)].value).find("BP_") != -1 and str(die['sheet_d'][die['net_name'] + str(i)].value).find("ATO") == -1 and str(die['sheet_d'][die['net_name'] + str(i)].value).find("DTO") == -1 and str(die['sheet_d'][die['net_name'] + str(i)].value).find("ZN") == -1):
                adp['sheet'][adp['resdef'] + str(adp['r'])].value = "BUMP"
                if(str(die['sheet_d'][die['net_name'] + str(i)].value).find("BP_TX") != -1):
                    adp['sheet'][adp['Pad_use'] + str(adp['r'])].value = "D2D"
                elif(str(die['sheet_d'][die['net_name'] + str(i)].value).find("BP_RX") != -1):
                    adp['sheet'][adp['Pad_use'] + str(adp['r'])].value = "D2D"
                
                # if(str(die['sheet_d'][die['net_name'] + str(i)].value).find(die['name']) != -1):
                if(str(die['die_side']) == "L"):
                    for k in range(mapping['row_min'], mapping['row_max']+1):
                        if (str(die['sheet_d'][die['net_name'] + str(i)].value).find(str(mapping['sheet_d'][mapping['die_L'] + str(k)].value)) != -1):
                            adp['sheet'][adp['pkg_name'] + str(adp['r'])].value = mapping_prefix + str(mapping['sheet_d'][mapping['die_L'] + str(k)].value).replace("BP_","").replace("[","_").replace("]","")
                            break
                        # else:
                          
                        #     adp['sheet'][adp['pkg_name'] + str(adp['r'])].value = "NA"
                        #     print("The net name is not found. Please recheck mapping table")
                    # elif (str(ws_d[die['net_name'] + str(i)].value).find("DIE7") != -1):
                elif(str(die['die_side']) == "R"):
                    for n in range(mapping['row_min'], mapping['row_max']+1):
                        if (str(die['sheet_d'][die['net_name'] + str(i)].value).find(str(mapping['sheet_d'][mapping['die_R'] + str(n)].value)) != -1):
                            adp['sheet'][adp['pkg_name'] + str(adp['r'])].value = mapping_prefix + "_" + str(mapping['sheet_d'][mapping['die_L'] + str(n)].value).replace("BP_","").replace("[","_").replace("]","")
                            break
                        # else:                          
                        #     adp['sheet'][adp['pkg_name'] + str(adp['r'])].value = "NA"
                        #     print("The net name is not found. Please recheck mapping table")
                            
                adp['sheet'][adp['X'] + str(adp['r'])].value = f"='{die['sheet_name']}'!{die['X'] + str(i)}"
                adp['sheet'][adp['Y'] + str(adp['r'])].value = f"='{die['sheet_name']}'!{die['Y'] + str(i)}"   
                adp['sheet'][adp['die_name'] + str(adp['r'])].value = f"='{die['sheet_name']}'!{die['net_name'] + str(i)}"
                adp['sheet'][adp['bga_pin'] + str(adp['r'])].value = "-"
                adp['r'] += 1
            #------------------------------Common RDI input Connections--------------------------------------------
            elif (str(die['sheet_d'][die['net_name'] + str(i)].value).find("RDI_LP_CFG") != -1 or str(die['sheet_d'][die['net_name'] + str(i)].value).find("RDI_CFG_CLK") != -1 or str(die['sheet_d'][die['net_name'] + str(i)].value).find("RDI_MODE") != -1):
                adp['sheet'][adp['Pad_use'] + str(adp['r'])].value = "I"
                # print(die['sheet_d'][die['net_name'] + str(i)].value)
            
                if(die['die_side'] == "L"):
                    diecell_val = str(die['sheet_d'][die['net_name'] + str(i)].value).replace(die['name'],"L")
               
                    for j in ball['list_row']:
                            if (diecell_val == ball['sheet_d'][ball['net_col']+ str(j)].value):

                                adp['sheet'][adp['resdef'] + str(adp['r'])].value = "BGA"
                                adp['sheet'][adp['X'] + str(adp['r'])].value = f"='{die['sheet_name']}'!{die['X'] + str(i)}"
                                adp['sheet'][adp['Y'] + str(adp['r'])].value = f"='{die['sheet_name']}'!{die['Y'] + str(i)}"   
                                adp['sheet'][adp['pkg_name'] + str(adp['r'])].value = f"='{ball['sheet_name']}'!{ball['net_col']+ str(j)}"
                                adp['sheet'][adp['die_name'] + str(adp['r'])].value = f"='{die['sheet_name']}'!{die['net_name'] + str(i)}"
                                # print((ball['sheet_d'][ball['Y'] + str(j)].value))
                                adp['sheet'][adp['bga_pin'] + str(adp['r'])].value = f"='{str(ball['sheet_name'])}'!{ball['Y'] + str(j)}&'{str(ball['sheet_name'])}'!{ball['X'] + str(j)}"

                                adp['r'] += 1
                             
                                break
                            elif (ball['list_row'].index(j) == len(ball['list_row']) - 1):
                                print(ball['list_row'].index(j))
                                # print(ball['sheet_d'][ball['net_col']+ str(j)].value)
                                adp['sheet'][adp['X'] + str(adp['r'])].value = f"='{die['sheet_name']}'!{die['X'] + str(i)}"
                                adp['sheet'][adp['Y'] + str(adp['r'])].value = f"='{die['sheet_name']}'!{die['Y'] + str(i)}"   
                                adp['sheet'][adp['pkg_name'] + str(adp['r'])].value = "N/A"
                                adp['sheet'][adp['die_name'] + str(adp['r'])].value = f"='{die['sheet_name']}'!{die['net_name'] + str(i)}"
                                adp['sheet'][adp['bga_pin'] + str(adp['r'])].value = "N/A"
                                adp['r'] += 1

                elif (die['die_side'] == "R"):
                    diecell_val = str(die['sheet_d'][die['net_name'] + str(i)].value).replace(die['name'],"R")
               
                    for j in ball['list_row']:
                                if (diecell_val == ball['sheet_d'][ball['net_col']+ str(j)].value):

                                    adp['sheet'][adp['resdef'] + str(adp['r'])].value = "BGA"
                                    adp['sheet'][adp['X'] + str(adp['r'])].value = f"='{die['sheet_name']}'!{die['X'] + str(i)}"
                                    adp['sheet'][adp['Y'] + str(adp['r'])].value = f"='{die['sheet_name']}'!{die['Y'] + str(i)}"   
                                    adp['sheet'][adp['pkg_name'] + str(adp['r'])].value = f"='{ball['sheet_name']}'!{ball['net_col']+ str(j)}"
                                    adp['sheet'][adp['die_name'] + str(adp['r'])].value = f"='{die['sheet_name']}'!{die['net_name'] + str(i)}"
                                    adp['sheet'][adp['bga_pin'] + str(adp['r'])].value = f"='{str(ball['sheet_name'])}'!{ball['Y'] + str(j)}&'{str(ball['sheet_name'])}'!{ball['X'] + str(j)}"
                                    # print(adp['sheet'][adp['pkg_name'] + str(adp['r'])].value)
                                    # print(adp['sheet'][adp['die_name'] + str(adp['r'])].value)

                                    adp['r'] += 1
                                  
                                    break
                                elif (ball['list_row'].index(j) == len(ball['list_row']) - 1):
                            
                                    adp['sheet'][adp['X'] + str(adp['r'])].value = f"='{die['sheet_name']}'!{die['X'] + str(i)}"
                                    adp['sheet'][adp['Y'] + str(adp['r'])].value = f"='{die['sheet_name']}'!{die['Y'] + str(i)}"   
                                    adp['sheet'][adp['pkg_name'] + str(adp['r'])].value = "N/A"
                                    adp['sheet'][adp['die_name'] + str(adp['r'])].value = f"='{die['sheet_name']}'!{die['net_name'] + str(i)}"
                                    adp['sheet'][adp['bga_pin'] + str(adp['r'])].value = "N/A"
                                    adp['r'] += 1
                 

            elif (str(die['sheet_d'][die['net_name'] + str(i)].value).find("VAA") != -1):
                adp['sheet'][adp['Pad_use'] + str(adp['r'])].value = "POWER"

                if (str(die['sheet_d'][die['net_name']  + str(i)].value).find("VAA2") != -1):
                    die_vaa2 = str(die['sheet_d'][die['net_name'] + str(i)].value).replace("VAA2","VDDA")
                   
                    for j in ball['list_row']:
                            # print(die['sheet_d'][die['net_name'] + str(i)].value)
                            # print(ball['sheet_d'][ball['net_col']+ str(j)].value)
                            if (die_vaa2 == ball['sheet_d'][ball['net_col']+ str(j)].value):
                                adp['sheet'][adp['resdef'] + str(adp['r'])].value = "BGA"
                                adp['sheet'][adp['X'] + str(adp['r'])].value = f"='{die['sheet_name']}'!{die['X'] + str(i)}"
                                adp['sheet'][adp['Y'] + str(adp['r'])].value = f"='{die['sheet_name']}'!{die['Y'] + str(i)}"    
                                adp['sheet'][adp['pkg_name'] + str(adp['r'])].value = f"='{ball['sheet_name']}'!{ball['net_col']+ str(j)}"
                                adp['sheet'][adp['die_name'] + str(adp['r'])].value = f"='{die['sheet_name']}'!{die['net_name'] + str(i)}"
                                adp['sheet'][adp['bga_pin'] + str(adp['r'])].value = f"='{str(ball['sheet_name'])}'!{ball['Y'] + str(j)}&'{str(ball['sheet_name'])}'!{ball['X'] + str(j)}"
                                # if(str(die['sheet_d'][die['net_name'] + str(i)].value).find("VDD") != -1):
                                
                                last_ball['vaa2'] = adp['sheet'][adp['bga_pin'] + str(adp['r'])].value
                                               
                                adp['r'] += 1
           
                                ball['list_row'].remove(j)
                          
                                break
                            elif (ball['list_row'].index(j) == len(ball['list_row']) - 1):
                                adp['sheet'][adp['resdef'] + str(adp['r'])].value = "BGA"
                                adp['sheet'][adp['X'] + str(adp['r'])].value = f"='{die['sheet_name']}'!{die['X'] + str(i)}"
                                adp['sheet'][adp['Y'] + str(adp['r'])].value = f"='{die['sheet_name']}'!{die['Y'] + str(i)}"    
                                adp['sheet'][adp['pkg_name'] + str(adp['r'])].value = f"='{die['sheet_name']}'!{die['net_name'] + str(i)}"
                                adp['sheet'][adp['die_name'] + str(adp['r'])].value = f"='{die['sheet_name']}'!{die['net_name'] + str(i)}"
                                adp['sheet'][adp['bga_pin'] + str(adp['r'])].value = last_ball['vaa2']
                else:
  
                    for j in ball['list_row']:
                            if (die['sheet_d'][die['net_name'] + str(i)].value == ball['sheet_d'][ball['net_col']+ str(j)].value):
                                adp['sheet'][adp['resdef'] + str(adp['r'])].value = "BGA"
                                adp['sheet'][adp['X'] + str(adp['r'])].value = die['sheet_d'][die['X']+ str(i)].value
                                adp['sheet'][adp['Y'] + str(adp['r'])].value = f"='{die['sheet_name']}'!{die['Y'] + str(i)}"    
                                adp['sheet'][adp['pkg_name'] + str(adp['r'])].value = f"='{ball['sheet_name']}'!{ball['net_col']+ str(j)}"
                                adp['sheet'][adp['die_name'] + str(adp['r'])].value = f"='{die['sheet_name']}'!{die['net_name'] + str(i)}"
                                adp['sheet'][adp['bga_pin'] + str(adp['r'])].value = f"='{str(ball['sheet_name'])}'!{ball['Y'] + str(j)}&'{str(ball['sheet_name'])}'!{ball['X'] + str(j)}"

 
                                last_ball['vaa'] = adp['sheet'][adp['bga_pin'] + str(adp['r'])].value
                                
                                ball['list_row'].remove(j)
                                adp['r'] += 1
                      
                                break
                            elif (ball['list_row'].index(j) == len(ball['list_row']) - 1 ):
                                adp['sheet'][adp['resdef'] + str(adp['r'])].value = "BGA"
                                adp['sheet'][adp['X'] + str(adp['r'])].value = die['sheet_d'][die['X']+ str(i)].value
                                adp['sheet'][adp['Y'] + str(adp['r'])].value = f"='{die['sheet_name']}'!{die['Y'] + str(i)}"    
                                adp['sheet'][adp['pkg_name'] + str(adp['r'])].value = f"='{die['sheet_name']}'!{die['net_name'] + str(i)}"
                                adp['sheet'][adp['die_name'] + str(adp['r'])].value = f"='{die['sheet_name']}'!{die['net_name'] + str(i)}"
                                adp['sheet'][adp['bga_pin'] + str(adp['r'])].value = last_ball['vaa']
                                adp['r'] += 1
            elif ((str(die['sheet_d'][die['net_name'] + str(i)].value).find("VDD") != -1) or str(die['sheet_d'][die['net_name'] + str(i)].value).find("VSS") != -1 or str(die['sheet_d'][die['net_name'] + str(i)].value).find("VCCIO") != -1 or str(die['sheet_d'][die['net_name'] + str(i)].value).find("VCCAON") != -1 or str(die['sheet_d'][die['net_name'] + str(i)].value).find("TC_VDDQ") != -1):

                if(str(die['sheet_d'][die['net_name'] + str(i)].value).find("VSS") != -1):
                    adp['sheet'][adp['Pad_use'] + str(adp['r'])].value = "GROUND"
                else:
                    adp['sheet'][adp['Pad_use'] + str(adp['r'])].value = "POWER"
              
                for j in ball['list_row']:
                    if (die['sheet_d'][die['net_name'] + str(i)].value == ball['sheet_d'][ball['net_col']+ str(j)].value):
                        adp['sheet'][adp['resdef'] + str(adp['r'])].value = "BGA"
                        adp['sheet'][adp['X'] + str(adp['r'])].value = f"='{die['sheet_name']}'!{die['X'] + str(i)}"
                        adp['sheet'][adp['Y'] + str(adp['r'])].value = f"='{die['sheet_name']}'!{die['Y'] + str(i)}"     
                        adp['sheet'][adp['pkg_name'] + str(adp['r'])].value = f"='{ball['sheet_name']}'!{ball['net_col']+ str(j)}"
                        adp['sheet'][adp['die_name'] + str(adp['r'])].value = f"='{die['sheet_name']}'!{die['net_name'] + str(i)}"
                        adp['sheet'][adp['bga_pin'] + str(adp['r'])].value = f"='{str(ball['sheet_name'])}'!{ball['Y'] + str(j)}&'{str(ball['sheet_name'])}'!{ball['X'] + str(j)}"
                        if(str(die['sheet_d'][die['net_name'] + str(i)].value).find("VDD") != -1):
                            
                            last_ball['vdd'] = adp['sheet'][adp['bga_pin'] + str(adp['r'])].value
                        elif (str(die['sheet_d'][die['net_name'] + str(i)].value).find("VCCIO") != -1):
                            last_ball['vccio'] = adp['sheet'][adp['bga_pin'] + str(adp['r'])].value
                        elif (str(die['sheet_d'][die['net_name'] + str(i)].value).find("VCCAON") != -1):
                            last_ball['vccaon'] = adp['sheet'][adp['bga_pin'] + str(adp['r'])].value
                        elif (str(die['sheet_d'][die['net_name'] + str(i)].value).find("TC_VDDQ") != -1):
                            last_ball['tc_vddq'] = adp['sheet'][adp['bga_pin'] + str(adp['r'])].value
                        elif (str(die['sheet_d'][die['net_name'] + str(i)].value).find("VSS") != -1):
                            last_ball['vss'] = adp['sheet'][adp['bga_pin'] + str(adp['r'])].value

                        adp['r'] += 1
                        
                        ball['list_row'].remove(j)
    
                        break
                    elif (ball['list_row'].index(j) == len(ball['list_row']) - 1):
                        adp['sheet'][adp['resdef'] + str(adp['r'])].value = "BGA"
                        adp['sheet'][adp['X'] + str(adp['r'])].value = f"='{die['sheet_name']}'!{die['X'] + str(i)}"
                        adp['sheet'][adp['Y'] + str(adp['r'])].value = f"='{die['sheet_name']}'!{die['Y'] + str(i)}"     
                        adp['sheet'][adp['pkg_name'] + str(adp['r'])].value = f"='{die['sheet_name']}'!{die['net_name'] + str(i)}"
                        adp['sheet'][adp['die_name'] + str(adp['r'])].value = f"='{die['sheet_name']}'!{die['net_name'] + str(i)}"

                        if(str(die['sheet_d'][die['net_name'] + str(i)].value).find("VDD") != -1):
                            
                            adp['sheet'][adp['bga_pin'] + str(adp['r'])].value = last_ball['vdd']
                        elif (str(die['sheet_d'][die['net_name'] + str(i)].value).find("VCCIO") != -1):
                            adp['sheet'][adp['bga_pin'] + str(adp['r'])].value = last_ball['vccio']
                        elif (str(die['sheet_d'][die['net_name'] + str(i)].value).find("VCCAON") != -1):
                            adp['sheet'][adp['bga_pin'] + str(adp['r'])].value = last_ball['vccaon']
                        elif (str(die['sheet_d'][die['net_name'] + str(i)].value).find("TC_VDDQ") != -1):
                            adp['sheet'][adp['bga_pin'] + str(adp['r'])].value = last_ball['tc_vddq']
                        elif (str(die['sheet_d'][die['net_name'] + str(i)].value).find("VSS") != -1):
                            adp['sheet'][adp['bga_pin'] + str(adp['r'])].value = last_ball['vss'] 

                        adp['r'] += 1
                            
            else:
                if(str(die['sheet_d'][die['net_name'] + str(i)].value).find("RDI_PL_CFG") != -1 or str(die['sheet_d'][die['net_name'] + str(i)].value).find("TDO") != -1):
                    temp = die['sheet_d'][die['net_name'] + str(i)].value
                    adp['sheet'][adp['Pad_use'] + str(adp['r'])].value = "O"
                elif(str(die['sheet_d'][die['net_name'] + str(i)].value).find("CLK") != -1 or str(die['sheet_d'][die['net_name'] + str(i)].value).find("DBG") != -1 or str(die['sheet_d'][die['net_name'] + str(i)].value).find("CHIP_RST") != -1 or str(die['sheet_d'][die['net_name'] + str(i)].value).find("TCK") != -1 or str(die['sheet_d'][die['net_name'] + str(i)].value).find("TRST") != -1 or str(die['sheet_d'][die['net_name'] + str(i)].value).find("TMS") != -1 or str(die['sheet_d'][die['net_name'] + str(i)].value).find("TDI") != -1):
                    temp = die['sheet_d'][die['net_name'] + str(i)].value
                    adp['sheet'][adp['Pad_use'] + str(adp['r'])].value = "I"
                elif(str(die['sheet_d'][die['net_name'] + str(i)].value).find("ATO") != -1):
                    temp = die['sheet_d'][die['net_name'] + str(i)].value
                    adp['sheet'][adp['Pad_use'] + str(adp['r'])].value = "BI"
                elif(str(die['sheet_d'][die['net_name'] + str(i)].value).find("DTO") != -1):
                    temp = die['sheet_d'][die['net_name'] + str(i)].value
                    adp['sheet'][adp['Pad_use'] + str(adp['r'])].value = "O"
                elif(str(die['sheet_d'][die['net_name'] + str(i)].value).find("ZN") != -1):
                    temp = die['sheet_d'][die['net_name'] + str(i)].value
                    adp['sheet'][adp['Pad_use'] + str(adp['r'])].value = "O"
                elif(str(die['sheet_d'][die['net_name'] + str(i)].value).find("ACK") != -1):
                    temp = die['sheet_d'][die['net_name'] + str(i)].value
                    adp['sheet'][adp['Pad_use'] + str(adp['r'])].value = "BI"
                elif(str(die['sheet_d'][die['net_name'] + str(i)].value).find("DCK") != -1):
                    temp = die['sheet_d'][die['net_name'] + str(i)].value
                    adp['sheet'][adp['Pad_use'] + str(adp['r'])].value = "O"
       
                for j in ball['list_row']:
                        if (die['sheet_d'][die['net_name'] + str(i)].value == ball['sheet_d'][ball['net_col']+ str(j)].value):
                            adp['sheet'][adp['resdef'] + str(adp['r'])].value = "BGA"
                            adp['sheet'][adp['X'] + str(adp['r'])].value = f"='{die['sheet_name']}'!{die['X'] + str(i)}"
                            adp['sheet'][adp['Y'] + str(adp['r'])].value = f"='{die['sheet_name']}'!{die['Y'] + str(i)}"     
                            adp['sheet'][adp['pkg_name'] + str(adp['r'])].value = f"='{ball['sheet_name']}'!{ball['net_col']+ str(j)}"
                            adp['sheet'][adp['die_name'] + str(adp['r'])].value = f"='{die['sheet_name']}'!{die['net_name'] + str(i)}"
                            adp['sheet'][adp['bga_pin'] + str(adp['r'])].value = f"='{str(ball['sheet_name'])}'!{ball['Y'] + str(j)}&'{str(ball['sheet_name'])}'!{ball['X'] + str(j)}"
                            
                            adp['r'] += 1
          
                            
                            ball['list_row'].remove(j)
                            break
                        elif (ball['list_row'].index(j) == len(ball['list_row']) - 1):

                            adp['sheet'][adp['X'] + str(adp['r'])].value = f"='{die['sheet_name']}'!{die['X'] + str(i)}"
                            adp['sheet'][adp['Y'] + str(adp['r'])].value = f"='{die['sheet_name']}'!{die['Y'] + str(i)}"    
                            adp['sheet'][adp['pkg_name'] + str(adp['r'])].value = "NA"
                            adp['sheet'][adp['die_name'] + str(adp['r'])].value = f"='{die['sheet_name']}'!{die['net_name'] + str(i)}"
                            adp['sheet'][adp['bga_pin'] + str(adp['r'])].value = "NA"

                            adp['r'] += 1
            adp['pin_num'] = int(adp['pin_num']) + 1

        return adp,mapping,ball,last_ball

def gen_apd():
    print("Getting config...")
    mynotif("Generating...")
    params = get_config()
    die_params = params[0]
    input_params = params[1]
    output_params = params[2]
    progress_bar(10)
    print("Creating the ball temple sheet...")
    # ball_tmp_tb = copy_table(input_params)
    # print("Refreshing the excelfile...")
    # refresh_excel(input_params['excel_file'])
    # progress_bar(20)
    

    last_ball ={
        "vss": "NULL",
        "vdd": "NULL",
        "vccio":"NULL",
        "vaa":"NULL",
        "vccaon":"NULL",
        "vaa2":"NULL",
        "tc_vddq":"NULL",
        "die_cmp": 0

    }

 
    print("Loading workbook")
    try:
        wb_d = load_workbook(input_params['excel_file'], data_only=True)
        wb_f = load_workbook(input_params['excel_file'])
    except:
        print("Ploc file doesn't exist or it is being opened. Please check and re-run!!!")
        show_error("Ploc file doesn't exist or it is being opened. Please check and re-run!!!")
    apd_sheet = output_params['sheet']
    sheet_list = wb_f.sheetnames
    if apd_sheet in sheet_list:
        
        ws_adp_f = wb_f[apd_sheet]
        
    else:
        msg_ws = messagebox.askquestion('Create Sheet', 'The sheet' + apd_sheet + ' doesn\'t exist. Do you want to create it?', icon='question')
            
        if(msg_ws == 'yes'):
            ws_adp_f = wb_f.create_sheet(apd_sheet)
            # mynotif("")
            # mynotif('Creating the sheet...')
        else:
            # mynotif("")
            # progress_bar(0)
            return
            

    
    pin_number = 1
    adp_tb_loc = output_params['tb_loc']
    adp_begincol = coordinate_to_tuple(adp_tb_loc)[1]
    adp_beginrow = coordinate_to_tuple(adp_tb_loc)[0]
    adp_Pin_number = get_column_letter(adp_begincol)
    adp_Pad_stack = get_column_letter(adp_begincol + 1)
    adp_Xcoor = get_column_letter(adp_begincol + 2)
    adp_Ycoor = get_column_letter(adp_begincol + 3)
    adp_rotation = get_column_letter(adp_begincol + 4)
    adp_Pad_use = get_column_letter(adp_begincol + 5)
    adp_pkg_name = get_column_letter(adp_begincol + 6)
    adp_die_name = get_column_letter(adp_begincol + 7)
    adp_resdef = get_column_letter(adp_begincol + 8)
    adp_bga_pin = get_column_letter(adp_begincol + 9)
    radp = adp_beginrow + 1

    adp ={
        "sheet":ws_adp_f,
        "pin_num":int(pin_number),
        "tb_loc": adp_tb_loc,
        "begincol": adp_begincol, 
        "beginrow" : adp_beginrow,
        "Pin_number": adp_Pin_number, 
        "Pad_stack": adp_Pad_stack, 
        "X": adp_Xcoor, 
        "Y": adp_Ycoor, 
        "rotation": adp_rotation, 
        "Pad_use": adp_Pad_use, 
        "pkg_name": adp_pkg_name, 
        "die_name": adp_die_name, 
        "resdef":  adp_resdef,
        "bga_pin":  adp_bga_pin,
        "r":radp
    }

   
    mapping_row_col = get_col_row_range(input_params['mapping_begin_cell'], input_params["mapping_end_cell"])
    mapping_row_min = mapping_row_col[0]
    mapping_row_max = mapping_row_col[1]
    mapping_col_min = mapping_row_col[2]
    mapping_col_max = mapping_row_col[3]
    mapping_die_L = get_column_letter(mapping_col_min)
    mapping_die_R = get_column_letter(mapping_col_min + 1)
    try:
        mapping_sheet_d = wb_d[input_params['mapping_sheet']]
    except:
        show_error("The sheet: " + input_params['mapping_sheet'] + " doesn't exist. Please recheck")
        print("Error: The sheet: " + input_params['mapping_sheet'] + " doesn't exist. Please recheck")
        mynotif("Error!!!")
        progress_bar(0)
        return

    mapping={
        "sheet_name":input_params['mapping_sheet'],
        "sheet_d":mapping_sheet_d,
        "row_min":mapping_row_min,
        "row_max":mapping_row_max,
        "col_min":mapping_col_min,
        "col_max":mapping_col_max,
        "die_L":mapping_die_L,
        "die_R":mapping_die_R
        
    }

    ball_row_col = get_col_row_range(input_params['ball_tb_begin_cell'],input_params['ball_tb_end_cell'])

    ball_begin_row = ball_row_col[0]
    ball_end_row = ball_row_col[1]
    ball_begin_col = ball_row_col[2]
    ball_end_col = ball_row_col[3]
    ball_net_col = get_column_letter(ball_end_col)
    ball_X = get_column_letter(ball_end_col - 2)
    ball_Y = get_column_letter(ball_end_col - 1)
    ball_list_row=[]
    for row in range(ball_begin_row + 2, ball_end_row + 1):
        ball_list_row.append(row) 
    try:
        ball_sheet_d = wb_d[input_params['ball_tb_sheet']]
        ball_sheet_f = wb_f[input_params['ball_tb_sheet']]
    except:
        show_error("The sheet: " + input_params['ball_tb_sheet'] + " doesn't exist. Please recheck")
        print("Error: The sheet: " + input_params['ball_tb_sheet'] + " doesn't exist. Please recheck")
        mynotif("Error!!!")
        progress_bar(0)
        return
    ball={
        "sheet_name":input_params['ball_tb_sheet'],
        "sheet_d":ball_sheet_d,
        "sheet_f": ball_sheet_f,
        "begin_row":ball_begin_row,
        "end_row":ball_end_row,
        "begin_col":ball_begin_col,
        "end_col":ball_end_col,
        "net_col":ball_net_col,
        "X":ball_X,
        "Y":ball_Y,
        "list_row": ball_list_row,
        "list_compared":[]
    }

    progress_bar(30)
    progresval = int(40/len(die_params['diel_list']))
    for die_cnt in range(0, len(die_params['diel_list'])):

        Die_L_row_col = get_col_row_range(die_params['diel_begin_list'][die_cnt], die_params['diel_end_list'][die_cnt])
        Die_L_name = die_params['diel_list'][die_cnt]

        die_L={
            "sheet_name": die_params['die_sheet'],
            "sheet_d": wb_d[die_params['die_sheet']],
            "sheet_f": wb_f[die_params['die_sheet']],
            "row_col" : Die_L_row_col,
            "row_min" : Die_L_row_col[0],
            "row_max" : Die_L_row_col[1],
            "col_min" : Die_L_row_col[2],
            "col_max" : Die_L_row_col[3],
            "X" : get_column_letter(Die_L_row_col[3] - 2),
            "Y" : get_column_letter(Die_L_row_col[3] - 1),
            "net_name" : get_column_letter(Die_L_row_col[3]),
            "name" : Die_L_name,
            "die_side":"L",
            "Pad_stack": "60x60"
        }

        Die_R_row_col = get_col_row_range(die_params['dier_begin_list'][die_cnt], die_params['dier_end_list'][die_cnt])
        Die_R_name = die_params['dier_list'][die_cnt]
        die_R={
            "sheet_name": die_params['die_sheet'],
            "sheet_d": wb_d[die_params['die_sheet']],
            "sheet_f": wb_f[die_params['die_sheet']],
            "row_col" : Die_R_row_col,
            "row_min" : Die_R_row_col[0],
            "row_max" : Die_R_row_col[1],
            "col_min" : Die_R_row_col[2],
            "col_max" : Die_R_row_col[3],
            "X" : get_column_letter(Die_R_row_col[3] - 2),
            "Y" : get_column_letter(Die_R_row_col[3] - 1),
            "net_name" : get_column_letter(Die_R_row_col[3]),
            "name": Die_R_name,
            "die_side" : "R",
            "Pad_stack": "60x60"
        }
        mapping_prefix = die_L['name']
        print("Processing for: L:" + Die_L_name + "  R:" + Die_R_name )
        get_gen = gen(adp,die_L,mapping,mapping_prefix,ball,last_ball)
        adp = get_gen[0]
        mapping = get_gen[1]
        ball = get_gen[2]
        last_ball = get_gen[3]
        print(ball)
        print(last_ball)
        print(len(ball['list_compared']))
        get_gen = gen(adp,die_R,mapping,mapping_prefix,ball,last_ball)
        adp = get_gen[0]
        mapping = get_gen[1]
        ball = get_gen[2]
        last_ball = get_gen[3]
        print(len(ball['list_compared']))
        progress_bar(30 + die_cnt*progresval)
    print(ball['list_row'])
    # for m in range (ball['begin_row'], ball['end_row'] + 1):
    print("Processing unmap Ball...")
    for m in ball['list_row']:
        adp['sheet'][adp['resdef'] + str(adp['r'])].value = "BGA"
        if(str(ball['sheet_d'][ball['net_col']+ str(m)].value).find("VSS") != -1):
            adp['sheet'][adp['Pad_use'] + str(adp['r'])] = "GROUND"
        elif(str(ball['sheet_d'][ball['net_col']+ str(m)].value).find("VDD") != -1 or str(ball['sheet_d'][ball['net_col']+ str(m)].value).find("VCCIO") != -1 or str(ball['sheet_d'][ball['net_col']+ str(m)].value).find("VAA") != -1 
             or str(ball['sheet_d'][ball['net_col']+ str(m)].value).find("TC_VDDQ") != -1 or str(ball['sheet_d'][ball['net_col']+ str(m)].value).find("VCCAON") != -1):
            adp['sheet'][adp['Pad_use'] + str(adp['r'])] = "POWER"
        elif(str(ball['sheet_d'][ball['net_col']+ str(m)].value).find("RDI_LP_CFG") != -1 or str(ball['sheet_d'][ball['net_col']+ str(m)].value).find("RDI_CFG_CLK") != -1 or str(ball['sheet_d'][ball['net_col']+ str(m)].value).find("RDI_MODE") != -1):
            pass
        #  (str(die['sheet_d'][die['net_name'] + str(i)].value).find("RDI_LP_CFG") != -1 or str(die['sheet_d'][die['net_name'] + str(i)].value).find("RDI_CFG_CLK") != -1 or str(die['sheet_d'][die['net_name'] + str(i)].value).find("RDI_MODE") != -1):
        else:
            adp['sheet'][adp['Pad_use'] + str(adp['r'])] = "NC"
        if(str(ball['sheet_d'][ball['net_col']+ str(m)].value).find("RDI_LP_CFG") != -1 or str(ball['sheet_d'][ball['net_col']+ str(m)].value).find("RDI_CFG_CLK") != -1 or str(ball['sheet_d'][ball['net_col']+ str(m)].value).find("RDI_MODE") != -1):
            pass
        else:
            adp['sheet'][adp['pkg_name'] + str(adp['r'])].value = ball['sheet_f'][ball['net_col'] + str(m)].value
            
            adp['sheet'][adp['bga_pin'] + str(adp['r'])].value = f"='{str(ball['sheet_name'])}'!{ball['Y'] + str(m)}&'{str(ball['sheet_name'])}'!{ball['X'] + str(m)}"   
            adp['r'] += 1
            # print(ball['end_row']) 
    progress_bar(80)
    print("Saving excel file...")
    wb_f.save(input_params['excel_file'])
    progress_bar(100)
    popup("ADP Nestlist generated successful!!!")
    print("Successful!!!")
    mynotif("Successful!!!")
    

progress = ttk.Progressbar(root, orient = 'horizontal',
              length = 100, mode = 'determinate')
progress_w = my_canvas.create_window(80,800, anchor="nw", window=progress, width= 800)


# Button
#Create style object
style = ttk.Style()

# #configure style
# style = ttk.Style()
# style.configure('TButton', font =
#                ('calibri', 20, 'bold'),
#                     borderwidth = '4',
#                     width = '80')
style.configure('TCheckbutton', font= ('System', 12, 'underline', 'bold'),
 foreground='black', border=50)
mediumFont = tkfont(
	family="System",
	size=12,
	weight="normal",
	slant="italic",
	underline=1,
	overstrike=0)
def hihi():
    button.configure(font=mediumFont, foreground='white', background='Green')
browse_btn = ttk.Button(root, text="Open File", image=open_imag, command=browse_file)
browse_btn_w = my_canvas.create_window(865, 40, anchor="nw", window=browse_btn)
# button = tk.Button(root, text="Generate",font=("System", 14, 'underline', 'bold'), foreground='white', background='#9b34eb', command=get_path, width=40)
button = tk.Button(root, text="Generate",font = mediumFont, foreground='white', background='#9b34eb', command=gen_apd, width=40)
# button = ttk.Button(root, text="Generate", command=get_path, width=80)

button_w = my_canvas.create_window(300, 860, anchor="nw", window=button)

def get_saved_params():
    try:
        with open(".adp_params_saved.txt",'r') as params_saved:
            line1 = [line.rstrip() for line in params_saved]
            params = {
                'theme_combo': line1[0],
                'excel_i': line1[1],
                'ball_sheet_i': line1[2],
                'ball_begin_cell_i': line1[3],
                'ball_end_cell': line1[4],
                'mapping_sheet_i': line1[5],
                'mapping_begin_cell_i': line1[6],
                'mapping_end_cell_i': line1[7],
                'dietb_sheet': line1[8],
                'DieL_name': line1[9],
                'DieR_name': line1[10],
                'DieL_begincell_i': line1[11],
                'DieL_endcell_i': line1[12],
                'DieR_begincell_i': line1[13],
                'DieR_endcell_i': line1[14],
                'adp_tb': line1[15],
                'adp_tb_loc': line1[16],

            }
        theme_combo.current(theme_list.index(params['theme_combo']))
        root.set_theme(params['theme_combo'])
        change_colour(theme_list.index(params['theme_combo']))
        excel_i.insert(0, params['excel_i'])
        ball_sheet_i.insert(0, params['ball_sheet_i'])
        ball_begin_cell_i.insert(0, params['ball_begin_cell_i'])
        ball_end_cell.insert(0, params['ball_end_cell'])
        mapping_sheet_i.insert(0, params['mapping_sheet_i'])
        mapping_begin_cell_i.insert(0, params['mapping_begin_cell_i'])
        mapping_end_cell_i.insert(0, params['mapping_end_cell_i'])
        dietb_sheet.insert(0, params['dietb_sheet'])
        DieL_name.insert(0, params['DieL_name'])
        DieR_name.insert(0, params['DieR_name'])
        DieL_begincell_i.insert(0, params['DieL_begincell_i'])
        DieL_endcell_i.insert(0, params['DieL_endcell_i'])
        DieR_begincell_i.insert(0, params['DieR_begincell_i'])
        DieR_endcell_i.insert(0, params['DieR_endcell_i'])
        adp_tb.insert(0, params['adp_tb'])
        adp_tb_loc.insert(0, params['adp_tb_loc'])
    except:
        theme_combo.current(0)
        # root.set_theme(params['theme_combo'])
        excel_i.insert(0, r"C:\Users\sytung\OneDrive - Synopsys, Inc\Desktop\py\Test3.xlsx")
        ball_sheet_i.insert(0, "BGA")
        ball_begin_cell_i.insert(0, "AQ1")
        ball_end_cell.insert(0, "AT1297")
        mapping_sheet_i.insert(0, "UCIe_Mapping_connection")
        mapping_begin_cell_i.insert(0, "F1")
        mapping_end_cell_i.insert(0, "G178")
        dietb_sheet.insert(0, "Package_substrate")
        DieL_name.insert(0, "DIE3 DIE4")
        DieR_name.insert(0, "DIE7 DIE8")
        DieL_begincell_i.insert(0, "T19 AB19")
        DieL_endcell_i.insert(0, "V791 AD791")
        DieR_begincell_i.insert(0, "X19 AF19")
        DieR_endcell_i.insert(0, "Z791 AH791")
        adp_tb.insert(0, "APD")
        adp_tb_loc.insert(0, "P4")

get_saved_params()

root.mainloop()
