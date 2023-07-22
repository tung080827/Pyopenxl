from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.utils import range_boundaries
from openpyxl.utils import column_index_from_string
#from openpyxl.utils import coordinate_from_string
from openpyxl.utils import coordinate_to_tuple
from openpyxl.worksheet.table import Table, TableStyleInfo
# from tkinter import *
from tkinter import messagebox
from tkinter import ttk
import getcolumn
from array import *
import tkinter as tk
from tkinter import *
from ttkthemes import ThemedTk, THEMES
from PIL import Image
from PIL import ImageTk, Image
from tkinter.font import Font as tkfont
from tkinter import filedialog
import gui_function as gui
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection
import os
import win32com.client
from pathlib import Path  # core library
# excel_file = r"C:\Users\sytung\OneDrive - Synopsys, Inc\Desktop\py\Test.xlsx"
from datamapping_gen import * 
import tempfile

root = ThemedTk()
# my_canvas=tk.Canvas(root)
# root.set_theme("scidpurple")

root.title("PLOC DATA CHANNEL VISUAL GENERATOR")
root.geometry("800x800+30+100")
root.resizable(width=False, height=False)
root.iconbitmap(r".\mylogo.ico")
root.option_add("*tearOff", False) # This is always a good idea

# bg = ImageTk.PhotoImage(file=r".\bg3_1.png")
try:
    temp_file =  os.path.join(tempfile.gettempdir(), ".datachannelgen_params_saved.txt")
    print(temp_file)
    tmp_flag = 0
except:
    messagebox.showerror("Can not find the User Temp dir")
    tmp_flag = 1

open_imag = PhotoImage(file = r".\open-folder.png")

img_path = r".\img\resize1000x1000"
# bg = ImageTk.PhotoImage(file=r".\img\mountain.png")
bgm = PhotoImage(file=img_path + r"\frog.png")

img_list = ["owl.png", "mountain.png","car.png", "penguin.png","sunset1.png", "flower3.png", "kid.png", "pug.png", "cat.png", "whale2.png", "elephant_grey.png", "snowman.png", "bee4.png", "elephant.png", "bee2.png", "fox.png", "beach.png", "frog.png", "cow.png", "forest.png", "owlpink2.png", "girl.png", "sand1.png", "baby2.png", "pig.png", "discord1.png" ]

lable_bg_list = ["#F0F0F0","#EDEDED","#EBECEE","#F0F0F0","#F0F0F0","#FCFCFC","#EFF0F1","#EFF0F1","#EFF0F1","#EAECEF","#EFF0F1","#EFF0F1","#FECDD9","#EFF0F1","#EFF0F1","#EFF0F1","#EFF0F1","#EFF0F1","#EFF0F1","#EFF0F1","#EFF0F1", "#EFF0F1","#EFF0F1", "#EFF0F1","#EFF0F1","#E6EBEF"]
# Define Canvas
my_canvas = tk.Canvas(root, width=1200, height=800, bd=0, highlightthickness=0)
my_canvas.pack(fill="both", expand=True)

# Put the image on the canvas
bg_img = my_canvas.create_image(0,0, image=bgm, anchor="nw")
stfont= ("Franklin Gothic Medium", 10, 'underline', "italic")
# Create lists for the Comboboxes
theme_list = ["adapta", "aquativo", "arc", "black","blue", "breeze", "clearlooks", "elegance", "equilux", "itft1", "keramik", "keramik_alt", "kroc", "plastik", "radiance", "ubuntu", "scidblue", "scidgreen", "scidgrey", "scidmint", "scidpink", "scidpurple", "scidsand", "smog", "winxpblue", "yaru" ]
colour_list = ["#09a5e8", "#292b33", "#1583eb", "#292a2b","#1a7cad", "#0664bd", "#8baac7", "#59564f", "#40454a", "#7aa7f5", "#7795b4", "#7795b4", "#ebab0c", "#0c99eb", "#eb830c", "#eb830c", "#0937ab", "#37ed80", "#707371", "#479403", "#d12a9f", "#9b34eb", "#787122", "#118cbd", "#505257", "#924d8b" ]


ch_number = [1,2,3,4,5,6,7,8,9, 10, 11, 12, 13, 14, 15, 16,17,18,19,20,21,22,23,24,25,26,27,28,29,30,31,32]
ch_sequence = ["Right to Left","Left to Right", "Center to Left first","Center to Right first"]

# Create control variables

tc_opt = tk.IntVar()
isIntp = tk.IntVar()
ismapgen = tk.IntVar()
def change_colour(index):
    listchange =[out_t,in_t,in_out_t]
    entry_list1 = [button, theme_combo_t,excel_t]
    #  sheet_t, theme_combo_t,excel_t, pkg_t
    for t in listchange:
        my_canvas.itemconfig(t, fill = colour_list[index])
    for l in entry_list1:
        l.config(background = colour_list[index])
    text.configure(foreground=colour_list[index],bg=lable_bg_list[index], highlightbackground=colour_list[index])
    global entry_list
    if (index == 8):
        excel_i.configure(foreground="white")
        for ent in entry_list:
            ent.configure(foreground="white")
    else:
        excel_i.configure(foreground="black")
        for ent in entry_list:
            ent.configure(foreground="black")
    global bgm
    p = os.path.join(img_path, img_list[index])
    pic = Image.open(p)
    picrsz = pic.resize((800,800)) 
    bgm = ImageTk.PhotoImage(picrsz)
    my_canvas.itemconfigure(bg_img, image=bgm)
def text_delete():
   text.delete("1.0","end")
def mynotif(content):
    text.insert(tk.END,content+"\n")
    text.see("end")
def process_notify(content):    
        root.update_idletasks()
        mynotif(content)
        root.update_idletasks()
def entry_disable(*entries):
    for entry in entries:
        entry.config(state='disable')
def entry_enable(*entries):
    for entry in entries:
        entry.config(state='normal')
# ------------------------------------------------------------------------------------------------------------------------------------------------

def myguide(entries, content):
    entries.insert(tk.END, content)
def progress_bar(value):
    progress['value'] = value
    root.update_idletasks()
def choosetheme(event):
    root.set_theme(theme_combo.get())
    change_colour(theme_list.index(theme_combo.get()))

def guide(gui_list):
    entry_enable(text)
    text_delete()
    for gui in gui_list:
       myguide(text,gui)
    entry_disable(text)
x1y1_guide = [
    "INFO: Die window begin cell\n\n ",
    "      * Example:   A0           "
]
map_s_guide = [
    "INFO: Sheet to put Channels Mapping table\n\n ",
    "      * Example: Mapping Data Channel "
]
map_loc_guide = [
    "INFO: This field to define the output mapping table location.  \n\n",
    "      * Example: O64 "
]
x2y2_guide = [
    "INFO: Reference channel visual window end cell.  \n\n",
    "      * Example: CU100  "
]
out_name_in_guide = [
    "INFO: This field to define the output table name.  \n\n",
    "      * Example: Mapping  "
]
out_visual_loc_guide = [
    "INFO: This field to define the output bump visual location.  \n\n",
    "      * Example: D10  "
]
outtb_s_guide = [
    "INFO: This field to define Sheet to put channels bump visual.  \n\n",
    "      * Example: Bump coordination  "
]
bit_num_guide = [
    "INFO: Number of bit of each channel.  \n\n",
    "      * Example: 16  "
]
DieL_name_guide = [
    "INFO: Name list of Left Die. (Die Flipped + Rotate -90) \n",
    "- The dies name are separated by spaces. \n",
    "      * Example: DIE1 DIE2 DIE3 DIE4  "
]
DieR_name_guide = [
    "INFO: Name list of Right Die. (Die Flipped then Rotate +90) \n",
    "- The dies name are separated by spaces. \n",
    "      * Example: DIE5 DIE6 DIE7 DIE8  "
]
sheet_guide = [
    "INFO: This field to define Sheet of reference channel bump visual.  \n\n",
    "      * Example: Bump coordination  "
]
def get_ch_seq(event):
    if (ch_combo.get()=="Left to Right"):
        # myguide(frame, "INFO:" + "The pin order will be indexed\n From Left to Right          ")
        # myguide(text, "INFO: The pin order will be indexed From Left to Right.  \n\n" )
        guide("INFO: The pin order will be indexed From Left to Right.  \n\n")

    else:
        guide("INFO: The pin order will be indexed From Right to Left.  \n\n" )

def get_ch_cnt(event):
    guide("INFO: The number of channels: " +  ch_combo.get() )
    

def gen_mapping_toggle():
    if(ismapgen.get() == 1):
        entry_enable(die_R_list, die_L_list, map_tb_sheet, map_loc_i)
       
        print("Gen interposer Die table: ON")
        
        guide("Gen interposer Die table: ON")
    elif(ismapgen.get() == 0):
        entry_disable(die_R_list, die_L_list, map_tb_sheet, map_loc_i)
        
        print("Gen interposer Die table: OFF")
        
        guide("Gen interposer Die table: OFF")

xfont = ("System", 12, "bold", 'underline', 'italic')
theme_combo_t = ttk.Label(root,text="Choose theme:",border=20, font=xfont, background='#b434eb', borderwidth=3)
theme_combo_t_w = my_canvas.create_window(580, 15, window=theme_combo_t)

theme_combo = ttk.Combobox(root, state="readonly", values=theme_list, width=15)
theme_combo_w = my_canvas.create_window(720,15, window=theme_combo)
theme_combo.current(0)
theme_combo.bind('<<ComboboxSelected>>', choosetheme)

# -------------------------excelpath input--------------------------#
pfont= ("Rosewood Std Regular", 12, "bold", 'underline' )
excel_t = ttk.Label(root,text="PLOC path:",border=20,font=pfont, borderwidth=5)
excel_t_w = my_canvas.create_window(30,40, anchor="nw", window=excel_t)
excel_i = ttk.Entry(root, width=94)

excel_i_w = my_canvas.create_window(150,40, anchor="nw", window=excel_i)

# -------------------------excel sheet_name input--------------------------#



sheet_i = ttk.Entry(root, background="#217346", width=20)

sheet_i_w = my_canvas.create_window(150,250, anchor="nw", window=sheet_i)
sheet_i.bind('<FocusIn>', lambda event: guide(sheet_guide))
# sheet_i.bind('<FocusOut>', un_guide)
in_out_t = my_canvas.create_text(30, 200, text="Input/Output Config:", anchor="nw",font=("Helvetica", 10, 'italic', 'underline', 'bold'), fill="#b434eb")



# ------------------------Die bump visual parameters input --------------------------#
in_t = my_canvas.create_text(150, 230, text="Input:", anchor="nw",font=("Helvetica", 10, 'italic', 'underline', 'bold'), fill="#b434eb")

ch_combo = ttk.Combobox(root, state="readonly", values=ch_number, width=17)
ch_combo_w = my_canvas.create_window(300,250, anchor="nw", window=ch_combo)

ch_combo.bind('<<ComboboxSelected>>', get_ch_cnt)

ch_seq_combo = ttk.Combobox(root, state="readonly", values=ch_sequence, width=17)
ch_seq_combo_w = my_canvas.create_window(300,290, anchor="nw", window=ch_seq_combo)

ch_seq_combo.bind('<<ComboboxSelected>>', get_ch_seq)

bit_num_i = ttk.Entry(root, width=20)
my_canvas.create_window(300, 330, anchor="nw", window=bit_num_i)

bit_num_i.bind('<FocusIn>', lambda event: guide(bit_num_guide))
# bit_num_i.bind('<FocusOut>', un_guide)

x1y1_i = ttk.Entry(root, width=20)
my_canvas.create_window(150, 290, anchor="nw", window=x1y1_i)

x1y1_i.bind('<FocusIn>', lambda event: guide(x1y1_guide))
# x1y1_i.bind('<FocusOut>', un_guide)


x2y2_i = ttk.Entry(root, width=20)
my_canvas.create_window(150, 330, anchor="nw", window=x2y2_i)

x2y2_i.bind('<FocusIn>', lambda event: guide(x2y2_guide))
# x2y2_i.bind('<FocusOut>', un_guide)

out_t = my_canvas.create_text(500, 230, text="Output:", anchor="nw",font=("Helvetica", 10, 'italic', 'underline', 'bold'), fill="#b434eb")
out_tb_sheet = ttk.Entry(root)
out_tb_sheet_w = my_canvas.create_window(500, 250, anchor="nw", window=out_tb_sheet)

out_tb_sheet.bind('<FocusIn>', lambda event: guide(outtb_s_guide))
# out_tb_sheet.bind('<FocusOut>', un_guide)

out_col_i = ttk.Entry(root, width=20)
out_col_i_w = my_canvas.create_window(500, 290, anchor="nw", window=out_col_i)

out_col_i.bind('<FocusIn>', lambda event: guide(out_visual_loc_guide))
# out_col_i.bind('<FocusOut>', un_guide)

# -------------------------pkg type input--------------------------#
die_map_t = gen_mapping = ttk.Checkbutton(root, text="Die Mapping generator", variable=ismapgen, onvalue=1, offvalue=0,command= gen_mapping_toggle)
gen_mapping_w =my_canvas.create_window(30, 360, anchor="nw", window=gen_mapping)

die_L_list = ttk.Entry(root, width=20)
my_canvas.create_window(150, 390, anchor="nw", window=die_L_list, width=275)

die_L_list.bind('<FocusIn>', lambda event: guide(DieL_name_guide))
# die_L_list.bind('<FocusOut>', un_guide)

die_R_list = ttk.Entry(root, width=20)
my_canvas.create_window(150, 430, anchor="nw", window=die_R_list, width=275)

die_R_list.bind('<FocusIn>', lambda event: guide(DieR_name_guide))
# die_R_list.bind('<FocusOut>', un_guide)

map_tb_sheet = ttk.Entry(root)
map_tb_sheet_w = my_canvas.create_window(500, 390, anchor="nw", window=map_tb_sheet)

map_tb_sheet.bind('<FocusIn>', lambda event: guide(map_s_guide))
# map_tb_sheet.bind('<FocusOut>', un_guide)

map_loc_i = ttk.Entry(root, width=20)
map_loc_i_w = my_canvas.create_window(500, 430, anchor="nw", window=map_loc_i)

map_loc_i.bind('<FocusIn>', lambda event: guide(map_loc_guide))
# map_loc_i.bind('<FocusOut>', un_guide)




myLabel = ttk.Label(root,text="---")
myLabel_w =my_canvas.create_window(80,500,anchor="nw", window=myLabel)



def browse_file():
	# global my_image
    root.filename = filedialog.askopenfilename(initialdir="./", title="Select A File", filetypes=(("Excel files", "*.xlsx"),("all files", "*.*")))
    excel_i.delete(0,END)
    print(root.filename) 
    excel_i.insert(0, root.filename)
	# my_image = ImageTk.PhotoImage(Image.open(root.filename))
	# my_image_label = Label(image=my_image).pack()

#---------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------#
def popup(notif):
    messagebox.showinfo("Notification", notif)

def show_error(error):
    messagebox.showerror("Error", error)
progress = ttk.Progressbar(root, orient = 'horizontal',
              length = 100, mode = 'determinate')
progress_w = my_canvas.create_window(80,550, anchor="nw", window=progress, width= 600)


# Button
#Create style object
style = ttk.Style()

# #configure style
# style = ttk.Style()
# style.configure('TButton', font =
#                ('calibri', 20, 'bold'),
#                     borderwidth = '4',
#                     width = '80')
style.configure('TCheckbutton', font= ('System', 12, 'underline', 'bold'),
 foreground='black', border=50)
mediumFont = tkfont(
	family="System",
	size=16,
	weight="normal",
	slant="italic",
	underline=1,
	overstrike=0)
def hihi():
    button.configure(font=mediumFont, foreground='white', background='Green')

def getstring(string: str,c1: str, c2: str):
    cell = string
    idx1 = cell.find(c1)
    idx2 = cell.find(c2)
    if(idx1 == -1 or idx2 == -1):
        return "NA","NA"
    else:
        str_wo_c = cell[idx1+1:idx2]
        str_w_c = cell[idx1:idx2+1]
        return str_wo_c,str_w_c

def get_params():
    entry_enable(text)
    text_delete()
    progress_bar(10)
    mynotif("Getting params...")
    try:
        input_params ={
            "excel_file": excel_i.get(),
            "ch_sheet": sheet_i.get(),
            "ch_cell_start": x1y1_i.get(),
            "ch_cell_end": x2y2_i.get(),

            "ch_cnt": ch_combo.get(),
            "data_bit":bit_num_i.get(),
            "ch_seq":ch_seq_combo.get(),
            'DieL_name': die_L_list.get(),
            'DieR_name': die_R_list.get()
        }

        output_params ={
            "tb_sheet": out_tb_sheet.get(),
            "tb_loc": out_col_i.get()
        }

        mapping_tb_out = {
        'sheet_name': map_tb_sheet.get(),
        # 'tb_ch2ch_name':"DWORD Mapping",
        'tb_ch2ch_loc':map_loc_i.get(),
        
        # 'tb_d2d_name': "DIE to DIE Mapping",
        
    }
        global tmp_flag, temp_file
        if(tmp_flag == 0):
            with open(temp_file,'w') as params_saved:
                params_saved.writelines(input_params['excel_file'] +"\n")
                params_saved.writelines(input_params['ch_sheet'] +"\n")
                params_saved.writelines(input_params['ch_cnt'] +"\n")
                params_saved.writelines(input_params['ch_seq'] +"\n")
                params_saved.writelines(input_params['data_bit'] +"\n")
                params_saved.writelines(input_params['ch_cell_start'] +"\n")
                params_saved.writelines(input_params['ch_cell_end'] +"\n")
                params_saved.writelines(output_params['tb_sheet'] +"\n")
                params_saved.writelines(output_params['tb_loc'] +"\n")
                params_saved.writelines(input_params['DieL_name'] + "\n")
                params_saved.writelines(input_params['DieR_name'] + "\n")
                params_saved.writelines(mapping_tb_out['sheet_name'] + "\n")
                params_saved.writelines(mapping_tb_out['tb_ch2ch_loc'] + "\n")
                params_saved.writelines(theme_combo.get() + "\n")

        progress_bar(20)
        gen_datachanel(input_params,output_params, mapping_tb_out)
    except:
        messagebox.showerror("Error", "Some things wrong. Please re-check")
        progress_bar(0)
        mynotif("Error")
def Right2left(params):
    ch_cnt = params['ch_cnt']   
    col_begin = params['col_begin']
    col_end = params['col_end']
    row_begin = params['row_begin']
    row_end = params['row_end']
    wsi_f = params['wsi_f']
    wso_f = params['wso_f']
    vdd_bg = params['vdd_bg']
    vccio_bg = params['vccio_bg']
    vss_bg = params['vss_bg']
    tx_bg = params['tx_bg']
    rx_bg = params['rx_bg']
    bit_cnt = params['bit_cnt']
    c = params['out_col']
    r = params['out_row']
    ch_begin = params['ch_begin']
    ch_end = params['ch_end']
    while(ch_end>=ch_begin):   

        for col in range(col_begin, col_end + 1):
            for row in range(row_begin, row_end + 1):       
                col_l = get_column_letter(col)
                cell_val = wsi_f[col_l + str(row)].value
                
                
                        
                if (cell_val != None):
                    wso_f[get_column_letter(c)+str(r)].alignment = Alignment(shrinkToFit=True, horizontal='center')
                    wso_f[get_column_letter(c)+str(r)].border = Border(left=Side(style='thin'),right=Side(style='thin'),top=Side(style='thin'),bottom=Side(style='thin'))
                    if(str(cell_val).find("BP_") != -1):
                        index =  getstring(str(cell_val),"[","]")
                        if(index[0] != "NA"):
                        
                            wso_f[get_column_letter(c)+str(r)].value = str(cell_val).replace(index[1],'') + "[" + str(ch_end * bit_cnt + int(index[0])) + "]"
                            
                        else:
                            wso_f[get_column_letter(c)+str(r)].value = str(cell_val) + "[" + str(ch_end) + "]"
                    else:
                            wso_f[get_column_letter(c)+str(r)].value = cell_val
                            
                    if(cell_val == "VSS"):
                        wso_f[get_column_letter(c)+str(r)].fill = vss_bg

                    elif(cell_val == "VDD"):
                        wso_f[get_column_letter(c)+str(r)].fill = vdd_bg
                    elif(cell_val == "VCCIO"):
                        wso_f[get_column_letter(c)+str(r)].fill = vccio_bg
                    elif(str(cell_val).find("BP_TX") != -1):
                        wso_f[get_column_letter(c)+str(r)].fill = tx_bg
                    elif(str(cell_val).find("BP_RX") != -1):
                        wso_f[get_column_letter(c)+str(r)].fill = rx_bg
                r += 1
                print("Processing at: "+col_l + str(row) )
                
                mynotif("Processing at: "+col_l + str(row))
            c += 1
            r = params['out_row']
        ch_end -= 1
    return c
def Left2Right(params):
    ch_cnt = params['ch_cnt']   
    col_begin = params['col_begin']
    col_end = params['col_end']
    row_begin = params['row_begin']
    row_end = params['row_end']
    wsi_f = params['wsi_f']
    wso_f = params['wso_f']
    vdd_bg = params['vdd_bg']
    vccio_bg = params['vccio_bg']
    vss_bg = params['vss_bg']
    tx_bg = params['tx_bg']
    rx_bg = params['rx_bg']
    bit_cnt = params['bit_cnt']
    c = params['out_col']
    r = params['out_row']
    ch_begin = params['ch_begin']
    ch_end = params['ch_end']
    for cnt in range(ch_begin,ch_end + 1):  
        
            for col in range(col_begin, col_end + 1):
                for row in range(row_begin, row_end + 1):       
                    col_l = get_column_letter(col)
                    cell_val = wsi_f[col_l + str(row)].value
    
                    if (cell_val != None):
                        wso_f[get_column_letter(c)+str(r)].alignment = Alignment(shrinkToFit=True, horizontal='center')
                        wso_f[get_column_letter(c)+str(r)].border = Border(left=Side(style='thin'),right=Side(style='thin'),top=Side(style='thin'),bottom=Side(style='thin'))
                        if(str(cell_val).find("BP_") != -1):
                            index =  getstring(str(cell_val),"[","]")
                            if(index[0] != "NA"):
                            
                                wso_f[get_column_letter(c)+str(r)].value = str(cell_val).replace(index[1],'') + "[" + str(cnt * bit_cnt + int(index[0])) + "]"
                            else:
                                wso_f[get_column_letter(c)+str(r)].value = str(cell_val) + "[" + str(cnt) + "]"
                        else:
                                wso_f[get_column_letter(c)+str(r)].value = cell_val
                        if(cell_val == "VSS"):
                            wso_f[get_column_letter(c)+str(r)].fill = vss_bg
                        elif(cell_val == "VDD"):
                            wso_f[get_column_letter(c)+str(r)].fill = vdd_bg
                        elif(cell_val == "VCCIO"):
                            wso_f[get_column_letter(c)+str(r)].fill = vccio_bg
                        elif(str(cell_val).find("BP_TX") != -1):
                            wso_f[get_column_letter(c)+str(r)].fill = tx_bg
                        elif(str(cell_val).find("BP_RX") != -1):
                            wso_f[get_column_letter(c)+str(r)].fill = rx_bg
                    r += 1
                    print("Processing at: "+col_l + str(row) )
                    mynotif("Processing at: "+col_l + str(row))
                c += 1
                r = params['out_row']
    return c
# for cnt in range(0,ch_cnt):
def gen_datachanel(input_params, output_params, mapping_tb_out):

    excel_file = input_params["excel_file"]
    # wb_d = load_workbook(excel_file, data_only=True)
    print("Opening excel file...")
    mynotif("")
    mynotif("Opening excel file...")
    try:
        wb_f = load_workbook(excel_file,data_only=True)
    except:
        messagebox.showerror("Error", "The PLOC file is openning or not exist. Please close/check it :(")
        progress_bar(0)
        mynotif("Error")
        return
    progress_bar(50)
    print("Generating data channel..")
    mynotif("")
    mynotif("Generating data channel..")
    
    wsi_name = input_params['ch_sheet']
    wso_name = output_params['tb_sheet']
    try:
        sheet_list = wb_f.sheetnames

        if wsi_name in sheet_list:
            wsi_f = wb_f[wsi_name]
        else:
            msg_ws = messagebox.showerror('Create Sheet', 'The sheet:' + wsi_name + ' doesn\'t exist.')
            
            mynotif("The " + wsi_name + " doesn't exist.")
        
            mynotif("Error!!!")
            progress_bar(0)
            return
        
        if wso_name in sheet_list:
            
            wso_f = wb_f[wso_name]
        else:
          
            mynotif("The " + wso_name + " doesn't exist.")
            msg_ws = messagebox.askquestion('Create Sheet', 'The sheet' + wso_name + ' doesn\'t exist. Do you want to create it?', icon='question')
        
            if(msg_ws == 'yes'):
                wso_f = wb_f.create_sheet(wso_name)
                mynotif("Creating sheet...")
                mynotif('Creating the sheet...')
            else:
                mynotif("Aborted!!!")
                progress_bar(0)
                return

       
    except:
        print("Sheet " + wsi_name + " doesn't exist")
        show_error("Sheet " + wsi_name + " doesn't exist")
        progress_bar(0)
        mynotif("Error")
        root.update_idletasks()
        return

    
    row_begin = coordinate_to_tuple(input_params['ch_cell_start'])[0]
    col_begin = coordinate_to_tuple(input_params['ch_cell_start'])[1]
    row_end = coordinate_to_tuple(input_params['ch_cell_end'])[0]
    col_end = coordinate_to_tuple(input_params['ch_cell_end'])[1]
    if(row_begin > row_end or col_begin > col_end):
            print("Data channel params input is wrong. Please re-check")
            mynotif("")
            mynotif("Data channel params input is wrong. Please re-check")
            return

    out_col_begin = coordinate_to_tuple(output_params['tb_loc'])[1]
    out_row_begin = coordinate_to_tuple(output_params['tb_loc'])[0]

    c = out_col_begin
    r = out_row_begin
    ch_cnt = int(input_params["ch_cnt"]) - 1
    bit_cnt = int(input_params["data_bit"])
    ch_seq = input_params["ch_seq"]

    vdd_bg = PatternFill(patternType='solid', fgColor='9e42f5')
    vccio_bg = PatternFill(patternType='solid',fgColor='fa5565') 
    vss_bg = PatternFill(patternType='solid',fgColor='32a83a')
    tx_bg = PatternFill(patternType='solid',fgColor='f5f373')
    rx_bg = PatternFill(patternType='solid',fgColor='0e7bf0')

    params = {
        'ch_cnt': ch_cnt,
        'col_begin': col_begin,
        'col_end': col_end,
        'row_begin': row_begin,
        'row_end': row_end,
        'wsi_f': wsi_f,
        'wso_f': wso_f,
        'vdd_bg': vdd_bg,
        'vccio_bg': vccio_bg,
        'vss_bg': vss_bg,
        'tx_bg': tx_bg,
        'rx_bg': rx_bg,
        'bit_cnt': bit_cnt,
        'out_col': c,
        'out_row': r,
        'ch_begin': 0,
        'ch_end': ch_cnt
    }
    ch = 0
    if(ch_seq == "Right to Left"):
        if((ch_cnt + 1)%2 != 0):
            msg = messagebox.askquestion('Number channels choose', 'The number chanels is not even. Do you want to continue?', icon='question')
            if(msg == 'yes'):
                params['ch_begin'] = 0
                params['ch_end'] = ch_cnt
                Right2left(params) 
            else:
                mynotif("Aborted!!")
                print("Aborted!!")
                progress_bar(0)
                return
        else:
            params['ch_begin'] = 0
            params['ch_end'] = ch_cnt
            Right2left(params)
    elif(ch_seq == "Left to Right"):
        if((ch_cnt + 1)%2 != 0):
            msg = messagebox.askquestion('Number channels choose', 'The number chanels is not even. Do you want to continue?', icon='question')
            if(msg == 'yes'):
                params['ch_begin'] = 0
                params['ch_end'] = ch_cnt
                Left2Right(params)
            else:
                mynotif("Aborted!!")
                print("Aborted!!")
                progress_bar(0)
                return
        else:
            params['ch_begin'] = 0
            params['ch_end'] = ch_cnt
            Left2Right(params)
    elif(ch_seq == "Center to Left first"):
        if((ch_cnt + 1)%2 != 0):
            msg = messagebox.askquestion('Number channels choose', 'The number chanel is not even. Do you want to continue?', icon='question')
            if(msg == 'yes'):
                msg2 = messagebox.askquestion('Number channels', '\"Yes\" means Number of Left Channels is more than Right Channels\n \"No\" means Number of Right Channels is more than Left Channels', icon='question')
                if (msg2 == 'yes'):
                   center_nu=int((ch_cnt + 1)/2)
                   params['ch_begin'] = 0
                   params['ch_end'] = center_nu
                   current_col = Right2left(params)
                   params['out_col'] = current_col
                   params['ch_begin'] = center_nu + 1
                   params['ch_end'] = ch_cnt
                   Left2Right(params)
                else:
                   center_nu=int((ch_cnt + 1)/2 -1)
                   params['ch_begin'] = 0
                   params['ch_end'] = center_nu
                   current_col = Right2left(params)
                   params['out_col'] = current_col
                   params['ch_begin'] = center_nu + 1
                   params['ch_end'] = ch_cnt
                   Left2Right(params)
    
            else:
                mynotif("Aborted!!")
                print("Aborted!!")
                progress_bar(0)
                return
        else:
            center_nu=int((ch_cnt + 1)/2 - 1)
            params['ch_begin'] = 0
            params['ch_end'] = center_nu
            current_col = Right2left(params)
            params['out_col'] = current_col
            params['ch_begin'] = center_nu + 1
            params['ch_end'] = ch_cnt
            Left2Right(params)
            
    elif(ch_seq == "Center to Right first"):
        if((ch_cnt + 1)%2 != 0):
            msg = messagebox.askquestion('Number channels choose', 'The number chanel is not even. Do you want to continue', icon='question')
            if(msg == 'yes'):
                msg2 = messagebox.askquestion('Number channels', '\"Yes\" means Number of Left Channels is more than Right Channels\n \"No\" means Number of Right Channels is more than Left Channels', icon='question')
                if (msg2 == 'yes'):
                   center_nu=int((ch_cnt + 1)/2)
                   params['ch_begin'] = 0
                   params['ch_end'] = center_nu
                   current_col = Left2Right(params)

                   params['out_col'] = current_col
                   params['ch_begin'] = center_nu + 1
                   params['ch_end'] = ch_cnt
                   Right2left(params) 
                else:
                   center_nu=int((ch_cnt + 1)/2 -1)
                   params['ch_begin'] = 0
                   params['ch_end'] = center_nu
                   current_col = Left2Right(params)
                   params['out_col'] = current_col
                   params['ch_begin'] = center_nu + 1
                   params['ch_end'] = ch_cnt
                   
                   Right2left(params)
    
            else:
                mynotif("Abotted!!")
                print("Abotted!!")
                progress_bar(0)
                return
        else:
            center_nu=int((ch_cnt + 1)/2 - 1)
            params['ch_begin'] = 0
            params['ch_end'] = center_nu
            current_col = Left2Right(params)
            params['out_col'] = current_col
            params['ch_begin'] = center_nu + 1
            params['ch_end'] = ch_cnt
            Right2left(params)
    if(ismapgen.get() == 1):
        mapping_input = {
            'ch_sheet_name': input_params['ch_sheet'],
            'ch_begin_cell': input_params['ch_cell_start'],
            'ch_end_cell': input_params['ch_cell_end'],
            'ch_num': input_params['ch_cnt'],
            'bit_num': input_params['data_bit'],
            'DieL_name': input_params['DieL_name'],
            'DieR_name': input_params['DieR_name']
        }
        mapping_output = {
            'sheet_name': mapping_tb_out['sheet_name'],
            # 'tb_ch2ch_name': mapping_tb_out['tb_ch2ch_name'],
            'tb_ch2ch_loc': mapping_tb_out['tb_ch2ch_loc'],
            
        }
        wb_f = mapping_connections(wb_f, mapping_input, mapping_output)
        mynotif("Generating Die to Die Mapping...")
    print("Saving excel file...")
    mynotif("")
    mynotif("Saving excel file...")
    progress_bar(80)
    wb_f.save(excel_file)
    progress_bar(100)
    mynotif("")
    mynotif("Successed!!")
    messagebox.showinfo("Notification", "Data channel has been generated successful!!!")
    entry_disable(text)	

browse_btn = ttk.Button(root, text="Open File", image=open_imag, command=browse_file)
browse_btn_w = my_canvas.create_window(720, 40, anchor="nw", window=browse_btn)
# button = tk.Button(root, text="Generate",font=("System", 14, 'underline', 'bold'), foreground='white', background='#9b34eb', command=get_path, width=40)
button = tk.Button(root, text="Generate",font = mediumFont, foreground='white', background='#9b34eb', command=get_params, width=40)
# button = ttk.Button(root, text="Generate", command=get_path, width=80)

button_w = my_canvas.create_window(170, 650, anchor="nw", window=button)

text = tk.Text(my_canvas,width = 50, height = 100,bd=5,relief='groove', wrap='word', font=('arial',10), highlightthickness=2 ) #yscrollcommand=scroll_y.set
scroll_y = ttk.Scrollbar(text)
my_canvas.create_window(150,80, anchor='nw', window=text, height=100, width=500)

text.config(yscrollcommand=scroll_y.set)
scroll_y.pack(side=RIGHT, fill=Y)
scroll_y.config(command=text.yview)
entry_list = [ch_combo,ch_seq_combo,sheet_i,bit_num_i,x1y1_i,x2y2_i,out_tb_sheet,out_col_i,die_L_list,die_R_list,map_tb_sheet,map_loc_i]
def get_saved_params():
    global temp_file
    try:
        with open(temp_file,'r') as params_saved:
            line1 = [line.rstrip() for line in params_saved]
            params = {
                'excel_file': line1[0],
                'sheet': line1[1],
                'ch_combo': line1[2],
                'ch_seq_combo': line1[3],
                'bit_num_i': line1[4],
                'x1y1_i': line1[5],
                'x2y2_i': line1[6],
                'out_tb_sheet': line1[7],
                'out_col_i': line1[8],
                'DieL_name': line1[9],
                'DieR_name': line1[10],
                'map_sheet_name': line1[11],
                'map_tb_ch2ch_loc': line1[12],
                'theme': line1[13]

            }
        
        root.set_theme(params['theme'])
        theme_combo.current(theme_list.index(params['theme']))
        change_colour(theme_list.index(params['theme']))
        excel_i.insert(0, params['excel_file'])   

        sheet_i.insert(0, params['sheet'])
        ch_combo.current(ch_number.index(int(params['ch_combo'])))
        ch_seq_combo.current(ch_sequence.index(params['ch_seq_combo']))
        bit_num_i.insert(0, params['bit_num_i'])
        x1y1_i.insert(0, params['x1y1_i'])
        x2y2_i.insert(0, params['x2y2_i'])
        out_tb_sheet.insert(0, params['out_tb_sheet'])
        out_col_i.insert(0, params['out_col_i'])
        die_L_list.insert(0,params['DieL_name'])
        die_R_list.insert(0, params['DieR_name'])
        map_tb_sheet.insert(0, params['map_sheet_name'])
        map_loc_i.insert(0, params['map_tb_ch2ch_loc'])


    except:
        root.set_theme("scidpurple")
        theme_combo.current(theme_list.index("scidpurple"))
        change_colour(theme_list.index("scidpurple"))
        excel_i.insert(0, r"C:\Users\sytung\OneDrive - Synopsys, Inc\Desktop\py\Test3.xlsx")
        sheet_i.insert(0, "DWORD")
        ch_combo.current(3)
        ch_seq_combo.current(0)
        bit_num_i.insert(0, "16")
        x1y1_i.insert(0, "S16")
        x2y2_i.insert(0, "AD30")
        out_tb_sheet.insert(0, "Data_Channelx")
        out_col_i.insert(0, "D10")
    myguide(text, "INFO: This field is for showing information or guidence")

get_saved_params()

for ent in entry_list:
    ent.configure(justify='center')
entry_disable(die_R_list, die_L_list, map_tb_sheet, map_loc_i)
root.mainloop()


    


